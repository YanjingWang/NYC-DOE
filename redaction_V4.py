"""
Reports 1-4 = Initials subtotal calculation:
E + F = G, H + I = J, G + J = K, D + K + L + M = C
Reports 5-7 = Reevaluations subtotal calculation:
E + F = G, H + I = J, G + J = K, D + K + L  = C
Report 8 = Registers subtotal calculation:
C + D + E + F = G, H + I + J + K = L, G + L = M
SWDs by School : C4+ ...+ C1602 = C1603
Report 8a = Disability class subtotal calculation:
C + D +E + F + G + H + I + J + K + L + M + N + O = P 
Report 9 = Placement: C6+...+C37 = C38, D6+...+D37 = D38, C42+...C46 = C47, D42+...D46 = D47, C51+C52 = C53, D51+D52 = D53, C57+C58 = C59, D57+D58 = D59, C63+C64 = C65, 
D63+D64 = D65, C69+C72 = C73, D69+D72 = D73, C77+C89=D90, D77+D89=D90
Report 10 = LRE-MRE: C6+...C37=C38,D6+...D37=D38,E6+...E37=E38,F6+...F37=F38, C42+...C46=C47,D42+...D46=D47,E42+...E46=E47,F42+...F46=F47,C51+C52=C53,D51+D52=D53,E51+E52=E53,F51+F52=F53,E57+E58=E59,F57+F58=F59,C63+C64=C65,D63+D64=D65,E63+E64=E65,F63+F64=F65,C69+C72=C73,D69+D72=D73,E69+E72=E73,F69+F72=F73,C77+...+C89=C90,D77+...+D89=D90,E77+...+E89=E90,F77+...+F89=F90
"""
import openpyxl
class Solution:
    def mask_value_initial(self,val):
        if isinstance(val, (int, float)) and 0 < val <= 5:
            return '<=5'
        return val

    def mask_value_secondary(self,val):
        if isinstance(val, (int, float)) and 0 < val <= 5:
            return '<=5'
        elif isinstance(val, (int, float)) and val > 5:
            return '>5'
        return val

    def initial_mask(self,ws, start_row, start_col, end_row, end_col):
        # Step 1: Initial Masking
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                cell.value = self.mask_value_initial(cell.value)
        
        # Mask smallest if only one '<=5' in column
        for col in ws.iter_cols(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            valid_vals = [cell.value for cell in col if isinstance(cell.value, (int, float)) and cell.value != '<=5']
            if sum([cell.value == '<=5' for cell in col]) == 1 and valid_vals:
                min_val = min(valid_vals)
                for cell in col:
                    if cell.value == min_val:
                        cell.value = '>5'

    # Step 2: Secondary Masking
    def secondary_mask(self,ws, start_row, end_row):
        groups = [(5, 6, 7), (8, 9, 10), (7, 10, 11)]  # E,F,G and H,I,J and G,J,K respectively

        for group in groups:
            for row_num in range(start_row, end_row + 1):
                masked_vals = [col for col in group if ws.cell(row=row_num, column=col).value in ['<=5', '>5']]
                if len(masked_vals) == 1:  # If only one value in the group is masked, need to mask the smallest of the unmasked values
                    numeric_vals = [ws.cell(row=row_num, column=col).value for col in group if isinstance(ws.cell(row=row_num, column=col).value, (int, float))]
                    if numeric_vals:  # Ensure there are numeric values
                        min_val = min(numeric_vals)
                        for col in group:
                            if ws.cell(row=row_num, column=col).value == min_val:
                                ws.cell(row=row_num, column=col).value = self.mask_value_secondary(min_val)
                                break
                elif len(masked_vals) >= 2:  # Two or more values are already masked, so no action needed
                    continue


    # Step 3: Masking based on Subtotals        
    def third_mask(self,ws, start_row, end_row):
        groups = [(3,4,11,12)]  # C,D,K,L
        for group in groups:
            for row_num in range(start_row, end_row + 1):
                masked_vals = [col for col in group if ws.cell(row=row_num, column=col).value in ['<=5', '>5']]
                if len(masked_vals) == 1:  # If only one value in the group is masked, need to mask the smallest of the unmasked values
                    numeric_vals = [ws.cell(row=row_num, column=col).value for col in group if isinstance(ws.cell(row=row_num, column=col).value, (int, float))]
                    if numeric_vals:  # Ensure there are numeric values
                        min_val = min(numeric_vals)
                        for col in group:
                            if ws.cell(row=row_num, column=col).value == min_val:
                                ws.cell(row=row_num, column=col).value = self.mask_value_secondary(min_val)
                                break
                elif len(masked_vals) >= 2:  # Two or more values are already masked, so no action needed
                    continue


    def mask_excel_file(self,filename):
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        ws = wb['Reports 5-7 = Reevaluations']

        # Mask data for the specific ranges
        ranges = [(5, 3, 36, 12), (41, 3, 45, 12), (50, 3, 51, 12), (56, 3, 63, 12), (62, 3, 73, 12), (68, 3, 71, 12), (76, 3, 88, 12)]

        for r in ranges:
            self.initial_mask(ws, *r)
            
        # Invoke mask_DKLMC
        self.secondary_mask(ws, 5, 88) 
        self.third_mask(ws, 5, 88)  # Adjust the range if necessary
        # Save the modified workbook
        wb.save(filename)


# Call the function with your filename
if __name__ == "__main__":
    Tab1 = Solution()
    Tab1.mask_excel_file('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Annual Special Education Data Report Unredacted SY21.xlsx')
    Tab1.mask_excel_file('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted Annual Special Education Data Report SY22.xlsx')
    Tab1.mask_excel_file('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted Annual Special Education Data Report SY23.xlsx')




