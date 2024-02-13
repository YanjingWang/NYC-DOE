import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted City Council Triennial Report_CW.xlsx')
        ws = wb.create_sheet("Related Service Delivery")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Related Services Recommendation Type',
                'Fully Receiving',
                'Percent Fully Receiving',
                'Partially Receiving',
                'Percent Partially Receiving',
                'Not Receiving',
                'Percent Not Receiving']
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "D9E1F2"
        column_fill_color = "D9E1F2"
        self.format_header(ws, 'A2', "Related Services Recommendation Type", columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        # self.format_header(ws, 'A10', "Primary Program Type", columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all, 

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        # params = ('CC_PSStudentR12_061523')
        cursor.execute("[Mike].[USPCCTriannualReportRSCitywide]")
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_tab5(self,cursor):
        query_bytab5 = '''
        Select *
        from ##RSCitywide
        Order By MandatesBilingual
        '''  # the bytab12 SQL query goes here
        cursor.execute(query_bytab5)
        results_bytab5 = cursor.fetchall()
        return results_bytab5
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_only_bottom = Border(bottom=black_border_thickside)
        black_boarder_right_and_bottom = Border(right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 1)  # +2 because data starts from column 'A'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['A3':'A11']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                    cell.font = Font(bold=True, size=12)
                cell.alignment = openpyxl.styles.Alignment(horizontal='left')

        for row in ws['B3':'G11']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['A1': 'G1']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['A11':'G11']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        cell_ranges = ['B3:B11', 'D3:D11', 'F3:F11']
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        try:
                            cell.value = int(cell.value)
                            cell.number_format = '#,##0'  # Apply comma format
                            print("Int converting")
                        except ValueError:
                            # If the value cannot be converted to int, keep the original value
                            print("Int converting Error")
                            pass
        cell_ranges = ['C3:C11', 'E3:E11','G3:G11']
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '0%'  # Apply percentage format
                            print("Float converting")
                        except ValueError:
                            # If the value cannot be converted to int, keep the original value
                            print("Float converting Error")
                            pass
                        
        # merge cell A12:G12 and adjust cell height to 50
        # ws.merge_cells('A12:G12')
        ws.row_dimensions[12].height = 200
        ws['A12'].alignment = openpyxl.styles.Alignment(wrap_text=True)
        ws['A12'].border = black_boarder_only_bottom
        ws['A12'].font = Font(bold=True, size=12)
        # insert text into A12 and aligh text to middle align
        ws['A12'] = "Local Law 900-A requires reporting on provision of assistive technology service. Assistive technology service is often provided through speech-language therapy or occupational therapy IEP recommendations.  This data is not included in the tables because it would be subject to redaction. For SY 2023–24, 12,791 students in DOE schools were provided assistive technology devices, per IEP recommendations."
        ws['A12'].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
        # cell B12, C12, D12, E12, F12, G12 ALL have border on the bottom
        ws['B12'].border = black_boarder_only_bottom
        ws['C12'].border = black_boarder_only_bottom
        ws['D12'].border = black_boarder_only_bottom
        ws['E12'].border = black_boarder_only_bottom
        ws['F12'].border = black_boarder_only_bottom
        ws['G12'].border = black_boarder_right_and_bottom
        

    def Report_Related_Service_Delivery(self):
        title_cells = [
            # {"cell": "B1", "value": "Report 12 Delivery of Special Education Programs", "merge_cells": "B1:H1"},
        ]

        subtitle_cells = [
            {"cell": "A1", "value": "October 31, 2023 Number & Percentage of Related Service Recommendations with Encounter Recorded by Service Type", "merge_cells": "A1:G1"},
            # {"cell": "A12", "value": "Local Law 900-A requires reporting on provision of assistive technology service. Assistive technology service is often provided through speech-language therapy or occupational therapy IEP recommendations. On 10/31/22 there were <=5 IEP recommendations specifically for assistive technology service all of which had been encountered. This data is not included in the tables because it would be subject to redaction. For SY 2022–23, 9,772 students in DOE schools were provided assistive technology devices, per IEP recommendations.", "merge_cells": "A12:G12"}            
        ]

        column_widths = [50, 25, 25, 25, 25, 25, 25]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_bytab5= self.fetch_data_by_tab5(cursor)
        # sort the results_bytab5 by the first column 1: Counseling Services, 2: Counseling Services Bilingual, 3: Hearing Education Services, 4: Occupational Therapy, 5: Physical Therapy 6: Speech-Language Therapy, 7:Speech-Language Therapy Bilingual, 8: Vision Education Services, 9:Total
        sort_order= { 'Counseling Services': 1, 'Counseling Services Bilingual': 2, 'Hearing Education Services': 3, 'Occupational Therapy': 4, 'Physical Therapy': 5, 'Speech-Language Therapy': 6, 'Speech-Language Therapy Bilingual': 7, 'Vision Education Services': 8, 'Total': 9}
        results_bytab5.sort(key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, results_bytab5, start_row=3)

        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted City Council Triennial Report_CW.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab5 = Solution()
        Tab5.Report_Related_Service_Delivery()                                                                  