import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def __init__(self):
        self.schoolyear = 'SY 2024-25'
        self.sqlsnapshottableschoolyear = '24'
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Report 11 = 3Yr Reevaluations"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Report 13 = 3Yr Reevaluations")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Total Three-Year Reevaluations Completed',
                'Three-Year Reevaluations Completed – Timely',
                'Three-Year Reevaluations Completed – Not Timely']
        column_letters = ['C', 'D', 'E']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "B8CCE4"
        self.format_header(ws, 'B5', 'District', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B41', 'Race/Ethnicity', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B50', 'Meal Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B56', 'Gender', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B63', 'ELL Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B69', 'Language of Instruction', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B77', 'Grade Level', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B95', 'Temporary Housing Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B102', 'Foster Care Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_ThreeYearReevalsR11_SY'+self.sqlsnapshottableschoolyear)
        cursor.execute("EXEC [dbo].[USPCC_AnnaulReport11] @tableNameCCThreeYearReevalsR11=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_race(self,cursor):
        query_byRace = '''
        select EthnicityGroupCC, c1,c2,c3 from (  select * from  (  Select   EthnicityGroupCC_sort as sort , EthnicityGroupCC 	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11   group by EthnicityGroupCC, EthnicityGroupCC_sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort  
        '''  # the byRace SQL query goes here
        cursor.execute(query_byRace)
        results_byRace = cursor.fetchall()
        return results_byRace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,cursor):
        query_byDistrict = '''
        select * from  (  Select    ReportingDistrict as sort  	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11   group by ReportingDistrict  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byDistrict SQL query goes here
        cursor.execute(query_byDistrict)
        results_byDistrict = cursor.fetchall()
        return results_byDistrict

    def fetch_data_by_mealstatus(self,cursor):
        query_byMealStatus = '''
        select * from  (  Select    MealStatusGrouping as sort  	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11   group by MealStatusGrouping  ) a  union all  select * from ##TotalRow  order by sort  
        '''  # the byMealStatus SQL query goes here
        cursor.execute(query_byMealStatus)
        results_byMealStatus = cursor.fetchall()
        return results_byMealStatus
    
    def fetch_data_by_gender(self,cursor):
        query_byGender = '''
        select * from  (  Select    GENDER as sort  	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11   group by GENDER  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byGender SQL query goes here
        cursor.execute(query_byGender)
        results_byGender = cursor.fetchall()
        return results_byGender
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byELLStatus = '''
        select * from  (  Select    ELLStatus as sort  	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11   group by ELLStatus  ) a  union all  select * from ##TotalRow  order by sort
        '''  # the byELLStatus SQL query goes here
        cursor.execute(query_byELLStatus)
        results_byELLStatus = cursor.fetchall()
        return results_byELLStatus
    
    def fetch_data_by_language(self,cursor):
        query_byLanguage = '''
        select OutcomeLanguageCC, c1,c2,c3 from (  select * from  (  Select   LanguageOfInstruction_Sort as sort , OutcomeLanguageCC 	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11   group by OutcomeLanguageCC, LanguageOfInstruction_Sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort  
        '''  # the byLanguage SQL query goes here
        cursor.execute(query_byLanguage)
        results_byLanguage = cursor.fetchall()
        return results_byLanguage
    
    def fetch_data_by_gradelevel(self,cursor):
        query_byGradeLevel = '''
        select GradeLevel, c1,c2,c3 from (  select * from  (  Select   GradeLevel_Sort as sort , GradeLevel 	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11   group by GradeLevel, GradeLevel_Sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''
        cursor.execute(query_byGradeLevel)
        results_byGradeLevel = cursor.fetchall()
        return results_byGradeLevel
    
    def fetch_data_by_tempResFlag(self,cursor):
        query_byTempResFlag = '''
        select STHFlag, c1,c2,c3 from (  select * from  (  Select   STHFlagSort as sort ,  case when STHFlag = 'Y' then 'Yes' else 'No'  end as STHFlag 	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11  where STHFlag in ('Y', 'N')  group by STHFlag, STHFlagSort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''
        cursor.execute(query_byTempResFlag)
        results_byTempResFlag = cursor.fetchall()
        return results_byTempResFlag
    
    def fetch_data_by_fosterCareStatus(self,cursor):
        query_byFosterCareStatus = '''
        select FosterCareFlag, c1,c2,c3 from (  select * from  (  Select   FosterCareFlagSort as sort , FosterCareFlag 	 ,FORMAT(Count(ComplianceStatus), '#,##0') as c1         ,FORMAT(sum(Conduct_on_time), '#,##0') as c2        ,FORMAT(sum(Not_conducT_on_t), '#,##0') as c3  FROM ##Report_Final11   group by FosterCareFlag, FosterCareFlagSort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''
        cursor.execute(query_byFosterCareStatus)
        results_byFosterCareStatus = cursor.fetchall()
        return results_byFosterCareStatus
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C5':'E38'] + ws['C42':'E47'] + ws['C51':'E53'] + ws['C57':'E60'] + ws['C64':'E66'] + ws['C70':'E74'] + ws['C78':'E91'] + ws['C96':'E98'] + ws['C103':'E105']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                if isinstance(cell.value, str):
                    try:
                        cell.value = int(cell.value)
                    except ValueError:
                        # If the value cannot be converted to int, keep the original value
                        pass
        # Function to check if a string represents a valid number
        def is_number(s):
            try:
                float(s.replace(',', ''))  # Try converting after removing commas
                return True
            except ValueError:
                return False

        # Formatting specific cell ranges
        cell_ranges = ['C5:E38', 'C42:E47', 'C51:E53', 'C57:E60', 'C64:E66', 'C70:E74', 'C78:E91', 'C96:E98', 'C103:E105']
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if isinstance(cell.value, str) and is_number(cell.value):
                        # Convert to float after removing commas
                        cell.value = float(cell.value.replace(',', ''))
                        # Apply number format with commas (optional)
                        cell.number_format = '#,##0'
        # for row in ws['C42':'E47']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C51':'E53']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C57':'E60']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C64':'E66']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C70':'E74']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C78':'E91']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C96':'E98']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C103':'E105']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')
        for row in ws['B1': 'E1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B38':'E38'] + ws['B47':'E47'] + ws['B53':'E53'] + ws['B60':'E60'] + ws['B66':'E66'] + ws['B74':'E74'] + ws['B91':'E91'] + ws['B98':'E98'] + ws['B105':'E105']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['B4':'E4'] + ws['B40':'E40'] + ws['B49':'E49'] + ws['B55':'E55'] + ws['B62':'E62'] + ws['B68':'E68'] + ws['B76':'E76'] + ws['B94':'E94'] + ws['B101':'E101']:
            for cell in row:
                cell.border = black_border_thick
                cell.font = Font(bold=True, size=12)
        # wrap text in C5:D5
        for row in ws['C5':'D5']:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        # wrap text in C5:F5
        for row in ws['C5':'F5']:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # wrap text of B50 cell having 'Eligible for the Free/Reduced Price Lunch Program' to fit the cell
        ws['B50'].alignment = Alignment(wrap_text=True)
        # make cell B1 higher to fit the text
        ws.row_dimensions[1].height = 30           
    def Report_13_3Yr_Reevaluations(self):
        title_cells = [
            {"cell": "B1", "value": "Report 13 Three-Year Reevaluations Disaggregated by:  District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; Grade Level;  Temp House Status and Foster Care Status.", "merge_cells": "B1:E1"},
            

        ]

        subtitle_cells = [
            {"cell": "B4", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by District", "merge_cells": "B4:E4"},
            {"cell": "B40", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by Race/Ethnicity", "merge_cells": "B40:E40"},
            {"cell": "B49", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by Meal Status", "merge_cells": "B49:E49"},
            {"cell": "B55", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by Gender", "merge_cells": "B55:E55"},
            {"cell": "B62", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by English Language Learner (ELL) Status", "merge_cells": "B62:E62"},
            {"cell": "B68", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by Recommended Language of Instruction", "merge_cells": "B68:E68"},
            {"cell": "B76", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by Grade Level", "merge_cells": "B76:E76"},
            {"cell": "B94", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by Temporary Housing Status", "merge_cells": "B94:E94"},
            {"cell": "B101", "value": self.schoolyear + " Timeliness of Three-Year Reevaluations by Foster Care Status", "merge_cells": "B101:E101"},
            

        ]

        column_widths = [5, 30, 50, 50, 50]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_byRace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byRace, start_row=42)
        
        # Step 4: Fetch and write data for "Report 8b = IEP Service Recs by District"
        results_byDistrict = self.fetch_data_by_district(cursor)
        # # replace 01 as 1, 02 as 2, etc.
        # results_byDistrict = [(x[0].replace('01', '1'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('02', '2'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('03', '3'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('04', '4'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('05', '5'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('06', '6'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('07', '7'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('08', '8'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('09', '9'), *x[1:]) for x in results_byDistrict]
        self.write_data_to_excel(ws, results_byDistrict, start_row=6)

        # Step 5: Fetch and write data for "Report 8b = IEP Service Recs by Meal Status"
        results_byMealStatus = self.fetch_data_by_mealstatus(cursor)
        # replace Free or Reduced Price Meal to Eligible for the Free/Reduced Price Lunch Program
        results_byMealStatus = [(x[0].replace('Free or Reduced Price Meal', 'Eligible for the Free/Reduced Price Lunch Program'), *x[1:]) for x in results_byMealStatus]
        self.write_data_to_excel(ws, results_byMealStatus, start_row=51)

        # Step 6: Fetch and write data for "Report 8b = IEP Service Recs by Gender"
        results_byMealStatus = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_byMealStatus, start_row=57)

        # Step 7: Fetch and write data for "Report 8b = IEP Service Recs by ELL Status"
        results_byELLStatus = self.fetch_data_by_ellstatus(cursor)
        # replace 'ELL' with 'Ell' and 'NON-ELL' with 'Non-Ell'
        results_byELLStatus = [('ELL' if x[0] == 'ELL' else ('NOT ELL' if x[0] == 'Non-Ell' else x[0]), *x[1:]) for x in results_byELLStatus]
        self.write_data_to_excel(ws, results_byELLStatus, start_row=64)
        
        # Step 8: Fetch and write data for "Report 8b = IEP Service Recs by Language"
        results_byLanguage = self.fetch_data_by_language(cursor)
        # Replace 'ENGLISH' with 'English' and 'SPANISH' with 'Spanish' and 'CHINESE' with 'Chinese' and 'OTHER' with 'Other'
        results_byLanguage = [('English' if x[0] == 'ENGLISH' else ('Spanish' if x[0] == 'SPANISH' else ('Chinese' if x[0] == 'CHINESE' else ('Other' if x[0] == 'OTHER' else x[0]))), *x[1:]) for x in results_byLanguage]
        self.write_data_to_excel(ws, results_byLanguage, start_row=70)

        # Step 9: Fetch and write data for "Report 8b = IEP Service Recs by Grade Level"
        results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        # replace 0K as KG, 01 as 1, 02 as 2, etc. 
        results_byGradeLevel = [(x[0].replace('0K', 'KG'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('01', '1'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('02', '2'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('03', '3'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('04', '4'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('05', '5'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('06', '6'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('07', '7'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('08', '8'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('09', '9'), *x[1:]) for x in results_byGradeLevel]
        self.write_data_to_excel(ws, results_byGradeLevel, start_row=78)

        # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Temporary Housing"
        results_byTempResFlag = self.fetch_data_by_tempResFlag(cursor)
        self.write_data_to_excel(ws, results_byTempResFlag, start_row=96)

        # Step 11: Fetch and write data for "Report 8b = IEP Service Recs by Foster Care Status"
        results_byFosterCareStatus = self.fetch_data_by_fosterCareStatus(cursor)
        # replace 'Y' with 'Yes' and 'N' with 'No'
        results_byFosterCareStatus = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_byFosterCareStatus]
        self.write_data_to_excel(ws, results_byFosterCareStatus, start_row=103)
        # wrap text in C5:E5
        for row in ws['C5':'E5']:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.Report_13_3Yr_Reevaluations()                                                                  