import openpyxl
from openpyxl.styles import PatternFill

# Define the file paths and the configuration for cell ranges to compare
path_sy23 = 'C:/Users/Ywang36/OneDrive - NYCDOE/Desktop/Annual Special Education Data Report SY23.xlsx'
path_sy24 = 'C:/Users/Ywang36/OneDrive - NYCDOE/Desktop/Redacted Annual Special Education Data Report SY24.xlsx'
REPORTS_CONFIG_SY24 = {
    'Reports 1-4 = Initials': {
        'ranges': [(5, 2, 37, 13), (41, 2, 46, 13), (50, 2, 52, 13), (56, 2, 59, 13), (63, 2, 65, 13), (69, 2, 74, 13),(79, 2, 92, 13), (99, 2, 101, 13),(107, 2, 109, 13)],#Adjust accordingly
    },
    'Reports 5-7 = Reevaluations': {
        'ranges': [(5, 2, 37, 12), (41, 2, 46, 12), (50, 2, 52, 12), (56, 2, 59, 12), (63, 2, 65, 12), (69, 2, 73, 12), (77, 2, 90, 12),(95, 2, 97, 12),(102, 2, 104, 12)],  # Adjust accordingly
    },
    # Add more configurations for other tabs
    "Report 8 = Registers": {
        "ranges": [ (5, 2, 37, 13),(41, 2, 46, 13), (50, 2, 52, 13), (56, 2, 59, 13), (63, 2, 76, 13), (80, 2, 93, 13),(97, 2, 99, 13),(104, 2, 106, 13)],
    },
    "Report 8a = SWDs by School" : {
        "ranges": [(4, 2, 1606, 3)]
    },
    # "Report 9 = Disability class" : {
    #     "ranges": [(5,  2, 37, 16),(41, 2, 46, 16),(50, 2, 52, 16),(56, 2, 59, 16),(63, 2, 65, 16),(69, 2, 73, 16),(77, 2, 90, 16),(94, 2, 96, 16),(100, 2, 102, 16)],
    # },
    "Report 10 = IEP Service Recs": {
        "ranges": [
            (6, 2, 38, 14),    # B6 to N38
            (43, 2, 48, 14),   # B43 to N48
            (53, 2, 55, 14),   # B53 to N55
            (60, 2, 63, 14),   # B60 to N63
            (68, 2, 70, 14),   # B68 to N70
            (75, 2, 79, 14),   # B75 to N79
            (84, 2, 97, 14),   # B84 to N97
            (103, 2, 116, 14), # B103 to N116
            (122, 2, 124, 14),  # B122 to N124
            (130, 2, 132, 14)  # B130 to N132            
        ],
    },
    "Report 11 = Placement" : {
        "ranges": [(6, 2,  38, 4),(42, 2, 47, 4),(51, 2, 53, 4),(57, 2, 60, 4),(64, 2, 66, 4),(70, 2, 74, 4),(78, 2, 91, 4),(96, 2, 98, 4),(103, 2, 105, 4)]
    },
    "Report 12 = LRE-MRE" : {
        "ranges": [(6, 2, 38, 6), (42, 2, 47, 6), (51, 2, 53, 6), (57, 2, 60, 6), (64, 2, 66, 6), (70, 2, 74, 6), (78, 2, 91, 6), (96, 2, 98, 6), (103, 2, 105, 6)]
    },
    "Report 13 = 3Yr Reevaluations" : {
    "ranges": [(6, 2, 38, 5),(42, 2, 47, 5),(51, 2, 53, 5),(57, 2, 60, 5),(64, 2, 66, 5),(70, 2, 74, 5),(78, 2, 91, 5),(96, 2, 98, 5),(103, 2, 105, 5)],
    },
     "Report 14 = Programs" : {
             "ranges": [
        (5, 2, 8, 8),    # C5 to H8
        (12, 2, 17, 8),  # C12 to H17
        (21, 2, 23, 8),  # C21 to H23
        (27, 2, 30, 8),  # C27 to H30
        (34, 2, 36, 8),  # C34 to H36
        (40, 2, 44, 8),  # C40 to H44
        (48, 2, 61, 8),  # C48 to H61
        (65, 2, 67, 8),  # C65 to H67
        (71, 2, 73, 8),  # C71 to H73
    ],
    },
    # "Report 14a = Bilingual Programs" : {
    #     "ranges": [(5, 2, 8, 8)],
    # },    
    "Report 15 = Related Services" : {
    "ranges": [
        (5, 2, 13, 8),   # B5 to H13
        (17, 2, 22, 8),  # B17 to H22
        (26, 2, 28, 8),  # B26 to H28
        (32, 2, 35, 8),  # B32 to H35
        (39, 2, 41, 8),  # B39 to H41
        (45, 2, 49, 8),  # B45 to H49
        (53, 2, 66, 8),  # B53 to H66
        (70, 2, 72, 8),  # B70 to H72
        (76, 2, 78, 8),  # B76 to H78
    ],
    },
    "Report 15a = Transportation" : {
    "ranges": [
        (5, 2, 37, 8),   # B5 to H37
        (41, 2, 46, 8),  # B41 to H46
        (50, 2, 52, 8),  # B50 to H52
        (56, 2, 59, 8),  # B56 to H59
        (63, 2, 65, 8),  # B63 to H65
        (69, 2, 73, 8),  # B69 to H73
        (77, 2, 90, 8),  # B77 to H90
        (94, 2, 96, 8),  # B94 to H96
        (100, 2, 102, 8),# B100 to H102
        (106, 2, 1614, 8)# B106 to H1614
    ],
    },
    # "Report 16 = BIP" : {
    # "ranges": [
    #     (5, 2, 9, 6),    # B5 to F9
    #     (13, 2, 45, 6),  # B13 to F45
    #     (49, 2, 54, 6),  # B49 to F54
    #     (58, 2, 60, 6),  # B58 to F60
    #     (64, 2, 67, 6),  # B64 to F67
    #     (71, 2, 73, 6),  # B71 to F73
    #     (77, 2, 81, 6),  # B77 to F81
    #     (85, 2, 98, 6),  # B85 to F98
    #     (102, 2, 104, 6),# B102 to F104
    #     (108, 2, 110, 6),# B108 to F110
    #     (114, 2, 1717, 6)# B114 to F1717
    # ],
    # },
    "Report 17 = Inclusion" : {
        "ranges": [(5, 2, 8, 4)],
    }
}

# Function to compare two cells and apply color if different
def compare_cells(cell1, cell2, color):
    val1 = str(cell1.value).strip() if cell1.value is not None else ""
    val2 = str(cell2.value).strip() if cell2.value is not None else ""
    if val1 != val2:
        cell1.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell2.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Open both workbooks
wb_sy23 = openpyxl.load_workbook(path_sy23, data_only=True)
wb_sy24 = openpyxl.load_workbook(path_sy24, data_only=True)

# Loop through the configuration dictionary
for sheet_name, config in REPORTS_CONFIG_SY24.items():
    ws_sy23 = wb_sy23[sheet_name]
    ws_sy24 = wb_sy24[sheet_name]
    
    # Iterate over the specified ranges
    for cell_range in config['ranges']:
        for row in range(cell_range[0], cell_range[2] + 1):
            for col in range(cell_range[1], cell_range[3] + 1):
                cell_sy23 = ws_sy23.cell(row, col)
                cell_sy24 = ws_sy24.cell(row, col)
                
                # Compare and highlight differences
                compare_cells(cell_sy23, cell_sy24, "FFFF00")  # Yellow for SY24
                compare_cells(cell_sy24, cell_sy23, "00FF00")  # Green for SY23

# Save the updated workbooks
wb_sy23.save(path_sy23)
wb_sy24.save(path_sy24)