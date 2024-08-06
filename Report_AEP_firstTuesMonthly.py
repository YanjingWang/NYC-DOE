import pandas as pd
from sqlalchemy import create_engine, text
import urllib
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from datetime import datetime

class DMLOfficeRport1:
    def __init__(self, schoolyear='SY 2024-2025'):
        # Database connection details
        server = 'ES00VPADOSQL180,51433'
        database = 'SEO_MART'
        username = 'your_username'
        password = 'your_password'

        # Connection string
        conn_str = (
            'DRIVER=SQL Server;'
            'SERVER=' + server + ';'
            'DATABASE=' + database + ';'
            # 'UID=' + username + ';'
            # 'PWD=' + password
        )
        params = urllib.parse.quote_plus(conn_str)

        # Create engine
        self.engine = create_engine(f'mssql+pyodbc:///?odbc_connect={params}')

        self.schoolyear = schoolyear
        self.lastrow = self.get_row_count()
        self.datestamp = self.get_latest_processed_date()
        self.folderpath = fr'R:\\'

    def get_row_count(self):
        query = "SELECT COUNT(*) AS row_count FROM [SEO_MART].[dbo].[RPT_AdaptedPEAccessibleNeeds] WHERE EnrolledSchoolSetting = 'csd'"
        with self.engine.connect() as connection:
            result = connection.execute(text(query)).fetchone()._mapping
            print(f"Number of rows in the table: {result['row_count']}")
            return result['row_count'] + 1

    def get_latest_processed_date(self):
        query = "SELECT MAX(ProcessedDate) AS latest_date FROM [SEO_MART].[dbo].[RPT_AdaptedPEAccessibleNeeds] WHERE EnrolledSchoolSetting = 'csd'"
        with self.engine.connect() as connection:
            result = connection.execute(text(query)).fetchone()._mapping
            latest_date = result['latest_date'] # 2024-08-05
            print(f"Latest processed date: {latest_date}")
            formatted_date = datetime.strptime(latest_date, '%Y-%m-%d').strftime('%m%d%Y')
            return formatted_date

    def create_excel_report(self):
        # SQL query
        query = """
        SELECT  
            [EnrolledDBN]
            ,AdminDistrict as [SchoolDistrict] 
            ,[LastName]
            ,[FirstName]
            ,[StudentID]
            ,[GradeLevel]
            ,[BirthDate]
            ,[OutcomeDocumentType]
            ,[OutcomeDocumentIDT]
            ,[RecommendedProgram]
            ,[RecommendedSubject]
            ,[ProjectedIEPImplementationDate]
            ,[RecommendedStartDate]
            ,[RecommendedEndDate]
            ,[RecommendedFrquency]
            ,[RecommendedDuration]
            ,[ServiceLocation]
            ,[AccessibleProgramFlag] 
            ,[ProcessedDate]
        FROM [SEO_MART].[dbo].[RPT_AdaptedPEAccessibleNeeds]
        WHERE EnrolledSchoolSetting = 'csd'
        """

        # Establish a connection and load data into pandas DataFrame
        with self.engine.connect() as connection:
            df = pd.read_sql(text(query), connection)

        # Specify the path for the output file
        # output_file = f'R:\\SEO Analytics\\Reporting\\Adapted PE Reports\\Adapted_PE_AccessibleNeeds_{self.datestamp}.xlsx'
        output_file = f'{self.folderpath}SEO Analytics\\Reporting\\Adapted PE Reports\\Adapted_PE_AccessibleNeeds_{self.datestamp}.xlsx'

        # Save the DataFrame to an Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=f'APE Report {self.datestamp}')

        # Load the workbook and select the active sheet
        workbook = load_workbook(output_file)
        sheet = workbook[f'APE Report {self.datestamp}']

        # Set the width of columns A to S as the header columns width that fits the content
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
            max_length = 0
            for row in sheet[col]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(row.value)
                except:
                    pass
            adjusted_width = (max_length + 4)
            sheet.column_dimensions[col].width = adjusted_width

        # Set A1 to H1 to bold and center and add filter to the first row
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
            sheet[col + '1'].font = sheet[col + '1'].font.copy(bold=True)
            sheet[col + '1'].alignment = sheet[col + '1'].alignment.copy(horizontal='left')

        # Add the filter for column A to S
        sheet.auto_filter.ref = f'A1:S{self.lastrow + 1}'

        # Save the changes to the workbook
        workbook.save(output_file)

        print(f"Excel report with borders has been created: {output_file}")

if __name__ == '__main__':
    report = DMLOfficeRport1()
    report.create_excel_report()
