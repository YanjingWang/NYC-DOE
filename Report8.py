import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def __init__(self):
        self.schoolyear = 'SY 2024-25'
        self.sqlsnapshottableschoolyear = '24'
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Report 8 = Registers"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Report 8 = Registers")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=False, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Non-ELL with English Language of Instruction', 'Non-ELL with Spanish Language of Instruction',
                'Non-ELL with Chinese Language of Instruction', 'Non-ELL with Other Language of Instruction', 'Total Non-ELL', 'ELL with English Language of Instruction','ELL with Spanish Language of Instruction','ELL with Chinese Language of Instruction','ELL with Other Language of Instruction','Total ELL','Total Register']
        column_letters = ['C', 'D', 'E', 'F', 'G', 'H','I', 'J', 'K', 'L', 'M']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=False, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "E0F0F8"
        self.format_header(ws, 'B4', 'District', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B40', 'Race/Ethnicity', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B49', 'Meal Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B55', 'Gender', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B62', 'Grade Level', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B79', 'Disability Classification', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B96', 'Temporary Housing Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B103', 'Foster Care Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_StudentRegisterR814_0615'+self.sqlsnapshottableschoolyear)
        cursor.execute("EXEC [dbo].[USPCC_AnnaulReport8] @tableNameCCStudentRegisterR814=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_race(self,cursor):
        query_byRace = '''
        select EthnicityGroupCC, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 ,  TotalRegister from (  select * from  (  Select   Ethnicity_sort as sort , EthnicityGroupCC ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other) + sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other) , '#,##0') as TotalRegister  FROM ##CCTotaltemp8  group by EthnicityGroupCC, Ethnicity_sort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort 
        '''  # the byRace SQL query goes here
        cursor.execute(query_byRace)
        results_byRace = cursor.fetchall()
        return results_byRace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,cursor):
        query_byDistrict = '''
        select * from  (  Select    ReportingDistrict as sort  ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other) + sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other) , '#,##0') as TotalRegister  FROM ##CCTotaltemp8  group by ReportingDistrict  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byDistrict SQL query goes here
        cursor.execute(query_byDistrict)
        results_byDistrict = cursor.fetchall()
        return results_byDistrict

    def fetch_data_by_mealstatus(self,cursor):
        query_byMealStatus = '''
        select * from  (  Select    MealStatusGrouping as sort  ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other) + sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other) , '#,##0') as TotalRegister  FROM ##CCTotaltemp8  group by MealStatusGrouping  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byMealStatus SQL query goes here
        cursor.execute(query_byMealStatus)
        results_byMealStatus = cursor.fetchall()
        return results_byMealStatus
    
    def fetch_data_by_gender(self,cursor):
        query_byGender = '''
        select * from  (  Select    Gender as sort  ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other) + sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other) , '#,##0') as TotalRegister  FROM ##CCTotaltemp8  group by Gender  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byGender SQL query goes here
        cursor.execute(query_byGender)
        results_byGender = cursor.fetchall()
        return results_byGender
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byELLStatus = '''
        Select Gender ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  FROM ##CCTotaltemp group by Gender  order by Gender
        '''  # the byELLStatus SQL query goes here
        cursor.execute(query_byELLStatus)
        results_byELLStatus = cursor.fetchall()
        return results_byELLStatus
    
    def fetch_data_by_classification(self,cursor):
        query_byClassification = '''
        select * from  (  Select    Classification as sort  ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other) + sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other) , '#,##0') as TotalRegister  FROM ##CCTotaltemp8  group by Classification  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byLanguage SQL query goes here
        cursor.execute(query_byClassification)
        results_byClassification = cursor.fetchall()
        return results_byClassification
    
    def fetch_data_by_gradelevel(self,cursor):
        query_byGradeLevel = '''
        select GradeLevel, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 ,  TotalRegister from (  select * from  (  Select   Grade_Sort as sort , GradeLevel ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other) + sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other) , '#,##0') as TotalRegister  FROM ##CCTotaltemp8  group by GradeLevel, Grade_Sort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort 
        '''
        cursor.execute(query_byGradeLevel)
        results_byGradeLevel = cursor.fetchall()
        return results_byGradeLevel
    
    def fetch_data_by_tempResFlag(self,cursor):
        query_byTempResFlag = '''
        select STHFlag, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 ,  TotalRegister from (  select * from  (  Select   STHFlagSort as sort ,  case when STHFlag = 'Y' then 'Yes' else 'No'  end as STHFlag ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other) + sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other) , '#,##0') as TotalRegister  FROM ##CCTotaltemp8 where STHFlag in ('Y', 'N')  group by STHFlag, STHFlagSort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort 
        '''
        cursor.execute(query_byTempResFlag)
        results_byTempResFlag = cursor.fetchall()
        return results_byTempResFlag
    
    def fetch_data_by_fosterCareStatus(self,cursor):
        query_byFosterCareStatus = '''
        select FostercareFlag, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 ,  TotalRegister from (  select * from  (  Select   FosterCareFlagSort as sort , FostercareFlag ,FORMAT(sum(Non_ELL_English), '#,##0') as c1 ,FORMAT(sum(Non_ELL_Spanish), '#,##0') as c2 ,FORMAT(sum(Non_ELL_Chinese), '#,##0') as c3 ,FORMAT(sum(Non_ELL_Other), '#,##0') as c4 ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other), '#,##0') as c5  ,FORMAT(sum(ELL_English), '#,##0') as c6 ,FORMAT(sum(ELL_Spanish), '#,##0') as c7 ,FORMAT(sum(ELL_Chinese), '#,##0') as c8 ,FORMAT(sum(ELL_Other), '#,##0') as  c9 ,FORMAT(sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other), '#,##0') as c10  ,FORMAT(sum(Non_ELL_English + Non_ELL_Spanish + Non_ELL_Chinese + Non_ELL_Other) + sum(ELL_English + ELL_Spanish + ELL_Chinese + ELL_Other) , '#,##0') as TotalRegister  FROM ##CCTotaltemp8  group by FostercareFlag, FosterCareFlagSort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort 
        '''
        cursor.execute(query_byFosterCareStatus)
        results_byFosterCareStatus = cursor.fetchall()
        return results_byFosterCareStatus
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C5':'M37'] + ws['C41':'M46'] + ws['C50':'M52'] + ws['C56':'M59'] +  ws['C63':'M76'] + ws['C80':'M93'] + ws['C97':'M99'] + ws['C104':'M106']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                if isinstance(cell.value, str):
                    try:
                        cell.value = int(cell.value)
                    except ValueError:
                        # If the value cannot be converted to int, keep the original value
                        pass
        # Function to check if a string represents a valid number
        def is_number(s):
            try:
                float(s.replace(',', ''))  # Try converting after removing commas
                return True
            except ValueError:
                return False

        # Formatting specific cell ranges
        cell_ranges = ['C5:M37', 'C41:M46', 'C50:M52', 'C56:M59', 'C63:M76', 'C80:M93', 'C97:M99', 'C104:M106']
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if isinstance(cell.value, str) and is_number(cell.value):
                        # Convert to float after removing commas
                        cell.value = float(cell.value.replace(',', ''))
                        # Apply number format with commas (optional)
                        cell.number_format = '#,##0'
        # for row in ws['C41':'M46']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C50':'M52']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C56':'M59']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C63':'M76']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C80':'M93']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C97':'M99']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C104':'M106']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['B1': 'M1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B37':'M37'] + ws['B46':'M46'] + ws['B52':'M52'] + ws['B59':'M59'] + ws['B76':'M76'] + ws['B93':'M93'] + ws['B99':'M99'] + ws['B106':'M106']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        for row in ws['B3':'M3'] + ws['B39':'M39'] + ws['B48':'M48'] + ws['B54':'M54'] + ws['B61':'M61'] +  ws['B78':'M78'] + ws['B95':'M95'] + ws['B102':'M102']:
            for cell in row:
                cell.border = black_border_thick
                cell.font = Font(bold=True, size=12)

        for row in ws['G5':'G37'] + ws['G41':'G46'] + ws['G50':'G52'] + ws['G56':'G59'] +  ws['G63':'G76'] + ws['G80':'G93'] + ws['G97':'G99'] + ws['G104':'G106']:
            for cell in row:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for row in ws['L5':'L37'] + ws['L41':'L46'] + ws['L50':'L52'] + ws['L56':'L59'] +  ws['L63':'L76'] + ws['L80':'L93'] + ws['L97':'L99'] + ws['L104':'L106']:
            for cell in row:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for row in ws['G4':'G4'] + ws['L4':'L4'] + ws['G40':'G40'] + ws['L40':'L40'] + ws['G49':'G49'] + ws['L49':'L49'] + ws['G55':'G55'] + ws['L55':'L55'] + ws['G62':'G62'] + ws['L62':'L62'] + ws['G79':'G79'] + ws['L79':'L79'] + ws['G96':'G96'] + ws['L96':'L96'] + ws['G103':'G103'] + ws['L103':'L103'] + ws['M5':'M37'] + ws['M41':'M46'] + ws['M50':'M52'] + ws['M56':'M59'] +  ws['M63':'M76'] + ws['M80':'M93'] + ws['M97':'M99'] + ws['M104':'M106']:
            for cell in row:
                # cell.fill = PatternFill(start_color='#DCE6F1', end_color='#DCE6F1', fill_type='solid')
                cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        # Make column M bold
        for row in ws['M5':'M37'] + ws['M41':'M46'] + ws['M50':'M52'] + ws['M56':'M59'] +  ws['M63':'M76'] + ws['M80':'M93'] + ws['M97':'M99'] + ws['M104':'M106']:
            for cell in row:
                cell.font = Font(bold=True, size=12)

        # wrap text of B50 cell having 'Eligible for the Free/Reduced Price Lunch Program' to fit the cell
        ws['B50'].alignment = Alignment(wrap_text=True)

        # make cell B4, G4, L4, M4, B40, G40, L40, M40, B49, G49, L49, M49, B55, G55, L55, M55, B62, G62, L62, M62, B79, G79, L79, M79, B96, G96, L96, M96, B103, G103, L103, M103 bold
        for row in ws['B4':'B4'] + ws['G4':'G4'] + ws['L4':'L4'] + ws['M4':'M4'] + ws['B40':'B40'] + ws['G40':'G40'] + ws['L40':'L40'] + ws['M40':'M40'] + ws['B49':'B49'] + ws['G49':'G49'] + ws['L49':'L49'] + ws['M49':'M49'] + ws['B55':'B55'] + ws['G55':'G55'] + ws['L55':'L55'] + ws['M55':'M55'] + ws['B62':'B62'] + ws['G62':'G62'] + ws['L62':'L62'] + ws['M62':'M62'] + ws['B79':'B79'] + ws['G79':'G79'] + ws['L79':'L79'] + ws['M79':'M79'] + ws['B96':'B96'] + ws['G96':'G96'] + ws['L96':'L96'] + ws['M96':'M96'] + ws['B103':'B103'] + ws['G103':'G103'] + ws['L103':'L103'] + ws['M103':'M103']:
            for cell in row:
                cell.font = Font(bold=True, size=12)
        
    def main_Report_8_Registers(self):
        title_cells = [
            {"cell": "B1", "value": "Report 8 Register Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; Grade Level; Disability Classification; Temp House Status and Foster Care Status.", "merge_cells": "B1:M1"},
            

        ]

        subtitle_cells = [
            {"cell": "B3", "value": self.schoolyear + " Students with IEPs by District", "merge_cells": "B3:M3"},
            {"cell": "B39", "value": self.schoolyear + " Students with IEPs by Race/Ethnicity", "merge_cells": "B39:M39"},
            {"cell": "B48", "value": self.schoolyear + " Students with IEPs by Meal Status", "merge_cells": "B48:M48"},
            {"cell": "B54", "value": self.schoolyear + " Students with IEPs by Gender", "merge_cells": "B54:M54"},
            {"cell": "B61", "value": self.schoolyear + " Students with IEPs by Grade Level", "merge_cells": "B61:M61"},
            {"cell": "B78", "value": self.schoolyear + " Students with IEPs by Disability Classification", "merge_cells": "B78:M78"},
            {"cell": "B95", "value": self.schoolyear + " Students with IEPs by Temporary Housing Status", "merge_cells": "B95:M95"},
            {"cell": "B102", "value": self.schoolyear + " Students with IEPs by Foster Care Status", "merge_cells": "B102:M102"},
            

        ]

        column_widths = [5, 30, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_byRace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byRace, start_row=41)
        
        # Step 4: Fetch and write data for "Report 8b = IEP Service Recs by District"
        results_byDistrict = self.fetch_data_by_district(cursor)
        # # replace 01 as 1, 02 as 2, etc.
        # results_byDistrict = [(x[0].replace('01', '1'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('02', '2'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('03', '3'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('04', '4'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('05', '5'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('06', '6'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('07', '7'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('08', '8'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('09', '9'), *x[1:]) for x in results_byDistrict]
        self.write_data_to_excel(ws, results_byDistrict, start_row=5)

        # Step 5: Fetch and write data for "Report 8b = IEP Service Recs by Meal Status"
        results_byMealStatus = self.fetch_data_by_mealstatus(cursor)
        # replace Free or Reduced Price Meal to Eligible for the Free/Reduced Price Lunch Program
        results_byMealStatus = [(x[0].replace('Free or Reduced Price Meal', 'Eligible for the Free/Reduced Price Lunch Program'), *x[1:]) for x in results_byMealStatus]
        self.write_data_to_excel(ws, results_byMealStatus, start_row=50)

        # Step 6: Fetch and write data for "Report 8b = IEP Service Recs by Gender"
        results_byMealStatus = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_byMealStatus, start_row=56)

        # # Step 7: Fetch and write data for "Report 8b = IEP Service Recs by ELL Status"
        # results_byELLStatus = self.fetch_data_by_ellstatus(cursor)
        # self.write_data_to_excel(ws, results_byELLStatus, start_row=63)
        
        # Step 8: Fetch and write data for "Report 8b = IEP Service Recs by Language"
        results_byClassification = self.fetch_data_by_classification(cursor)
        #define a dictionary for sorting order
        sort_order= {'Autism': 1, 'Deaf-Blindness': 2, 'Deafness': 3, 'Emotional Disability': 4, 'Hearing Impairment': 5, 'Intellectual Disability': 6, 'Learning Disability': 7, 'Multiple Disabilities': 8, 'Orthopedic Impairment': 9, 'Other Health Impairment': 10, 'Speech or Language Impairment': 11, 'Traumatic Brain Injury': 12, 'Visual Impairment': 13, 'Total': 14}
        #sort the list using a lambda function that references the sort_order dictionary
        sort_results_byClassification = sorted(results_byClassification, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, sort_results_byClassification, start_row=80)

        # Step 9: Fetch and write data for "Report 8b = IEP Service Recs by Grade Level"
        results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        # replace 01 as 1, 02 as 2, etc.
        results_byGradeLevel = [(x[0].replace('0K', 'KG'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('01', '1'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('02', '2'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('03', '3'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('04', '4'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('05', '5'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('06', '6'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('07', '7'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('08', '8'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('09', '9'), *x[1:]) for x in results_byGradeLevel]   
        # Define a dictionary for sorting order
        sort_order= {'KG': 1, '1': 2, '2': 3, '3': 4, '4': 5, '5': 6, '6': 7, '7': 8, '8': 9, '9': 10, '10': 11, '11': 12, '12': 13, 'Total': 14}
        # Sort the list using a lambda function that references the sort_order dictionary KG,01,02...12
        sort_results_byGradeLevel = sorted(results_byGradeLevel, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, sort_results_byGradeLevel, start_row=63)

        # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Temporary Housing"
        results_byTempResFlag = self.fetch_data_by_tempResFlag(cursor)
        self.write_data_to_excel(ws, results_byTempResFlag, start_row=97)

        # Step 11: Fetch and write data for "Report 8b = IEP Service Recs by Foster Care Status"
        results_byFosterCareStatus = self.fetch_data_by_fosterCareStatus(cursor)
        self.write_data_to_excel(ws, results_byFosterCareStatus, start_row=104)
        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.main_Report_8_Registers()                                                                  