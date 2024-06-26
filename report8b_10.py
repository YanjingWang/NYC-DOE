import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def __init__(self):
        self.schoolyear = 'SY 2022-23'
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        # ws.merge_cells(header_start_cell + ':' + chr(ord(column_letters[-1]) + 1) + str(int(header_start_cell[1:])+1))
        # Merge the header_start_cell with the cell directly below it
        ws.merge_cells(start_row=ws[header_start_cell].row,
                    start_column=ws[header_start_cell].column,
                    end_row=ws[header_start_cell].row + 1,
                    end_column=ws[header_start_cell].column)

        print('headercell:'+header_start_cell)
        print('mergecell:'+ header_start_cell + ':' +str(int(header_start_cell[1:])+1))
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            ws[col + str(int(cell_number)+1)] = '#'
            ws[col + str(int(cell_number)+1)].font = font_style
            ws[col + str(int(cell_number)+1)].border = border_style
            ws[col + str(int(cell_number)+1)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
            ws[col + str(int(cell_number)+1)].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)] = '%'
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)].font = font_style
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)].border = border_style
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)].alignment = Alignment(horizontal='center', vertical='center')
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            # Merge header cells for '%' and '#' under each main column
            ws.merge_cells(col + cell_number + ':' + chr(ord(col) + 1) + cell_number)
            print('col: '+col)
            print('col+cell_number: '+col + cell_number) 
            print('#: '+ col + str(int(cell_number)+1))
            print('%: '+chr(ord(col) + 1) + str(int(cell_number)+1))
            print('Sub mergecell:'+ col + cell_number + ':' + chr(ord(col) + 1) + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c) + 1) for c in column_letters]:
            ws[col + cell_number].border = border_style
            ws[col + str(int(cell_number)+1)].border = border_style



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Report 8b = IEP Service Recs"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Report 10 = IEP Service Recs")


        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=140, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            # cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Related services only', 'Special Education Teacher Support Services (SETSS)',
                'Integrated Co-Teaching Services', 'Special Class in a Community School', 
                'Special Class in a District 75 school', 'Special Class in a Non-public School Placement']
        column_letters = ['C', 'E', 'G', 'I', 'K', 'M']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=False, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        header_fill_color = "B8CCE4"
        column_fill_color = "E0F0F8"
        self.format_header(ws, 'B4', 'District', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B41', 'Ethnicity', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B51', 'Meal Status', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B58', 'Gender', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B66', 'ELL Status', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B73', 'Language of Instruction', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B82', 'Grade Level', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B101', 'Disability Classification', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B120', 'Temporary Housing Status', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B128', 'Foster Care Status', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_StudentRegisterR814_061523')
        cursor.execute("EXEC [dbo].[USPCC_AnnaulReport8b] @tableNameCCStudentRegisterR814=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_race(self,cursor):
        query_byRace = '''
        select EthnicityGroupCC, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11,c12 from (  select * from  ( Select Ethnicity_sort as sort , EthnicityGroupCC,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by EthnicityGroupCC, Ethnicity_sort ) a  union all  select * from ##TotalRow_Sort8b  ) a order by sort
        '''  # the byRace SQL query goes here
        cursor.execute(query_byRace)
        results_byRace = cursor.fetchall()
        return results_byRace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,cursor):
        query_byDistrict = '''
        select * from  ( Select  ReportingDistrict as sort ,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by ReportingDistrict  ) a  union all  select * from ##TotalRow8b  order by sort 
        '''  # the byDistrict SQL query goes here
        cursor.execute(query_byDistrict)
        results_byDistrict = cursor.fetchall()
        return results_byDistrict

    def fetch_data_by_mealstatus(self,cursor):
        query_byMealStatus = '''
        select * from  ( Select  MealStatusGrouping as sort ,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by MealStatusGrouping  ) a  union all  select * from ##TotalRow8b  order by sort
        '''  # the byMealStatus SQL query goes here
        cursor.execute(query_byMealStatus)
        results_byMealStatus = cursor.fetchall()
        return results_byMealStatus
    
    def fetch_data_by_gender(self,cursor):
        query_byGender = '''
        select * from  ( Select  Gender as sort ,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by Gender  ) a  union all  select * from ##TotalRow8b  order by sort 
        '''  # the byGender SQL query goes here
        cursor.execute(query_byGender)
        results_byGender = cursor.fetchall()
        return results_byGender
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byELLStatus = '''
        select * from  ( Select  ELLStatus as sort ,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by ELLStatus  ) a  union all  select * from ##TotalRow8b  order by sort       
        '''  # the byELLStatus SQL query goes here
        cursor.execute(query_byELLStatus)
        results_byELLStatus = cursor.fetchall()
        return results_byELLStatus
    
    def fetch_data_by_language(self,cursor):
        query_byLanguage = '''
        select OutcomeLanguageCC, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11,c12 from (  select * from  ( Select OutcomeLanguageCCSort as sort , OutcomeLanguageCC,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by OutcomeLanguageCC, OutcomeLanguageCCSort ) a  union all  select * from ##TotalRow_Sort8b  ) a order by sort  
        '''  # the byLanguage SQL query goes here
        cursor.execute(query_byLanguage)
        results_byLanguage = cursor.fetchall()
        return results_byLanguage
    
    def fetch_data_by_gradelevel(self,cursor):
        query_byGradeLevel = '''
        select GradeLevel, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11,c12 from (  select * from  ( Select GradeSort as sort , GradeLevel,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by GradeLevel, GradeSort ) a  union all  select * from ##TotalRow_Sort8b  ) a order by sort
        '''
        cursor.execute(query_byGradeLevel)
        results_byGradeLevel = cursor.fetchall()
        return results_byGradeLevel
    
    def fetch_data_by_tempResFlag(self,cursor):
        query_byTempResFlag = '''
        select STHFlag, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11,c12 from (  select * from  ( Select STHFlagSort as sort , STHFlag,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by STHFlag, STHFlagSort ) a  union all  select * from ##TotalRow_Sort8b  ) a order by sort 
        '''
        cursor.execute(query_byTempResFlag)
        results_byTempResFlag = cursor.fetchall()
        return results_byTempResFlag
    
    def fetch_data_by_fosterCareStatus(self,cursor):
        query_byFosterCareStatus = '''
        select FostercareFlag, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11,c12 from (  select * from  ( Select FosterCareFlagSort as sort , FostercareFlag,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by FostercareFlag, FosterCareFlagSort ) a  union all  select * from ##TotalRow_Sort8b  ) a order by sort 
        '''
        cursor.execute(query_byFosterCareStatus)
        results_byFosterCareStatus = cursor.fetchall()
        return results_byFosterCareStatus
    
    def fetch_data_by_disabilityclassification(self,cursor):
        query_byDisabilityClassification = '''
        select * from  ( Select  Classification as sort ,FORMAT(sum(RSOnly), '#,##0') as c1 ,CONCAT(cast((sum(RSOnly)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c2 ,FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c4 ,FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c6 ,FORMAT(sum(SpecialClass), '#,##0') as c7 ,CONCAT(cast((sum(SpecialClass)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,CONCAT(cast((sum(SpecialClassD75)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,CONCAT(cast((sum(SpecialClassNPS)*1.0)/(nullif(Count(studentid),0))*100 as numeric(7)), '%') as c12  FROM ##CCTotaltemp8b  group by Classification  ) a  union all  select * from ##TotalRow8b  
        '''
        cursor.execute(query_byDisabilityClassification)
        results_byDisabilityClassification = cursor.fetchall()
        return results_byDisabilityClassification

    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel
        for row_num, row_data in enumerate(data, start=start_row):
            # Check if the cell is a merged cell and skip if it is
            if isinstance(ws.cell(row=row_num, column=2), openpyxl.cell.cell.MergedCell):
                continue

            ws.cell(row=row_num, column=2).value = row_data[0]  # Assuming 'B' column is not merged for data rows
            col_pointer = 3  # Start at column 'C'

            for i in range(1, len(row_data), 2):
                # Check for each cell if it's a MergedCell before writing the value
                if not isinstance(ws.cell(row=row_num, column=col_pointer), openpyxl.cell.cell.MergedCell):
                    ws.cell(row=row_num, column=col_pointer).value = row_data[i]  # Number data
                col_pointer += 1

                if not isinstance(ws.cell(row=row_num, column=col_pointer), openpyxl.cell.cell.MergedCell):
                    ws.cell(row=row_num, column=col_pointer).value = row_data[i+1]  # Percentage data
                col_pointer += 1
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M','N']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C6':'N38'] + ws['C43':'N48'] + ws['C53':'N55'] + ws['C60':'N63'] + ws['C68':'N70'] + ws['C75':'N79'] + ws['C84':'N97'] + ws['C103':'N116'] + ws['C122':'N124'] + ws['C130':'N132']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                if isinstance(cell.value, str):
                    try:
                        cell.value = int(cell.value)
                    except ValueError:
                        # If the value cannot be converted to int, keep the original value
                        pass
        # Function to check if a string represents a valid number
        def is_number(s):
            try:
                float(s.replace(',', ''))  # Try converting after removing commas
                return True
            except ValueError:
                return False

        # Formatting specific cell ranges
        cell_ranges = ['C6:N38', 'C43:N48', 'C53:N55', 'C60:N63', 'C68:N70', 'C75:N79', 'C84:N97', 'C103:N116', 'C122:N124', 'C130:N132']
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if isinstance(cell.value, str) and is_number(cell.value):
                        # Convert to float after removing commas
                        cell.value = float(cell.value.replace(',', ''))
                        # Apply number format with commas (optional)
                        cell.number_format = '#,##0'
        # for row in ws['C43':'N48']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C53':'N55']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C60':'N63']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C68':'N70']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C75':'N79']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C84':'N97']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C103':'N105']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C111':'N113']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['B1': 'N1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B38':'N38'] + ws['B48':'N48'] + ws['B55':'N55'] + ws['B63':'N63'] + ws['B70':'N70'] + ws['B79':'N79'] + ws['B97':'N97'] + ws['B116':'N116'] + ws['B124':'N124'] + ws['B132':'N132']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['B3':'N3'] + ws['B40':'N40'] + ws['B50':'N50'] + ws['B57':'N57'] + ws['B65':'N65'] + ws['B72':'N72'] + ws['B81':'N81'] + ws['B100':'N100'] + ws['B119':'N119'] + ws['B127':'N127']:
            for cell in row:
                cell.border = black_border_thick
                cell.font = Font(bold=True, size=12)

        for row in ws['B4': 'B5'] + ws['C4':'D4'] + ws['E4':'F4'] + ws['G4':'H4'] + ws['I4':'J4'] + ws['K4':'L4'] + ws['M4':'N4'] + ws['C5':'C5'] + ws['D5':'D5'] + ws['E5':'E5'] + ws['F5':'F5'] + ws['G5':'G5'] + ws['H5':'H5'] + ws['I5':'I5'] + ws['J5':'J5'] + ws['K5':'K5'] + ws['L5':'L5'] + ws['M5':'M5'] + ws['N5':'N5'] + ws ['B41':'B42'] + ws['C41':'D41'] + ws['E41':'F41'] + ws['G41':'H41'] + ws['I41':'J41'] + ws['K41':'L41'] + ws['M41':'N41'] + ws['C42': 'C42'] + ws['D42':'D42'] + ws['E42':'E42'] + ws['F42':'F42'] + ws['G42':'G42'] + ws['H42':'H42'] + ws['I42':'I42'] + ws['J42':'J42'] + ws['K42':'K42'] + ws['L42':'L42'] + ws['M42':'M42'] + ws['N42':'N42'] + ws['B51':'B52'] + ws['C51':'D51'] + ws['E51':'F51'] + ws['G51':'H51'] + ws['I51':'J51'] + ws['K51':'L51'] + ws['M51':'N51'] + ws['C52':'C52'] + ws['D52':'D52'] + ws['E52':'E52'] + ws['F52':'F52'] + ws['G52':'G52'] + ws['H52':'H52'] + ws['I52':'I52'] + ws['J52':'J52'] + ws['K52':'K52'] + ws['L52':'L52'] + ws['M52':'M52'] + ws['N52':'N52'] + ws['B58':'B59'] + ws['C58':'D58'] + ws['E58':'F58'] + ws['G58':'H58'] + ws['I58':'J58'] + ws['K58':'L58'] + ws['M58':'N58'] + ws['C59':'C59'] + ws['D59':'D59'] + ws['E59':'E59'] + ws['F59':'F59'] + ws['G59':'G59'] + ws['H59':'H59'] + ws['I59':'I59'] + ws['J59':'J59'] + ws['K59':'K59'] + ws['L59':'L59'] + ws['M59':'M59'] + ws['N59':'N59'] + ws['B66':'B67'] + ws['C66':'D66'] + ws['E66':'F66'] + ws['G66':'H66'] + ws['I66':'J66'] + ws['K66':'L66'] + ws['M66':'N66'] + ws['C67':'C67'] + ws['D67':'D67'] + ws['E67':'E67'] + ws['F67':'F67'] + ws['G67':'G67'] + ws['H67':'H67'] + ws['I67':'I67'] + ws['J67':'J67'] + ws['K67':'K67'] + ws['L67':'L67'] + ws['M67':'M67'] + ws['N67':'N67'] + ws['B73':'B74'] + ws['C73':'D73'] + ws['E73':'F73'] + ws['G73':'H73'] + ws['I73':'J73'] + ws['K73':'L73'] + ws['M73':'N73'] + ws['C74':'C74'] + ws['D74':'D74'] + ws['E74':'E74'] + ws['F74':'F74'] + ws['G74':'G74'] + ws['H74':'H74'] + ws['I74':'I74'] + ws['J74':'J74'] + ws['K74':'K74'] + ws['L74':'L74'] + ws['M74':'M74'] + ws['N74':'N74'] + ws['B82':'B83'] + ws['C82':'D82'] + ws['E82':'F82'] + ws['G82':'H82'] + ws['I82':'J82'] + ws['K82':'L82'] + ws['M82':'N82'] + ws['C83':'C83'] + ws['D83':'D83'] + ws['E83':'E83'] + ws['F83':'F83'] + ws['G83':'G83'] + ws['H83':'H83'] + ws['I83':'I83'] + ws['J83':'J83'] + ws['K83':'K83'] + ws['L83':'L83'] + ws['M83':'M83'] + ws['N83':'N83'] + ws['B101':'B102'] + ws['C101':'D101'] + ws['E101':'F101'] + ws['G101':'H101'] + ws['I101':'J101'] + ws['K101':'L101'] + ws['M101':'N101'] + ws['C102':'C102'] + ws['D102':'D102'] + ws['E102':'E102'] + ws['F102':'F102'] + ws['G102':'G102'] + ws['H102':'H102'] + ws['I102':'I102'] + ws['J102':'J102'] + ws['K102':'K102'] + ws['L102':'L102'] + ws['M102':'M102'] + ws['N102':'N102'] + ws['B120':'B121'] + ws['C120':'D120'] + ws['E120':'F120'] + ws['G120':'H120'] + ws['I120':'J120'] + ws['K120':'L120'] + ws['M120':'N120'] + ws['C121':'C121'] + ws['D121':'D121'] + ws['E121':'E121'] + ws['F121':'F121'] + ws['G121':'G121'] + ws['H121':'H121'] + ws['I121':'I121'] + ws['J121':'J121'] + ws['K121':'K121'] + ws['L121':'L121'] + ws['M121':'M121'] + ws['N121':'N121'] + ws['B128':'B129'] + ws['C128':'D128'] + ws['E128':'F128'] + ws['G128':'H128'] + ws['I128':'J128'] + ws['K128':'L128'] + ws['M128':'N128'] + ws['C129':'C129'] + ws['D129':'D129'] + ws['E129':'E129'] + ws['F129':'F129'] + ws['G129':'G129'] + ws['H129':'H129'] + ws['I129':'I129'] + ws['J129':'J129'] + ws['K129':'K129'] + ws['L129':'L129'] + ws['M129':'M129'] + ws['N129':'N129']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=False, size=12)

        # make B4, B41, B51, B58, B66, B73, B82, B101, B120, B128 bold
        for row in  ws['B4': 'B5'] + ws['B41':'B42'] + ws['B51':'B52'] + ws['B58':'B59'] + ws['B66':'B67'] + ws['B73':'B74'] + ws['B82':'B83'] + ws['B101':'B102'] + ws['B120':'B121'] + ws['B128':'B129']:
            for cell in row:
                cell.font = Font(bold=True, size=12)
                
        # wrap text of B53 cell having 'Eligible for the Free/Reduced Price Lunch Program' to fit the cell
        ws['B53'].alignment = Alignment(wrap_text=True)
        # make cell B1 higher to fit the text
        ws.row_dimensions[1].height = 30   
    def main_Report_10_IEP_Service_Recs(self):
        title_cells = [
            {"cell": "B1", "value": "Report 10 IEP Service Recommendations Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; Grade Level; Disability Classification; Temp House Status and Foster Care Status.", "merge_cells": "B1:N1"},
            

        ]

        subtitle_cells = [
            {"cell": "B3", "value": self.schoolyear + " Students with IEP Recommended Programs and Services by District", "merge_cells": "B3:N3"},
            {"cell": "B40", "value": self.schoolyear + " Students with IEP Recommended Programs and Services by Ethnicity", "merge_cells": "B40:N40"},
            {"cell": "B50", "value": self.schoolyear + " Students with IEP Recommended Programs and Services by Meal Status", "merge_cells": "B50:N50"},
            {"cell": "B57", "value": self.schoolyear + " Students with IEP Recommended Programs and Services by Gender", "merge_cells": "B57:N57"},
            {"cell": "B65", "value": self.schoolyear + " Students with IEP Recommended Programs and Services by ELL Status", "merge_cells": "B65:N65"},
            {"cell": "B72", "value": self.schoolyear + " Students with IEP Recommended Programs and Services  by Recommended Language of Instruction", "merge_cells": "B72:N72"},
            {"cell": "B81", "value": self.schoolyear + " Students with IEP Recommended Programs and Services  by Grade Level", "merge_cells": "B81:N81"},
            {"cell": "B100", "value": self.schoolyear + " Students with IEP Recommended Programs and Services  by Disability Classification", "merge_cells": "B100:N100"},
            {"cell": "B119", "value": self.schoolyear + " Students with IEP Recommended Programs and Services by Temporary Housing Status", "merge_cells": "B119:N119"},            
            {"cell": "B127", "value": self.schoolyear + " Students with IEP Recommended Programs and Services by Foster Care Status", "merge_cells": "B127:N127"},
            

        ]

        column_widths = [5, 30, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_byRace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byRace, start_row=43)
        
        # Step 4: Fetch and write data for "Report 8b = IEP Service Recs by District"
        results_byDistrict = self.fetch_data_by_district(cursor)
        # # replace 01 as 1, 02 as 2, etc.
        # results_byDistrict = [(x[0].replace('01', '1'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('02', '2'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('03', '3'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('04', '4'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('05', '5'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('06', '6'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('07', '7'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('08', '8'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('09', '9'), *x[1:]) for x in results_byDistrict]
        self.write_data_to_excel(ws, results_byDistrict, start_row=6)

        # Step 5: Fetch and write data for "Report 8b = IEP Service Recs by Meal Status"
        results_byMealStatus = self.fetch_data_by_mealstatus(cursor)
        results_byMealStatus = [('Eligible for the Free/Reduced Price Lunch Program' if x[0] == 'Free or Reduced Price Meal' else x[0], *x[1:]) for x in results_byMealStatus]
        self.write_data_to_excel(ws, results_byMealStatus, start_row=53)

        # Step 6: Fetch and write data for "Report 8b = IEP Service Recs by Gender"
        results_byGender = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_byGender, start_row=60)

        # Step 7: Fetch and write data for "Report 8b = IEP Service Recs by ELL Status"
        results_byELLStatus = self.fetch_data_by_ellstatus(cursor)
        # replace 'ELL' with 'Ell' and 'NON-ELL' with 'Non-Ell'
        results_byELLStatus = [('ELL' if x[0] == 'ELL' else ('NOT ELL' if x[0] == 'Non-Ell' else x[0]), *x[1:]) for x in results_byELLStatus]
        self.write_data_to_excel(ws, results_byELLStatus, start_row=68)
        
        # Step 8: Fetch and write data for "Report 8b = IEP Service Recs by Language"
        results_byLanguage = self.fetch_data_by_language(cursor)
        # Replace 'ENGLISH' with 'English' and 'SPANISH' with 'Spanish' and 'CHINESE' with 'Chinese' and 'OTHER' with 'Other'
        results_byLanguage = [('English' if x[0] == 'ENGLISH' else ('Spanish' if x[0] == 'SPANISH' else ('Chinese' if x[0] == 'CHINESE' else ('Other' if x[0] == 'OTHER' else x[0]))), *x[1:]) for x in results_byLanguage]
        # Define a dictionary for sorting order
        sort_order = {'English': 1, 'Spanish': 2, 'Chinese': 3, 'Other': 4, 'Total': 5}
        # Sort the list using a lambda function that references the sort_order dictionary
        sorted_results_byLanguage = sorted(results_byLanguage, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, sorted_results_byLanguage, start_row=75)

        # Step 9: Fetch and write data for "Report 8b = IEP Service Recs by Grade Level"
        results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        # replace 0K as KG, 01 as 1, 02 as 2, etc. 
        results_byGradeLevel = [(x[0].replace('0K', 'KG'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('01', '1'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('02', '2'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('03', '3'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('04', '4'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('05', '5'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('06', '6'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('07', '7'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('08', '8'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('09', '9'), *x[1:]) for x in results_byGradeLevel]
        # Define a dictionary for sorting order
        sort_order= {'KG': 1, '1': 2, '2': 3, '3': 4, '4': 5, '5': 6, '6': 7, '7': 8, '8': 9, '9': 10, '10': 11, '11': 12, '12': 13, 'Total': 14}
        # Sort the list using a lambda function that references the sort_order dictionary KG,01,02...12
        sort_results_byGradeLevel = sorted(results_byGradeLevel, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, sort_results_byGradeLevel, start_row=84)

        # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Disability Classification"
        results_byDisabilityClassification = self.fetch_data_by_disabilityclassification(cursor)
        # Define a dictionary for sorting order
        sort_order = {'Autism': 1, 'Deaf-Blindness': 2, 'Deafness': 3, 'Emotional Disability': 4, 'Hearing Impairment': 5, 'Intellectual Disability': 6, 'Learning Disability': 7, 'Multiple Disabilities': 8, 'Orthopedic Impairment': 9, 'Other Health Impairment': 10, 'Speech or Language Impairment': 11, 'Traumatic Brain Injury': 12, 'Visual Impairment': 13, 'Total': 14}
        # Sort the list using a lambda function that references the sort_order dictionary
        sorted_results_byDisabilityClassification = sorted(results_byDisabilityClassification, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, sorted_results_byDisabilityClassification, start_row=103)

        # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Temporary Housing"
        results_byTempResFlag = self.fetch_data_by_tempResFlag(cursor)
        # Replace 'Y' with 'Yes' and 'N' with 'No'
        results_byTempResFlag = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_byTempResFlag]
        # Define a dictionary for sorting order
        sort_order = {'Yes': 1, 'No': 2, 'Total': 3}
        # Sort the list using a lambda function that references the sort_order dictionary
        sorted_results_byTempResFlag = sorted(results_byTempResFlag, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, sorted_results_byTempResFlag, start_row=122)

        # Step 11: Fetch and write data for "Report 8b = IEP Service Recs by Foster Care Status"
        results_byFosterCareStatus = self.fetch_data_by_fosterCareStatus(cursor)
        # Replace 'Y' with 'Yes' and 'N' with 'No'
        results_byFosterCareStatus = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_byFosterCareStatus]
        # Define a dictionary for sorting order
        sort_order = {'Yes': 1, 'No': 2, 'Total': 3}
        # Sort the list using a lambda function that references the sort_order dictionary
        sorted_results_byFosterCareStatus = sorted(results_byFosterCareStatus, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, sorted_results_byFosterCareStatus, start_row=130)
        # Step 12: close the database connection
        cursor.close()
        # Step 13: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.main_Report_10_IEP_Service_Recs()                                                            