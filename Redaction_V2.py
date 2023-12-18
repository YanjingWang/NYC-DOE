"""
Reports 1-4 = Initials subtotal calculation:
E + F = G, H + I = J, G + J = K, D + K + L + M = C
Reports 5-7 = Reevaluations subtotal calculation:
E + F = G, H + I = J, G + J = K, D + K + L  = C
Report 8 = Registers subtotal calculation:
C + D + E + F = G, H + I + J + K = L, G + L = M
SWDs by School : C4+ ...+ C1602 = C1603
Report 8a = Disability class subtotal calculation:
C + D +E + F + G + H + I + J + K + L + M + N + O = P 
Report 9 = Placement: C6+...+C37 = C38, D6+...+D37 = D38, C42+...C46 = C47, D42+...D46 = D47, C51+C52 = C53, D51+D52 = D53, C57+C58 = C59, D57+D58 = D59, C63+C64 = C65, 
D63+D64 = D65, C69+C72 = C73, D69+D72 = D73, C77+C89=D90, D77+D89=D90
Report 10 = LRE-MRE: C6+...C37=C38,D6+...D37=D38,E6+...E37=E38,F6+...F37=F38, C42+...C46=C47,D42+...D46=D47,E42+...E46=E47,F42+...F46=F47,C51+C52=C53,D51+D52=D53,E51+E52=E53,F51+F52=F53,E57+E58=E59,F57+F58=F59,C63+C64=C65,D63+D64=D65,E63+E64=E65,F63+F64=F65,C69+C72=C73,D69+D72=D73,E69+E72=E73,F69+F72=F73,C77+...+C89=C90,D77+...+D89=D90,E77+...+E89=E90,F77+...+F89=F90
"""
import openpyxl
from redaction_config import REPORTS_CONFIG
from redaction_config_SY23 import REPORTS_CONFIG_SY23
class Solution:
    # def is_percentage(self, val):
    #     return isinstance(val, str) and val.endswith('%')
    def is_percentage(self, cell):
        if isinstance(cell.value, float) and '0%' in cell.number_format:
            return True
        return False

    
    def mask_value_initial(self,val):
        if isinstance(val, (int, float)) and 0 < val <= 5:
            return '<=5'
        return val

    def mask_value_secondary(self,val):
        if isinstance(val, (int, float)) and 0 < val <= 5:
            return '<=5'
        elif isinstance(val, (int, float)) and val > 5:
            return '>5'
        return val
    
    def mask_value_zero(self,val):
        if isinstance(val, (int, float)) and val == 0:
            return '<=5'
        return val

    def initial_mask(self,ws, start_row, start_col, end_row, end_col):
        # Step 1: Initial Masking
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                if not self.is_percentage(cell):
                    cell.value = self.mask_value_initial(cell.value)
        
        # Mask smallest if only one '<=5' in column
        for col in ws.iter_cols(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            valid_vals = [cell.value for cell in col if isinstance(cell.value, (int, float)) and cell.value != '<=5']
            if sum([cell.value == '<=5' for cell in col]) == 1 and valid_vals:
                min_val = min(valid_vals)
                for cell in col:
                    if cell.value == min_val:
                        cell.value = '>5'

    # Step 2: Secondary Masking
    def secondary_mask(self, ws, total_col_indexes,start_row, end_row, groups):
        # groups = [(5, 6, 7), (8, 9, 10), (7, 10, 11)]  # E,F,G and H,I,J and G,J,K respectively

        for group in groups:
            for row_num in range(start_row, end_row + 1):
                masked_vals = [col for col in group if ws.cell(row=row_num, column=col).value in ['<=5', '>5']]
                if len(masked_vals) == 1:  # If only one value in the group is masked
                    numeric_vals = [ws.cell(row=row_num, column=col).value for col in group if isinstance(ws.cell(row=row_num, column=col).value, (int, float)) and not self.is_percentage(ws.cell(row=row_num, column=col))]
                    if numeric_vals:  # Ensure there are numeric values
                        min_val = min(numeric_vals)
                        for col in group:
                            if ws.cell(row=row_num, column=col).value == min_val:
                                if min_val == 0:
                                    ws.cell(row=row_num, column=col).value = self.mask_value_zero(min_val)
                                else:
                                    ws.cell(row=row_num, column=col).value = self.mask_value_secondary(min_val)
                                break
                elif len(masked_vals) > 2:  # Two or more values are already masked, if total column == `>5` and one of the other columns is `<=5` or `>5`, then unmask the rest `>5` column in this group
                    total_col = [col for col in group if col in total_col_indexes]
                    if total_col:
                        total_col = total_col[0]
                        if ws.cell(row=row_num, column=total_col).value == '>5':
                            for col in group:
                                if ws.cell(row=row_num, column=col).value == '<=5' or ws.cell(row=row_num, column=col).value == '>5':
                                    ws.cell(row=row_num, column=col).value = self.mask_value_initial(ws.cell(row=row_num, column=col).value)

                elif len(masked_vals) == 2: # Two or more values are already masked, no action needed
                    continue


    # Step 3: Masking based on Subtotals        
    def third_mask(self,ws, total_col_indexes, start_row, end_row,groups):
        # groups = [(3,4,11,12)]  # C,D,K,L
        for group in groups:
            for row_num in range(start_row, end_row + 1):
                masked_vals = [col for col in group if ws.cell(row=row_num, column=col).value in ['<=5', '>5']]
                if len(masked_vals) == 1:  # If only one value in the group is masked, need to mask the smallest of the unmasked values
                    numeric_vals = [ws.cell(row=row_num, column=col).value for col in group if isinstance(ws.cell(row=row_num, column=col).value, (int, float)) and not self.is_percentage(ws.cell(row=row_num, column=col))]
                    if numeric_vals:  # Ensure there are numeric values
                        min_val = min(numeric_vals)
                        for col in group:
                            if ws.cell(row=row_num, column=col).value == min_val:
                                if min_val == 0:
                                    ws.cell(row=row_num, column=col).value = self.mask_value_zero(min_val)
                                else:
                                    ws.cell(row=row_num, column=col).value = self.mask_value_secondary(min_val)
                                break
                elif len(masked_vals) > 2:  # Two or more values are already masked, if total column == `>5` and one of the other columns is `<=5` or `>5`, then unmask the rest `>5` column in this group
                    total_col = [col for col in group if col in total_col_indexes]
                    if total_col:
                        total_col = total_col[0]
                        if ws.cell(row=row_num, column=total_col).value == '>5':
                            for col in group:
                                if ws.cell(row=row_num, column=col).value == '<=5' or ws.cell(row=row_num, column=col).value == '>5':
                                    ws.cell(row=row_num, column=col).value = self.mask_value_initial(ws.cell(row=row_num, column=col).value)
                elif len(masked_vals) >= 2:  # Two or more values are already masked, so no action needed
                    continue

    def highlight_overredaction(self, ws, total_col_indexes, start_row, end_row, groups):
        for col_index in total_col_indexes:
            for row_num in range(start_row, end_row + 1):
                cell = ws.cell(row=row_num, column=col_index)
                
                # First, we check if we're in a 'total' column and if so, count the occurrences of '>5' and '<=5'.
                if col_index in total_col_indexes:
                    gt5_count = sum(ws.cell(row=r, column=col_index).value == '>5' for r in range(start_row, end_row + 1))
                    lte5_count = sum(ws.cell(row=r, column=col_index).value == '<=5' for r in range(start_row, end_row + 1))
                    # print(ws.title,'startrow:', str(start_row), 'endrow:', str(end_row), gt5_count, lte5_count)


                    # If more than two '>5' in the 'total' column and it's not a case of acceptable overredaction, we highlight.
                    if gt5_count > 2 and cell.value == '>5':
                        self.highlight_cell(cell)
                        # print the cell coordinate
                        print('1',ws.title, cell.coordinate)
                    
        # If the cell is '>5', check if it's part of a group that has two masked values already.
        if cell.value == '>5':
            for group in groups:
                if row_num in range(group[0], group[1] + 1):
                    masked_count = sum(ws.cell(row=row_num, column=gc).value in ['<=5', '>5'] for gc in group[2:])
                    if masked_count >= 2:
                        self.highlight_cell(cell)
                        # print the cell coordinate
                        print('2',ws.title, cell.coordinate)
                        break

    def highlight_cell(self, cell):
        cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')




    def mask_excel_file(self,filename,tab_name,configurations):
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        try:
            ws = wb[tab_name]
        except KeyError:
            print(f"Warning: Worksheet {tab_name} does not exist in the file {filename}. Skipping...") #SWDs by School is not in SY23
            return
        # Mask data for the specific ranges
        for r in configurations['ranges']:
            self.initial_mask(ws, *r)
            
        # Invoke secondary masking if present in configurations, otherwise skip
        if 'secondary_mask' in configurations and 'secondary_mask_kwargs' in configurations and 'total_col_indexes' in configurations:
            self.secondary_mask(ws,  configurations['total_col_indexes'], *configurations['secondary_mask'], **configurations.get('secondary_mask_kwargs', {}))
        
        # Conditionally invoke third masking if present in configurations
        if 'third_mask' in configurations and 'third_mask_kwargs' in configurations:
            self.third_mask(ws,  configurations['total_col_indexes'],*configurations['third_mask'], **configurations['third_mask_kwargs'])
        # print('all masking is done, now highlight overredaction')
        # Highlight overredaction
        # if 'total_col_indexes' in configurations:
        #     start_row, end_row = configurations['secondary_mask']  # Assuming secondary_mask defines the row range
        #     # print(start_row, end_row)
        #     self.highlight_overredaction(ws, configurations['total_col_indexes'], start_row, end_row)
        # ... inside mask_excel_file after all masking operations
        if 'total_col_indexes' in configurations and 'groups' in configurations:
            start_row, end_row = configurations['secondary_mask']  # Assuming secondary_mask defines the row range
            # print(start_row, end_row)
            self.highlight_overredaction(ws, configurations['total_col_indexes'], start_row, end_row, configurations['groups'])
        # print('highlight overredaction is done')

        # Save the modified workbook
        wb.save(filename) # Adjust the range if necessary


# Call the function with your filename
if __name__ == "__main__":
    processor = Solution()
    filenames = [
        'C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Annual Special Education Data Report Unredacted SY21.xlsx',
        'C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted Annual Special Education Data Report SY22.xlsx'
    ]
    
    for fname in filenames:
        for report, config in REPORTS_CONFIG.items():
            processor.mask_excel_file(fname, report, config)
            
    # filenames = ['C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted Annual Special Education Data Report SY23.xlsx']
    # for fname in filenames:
    #     for report, config in REPORTS_CONFIG_SY23.items():
    #         processor.mask_excel_file(fname, report, config)