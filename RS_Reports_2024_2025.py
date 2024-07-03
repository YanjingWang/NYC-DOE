import pandas as pd
import sqlalchemy as sa
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
class RSCSD():
    def __init__(self,schoolyear="SY 24-25"):
        self.schoolyear = schoolyear
    def RSWeeklyCSDExcel(self):
        start = datetime.now()
        # Set up database connection
        engine = sa.create_engine('mssql+pyodbc://username:password@SEO_MART?driver=ODBC+Driver+17+for+SQL+Server')

        # Define SQL queries
        qry_report = sa.text("SELECT * from [SEO_MART].[dbo].[RPT_RSProvisioning] where SUBSTRING(EnrolledDBN,1,2)<'84'")
        comp_report = sa.text("SELECT * from [SEO_MART].[dbo].[RPT_RSCompliance] where SchoolDistrict<84")
        comp_report2 = sa.text("SELECT * from [SEO_MART].[dbo].[RPT_RSCompliancebyGroup]")
        comp_report3 = sa.text("SELECT * from [SEO_MART].[dbo].[RPT_Locations]")

        # Execute SQL queries and store results in DataFrames
        with engine.connect() as conn:
            report_RS = pd.read_sql(qry_report, conn)
            report_citywide = pd.read_sql(comp_report, conn)
            report_citywide2 = pd.read_sql(comp_report2, conn)
            report_lcgms = pd.read_sql(comp_report3, conn)

        # Convert datetime columns to date format
        report_RS = report_RS.apply(lambda x: pd.to_datetime(x).dt.date if x.dtype == 'datetime64[ns]' else x)

        # Additional DataFrames preparation using SQL queries
        report_RS_Temp1 = report_RS.copy()
        report_RS_Temp1['STUDENTNAME'] = report_RS_Temp1['LastName'] + ',' + report_RS_Temp1['FirstName']
        report_RS_Temp1['Attendrate'] = (report_RS_Temp1['AttendRate'] * 100).round(0).astype(str) + '%'
        report_RS_Temp1['district'] = report_RS_Temp1['EnrolledDBN'].str[:2].astype(int)

        report_RS_Temp = report_RS_Temp1.query('district < 84')

        report_RS_comp = report_citywide[['EnrolledDBN', 'FieldSupportCenterReportingName', 'SuperintendentName', 'NotEncounteredCounsel', 'EncounteredCounsel', 'PercentageEncounteredCounsel', 'NotEncounteredMajor', 'EncounteredMajor', 'PercentageEncounteredMajor', 'PercentageEncounteredAllRS']]

        # Creating different tables for various levels
        report_RS_dbn = report_RS_comp.copy()
        report_RS_dst = report_citywide2[report_citywide2['ReportGroupDesc'] == 'SchoolDistrict'].sort_values('SchoolDistrict')
        report_RS_sup = report_citywide2[report_citywide2['ReportGroupDesc'] == 'Superintendent'].sort_values('SuperintendentName')
        report_RS_fsc = report_citywide2[report_citywide2['ReportGroupDesc'] == 'FieldSupportCenter'].sort_values('FieldSupportCenterReportingName')

        report_asofdt = report_RS['ProcessedDate'].iloc[0]

        # Output Excel reports - WEEKLY SCHOOL LEVEL WITH 2 TABS
        dt = datetime.now().strftime("%Y%m%d")
        pth2 = f"R:/SEO Analytics/Reporting/Related Services/Output Files/{self.schoolyear}/MandatedServices_{dt}"
        os.makedirs(pth2, exist_ok=True)

        mydata = report_RS_Temp
        mycomp = report_RS_comp
        asofdt = report_asofdt

        varNames = mycomp['EnrolledDBN'].unique()

        for i in varNames:
            print(i)
            mydata2 = mydata[mydata['EnrolledDBN'] == i]
            mycomp2 = mycomp[mycomp['EnrolledDBN'] == i]
            wb = load_workbook("C:/Template/RS_Template_new.xlsx")

            ws_data = wb['Data']
            for r_idx, row in enumerate(dataframe_to_rows(mydata2, index=False, header=False), 6):
                for c_idx, value in enumerate(row, 1):
                    ws_data.cell(row=r_idx, column=c_idx, value=value)
            ws_data.cell(row=1, column=2, value=asofdt)

            ws_comp = wb['Completion Reports']
            for r_idx, row in enumerate(dataframe_to_rows(mycomp2, index=False, header=False), 5):
                for c_idx, value in enumerate(row, 1):
                    ws_comp.cell(row=r_idx, column=c_idx, value=value)
            ws_comp.cell(row=1, column=3, value=asofdt)

            pth = f"{pth2}/{i}_MandatedServices_{dt}.xlsx"
            wb.save(pth)

        # Output Consolidated Excel reports - WITH 4 TABS
        pth2 = f"R:/SEO Analytics/Share/Related Services/{dt}"
        os.makedirs(pth2, exist_ok=True)

        mycomp1 = report_RS_dbn
        mycomp2 = report_RS_dst
        mycomp3 = report_RS_sup

        wb = load_workbook("C:/Template/RS_Compliance_new.xlsx")

        ws_supt = wb['RS Supt Citywide Summary']
        for r_idx, row in enumerate(dataframe_to_rows(mycomp1, index=False, header=False), 8):
            for c_idx, value in enumerate(row, 1):
                ws_supt.cell(row=r_idx, column=c_idx, value=value)
        ws_supt.cell(row=1, column=3, value=asofdt)

        ws_dst = wb['RS District Summary']
        for r_idx, row in enumerate(dataframe_to_rows(mycomp2, index=False, header=False), 8):
            for c_idx, value in enumerate(row, 1):
                ws_dst.cell(row=r_idx, column=c_idx, value=value)
        ws_dst.cell(row=1, column=3, value=asofdt)

        ws_sup = wb['RS Superintendent Summary']
        for r_idx, row in enumerate(dataframe_to_rows(mycomp3, index=False, header=False), 8):
            for c_idx, value in enumerate(row, 1):
                ws_sup.cell(row=r_idx, column=c_idx, value=value)
        ws_sup.cell(row=1, column=3, value=asofdt)

        pth1 = f"{pth2}/RS Compliance Report_{dt}.xlsx"
        wb.save(pth1)
        end = datetime.now()
        print("Mandated Services CSD report completed")
        print("Time Taken:", end - start)

if __name__ == "__main__":
    RSCSD().RSWeeklyCSDExcel()
