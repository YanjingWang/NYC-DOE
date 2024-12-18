"""
Reports 1-4 = Initials subtotal calculation:
E + F = G, H + I = J, G + J = K, D + K + L + M = C
Reports 5-7 = Reevaluations subtotal calculation:
E + F = G, H + I = J, G + J = K, D + K + L  = C
Report 8 = Registers subtotal calculation:
C + D + E + F = G, H + I + J + K = L, G + L = M
SWDs by School : C4+ ...+ C1602 = C1603
Report 8a = Disability class subtotal calculation:
C + D +E + F + G + H + I + J + K + L + M + N + O = P 
Report 9 = Placement: C6+...+C37 = C38, D6+...+D37 = D38, C42+...C46 = C47, D42+...D46 = D47, C51+C52 = C53, D51+D52 = D53, C57+C58 = C59, D57+D58 = D59, C63+C64 = C65, 
D63+D64 = D65, C69+C72 = C73, D69+D72 = D73, C77+C89=D90, D77+D89=D90
Report 10 = LRE-MRE: C6+...C37=C38,D6+...D37=D38,E6+...E37=E38,F6+...F37=F38, C42+...C46=C47,D42+...D46=D47,E42+...E46=E47,F42+...F46=F47,C51+C52=C53,D51+D52=D53,E51+E52=E53,F51+F52=F53,E57+E58=E59,F57+F58=F59,C63+C64=C65,D63+D64=D65,E63+E64=E65,F63+F64=F65,C69+C72=C73,D69+D72=D73,E69+E72=E73,F69+F72=F73,C77+...+C89=C90,D77+...+D89=D90,E77+...+E89=E90,F77+...+F89=F90

Redaction rules: 
Column based masking:
Initial mask: find all data 0<data<=5 and make them as <=5
Then since there are Total calculations for every column, iterate each column if there is only one <=5 in this column to mask the smallest data of rest unmasked data for this column as >5. if min_val == 0 then skip it and mask the next smallest value

Row based masking:
Secondary mask: for all the 3 column subtotal calculation, if there is only one masked data (<=5 or >5) in this row, mask the smallest data of the rest unmasked data . If there are two or more masked data, no need to data masking.


Row based masking:
for all the 5 column subtotal calculation, if there is only one masked data in the row, mask the smallest data of the rest unmasked data. If there are two or more masked data, no need to do data masking.

Why always mask the smallest of rest unmasked data? Avoid overredaction. For example, in the secondary masking, if we masked E5 and G5, since G5 being to G +   = , it may trigger third masking and more data will be masked. However, if we mask E5 and F5 no need to mask G5 and it won't trigger third layer masking.

For overredaction, if there are more than two` >5 ` within the same group the same row or there are two `>5` within the same group the same row but also has at least one `<=5`  within the same group the same row and this `>5` is not the only `>5` in the column within its range it belongs to,  its' an overredaction, the first`>5` within the same group should be highlighted as green, but here my code the overredacted value has not been highlighted as green, for example, E89 needs to be highlighted as green since E89 + F89 = G89 and there are already two `>5` within the same group, then the first `>5` which is E89 needs to be highlighted as green and unmasked later. For example, check attached picture, for tab `Reports 1-4 = Initials`, D89 is the only `>5 ` of its column within its range it belongs to avoid the back track D90, D89's range is   (78, 3, 91, 13) according to redaction_config.py
For underredaction,  interate each row and for the same row if there is only one masked cell within the same group then highlight this cell, for example, H89 + I89 = J89, since there is only J89 was masked as `>5` we need to mask the smallest value within this group which is I89 here, I89 is 16 so it should be masked as `>5`.

For percentage redaction, if the number is masked as `<=5` or `>5` then the percentage should be masked as `*`.

100% percentage sum redaction, if the sum of the percentage is 100% then mask the smallest numeric and percentage within the same row. Since
Tab `Report 10 = IEP Service Recs`:  D+F+H+J+L+N = 100%
Tab `Report 14 = Programs`: D+F+H = 100%
Tab `Report 14a = Bilingual Programs`: D+F+H = 100%
Tab `Report 15 = Related Services`: D+F+H = 100%
Tab `Report 15a = Transportation`: D+F+H = 100%
Tab `Report 16 = BIP`: D+F = 100% I want to change percentage redaction rule: if there is one percentage cell is redacted as `*` and its adjacent numeric cell is masked as `<=5` or `>5` , then we redact the smallest numeric cell for the rest of the same row of that `*` redacted cell belongs to based on its value, if its value >5 then mask it as `>5` if its value <=5 then mask it as `<=5` and then mask its adjacent percentage cell. For example, check attached picture, for Tab `Report 10 = IEP Service Recs` since   D21+F21+H21+J21+L21+N21 = 100% and M21 is redacted as `<=5` and N21 is redacted as `*`, we need to mask K21 as `<=5` and its adjacent percentage cell * to avoid back track. 

For overall column in every range check, if there is only one masked cell (`<=5` or `>5`) in that column then it's underredaction, we need to mask the smallest value of the rest cell in that column.For example, in tab `Report 9 = Disability class`, N102 is the only `>5 ` of its column within its range it belongs to, in order to avoid the back track we need to redact the smallest value of this column which is N100 (0), N102's range is (100, 3, 102, 16) according to redaction_config_SY24.py


"""
"""
Author: Yanjing(Charlotte) Wang 
Date: 2023-09-29
Email: YWang36@schools.nyc.gov
"""
import openpyxl
import os, shutil
from redaction_config import REPORTS_CONFIG
from redaction_config_SY23 import REPORTS_CONFIG_SY23
from redaction_config_SY24 import REPORTS_CONFIG_SY24
class Solution:
    def __init__(self):
        self.schoolyear = 'SY24'
        self.folderpath = r'R:\\'
        self.mylocalCCfolder = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\CCUnredacted' # to store the unredacted file
        self.desktop_folder = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop' # to store the redacted file

    def copyonefile(self, src, dst):
        shutil.copy(src, dst)
        print(f'Copying one file from {src} to {dst} is complete')

    def copy_reports(self, schoolyears):
        """
        Copies annual reports for the given school years.

        :param years: List of school years to process (e.g., ['SY21', 'SY22', 'SY23', 'SY24']).
        """
        for schoolyear in schoolyears:
            src_path = os.path.join(
                self.folderpath,
                f'SEO Analytics\\Reporting\\City Council\\City Council {schoolyear}\\Annual Reports\\Non-Redacted Annual Special Education Data Report {schoolyear}.xlsx'
            )
            # Copy to the predefined folders
            self.copyonefile(src_path, self.mylocalCCfolder)
            self.copyonefile(src_path, self.desktop_folder)

    def is_percentage(self, cell):
        if isinstance(cell.value, float) and '0%' in cell.number_format:
            return True
        return False

    
    def mask_value_initial(self,val):
        if isinstance(val, (int, float)) and 0 < val <= 5:
            return '<=5'
        return val

    def mask_value_secondary(self,val):
        if isinstance(val, (int, float)) and 0 < val <= 5:
            return '<=5'
        elif isinstance(val, (int, float)) and val > 5:
            return '>5'
        return val
    
    def mask_value_zero(self,val):
        if isinstance(val, (int, float)) and val == 0:
            return '<=5'
        return val

    def initial_mask(self,ws, start_row, start_col, end_row, end_col):
        # Step 1: Initial Masking
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                if not self.is_percentage(cell):
                    cell.value = self.mask_value_initial(cell.value)
        
        # Mask smallest if only one '<=5' in column
        for col in ws.iter_cols(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            valid_vals = [cell.value for cell in col if isinstance(cell.value, (int, float)) and cell.value != '<=5']
            if sum([cell.value == '<=5' for cell in col]) == 1 and valid_vals:
                min_val = min(valid_vals)
                # if min_val == 0 then mask it as <=5 if not then mask it as >5
                if min_val == 0:
                    for cell in col:
                        if cell.value == min_val:
                            cell.value = '<=5'
                            break
                else:
                    for cell in col:
                        if cell.value == min_val:
                            cell.value = '>5'
                            break


    # Step 2: Secondary Masking
    def secondary_mask(self, ws, start_row, end_row, groups):
        # groups = [(5, 6, 7), (8, 9, 10), (7, 10, 11)]  # E,F,G and H,I,J and G,J,K respectively

        for group in groups:
            for row_num in range(start_row, end_row + 1):
                masked_vals = [col for col in group if ws.cell(row=row_num, column=col).value in ['<=5', '>5']]
                
                # If only one value in the group is masked
                if len(masked_vals) == 1:  
                    candidate_cols = []
                    for col in group:
                        col_values = [ws.cell(row=r, column=col).value for r in range(start_row, end_row + 1)]
                        if col_values.count('<=5') + col_values.count('>5') > 1:
                            # Prioritize this column if there are already masked values
                            candidate_cols.append(col)

                    if candidate_cols:
                        # If multiple candidate columns, choose the one with the smallest value in that row
                        min_val = float('inf')
                        selected_col = None
                        for col in candidate_cols:
                            cell_value = ws.cell(row=row_num, column=col).value
                            if isinstance(cell_value, (int, float)) and cell_value < min_val:
                                min_val = cell_value
                                selected_col = col

                        # Redact the cell in the selected column
                        if selected_col is not None:
                            if min_val == 0:
                                ws.cell(row=row_num, column=selected_col).value = self.mask_value_zero(min_val)
                            else:
                                ws.cell(row=row_num, column=selected_col).value = self.mask_value_secondary(min_val)

                    else:
                        # No columns with existing redacted cells, redact smallest value in the row
                        numeric_vals = [ws.cell(row=row_num, column=col).value for col in group if isinstance(ws.cell(row=row_num, column=col).value, (int, float)) and not self.is_percentage(ws.cell(row=row_num, column=col))]
                        if numeric_vals: # Ensure there are numeric values
                            min_val = min(numeric_vals)
                            for col in group:
                                if ws.cell(row=row_num, column=col).value == min_val:
                                    if min_val == 0:
                                        ws.cell(row=row_num, column=col).value = self.mask_value_zero(min_val)
                                    else:
                                        ws.cell(row=row_num, column=col).value = self.mask_value_secondary(min_val)
                                    break

                elif len(masked_vals) >= 2: # Two or more values are already masked, no action needed
                    continue



    # Step 3: Masking based on Subtotals        
    def third_mask(self,ws, start_row, end_row,groups):
        # groups = [(3,4,11,12)]  # C,D,K,L
        for group in groups:
            for row_num in range(start_row, end_row + 1):
                masked_vals = [col for col in group if ws.cell(row=row_num, column=col).value in ['<=5', '>5']]
                if len(masked_vals) == 1:  # If only one value in the group is masked, need to mask the smallest of the unmasked values
                    numeric_vals = [ws.cell(row=row_num, column=col).value for col in group if isinstance(ws.cell(row=row_num, column=col).value, (int, float)) and not self.is_percentage(ws.cell(row=row_num, column=col))]
                    if numeric_vals:  # Ensure there are numeric values
                        min_val = min(numeric_vals)
                        for col in group:
                            if ws.cell(row=row_num, column=col).value == min_val:
                                if min_val == 0:
                                    ws.cell(row=row_num, column=col).value = self.mask_value_zero(min_val)
                                else:
                                    ws.cell(row=row_num, column=col).value = self.mask_value_secondary(min_val)
                                break
                elif len(masked_vals) >= 2:  # Two or more values are already masked, so no action needed
                    continue


    # def highlight_overredaction(self, ws, start_row, end_row, groups):
    #     for group in groups:
    #         for row_num in range(start_row, end_row + 1):
    #             gt5_cells = []
    #             lte5_exists = False
    #             # Check each cell in the group for the current row
    #             for col_index in group:
    #                 cell = ws.cell(row=row_num, column=col_index)
    #                 if cell.value == '>5':
    #                     gt5_cells.append(cell)
    #                 elif cell.value == '<=5':
    #                     lte5_exists = True
    #             # Determine if overredaction rules are met
    #             if len(gt5_cells) > 2 or (len(gt5_cells) == 2 and lte5_exists):
    #                 # Highlight the first '>5' cell in green
    #                 self.green_cell(gt5_cells[0])
    #                 # Optionally, here you could unmask the green cell if needed
    #                 # gt5_cells[0].value = 'Unmasked Value'
    #                 print(ws.title, f"Overredaction: Highlighting cell {gt5_cells[0].coordinate} in green")
    # def highlight_overredaction(self, ws, groups, ranges):
    #     for group in groups:
    #         for (start_row, start_col, end_row, end_col) in ranges:  # Extracting the range from the tuple
    #             for row_num in range(start_row, end_row + 1):
    #                 gt5_cells = []
    #                 lte5_exists = False
    #                 for col_index in group:
    #                     cell = ws.cell(row=row_num, column=col_index)
    #                     if cell.value == '>5':
    #                         gt5_cells.append(cell)
    #                     elif cell.value == '<=5':
    #                         lte5_exists = True

    #                 # Check if the first '>5' is not the only one in its column within the range
    #                 if gt5_cells and (len(gt5_cells) > 2 or (len(gt5_cells) == 2 and lte5_exists)):
    #                     first_gt5_cell = gt5_cells[0]
    #                     col_values = [ws.cell(row=r, column=first_gt5_cell.column).value for r in range(start_row, end_row + 1)]
    #                     print(col_values)
    #                     if col_values.count('>5') > 1:
    #                         self.green_cell(first_gt5_cell)
    #                         print(ws.title, f"Overredaction: Highlighting cell {first_gt5_cell.coordinate} in green")
    def highlight_overredaction(self, ws, groups, ranges, unredacted_ws):
        for group in groups:
            for (start_row, start_col, end_row, end_col) in ranges:  # Extracting the range from the tuple
                for row_num in range(start_row, end_row + 1):
                    gt5_cells = []
                    lte5_exists = False
                    for col_index in group:
                        cell = ws.cell(row=row_num, column=col_index)
                        if cell.value == '>5':
                            gt5_cells.append(cell)
                        elif cell.value == '<=5':
                            lte5_exists = True

                    # Check if the first '>5' is not the only one in its column within the range
                    if gt5_cells and (len(gt5_cells) > 2 or (len(gt5_cells) == 2 and lte5_exists)):
                        first_gt5_cell = gt5_cells[0]
                        col_values = [ws.cell(row=r, column=first_gt5_cell.column).value for r in range(start_row, end_row + 1)]
                        if col_values.count('>5') > 1:
                            # self.green_cell(first_gt5_cell) # Highligh the overredaction cell in green
                            # Get the original value from the unredacted worksheet
                            original_value = unredacted_ws.cell(row=first_gt5_cell.row, column=first_gt5_cell.column).value
                            # Unmask the cell by setting its value to the original value
                            first_gt5_cell.value = original_value
                            print(ws.title, f"Unmasking overredacted cell {first_gt5_cell.coordinate} with original value")

    def highlight_underredaction(self, ws, start_row, end_row, groups):
        for group in groups:
            for row_num in range(start_row, end_row + 1):
                masked_cells = [ws.cell(row=row_num, column=col) for col in group if ws.cell(row=row_num, column=col).value in ['<=5', '>5']]
                if len(masked_cells) == 1:
                    # Find the smallest unmasked value within the group
                    unmasked_cells = [(cell, cell.value) for cell in [ws.cell(row=row_num, column=col) for col in group] if cell not in masked_cells]
                    if unmasked_cells:
                        smallest_cell = min(unmasked_cells, key=lambda x: x[1])
                        # Mask the smallest cell based on its value, if <=5, mask as <=5, otherwise mask as >5
                        if smallest_cell[1] == 0:
                            smallest_cell[0].value = self.mask_value_zero(smallest_cell[1])
                        else:
                            smallest_cell[0].value = self.mask_value_secondary(smallest_cell[1])
                        # self.yellow_cell(smallest_cell[0])  # Highlight this cell
                        print(ws.title, f"Underredaction: Highlighting cell {smallest_cell[0].coordinate} in yellow")

    def green_cell(self, cell):
        cell.fill = openpyxl.styles.PatternFill(start_color='00ff15', end_color='00ff15', fill_type='solid')
    def yellow_cell(self, cell):
        cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    def redact_percentage_based_on_number(self,ws, numeric_col, perc_col, start_row, end_row):
        for row in range(start_row, end_row + 1):
            number_cell = ws.cell(row=row, column=numeric_col)
            percentage_cell = ws.cell(row=row, column=perc_col)
            if number_cell.value in ['<=5', '>5']:
                percentage_cell.value = '*'

    def apply_percentage_redaction(self,ws, config, start_row, end_row):
        for numeric_col, perc_col in config['numeric_percentage_pairs']:
            self.redact_percentage_based_on_number(ws, numeric_col, perc_col, start_row, end_row)

    def mask_smallest_numeric_and_percentage(self, ws, start_row, end_row, groups):
        for row_num in range(start_row, end_row + 1):
            numeric_cells = []
            percentage_cells = []
            
            # Extract numeric and percentage cells from the row
            for col_index in range(1, ws.max_column + 1, 2):
                numeric_cell = ws.cell(row=row_num, column=col_index)
                percentage_cell = ws.cell(row=row_num, column=col_index + 1)
                # Ensure that only cells with numeric values are added
                if isinstance(numeric_cell.value, (int, float)):
                    numeric_cells.append(numeric_cell)
                if percentage_cell.value == '*':
                    percentage_cells.append(percentage_cell)
            
            # Only proceed if there's one masked numeric value and its adjacent percentage cell
            if len(numeric_cells) >= 1 and len(percentage_cells) == 1:
                # Filter out numeric cells that are already masked
                unmasked_numeric_cells = [cell for cell in numeric_cells if cell.value not in ['<=5', '>5']]
                if unmasked_numeric_cells:
                    # Find the smallest numeric cell by value
                    smallest_numeric_cell = min(unmasked_numeric_cells, key=lambda cell: cell.value)
                    # Mask the smallest numeric cell and its adjacent percentage cell
                    smallest_numeric_value = smallest_numeric_cell.value
                    if smallest_numeric_value <= 5:
                        smallest_numeric_cell.value = '<=5'
                    else:
                        smallest_numeric_cell.value = '>5'
                    # Mask the adjacent percentage cell
                    adjacent_percentage_cell = ws.cell(row=row_num, column=smallest_numeric_cell.column + 1)
                    adjacent_percentage_cell.value = '*'

    """
    for tab `Report 8a = SWDs by School` if there is only one redatced cell in the row of the same district, then mask the smallest value of the rest unmasked data in the same column in the same district of this redacted cell, district column is B,redatced column is C, to find the same district, the first two number of the district 
    """

    def tab8a_redact_additional_row_for_same_district(self, ws, start_row, end_row, district_column, redacted_column):
        # Dictionary to track redacted rows per district
        redacted_rows_per_district = {}

        for row in range(start_row, end_row + 1):  # Loop through the specified range
            district_value = ws.cell(row=row, column=district_column).value
            redacted_value = ws.cell(row=row, column=redacted_column).value

            # Check if redacted value is <=5 or >5
            if redacted_value in ['<=5', '>5']:
                # Extract the district number (first two characters of the district value)
                district_number = district_value[:2]
                print(f"Redacted row {row} in district {district_number}")

                # Track the redacted rows per district
                if district_number not in redacted_rows_per_district:
                    redacted_rows_per_district[district_number] = []
                redacted_rows_per_district[district_number].append(row)

        # Now redact additional rows if there's only one redacted row in a district
        for district, rows in redacted_rows_per_district.items():
            if len(rows) == 1:
                smallest_row = None
                smallest_value = float('inf')

                # Find the smallest unredacted value in the same district
                for row in range(start_row, end_row + 1):
                    district_value = ws.cell(row=row, column=district_column).value
                    redacted_value = ws.cell(row=row, column=redacted_column).value

                    if district_value.startswith(district) and isinstance(redacted_value, (int, float)):
                        if redacted_value < smallest_value:
                            smallest_value = redacted_value
                            smallest_row = row

                # Redact the smallest value found in the same district
                if smallest_row:
                    ws.cell(row=smallest_row, column=redacted_column).value = '>5' if smallest_value > 5 else '<=5'

                    print("Additional redaction completed.")
                    
    def detect_backtrackable_cells(self, ws, start_row, start_col, end_row, end_col):
        """
        Function to detect cells that can be backtracked. 
        A backtrackable cell is the only redacted cell in its column within the specified range.
        """
        backtrackable_cells = []

        # Iterate through each column within the specified range
        for col in range(start_col, end_col + 1):
            redacted_cells = [ws.cell(row=row, column=col) for row in range(start_row, end_row + 1)
                            if ws.cell(row=row, column=col).value in ['<=5', '>5']]

            # If there is only one redacted cell in this column, it's backtrackable
            if len(redacted_cells) == 1:
                backtrackable_cells.append(redacted_cells[0])

        return backtrackable_cells
    
    def detect_and_unmask_overredaction(self, ws, unredacted_ws, groups, ranges):
        """
        This function detects overredaction when:
        - There is one `<5` cell, and two or more redacted cells (`<=5` or `>5`) in its column of its range.
        - The cell next to it in the same row and the same group is also `>5` and there are two or more redacted cells in its column.
        """
        for group in groups:
            for (start_row, start_col, end_row, end_col) in ranges:  # Extracting the range from the tuple
                for row_num in range(start_row, end_row + 1):
                    for col_index in group:  # For each group of columns (E, F, G, etc.)
                        cell = ws.cell(row=row_num, column=col_index)

                        # Check if the current cell is `>5`
                        if cell.value == '>5':
                            # Get all redacted cells in the same column within the range
                            redacted_cells_in_column = [
                                ws.cell(row=r, column=col_index)
                                for r in range(start_row, end_row + 1)
                                if ws.cell(row=r, column=col_index).value in ['<=5', '>5']
                            ]
                            if len(redacted_cells_in_column) > 2:  # If there are two or more redacted cells in this column
                                # Check the adjacent cell in the same row and group
                                adjacent_cell = ws.cell(row=row_num, column=col_index + 1)  # Next cell in the group
                                if adjacent_cell.value == '>5':
                                    # Check if there are also two or more redacted cells in the adjacent cell's column
                                    redacted_cells_in_adjacent_column = [
                                        ws.cell(row=r, column=col_index + 1)
                                        for r in range(start_row, end_row + 1)
                                        if ws.cell(row=r, column=col_index + 1).value in ['<=5', '>5']
                                    ]
                                    if len(redacted_cells_in_adjacent_column) > 2:
                                        # This is an overredaction, unmask both cells
                                        original_value = unredacted_ws.cell(row=row_num, column=col_index).value
                                        cell.value = original_value
                                        print(ws.title, f"Unmasking overredacted cell {cell.coordinate} with value {original_value}")

                                        original_value_adjacent = unredacted_ws.cell(row=row_num, column=col_index + 1).value
                                        adjacent_cell.value = original_value_adjacent
                                        print(ws.title, f"Unmasking overredacted adjacent cell {adjacent_cell.coordinate} with value {original_value_adjacent}")
    
    def mask_excel_file(self,filename,tab_name,configurations):
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        try:
            ws = wb[tab_name]
        except KeyError:
            print(f"Warning: Worksheet {tab_name} does not exist in the file {filename}. Skipping...") #SWDs by School is not in SY23
            return

        # Convert string to int where possible
        for r in configurations['ranges']:
            for row in ws.iter_rows(min_row=r[0], max_row=r[2], min_col=r[1], max_col=r[3]):
                for cell in row:
                    if isinstance(cell.value, str):
                        try:
                            cell.value = int(cell.value)
                        except ValueError:
                            # If the value cannot be converted to int, keep the original value
                            pass
                        
        # Mask data for the specific ranges
        for r in configurations['ranges']:
            self.initial_mask(ws, *r)
            
        # Invoke secondary masking if present in configurations, otherwise skip
        if 'secondary_mask' in configurations and 'secondary_mask_kwargs' in configurations:
            self.secondary_mask(ws, *configurations['secondary_mask'], **configurations.get('secondary_mask_kwargs', {}))
        
        # Conditionally invoke third masking if present in configurations
        if 'third_mask' in configurations and 'third_mask_kwargs' in configurations:
            self.third_mask(ws, *configurations['third_mask'], **configurations['third_mask_kwargs'])

        if 'total_col_indexes' in configurations and 'groups' in configurations and 'ranges' in configurations:
            start_row, end_row = configurations['secondary_mask']  # Assuming secondary_mask defines the row range
            self.highlight_underredaction(ws, start_row, end_row, configurations['groups'])
        # print('highlight overredaction is done')
            
        # Apply the percentage redaction based on configuration if the key exists
        for r in configurations['ranges']:
            if 'numeric_percentage_pairs' in configurations:
                self.apply_percentage_redaction(ws, configurations, r[0], r[2])

        for r in configurations['ranges']:
            if '100_percentage_sum' in configurations:
                self.mask_smallest_numeric_and_percentage(ws, r[0], r[2], configurations.get('100_percentage_sum', []))

        # Mask underredacted columns
        for r in configurations['ranges']:
            self.check_and_mask_underredacted_columns(ws, r[0], r[2], r[1], r[3])

        # if tab_name == "Report 8a = SWDs by School" then apply yo function tab8a_redact_additional_row_for_same_district:
        if tab_name == "Report 8a = SWDs by School":
            for r in configurations['ranges']:
                self.tab8a_redact_additional_row_for_same_district(ws,r[0],r[2],r[1]-1,r[3])

        # Apply highlight_overredaction after redaction processing
        if 'groups' in configurations and 'ranges' in configurations:
            unredacted_filename = rf'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\CCUnredacted\Non-Redacted Annual Special Education Data Report {self.schoolyear}.xlsx'
            unredacted_wb = openpyxl.load_workbook(unredacted_filename, data_only=True)
            unredacted_ws = unredacted_wb[tab_name]
            self.highlight_overredaction(ws, configurations['groups'], configurations['ranges'], unredacted_ws)

        # apply detect_backtrackable_cells function
        for r in configurations['ranges']:
            backtrackable_cells = self.detect_backtrackable_cells(ws, r[0], r[1], r[2], r[3])
            for cell in backtrackable_cells:
                print(ws.title, f"Backtrackable cell found: {cell.coordinate}")
                # unmask the backtrackable cell
                cell.value = unredacted_ws[cell.coordinate].value
                print(ws.title, f"Unmasking backtrackable cell {cell.coordinate} with original value")

        # Detect and unmask overredaction
        if 'groups' in configurations and 'ranges' in configurations:
            self.detect_and_unmask_overredaction(ws, unredacted_ws, configurations['groups'], configurations['ranges'])
            
        # Save the modified workbook
        wb.save(filename) # Adjust the range if necessary
        wb.close()
    def unmask_green_cells(self, redacted_filename, unredacted_filename, tab_name):
        # Load both workbooks
        redacted_wb = openpyxl.load_workbook(redacted_filename, data_only=True)
        unredacted_wb = openpyxl.load_workbook(unredacted_filename, data_only=True)

        # Access the specific tab in both workbooks
        redacted_ws = redacted_wb[tab_name]
        unredacted_ws = unredacted_wb[tab_name]
        print(f"Unmasking green cells in {tab_name}...")
        # Iterate through the worksheet and find green cells
        for row in redacted_ws.iter_rows():
            for cell in row:
                if cell.fill.start_color.index == '00ff15':  # Check if the cell is highlighted in green
                    # Get the corresponding cell from the unredacted worksheet
                    original_cell = unredacted_ws[cell.coordinate]
                    print(f"Found green cell {cell.coordinate} with value {cell.value}")
                    # Replace the value in the redacted worksheet with the original value
                    cell.value = original_cell.value
                    print(f"Unmasking cell {cell.coordinate} with value {cell.value}")
                    # Remove the green fill
                    cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                else:
                    print("No green cells found in this worksheet")
                    continue

        # Save the modified redacted workbook
        redacted_wb.save(redacted_filename)

    # Helper function to print cell values, types, and formats
    def print_cell_formats(file_name, sheet_name, cell_range):
        wb = openpyxl.load_workbook(file_name)
        ws = wb[sheet_name]
        for row in ws[cell_range]:
            for cell in row:
                print(f"Cell {cell.coordinate} - Value: {cell.value} - Type: {type(cell.value)} - Format: {cell.number_format}")

    # Example usage
    # print_cell_formats(r'R:\SEO Analytics\Reporting\City Council\City Council SY24\Annual Reports\Non-Redacted Annual Special Education Data Report SY24.xlsx', 'Reports 5-7 = Reevaluations', 'C5:M37')



    def check_and_mask_underredacted_columns(self, ws, start_row, end_row, start_col, end_col):
        # Iterate over columns within the specified range
        for col_index in range(start_col, end_col + 1):
            column_cells = [ws.cell(row=row_index, column=col_index) for row_index in range(start_row, end_row + 1)]
            masked_cells = [cell for cell in column_cells if cell.value in ['<=5', '>5']]
            unmasked_cells = [cell for cell in column_cells if cell not in masked_cells and isinstance(cell.value, (int, float))]

            # If there is only one masked cell in the column
            if len(masked_cells) == 1 and unmasked_cells:
                # Find the smallest unmasked value
                smallest_unmasked_cell = min(unmasked_cells, key=lambda cell: cell.value)
                # Mask it based on its value
                if 0 <= smallest_unmasked_cell.value <= 5:
                    smallest_unmasked_cell.value = '<=5'
                elif smallest_unmasked_cell.value > 5:
                    smallest_unmasked_cell.value = '>5'
                # Optionally highlight the cell
                # self.yellow_cell(smallest_unmasked_cell)
                print(ws.title,f"Underredaction: Masking cell {smallest_unmasked_cell.coordinate} with {'<=5' if smallest_unmasked_cell.value == '<=5' else '>5'}")

    def process_reports(self, file_configs):
        """
        Processes each file and applies redaction configurations.

        :param filenames: List of filenames to process.
        :param configs: Corresponding configuration dictionary for each year.
        """
        for fname, config in file_configs:
            for report, report_config in config.items():
                self.mask_excel_file(fname, report, report_config)


# Call the function with your filename
if __name__ == "__main__":
    processor = Solution()
    processor.copy_reports(['SY24']) # ['SY21', 'SY22', 'SY23', 'SY24']
    
    # Define filenames and their configurations
    file_configs = [
        # ('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Annual Special Education Data Report Unredacted SY21.xlsx', REPORTS_CONFIG),
        # ('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted Annual Special Education Data Report SY22.xlsx', REPORTS_CONFIG),
        # ('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted Annual Special Education Data Report SY23.xlsx', REPORTS_CONFIG_SY23),
        ('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted Annual Special Education Data Report SY24.xlsx', REPORTS_CONFIG_SY24),
    ]

    # Process reports
    processor.process_reports(file_configs)
    