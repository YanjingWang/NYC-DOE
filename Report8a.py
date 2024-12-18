import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl
class Solution:
    # Existing code...
    # Function to format headers
    def __init__(self):
        self.sqlsnapshottableschoolyear = '24'
        self.lastrow = 1608
    def tab_8a(self):
        # 1. Connect to the SQL Server database
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # 2. Execute the stored procedure and fetch results
        params = ('CC_StudentRegisterR814_0615'+self.sqlsnapshottableschoolyear,)
        cursor.execute("EXEC [dbo].[USPCC_AnnaulReport8c] @tableName=?", params)
        results = cursor.fetchall()

        # 3. Create and format the Excel report
        # wb = Workbook()
        # ws = wb.active
        # ws.title = "Report 8 Count of Student with Disabilities by School"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Report 8a = SWDs by School")
        header_font = Font(bold=True)
        black_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=self.lastrow, min_col=1, max_col=26):  # max_row=1048576 is the last row in an Excel worksheet
            for cell in row:
                cell.fill = white_fill



        # Set the width of columns B and C
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25

        # Create solid black border
        black_border_side = Side(style='thin', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_no_bottom = Border(left=black_border_side, right=black_border_side)
        black_border_with_bottom = Border(left=black_border_side, right=black_border_side,bottom=black_border_side)
        # Add and format headers
        ws.merge_cells('B1:C1')
        ws['B1'] = "Report 8a Count of Student with Disabilities by School"
        ws['B1'].font = header_font
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')  # Centering the text
        ws['B1'].border = black_border  # Applying the black border to the merged cells
        ws['C1'].border = black_border  # Applying the black border to the merged cells

        # Adjust height for B1:C1
        ws.row_dimensions[1].height = 40

        ws['B3'] = "School DBN"
        ws['B3'].font = header_font
        ws['B3'].fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        ws['B3'].border = black_border

        ws['C3'] = "Students with IEPs"
        ws['C3'].font = header_font
        ws['C3'].fill = PatternFill(start_color="E0F0F8", end_color="E0F0F8", fill_type="solid")
        ws['C3'].border = black_border

        # 4. Write the results to the Excel report
        for idx, row in enumerate(results, start=4):  # Start from row 4 (1-based index)
            ws[f'B{idx}'] = row[0]  # Assuming first column is School DBN
            ws[f'C{idx}'] = row[1]  # Assuming second column is Students with IEPs

            # Apply black borders
            ws[f'B{idx}'].border = black_border_no_bottom
            ws[f'C{idx}'].border = black_border_no_bottom

        # Apply black borders to the last row
        ws[f'B{idx}'].border = black_border_with_bottom
        ws[f'C{idx}'].border = black_border_with_bottom

        for row in ws['C4':'C'+str(self.lastrow)]:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                if isinstance(cell.value, str):
                    try:
                        cell.value = int(cell.value)
                    except ValueError:
                        # If the value cannot be converted to int, keep the original value
                        pass
        # Function to check if a string represents a valid number
        def is_number(s):
            try:
                float(s.replace(',', ''))  # Try converting after removing commas
                return True
            except ValueError:
                return False

        # Formatting specific cell ranges
        cell_ranges = ['C4:C'+str(self.lastrow)]
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if isinstance(cell.value, str) and is_number(cell.value):
                        # Convert to float after removing commas
                        cell.value = float(cell.value.replace(',', ''))
                        # Apply number format with commas (optional)
                        cell.number_format = '#,##0'
        # Save the report to the specified path
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

        # Close the database connection
        conn.close()

if __name__ == '__main__':
    solution = Solution()
    solution.tab_8a()
        
