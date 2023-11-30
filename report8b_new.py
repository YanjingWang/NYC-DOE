import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
import pandas as pd
class Solution:
    # Existing code...
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        # ws.merge_cells(header_start_cell + ':' + chr(ord(column_letters[-1]) + 1) + str(int(header_start_cell[1:])+1))
        # Merge the header_start_cell with the cell directly below it
        ws.merge_cells(start_row=ws[header_start_cell].row,
                    start_column=ws[header_start_cell].column,
                    end_row=ws[header_start_cell].row + 1,
                    end_column=ws[header_start_cell].column)

        print('headercell:'+header_start_cell)
        print('mergecell:'+ header_start_cell + ':' +str(int(header_start_cell[1:])+1))
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            ws[col + str(int(cell_number)+1)] = '#'
            ws[col + str(int(cell_number)+1)].font = font_style
            ws[col + str(int(cell_number)+1)].border = border_style
            ws[col + str(int(cell_number)+1)].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True) 
            ws[col + str(int(cell_number)+1)].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)] = '%'
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)].font = font_style
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)].border = border_style
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)].alignment = Alignment(horizontal='center', vertical='center')
            ws[chr(ord(col) + 1) + str(int(cell_number)+1)].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            # Merge header cells for '%' and '#' under each main column
            ws.merge_cells(col + cell_number + ':' + chr(ord(col) + 1) + cell_number)
            print('col: '+col)
            print('col+cell_number: '+col + cell_number) 
            print('#: '+ col + str(int(cell_number)+1))
            print('%: '+chr(ord(col) + 1) + str(int(cell_number)+1))
            print('Sub mergecell:'+ col + cell_number + ':' + chr(ord(col) + 1) + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c) + 1) for c in column_letters]:
            ws[col + cell_number].border = border_style
            ws[col + str(int(cell_number)+1)].border = border_style



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Report 8b = IEP Service Recs"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Report 8b = IEP Service Recs")


        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Related services only', 'Special Education Teacher Support Services (SETSS)',
                'Integrated Co-Teaching Services', 'Integrated Co-Teaching Services', 'Special Class in a Community School', 
                'Special Class in a District 75 school', 'Special Class in a Non-public School Placement']
        column_letters = ['C', 'E', 'G', 'I', 'K', 'M']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        header_fill_color = "B8CCE4"
        column_fill_color = "E0F0F8"
        self.format_header(ws, 'B4', 'District', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B41', 'Ethnicity', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B51', 'Meal Status', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B58', 'Gender', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B66', 'ELL Status', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B73', 'Language of Instruction', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B82', 'Grade Level', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B101', 'Temporary Housing Status', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)
        self.format_header(ws, 'B109', 'Foster Care Status', columns, column_letters, 30, header_fill_color, column_fill_color, border_bottom_thin, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        return conn
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_race(self,conn):
        query_byRace = '''
        IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
            DROP TABLE #Report8b;

        Select a.StudentID
                ,a.Classification
                ,a.EnrolledDBN
                ,a.ReportingDistrict
                ,a.TempResFlag
                ,case when a.TempResFlag='Y' then 1 else 2
                end as TempResFlagSort
            ,b.RecommendPlacementDesc 
                ,CASE WHEN a.MRERecommendation = 'Special Class' 
                and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Residential',
                'NYS Supported Non Public School – 4201 - Day')
                Then  'SpecialClassNPS'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
                    THEN  'SpecialClassD75'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
                    THEN 'SpecialClass'
                WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
                    THEN 'RSOnly'
                WHEN a.MRERecommendation like '%SETSS%'
                    THEN 'SETSS'
                WHEN a.MRERecommendation like '%Co-%'
                    THEN 'ICT' 
                    end as IEPRecFlag
            
            ,a.[MealStatusGrouping]
            ,case when a.[EthnicityGroupCC] is null then 'Other'
            else a.[EthnicityGroupCC]
            end as [EthnicityGroupCC] 
            ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
            when a.[EthnicityGroupCC]  = 'Black' then 2
            when a.[EthnicityGroupCC]  = 'Hispanic' then 3
            when a.[EthnicityGroupCC] = 'White' then 4
            when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
            end as Ethnicity_sort 
            ,a.GradeLevel
            ,a.OutcomeLanguageCC
                ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
            when a.OutcomeLanguageCC = 'SPANISH' then 2
            when a.OutcomeLanguageCC = 'CHINESE' then 3
            when a.OutcomeLanguageCC = 'OTHER' then 4
                end as OutcomeLanguageCCSort
            ,a.Gender
            ,a.[ELLStatus]
            -- ,a.[EnrolledDBN]
            ,case when a.GradeLevel = '0K' then 1 
            when a.GradeLevel = '01' then 2
            when a.GradeLevel = '02' then 3
            when a.GradeLevel = '03' then 4
            when a.GradeLevel = '04' then 5
            when a.GradeLevel = '05' then 6
            when a.GradeLevel = '06' then 7
            when a.GradeLevel = '07' then 8
            when a.GradeLevel = '08' then 9
            when a.GradeLevel = '09' then 10
            when a.GradeLevel = '10' then 11
            when a.GradeLevel = '11' then 12
            when a.GradeLevel = '12' then 13
            end as GradeSort
            ,  case when c.studentid is not null then 'Y' else 'N'
                end as FosterCareFlag
                ,  case when c.studentid is not null then 1 else 2
                end as FosterCareFlagSort
            into #Report8b
            FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
                left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
            left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
            
        IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
            DROP TABLE #CCTotaltemp;


        select *, CASE when IEPRecFlag='SpecialClassNPS'
                    THEN 1 ELSE 0 
                            END AS SpecialClassNPS
                ,CASE when IEPRecFlag='SpecialClassD75'
                    THEN 1 ELSE 0 
                            END AS SpecialClassD75
                ,CASE when IEPRecFlag='SpecialClass'
                    THEN 1 ELSE 0 
                            END AS SpecialClass
                ,CASE when IEPRecFlag='RSOnly'
                    THEN 1 ELSE 0 
                            END AS RSOnly
                ,CASE when IEPRecFlag='SETSS'
                    THEN 1 ELSE 0 
                            END AS SETSS
                ,CASE when IEPRecFlag='ICT'
                    THEN 1 ELSE 0 
                            END AS ICT

            Into #CCTotaltemp from #Report8b
            --EthnicityGroupCC
        Select  
            EthnicityGroupCC, Ethnicity_sort
            ,cast(Sum(RSOnly) as varchar) as c1
        ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
        ,cast(sum(SETSS)as varchar) as c3
        ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
            ,cast(sum(ICT)as varchar) as c5
        ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
            ,cast(sum(SpecialClass)as varchar) as c7
        ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
            ,cast(sum(SpecialClassD75)as varchar) as c9
        ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
            ,cast(sum(SpecialClassNPS)as varchar) as c11
        ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12
        FROM #CCTotaltemp
        group by EthnicityGroupCC, Ethnicity_sort
        order by Ethnicity_sort
        '''  # the byRace SQL query goes here
        df_byRace = pd.read_sql_query(query_byRace, conn)
        conn.close()
        return df_byRace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,conn):
        query_byDistrict = '''
        IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
            DROP TABLE #Report8b;

        Select a.StudentID
                ,a.Classification
                ,a.EnrolledDBN
                ,a.ReportingDistrict
                ,a.TempResFlag
                ,case when a.TempResFlag='Y' then 1 else 2
                end as TempResFlagSort
            ,b.RecommendPlacementDesc 
                ,CASE WHEN a.MRERecommendation = 'Special Class' 
                and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Residential',
                'NYS Supported Non Public School – 4201 - Day')
                Then  'SpecialClassNPS'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
                    THEN  'SpecialClassD75'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
                    THEN 'SpecialClass'
                WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
                    THEN 'RSOnly'
                WHEN a.MRERecommendation like '%SETSS%'
                    THEN 'SETSS'
                WHEN a.MRERecommendation like '%Co-%'
                    THEN 'ICT' 
                    end as IEPRecFlag
            
            ,a.[MealStatusGrouping]
            ,case when a.[EthnicityGroupCC] is null then 'Other'
            else a.[EthnicityGroupCC]
            end as [EthnicityGroupCC] 
            ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
            when a.[EthnicityGroupCC]  = 'Black' then 2
            when a.[EthnicityGroupCC]  = 'Hispanic' then 3
            when a.[EthnicityGroupCC] = 'White' then 4
            when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
            end as Ethnicity_sort 
            ,a.GradeLevel
            ,a.OutcomeLanguageCC
                ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
            when a.OutcomeLanguageCC = 'SPANISH' then 2
            when a.OutcomeLanguageCC = 'CHINESE' then 3
            when a.OutcomeLanguageCC = 'OTHER' then 4
                end as OutcomeLanguageCCSort
            ,a.Gender
            ,a.[ELLStatus]
            -- ,a.[EnrolledDBN]
            ,case when a.GradeLevel = '0K' then 1 
            when a.GradeLevel = '01' then 2
            when a.GradeLevel = '02' then 3
            when a.GradeLevel = '03' then 4
            when a.GradeLevel = '04' then 5
            when a.GradeLevel = '05' then 6
            when a.GradeLevel = '06' then 7
            when a.GradeLevel = '07' then 8
            when a.GradeLevel = '08' then 9
            when a.GradeLevel = '09' then 10
            when a.GradeLevel = '10' then 11
            when a.GradeLevel = '11' then 12
            when a.GradeLevel = '12' then 13
            end as GradeSort
            ,  case when c.studentid is not null then 'Y' else 'N'
                end as FosterCareFlag
                ,  case when c.studentid is not null then 1 else 2
                end as FosterCareFlagSort
            into #Report8b
            FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
                left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
            left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
            
        IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
            DROP TABLE #CCTotaltemp;


        select *, CASE when IEPRecFlag='SpecialClassNPS'
                    THEN 1 ELSE 0 
                            END AS SpecialClassNPS
                ,CASE when IEPRecFlag='SpecialClassD75'
                    THEN 1 ELSE 0 
                            END AS SpecialClassD75
                ,CASE when IEPRecFlag='SpecialClass'
                    THEN 1 ELSE 0 
                            END AS SpecialClass
                ,CASE when IEPRecFlag='RSOnly'
                    THEN 1 ELSE 0 
                            END AS RSOnly
                ,CASE when IEPRecFlag='SETSS'
                    THEN 1 ELSE 0 
                            END AS SETSS
                ,CASE when IEPRecFlag='ICT'
                    THEN 1 ELSE 0 
                            END AS ICT

            Into #CCTotaltemp from #Report8b

        /*Disaggregations by Demogrphics*/

        --ReportingDistrict

        Select  ReportingDistrict
            ,cast(Sum(RSOnly) as varchar) as c1
        ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
        ,cast(sum(SETSS)as varchar) as c3
        ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
            ,cast(sum(ICT)as varchar) as c5
        ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
            ,cast(sum(SpecialClass)as varchar) as c7
        ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
            ,cast(sum(SpecialClassD75)as varchar) as c9
        ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
            ,cast(sum(SpecialClassNPS)as varchar) as c11
        ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12

        from
        #CCTotaltemp
        group by ReportingDistrict
        order by ReportingDistrict
        '''  # the byDistrict SQL query goes here
        df_byDistrict = pd.read_sql_query(query_byDistrict, conn)
        conn.close()
        return df_byDistrict

    def fetch_data_by_mealstatus(self,conn):
        query_byMealStatus = '''
        IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
            DROP TABLE #Report8b;

        Select a.StudentID
                ,a.Classification
                ,a.EnrolledDBN
                ,a.ReportingDistrict
                ,a.TempResFlag
                ,case when a.TempResFlag='Y' then 1 else 2
                end as TempResFlagSort
            ,b.RecommendPlacementDesc 
                ,CASE WHEN a.MRERecommendation = 'Special Class' 
                and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Residential',
                'NYS Supported Non Public School – 4201 - Day')
                Then  'SpecialClassNPS'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
                    THEN  'SpecialClassD75'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
                    THEN 'SpecialClass'
                WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
                    THEN 'RSOnly'
                WHEN a.MRERecommendation like '%SETSS%'
                    THEN 'SETSS'
                WHEN a.MRERecommendation like '%Co-%'
                    THEN 'ICT' 
                    end as IEPRecFlag
            
            ,a.[MealStatusGrouping]
            ,case when a.[EthnicityGroupCC] is null then 'Other'
            else a.[EthnicityGroupCC]
            end as [EthnicityGroupCC] 
            ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
            when a.[EthnicityGroupCC]  = 'Black' then 2
            when a.[EthnicityGroupCC]  = 'Hispanic' then 3
            when a.[EthnicityGroupCC] = 'White' then 4
            when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
            end as Ethnicity_sort 
            ,a.GradeLevel
            ,a.OutcomeLanguageCC
                ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
            when a.OutcomeLanguageCC = 'SPANISH' then 2
            when a.OutcomeLanguageCC = 'CHINESE' then 3
            when a.OutcomeLanguageCC = 'OTHER' then 4
                end as OutcomeLanguageCCSort
            ,a.Gender
            ,a.[ELLStatus]
            -- ,a.[EnrolledDBN]
            ,case when a.GradeLevel = '0K' then 1 
            when a.GradeLevel = '01' then 2
            when a.GradeLevel = '02' then 3
            when a.GradeLevel = '03' then 4
            when a.GradeLevel = '04' then 5
            when a.GradeLevel = '05' then 6
            when a.GradeLevel = '06' then 7
            when a.GradeLevel = '07' then 8
            when a.GradeLevel = '08' then 9
            when a.GradeLevel = '09' then 10
            when a.GradeLevel = '10' then 11
            when a.GradeLevel = '11' then 12
            when a.GradeLevel = '12' then 13
            end as GradeSort
            ,  case when c.studentid is not null then 'Y' else 'N'
                end as FosterCareFlag
                ,  case when c.studentid is not null then 1 else 2
                end as FosterCareFlagSort
            into #Report8b
            FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
                left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
            left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
            
        IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
            DROP TABLE #CCTotaltemp;


        select *, CASE when IEPRecFlag='SpecialClassNPS'
                    THEN 1 ELSE 0 
                            END AS SpecialClassNPS
                ,CASE when IEPRecFlag='SpecialClassD75'
                    THEN 1 ELSE 0 
                            END AS SpecialClassD75
                ,CASE when IEPRecFlag='SpecialClass'
                    THEN 1 ELSE 0 
                            END AS SpecialClass
                ,CASE when IEPRecFlag='RSOnly'
                    THEN 1 ELSE 0 
                            END AS RSOnly
                ,CASE when IEPRecFlag='SETSS'
                    THEN 1 ELSE 0 
                            END AS SETSS
                ,CASE when IEPRecFlag='ICT'
                    THEN 1 ELSE 0 
                            END AS ICT

            Into #CCTotaltemp from #Report8b
        --MealStatusGrouping
        Select  
            MealStatusGrouping
            ,cast(Sum(RSOnly) as varchar) as c1
        ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
        ,cast(sum(SETSS)as varchar) as c3
        ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
            ,cast(sum(ICT)as varchar) as c5
        ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
            ,cast(sum(SpecialClass)as varchar) as c7
        ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
            ,cast(sum(SpecialClassD75)as varchar) as c9
        ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
            ,cast(sum(SpecialClassNPS)as varchar) as c11
        ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12
        FROM #CCTotaltemp
        group by MealStatusGrouping
        order by MealStatusGrouping
        '''  # the byMealStatus SQL query goes here
        df_byMealStatus = pd.read_sql_query(query_byMealStatus, conn)
        conn.close()
        return df_byMealStatus
    
    def fetch_data_by_gender(self,conn):
        query_byGender = '''
        IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
            DROP TABLE #Report8b;

        Select a.StudentID
                ,a.Classification
                ,a.EnrolledDBN
                ,a.ReportingDistrict
                ,a.TempResFlag
                ,case when a.TempResFlag='Y' then 1 else 2
                end as TempResFlagSort
            ,b.RecommendPlacementDesc 
                ,CASE WHEN a.MRERecommendation = 'Special Class' 
                and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Residential',
                'NYS Supported Non Public School – 4201 - Day')
                Then  'SpecialClassNPS'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
                    THEN  'SpecialClassD75'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
                    THEN 'SpecialClass'
                WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
                    THEN 'RSOnly'
                WHEN a.MRERecommendation like '%SETSS%'
                    THEN 'SETSS'
                WHEN a.MRERecommendation like '%Co-%'
                    THEN 'ICT' 
                    end as IEPRecFlag
            
            ,a.[MealStatusGrouping]
            ,case when a.[EthnicityGroupCC] is null then 'Other'
            else a.[EthnicityGroupCC]
            end as [EthnicityGroupCC] 
            ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
            when a.[EthnicityGroupCC]  = 'Black' then 2
            when a.[EthnicityGroupCC]  = 'Hispanic' then 3
            when a.[EthnicityGroupCC] = 'White' then 4
            when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
            end as Ethnicity_sort 
            ,a.GradeLevel
            ,a.OutcomeLanguageCC
                ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
            when a.OutcomeLanguageCC = 'SPANISH' then 2
            when a.OutcomeLanguageCC = 'CHINESE' then 3
            when a.OutcomeLanguageCC = 'OTHER' then 4
                end as OutcomeLanguageCCSort
            ,a.Gender
            ,a.[ELLStatus]
            -- ,a.[EnrolledDBN]
            ,case when a.GradeLevel = '0K' then 1 
            when a.GradeLevel = '01' then 2
            when a.GradeLevel = '02' then 3
            when a.GradeLevel = '03' then 4
            when a.GradeLevel = '04' then 5
            when a.GradeLevel = '05' then 6
            when a.GradeLevel = '06' then 7
            when a.GradeLevel = '07' then 8
            when a.GradeLevel = '08' then 9
            when a.GradeLevel = '09' then 10
            when a.GradeLevel = '10' then 11
            when a.GradeLevel = '11' then 12
            when a.GradeLevel = '12' then 13
            end as GradeSort
            ,  case when c.studentid is not null then 'Y' else 'N'
                end as FosterCareFlag
                ,  case when c.studentid is not null then 1 else 2
                end as FosterCareFlagSort
            into #Report8b
            FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
                left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
            left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
            
        IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
            DROP TABLE #CCTotaltemp;


        select *, CASE when IEPRecFlag='SpecialClassNPS'
                    THEN 1 ELSE 0 
                            END AS SpecialClassNPS
                ,CASE when IEPRecFlag='SpecialClassD75'
                    THEN 1 ELSE 0 
                            END AS SpecialClassD75
                ,CASE when IEPRecFlag='SpecialClass'
                    THEN 1 ELSE 0 
                            END AS SpecialClass
                ,CASE when IEPRecFlag='RSOnly'
                    THEN 1 ELSE 0 
                            END AS RSOnly
                ,CASE when IEPRecFlag='SETSS'
                    THEN 1 ELSE 0 
                            END AS SETSS
                ,CASE when IEPRecFlag='ICT'
                    THEN 1 ELSE 0 
                            END AS ICT

            Into #CCTotaltemp from #Report8b
        --Gender
        Select  
            Gender
            ,cast(Sum(RSOnly) as varchar) as c1
        ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
        ,cast(sum(SETSS)as varchar) as c3
        ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
            ,cast(sum(ICT)as varchar) as c5
        ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
            ,cast(sum(SpecialClass)as varchar) as c7
        ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
            ,cast(sum(SpecialClassD75)as varchar) as c9
        ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
            ,cast(sum(SpecialClassNPS)as varchar) as c11
        ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12
        FROM #CCTotaltemp
        group by Gender
        order by Gender
        '''  # the byGender SQL query goes here
        df_byGender = pd.read_sql_query(query_byGender, conn)
        conn.close()
        return df_byGender
    
    def fetch_data_by_ellstatus(self,conn):
        query_byELLStatus = '''
        IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
            DROP TABLE #Report8b;

        Select a.StudentID
                ,a.Classification
                ,a.EnrolledDBN
                ,a.ReportingDistrict
                ,a.TempResFlag
                ,case when a.TempResFlag='Y' then 1 else 2
                end as TempResFlagSort
            ,b.RecommendPlacementDesc 
                ,CASE WHEN a.MRERecommendation = 'Special Class' 
                and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Residential',
                'NYS Supported Non Public School – 4201 - Day')
                Then  'SpecialClassNPS'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
                    THEN  'SpecialClassD75'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
                    THEN 'SpecialClass'
                WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
                    THEN 'RSOnly'
                WHEN a.MRERecommendation like '%SETSS%'
                    THEN 'SETSS'
                WHEN a.MRERecommendation like '%Co-%'
                    THEN 'ICT' 
                    end as IEPRecFlag
            
            ,a.[MealStatusGrouping]
            ,case when a.[EthnicityGroupCC] is null then 'Other'
            else a.[EthnicityGroupCC]
            end as [EthnicityGroupCC] 
            ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
            when a.[EthnicityGroupCC]  = 'Black' then 2
            when a.[EthnicityGroupCC]  = 'Hispanic' then 3
            when a.[EthnicityGroupCC] = 'White' then 4
            when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
            end as Ethnicity_sort 
            ,a.GradeLevel
            ,a.OutcomeLanguageCC
                ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
            when a.OutcomeLanguageCC = 'SPANISH' then 2
            when a.OutcomeLanguageCC = 'CHINESE' then 3
            when a.OutcomeLanguageCC = 'OTHER' then 4
                end as OutcomeLanguageCCSort
            ,a.Gender
            ,a.[ELLStatus]
            -- ,a.[EnrolledDBN]
            ,case when a.GradeLevel = '0K' then 1 
            when a.GradeLevel = '01' then 2
            when a.GradeLevel = '02' then 3
            when a.GradeLevel = '03' then 4
            when a.GradeLevel = '04' then 5
            when a.GradeLevel = '05' then 6
            when a.GradeLevel = '06' then 7
            when a.GradeLevel = '07' then 8
            when a.GradeLevel = '08' then 9
            when a.GradeLevel = '09' then 10
            when a.GradeLevel = '10' then 11
            when a.GradeLevel = '11' then 12
            when a.GradeLevel = '12' then 13
            end as GradeSort
            ,  case when c.studentid is not null then 'Y' else 'N'
                end as FosterCareFlag
                ,  case when c.studentid is not null then 1 else 2
                end as FosterCareFlagSort
            into #Report8b
            FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
                left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
            left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
            
        IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
            DROP TABLE #CCTotaltemp;


        select *, CASE when IEPRecFlag='SpecialClassNPS'
                    THEN 1 ELSE 0 
                            END AS SpecialClassNPS
                ,CASE when IEPRecFlag='SpecialClassD75'
                    THEN 1 ELSE 0 
                            END AS SpecialClassD75
                ,CASE when IEPRecFlag='SpecialClass'
                    THEN 1 ELSE 0 
                            END AS SpecialClass
                ,CASE when IEPRecFlag='RSOnly'
                    THEN 1 ELSE 0 
                            END AS RSOnly
                ,CASE when IEPRecFlag='SETSS'
                    THEN 1 ELSE 0 
                            END AS SETSS
                ,CASE when IEPRecFlag='ICT'
                    THEN 1 ELSE 0 
                            END AS ICT

            Into #CCTotaltemp from #Report8b
        --ELLStatus
        Select  
            ELLStatus
            ,cast(Sum(RSOnly) as varchar) as c1
        ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
        ,cast(sum(SETSS)as varchar) as c3
        ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
            ,cast(sum(ICT)as varchar) as c5
        ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
            ,cast(sum(SpecialClass)as varchar) as c7
        ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
            ,cast(sum(SpecialClassD75)as varchar) as c9
        ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
            ,cast(sum(SpecialClassNPS)as varchar) as c11
        ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12
        FROM #CCTotaltemp
        group by ELLStatus
        order by ELLStatus
        '''  # the byELLStatus SQL query goes here
        df_byELLStatus = pd.readsql(query_byELLStatus, conn)
        conn.close()
        return df_byELLStatus
    
    def fetch_data_by_language(self,conn):
        query_byLanguage = '''
        IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
            DROP TABLE #Report8b;

        Select a.StudentID
                ,a.Classification
                ,a.EnrolledDBN
                ,a.ReportingDistrict
                ,a.TempResFlag
                ,case when a.TempResFlag='Y' then 1 else 2
                end as TempResFlagSort
            ,b.RecommendPlacementDesc 
                ,CASE WHEN a.MRERecommendation = 'Special Class' 
                and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Residential',
                'NYS Supported Non Public School – 4201 - Day')
                Then  'SpecialClassNPS'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
                    THEN  'SpecialClassD75'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
                    THEN 'SpecialClass'
                WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
                    THEN 'RSOnly'
                WHEN a.MRERecommendation like '%SETSS%'
                    THEN 'SETSS'
                WHEN a.MRERecommendation like '%Co-%'
                    THEN 'ICT' 
                    end as IEPRecFlag
            
            ,a.[MealStatusGrouping]
            ,case when a.[EthnicityGroupCC] is null then 'Other'
            else a.[EthnicityGroupCC]
            end as [EthnicityGroupCC] 
            ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
            when a.[EthnicityGroupCC]  = 'Black' then 2
            when a.[EthnicityGroupCC]  = 'Hispanic' then 3
            when a.[EthnicityGroupCC] = 'White' then 4
            when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
            end as Ethnicity_sort 
            ,a.GradeLevel
            ,a.OutcomeLanguageCC
                ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
            when a.OutcomeLanguageCC = 'SPANISH' then 2
            when a.OutcomeLanguageCC = 'CHINESE' then 3
            when a.OutcomeLanguageCC = 'OTHER' then 4
                end as OutcomeLanguageCCSort
            ,a.Gender
            ,a.[ELLStatus]
            -- ,a.[EnrolledDBN]
            ,case when a.GradeLevel = '0K' then 1 
            when a.GradeLevel = '01' then 2
            when a.GradeLevel = '02' then 3
            when a.GradeLevel = '03' then 4
            when a.GradeLevel = '04' then 5
            when a.GradeLevel = '05' then 6
            when a.GradeLevel = '06' then 7
            when a.GradeLevel = '07' then 8
            when a.GradeLevel = '08' then 9
            when a.GradeLevel = '09' then 10
            when a.GradeLevel = '10' then 11
            when a.GradeLevel = '11' then 12
            when a.GradeLevel = '12' then 13
            end as GradeSort
            ,  case when c.studentid is not null then 'Y' else 'N'
                end as FosterCareFlag
                ,  case when c.studentid is not null then 1 else 2
                end as FosterCareFlagSort
            into #Report8b
            FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
                left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
            left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
            
        IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
            DROP TABLE #CCTotaltemp;


        select *, CASE when IEPRecFlag='SpecialClassNPS'
                    THEN 1 ELSE 0 
                            END AS SpecialClassNPS
                ,CASE when IEPRecFlag='SpecialClassD75'
                    THEN 1 ELSE 0 
                            END AS SpecialClassD75
                ,CASE when IEPRecFlag='SpecialClass'
                    THEN 1 ELSE 0 
                            END AS SpecialClass
                ,CASE when IEPRecFlag='RSOnly'
                    THEN 1 ELSE 0 
                            END AS RSOnly
                ,CASE when IEPRecFlag='SETSS'
                    THEN 1 ELSE 0 
                            END AS SETSS
                ,CASE when IEPRecFlag='ICT'
                    THEN 1 ELSE 0 
                            END AS ICT

            Into #CCTotaltemp from #Report8b
        --OutcomeLanguageCC
        Select  
            OutcomeLanguageCC, OutcomeLanguageCCSort
            ,cast(Sum(RSOnly) as varchar) as c1
        ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
        ,cast(sum(SETSS)as varchar) as c3
        ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
            ,cast(sum(ICT)as varchar) as c5
        ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
            ,cast(sum(SpecialClass)as varchar) as c7
        ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
            ,cast(sum(SpecialClassD75)as varchar) as c9
        ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
            ,cast(sum(SpecialClassNPS)as varchar) as c11
        ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12
        FROM #CCTotaltemp
        group by OutcomeLanguageCC,OutcomeLanguageCCSort
        order by OutcomeLanguageCCSort
        '''  # the byLanguage SQL query goes here
        df_byLanguage = pd.read_sql_query(query_byLanguage, conn)
        conn.close()
        return df_byLanguage
    
    def fetch_data_by_gradelevel(self,conn):
        query_byGradeLevel = '''
        IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
            DROP TABLE #Report8b;

        Select a.StudentID
                ,a.Classification
                ,a.EnrolledDBN
                ,a.ReportingDistrict
                ,a.TempResFlag
                ,case when a.TempResFlag='Y' then 1 else 2
                end as TempResFlagSort
            ,b.RecommendPlacementDesc 
                ,CASE WHEN a.MRERecommendation = 'Special Class' 
                and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Residential',
                'NYS Supported Non Public School – 4201 - Day')
                Then  'SpecialClassNPS'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
                    THEN  'SpecialClassD75'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
                    THEN 'SpecialClass'
                WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
                    THEN 'RSOnly'
                WHEN a.MRERecommendation like '%SETSS%'
                    THEN 'SETSS'
                WHEN a.MRERecommendation like '%Co-%'
                    THEN 'ICT' 
                    end as IEPRecFlag
            
            ,a.[MealStatusGrouping]
            ,case when a.[EthnicityGroupCC] is null then 'Other'
            else a.[EthnicityGroupCC]
            end as [EthnicityGroupCC] 
            ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
            when a.[EthnicityGroupCC]  = 'Black' then 2
            when a.[EthnicityGroupCC]  = 'Hispanic' then 3
            when a.[EthnicityGroupCC] = 'White' then 4
            when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
            end as Ethnicity_sort 
            ,a.GradeLevel
            ,a.OutcomeLanguageCC
                ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
            when a.OutcomeLanguageCC = 'SPANISH' then 2
            when a.OutcomeLanguageCC = 'CHINESE' then 3
            when a.OutcomeLanguageCC = 'OTHER' then 4
                end as OutcomeLanguageCCSort
            ,a.Gender
            ,a.[ELLStatus]
            -- ,a.[EnrolledDBN]
            ,case when a.GradeLevel = '0K' then 1 
            when a.GradeLevel = '01' then 2
            when a.GradeLevel = '02' then 3
            when a.GradeLevel = '03' then 4
            when a.GradeLevel = '04' then 5
            when a.GradeLevel = '05' then 6
            when a.GradeLevel = '06' then 7
            when a.GradeLevel = '07' then 8
            when a.GradeLevel = '08' then 9
            when a.GradeLevel = '09' then 10
            when a.GradeLevel = '10' then 11
            when a.GradeLevel = '11' then 12
            when a.GradeLevel = '12' then 13
            end as GradeSort
            ,  case when c.studentid is not null then 'Y' else 'N'
                end as FosterCareFlag
                ,  case when c.studentid is not null then 1 else 2
                end as FosterCareFlagSort
            into #Report8b
            FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
                left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
            left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
            
        IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
            DROP TABLE #CCTotaltemp;


        select *, CASE when IEPRecFlag='SpecialClassNPS'
                    THEN 1 ELSE 0 
                            END AS SpecialClassNPS
                ,CASE when IEPRecFlag='SpecialClassD75'
                    THEN 1 ELSE 0 
                            END AS SpecialClassD75
                ,CASE when IEPRecFlag='SpecialClass'
                    THEN 1 ELSE 0 
                            END AS SpecialClass
                ,CASE when IEPRecFlag='RSOnly'
                    THEN 1 ELSE 0 
                            END AS RSOnly
                ,CASE when IEPRecFlag='SETSS'
                    THEN 1 ELSE 0 
                            END AS SETSS
                ,CASE when IEPRecFlag='ICT'
                    THEN 1 ELSE 0 
                            END AS ICT

            Into #CCTotaltemp from #Report8b
        --GradeLevel
            Select  
            GradeLevel, GradeSort
            ,cast(Sum(RSOnly) as varchar) as c1
        ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
        ,cast(sum(SETSS)as varchar) as c3
        ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
            ,cast(sum(ICT)as varchar) as c5
        ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
            ,cast(sum(SpecialClass)as varchar) as c7
        ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
            ,cast(sum(SpecialClassD75)as varchar) as c9
        ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
            ,cast(sum(SpecialClassNPS)as varchar) as c11
        ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12
        FROM #CCTotaltemp
        group by GradeLevel, GradeSort
        order by GradeSort
        '''
        df_byGradeLevel = pd.read_sql_query(query_byGradeLevel, conn)
        conn.close()
        return df_byGradeLevel
    
    def fetch_data_by_tempResFlag(self,conn):
        query_byTempResFlag = '''
        IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
            DROP TABLE #Report8b;

        Select a.StudentID
                ,a.Classification
                ,a.EnrolledDBN
                ,a.ReportingDistrict
                ,a.TempResFlag
                ,case when a.TempResFlag='Y' then 1 else 2
                end as TempResFlagSort
            ,b.RecommendPlacementDesc 
                ,CASE WHEN a.MRERecommendation = 'Special Class' 
                and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Residential',
                'NYS Supported Non Public School – 4201 - Day')
                Then  'SpecialClassNPS'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
                    THEN  'SpecialClassD75'
                WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
                    THEN 'SpecialClass'
                WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
                    THEN 'RSOnly'
                WHEN a.MRERecommendation like '%SETSS%'
                    THEN 'SETSS'
                WHEN a.MRERecommendation like '%Co-%'
                    THEN 'ICT' 
                    end as IEPRecFlag
            
            ,a.[MealStatusGrouping]
            ,case when a.[EthnicityGroupCC] is null then 'Other'
            else a.[EthnicityGroupCC]
            end as [EthnicityGroupCC] 
            ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
            when a.[EthnicityGroupCC]  = 'Black' then 2
            when a.[EthnicityGroupCC]  = 'Hispanic' then 3
            when a.[EthnicityGroupCC] = 'White' then 4
            when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
            end as Ethnicity_sort 
            ,a.GradeLevel
            ,a.OutcomeLanguageCC
                ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
            when a.OutcomeLanguageCC = 'SPANISH' then 2
            when a.OutcomeLanguageCC = 'CHINESE' then 3
            when a.OutcomeLanguageCC = 'OTHER' then 4
                end as OutcomeLanguageCCSort
            ,a.Gender
            ,a.[ELLStatus]
            -- ,a.[EnrolledDBN]
            ,case when a.GradeLevel = '0K' then 1 
            when a.GradeLevel = '01' then 2
            when a.GradeLevel = '02' then 3
            when a.GradeLevel = '03' then 4
            when a.GradeLevel = '04' then 5
            when a.GradeLevel = '05' then 6
            when a.GradeLevel = '06' then 7
            when a.GradeLevel = '07' then 8
            when a.GradeLevel = '08' then 9
            when a.GradeLevel = '09' then 10
            when a.GradeLevel = '10' then 11
            when a.GradeLevel = '11' then 12
            when a.GradeLevel = '12' then 13
            end as GradeSort
            ,  case when c.studentid is not null then 'Y' else 'N'
                end as FosterCareFlag
                ,  case when c.studentid is not null then 1 else 2
                end as FosterCareFlagSort
            into #Report8b
            FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
                left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
            left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
            
        IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
            DROP TABLE #CCTotaltemp;


        select *, CASE when IEPRecFlag='SpecialClassNPS'
                    THEN 1 ELSE 0 
                            END AS SpecialClassNPS
                ,CASE when IEPRecFlag='SpecialClassD75'
                    THEN 1 ELSE 0 
                            END AS SpecialClassD75
                ,CASE when IEPRecFlag='SpecialClass'
                    THEN 1 ELSE 0 
                            END AS SpecialClass
                ,CASE when IEPRecFlag='RSOnly'
                    THEN 1 ELSE 0 
                            END AS RSOnly
                ,CASE when IEPRecFlag='SETSS'
                    THEN 1 ELSE 0 
                            END AS SETSS
                ,CASE when IEPRecFlag='ICT'
                    THEN 1 ELSE 0 
                            END AS ICT

            Into #CCTotaltemp from #Report8b
        --TempResFlag
        Select  
            TempResFlag ,TempResFlagSort
            ,cast(Sum(RSOnly) as varchar) as c1
        ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
        ,cast(sum(SETSS)as varchar) as c3
        ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
            ,cast(sum(ICT)as varchar) as c5
        ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
            ,cast(sum(SpecialClass)as varchar) as c7
        ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
            ,cast(sum(SpecialClassD75)as varchar) as c9
        ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
            ,cast(sum(SpecialClassNPS)as varchar) as c11
        ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12
        FROM #CCTotaltemp
        group by TempResFlag, TempResFlagSort
        order by TempResFlagSort
        '''
        df_byTempResFlag = pd.read_sql_query(query_byTempResFlag, conn)
        conn.close()
        return df_byTempResFlag
    
    def fetch_data_by_fosterCareStatus(self,conn):
        query_byFosterCareStatus = '''
IF OBJECT_ID('tempdb..#Report8b') IS NOT NULL
	DROP TABLE #Report8b;

Select a.StudentID
		,a.Classification
		,a.EnrolledDBN
	    ,a.ReportingDistrict
		,a.TempResFlag
		,case when a.TempResFlag='Y' then 1 else 2
		end as TempResFlagSort
	  ,b.RecommendPlacementDesc 
	  	,CASE WHEN a.MRERecommendation = 'Special Class' 
		and RecommendPlacementDesc in ('NYSED-Approved Non Public School - Day',
		'NYSED-Approved Non Public School - Residential',
		'NYS Supported Non Public School – 4201 - Day')
          Then  'SpecialClassNPS'
		 WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict='75' 
             THEN  'SpecialClassD75'
		 WHEN a.MRERecommendation = 'Special Class' and a.AdminDistrict<>'75' 
             THEN 'SpecialClass'
		 WHEN a.MRERecommendation in ('RS Only', 'Other Program', 'No Services')
             THEN 'RSOnly'
		 WHEN a.MRERecommendation like '%SETSS%'
             THEN 'SETSS'
		 WHEN a.MRERecommendation like '%Co-%'
             THEN 'ICT' 
			end as IEPRecFlag
	   
	   ,a.[MealStatusGrouping]
	   ,case when a.[EthnicityGroupCC] is null then 'Other'
	   else a.[EthnicityGroupCC]
	   end as [EthnicityGroupCC] 
	   ,case when a.[EthnicityGroupCC]  = 'Asian' then 1
	   when a.[EthnicityGroupCC]  = 'Black' then 2
	   when a.[EthnicityGroupCC]  = 'Hispanic' then 3
	   when a.[EthnicityGroupCC] = 'White' then 4
	   when (a.[EthnicityGroupCC] is Null or a.[EthnicityGroupCC] = 'Other') then 5
	   end as Ethnicity_sort 
	   ,a.GradeLevel
	   ,a.OutcomeLanguageCC
	     ,case when (a.OutcomeLanguageCC = 'English' OR a.OutcomeLanguageCC IS NULL) then 1
	   when a.OutcomeLanguageCC = 'SPANISH' then 2
	   when a.OutcomeLanguageCC = 'CHINESE' then 3
	   when a.OutcomeLanguageCC = 'OTHER' then 4
	    end as OutcomeLanguageCCSort
	   ,a.Gender
	   ,a.[ELLStatus]
	 -- ,a.[EnrolledDBN]
	   ,case when a.GradeLevel = '0K' then 1 
	   when a.GradeLevel = '01' then 2
	   when a.GradeLevel = '02' then 3
	   when a.GradeLevel = '03' then 4
	   when a.GradeLevel = '04' then 5
	   when a.GradeLevel = '05' then 6
	   when a.GradeLevel = '06' then 7
	   when a.GradeLevel = '07' then 8
	   when a.GradeLevel = '08' then 9
	   when a.GradeLevel = '09' then 10
	   when a.GradeLevel = '10' then 11
	   when a.GradeLevel = '11' then 12
	   when a.GradeLevel = '12' then 13
	   end as GradeSort
	   ,  case when c.studentid is not null then 'Y' else 'N'
		end as FosterCareFlag
		,  case when c.studentid is not null then 1 else 2
		end as FosterCareFlagSort
	   into #Report8b
      FROM [SEO_MART].[snap].[CC_StudentRegisterR814_061523] a
	    left join [SEO_MART].[snap].[RPT_StudentRegister_061523] b on a.studentid=b.studentid
	   left join SEO_Mart.dbo.lk_FosterCare c on a.studentid=c.studentid
	 
IF OBJECT_ID('tempdb..#CCTotaltemp') IS NOT NULL
	DROP TABLE #CCTotaltemp;


select *, CASE when IEPRecFlag='SpecialClassNPS'
             THEN 1 ELSE 0 
                     END AS SpecialClassNPS
		,CASE when IEPRecFlag='SpecialClassD75'
             THEN 1 ELSE 0 
                     END AS SpecialClassD75
		,CASE when IEPRecFlag='SpecialClass'
             THEN 1 ELSE 0 
                     END AS SpecialClass
		,CASE when IEPRecFlag='RSOnly'
             THEN 1 ELSE 0 
                     END AS RSOnly
		,CASE when IEPRecFlag='SETSS'
             THEN 1 ELSE 0 
                     END AS SETSS
		,CASE when IEPRecFlag='ICT'
             THEN 1 ELSE 0 
                     END AS ICT

	Into #CCTotaltemp from #Report8b
 --FostercareFlag

 Select  
      FostercareFlag, FostercareFlagsort
  ,cast(Sum(RSOnly) as varchar) as c1
  ,cast(Sum(RSOnly)*1.0/nullif(Count(studentid),0) as varchar) as c2
  ,cast(sum(SETSS)as varchar) as c3
  ,cast(Sum(SETSS)*1.0/nullif(Count(studentid),0) as varchar) as c4
    ,cast(sum(ICT)as varchar) as c5
  ,cast(Sum(ICT)*1.0/nullif(Count(studentid),0) as varchar) as c6
    ,cast(sum(SpecialClass)as varchar) as c7
  ,cast(Sum(SpecialClass)*1.0/nullif(Count(studentid),0) as varchar) as c8
      ,cast(sum(SpecialClassD75)as varchar) as c9
  ,cast(Sum(SpecialClassD75)*1.0/nullif(Count(studentid),0) as varchar) as c10
      ,cast(sum(SpecialClassNPS)as varchar) as c11
  ,cast(Sum(SpecialClassNPS)*1.0/nullif(Count(studentid),0) as varchar) as c12
 FROM #CCTotaltemp
 group by FostercareFlag, FostercareFlagsort
 order by FostercareFlagsort
        '''
        df_byFosterCareStatus = pd.read_sql_query(query_byFosterCareStatus, conn)
        conn.close()
        return df_byFosterCareStatus
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel
        for row_num, row_data in enumerate(data, start=start_row):
            # Check if the cell is a merged cell and skip if it is
            if isinstance(ws.cell(row=row_num, column=2), openpyxl.cell.cell.MergedCell):
                continue

            ws.cell(row=row_num, column=2).value = row_data[0]  # Assuming 'B' column is not merged for data rows
            col_pointer = 3  # Start at column 'C'

            for i in range(1, len(row_data), 2):
                # Check for each cell if it's a MergedCell before writing the value
                if not isinstance(ws.cell(row=row_num, column=col_pointer), openpyxl.cell.cell.MergedCell):
                    ws.cell(row=row_num, column=col_pointer).value = row_data[i]  # Number data
                col_pointer += 1

                if not isinstance(ws.cell(row=row_num, column=col_pointer), openpyxl.cell.cell.MergedCell):
                    ws.cell(row=row_num, column=col_pointer).value = row_data[i+1]  # Percentage data
                col_pointer += 1
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M','N']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C6':'N38']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C43':'N48']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C53':'N55']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C60':'N63']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C68':'N70']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C75':'N79']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C84':'N97']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C103':'N105']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C111':'N113']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['B1': 'N1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B38':'N38'] + ws['B48':'N48'] + ws['B55':'N55'] + ws['B63':'N63'] + ws['B70':'N70'] + ws['B79':'N79'] + ws['B97':'N97'] + ws['B105':'N105'] + ws['B113':'N113']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['B3':'N3'] + ws['B40':'N40'] + ws['B50':'N50'] + ws['B57':'N57'] + ws['B65':'N65'] + ws['B72':'N72'] + ws['B81':'N81'] + ws['B100':'N100'] + ws['B108':'N108']:
            for cell in row:
                cell.border = black_border_thick
                cell.font = Font(bold=True, size=12)

        for row in ws['B4': 'B5'] + ws['C4':'D4'] + ws['E4':'F4'] + ws['G4':'H4'] + ws['I4':'J4'] + ws['K4':'L4'] + ws['M4':'N4'] + ws['C5':'C5'] + ws['D5':'D5'] + ws['E5':'E5'] + ws['F5':'F5'] + ws['G5':'G5'] + ws['H5':'H5'] + ws['I5':'I5'] + ws['J5':'J5'] + ws['K5':'K5'] + ws['L5':'L5'] + ws['M5':'M5'] + ws['N5':'N5'] + ws ['B41':'B42'] + ws['C41':'D41'] + ws['E41':'F41'] + ws['G41':'H41'] + ws['I41':'J41'] + ws['K41':'L41'] + ws['M41':'N41'] + ws['C42': 'C42'] + ws['D42':'D42'] + ws['E42':'E42'] + ws['F42':'F42'] + ws['G42':'G42'] + ws['H42':'H42'] + ws['I42':'I42'] + ws['J42':'J42'] + ws['K42':'K42'] + ws['L42':'L42'] + ws['M42':'M42'] + ws['N42':'N42'] + ws['B51':'B52'] + ws['C51':'D51'] + ws['E51':'F51'] + ws['G51':'H51'] + ws['I51':'J51'] + ws['K51':'L51'] + ws['M51':'N51'] + ws['C52':'C52'] + ws['D52':'D52'] + ws['E52':'E52'] + ws['F52':'F52'] + ws['G52':'G52'] + ws['H52':'H52'] + ws['I52':'I52'] + ws['J52':'J52'] + ws['K52':'K52'] + ws['L52':'L52'] + ws['M52':'M52'] + ws['N52':'N52'] + ws['B58':'B59'] + ws['C58':'D58'] + ws['E58':'F58'] + ws['G58':'H58'] + ws['I58':'J58'] + ws['K58':'L58'] + ws['M58':'N58'] + ws['C59':'C59'] + ws['D59':'D59'] + ws['E59':'E59'] + ws['F59':'F59'] + ws['G59':'G59'] + ws['H59':'H59'] + ws['I59':'I59'] + ws['J59':'J59'] + ws['K59':'K59'] + ws['L59':'L59'] + ws['M59':'M59'] + ws['N59':'N59'] + ws['B66':'B67'] + ws['C66':'D66'] + ws['E66':'F66'] + ws['G66':'H66'] + ws['I66':'J66'] + ws['K66':'L66'] + ws['M66':'N66'] + ws['C67':'C67'] + ws['D67':'D67'] + ws['E67':'E67'] + ws['F67':'F67'] + ws['G67':'G67'] + ws['H67':'H67'] + ws['I67':'I67'] + ws['J67':'J67'] + ws['K67':'K67'] + ws['L67':'L67'] + ws['M67':'M67'] + ws['N67':'N67'] + ws['B73':'B74'] + ws['C73':'D73'] + ws['E73':'F73'] + ws['G73':'H73'] + ws['I73':'J73'] + ws['K73':'L73'] + ws['M73':'N73'] + ws['C74':'C74'] + ws['D74':'D74'] + ws['E74':'E74'] + ws['F74':'F74'] + ws['G74':'G74'] + ws['H74':'H74'] + ws['I74':'I74'] + ws['J74':'J74'] + ws['K74':'K74'] + ws['L74':'L74'] + ws['M74':'M74'] + ws['N74':'N74'] + ws['B82':'B83'] + ws['C82':'D82'] + ws['E82':'F82'] + ws['G82':'H82'] + ws['I82':'J82'] + ws['K82':'L82'] + ws['M82':'N82'] + ws['C83':'C83'] + ws['D83':'D83'] + ws['E83':'E83'] + ws['F83':'F83'] + ws['G83':'G83'] + ws['H83':'H83'] + ws['I83':'I83'] + ws['J83':'J83'] + ws['K83':'K83'] + ws['L83':'L83'] + ws['M83':'M83'] + ws['N83':'N83'] + ws['B101':'B102'] + ws['C101':'D101'] + ws['E101':'F101'] + ws['G101':'H101'] + ws['I101':'J101'] + ws['K101':'L101'] + ws['M101':'N101'] + ws['C102':'C102'] + ws['D102':'D102'] + ws['E102':'E102'] + ws['F102':'F102'] + ws['G102':'G102'] + ws['H102':'H102'] + ws['I102':'I102'] + ws['J102':'J102'] + ws['K102':'K102'] + ws['L102':'L102'] + ws['M102':'M102'] + ws['N102':'N102'] + ws['B109':'B110'] + ws['C109':'D109'] + ws['E109':'F109'] + ws['G109':'H109'] + ws['I109':'J109'] + ws['K109':'L109'] + ws['M109':'N109'] + ws['C110':'C110'] + ws['D110':'D110'] + ws['E110':'E110'] + ws['F110':'F110'] + ws['G110':'G110'] + ws['H110':'H110'] + ws['I110':'I110'] + ws['J110':'J110'] + ws['K110':'K110'] + ws['L110':'L110'] + ws['M110':'M110'] + ws['N110':'N110'] :
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)
    def main_report8b(self):
        title_cells = [
            {"cell": "B1", "value": "Report 8b IEP Service Recommendations Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; and Grade Level.", "merge_cells": "B1:L1"},
            

        ]

        subtitle_cells = [
            {"cell": "B3", "value": "SY 2022-23 Students with IEP Recommended Services by District", "merge_cells": "B3:N3"},
            {"cell": "B40", "value": "SY 2022-23 Students with IEP Recommended Services by Ethnicity", "merge_cells": "B40:N40"},
            {"cell": "B50", "value": "SY 2022-23 Students with IEP Recommended Services by Meal Status", "merge_cells": "B50:N50"},
            {"cell": "B57", "value": "SY 2022-23 Students with IEP Recommended Services by Gender", "merge_cells": "B57:N57"},
            {"cell": "B65", "value": "SY 2022-23 Students with IEP Recommended Services by ELL Status", "merge_cells": "B65:N65"},
            {"cell": "B72", "value": "SY 2022-23 Students with IEP Recommended Services  by Recommended Language of Instruction", "merge_cells": "B72:N72"},
            {"cell": "B81", "value": "SY 2022-23 Students with IEP Recommended Services  by Grade Level", "merge_cells": "B81:N81"},
            {"cell": "B100", "value": "SY 2022-23 Students with IEP Recommended Services  by Temporary Housing", "merge_cells": "B100:N100"},
            {"cell": "B108", "value": "SY 2022-23 Students with IEP Recommended Services  by Foster Care Status", "merge_cells": "B108:N108"},
            

        ]

        column_widths = [5, 30, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_byRace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byRace, start_row=43)
        
        # Step 4: Fetch and write data for "Report 8b = IEP Service Recs by District"
        results_byDistrict = self.fetch_data_by_district(cursor)
        # replace 01 as 1, 02 as 2, etc.
        results_byDistrict = [(x[0].replace('01', '1'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('02', '2'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('03', '3'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('04', '4'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('05', '5'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('06', '6'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('07', '7'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('08', '8'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('09', '9'), *x[1:]) for x in results_byDistrict]
        self.write_data_to_excel(ws, results_byDistrict, start_row=6)

        # Step 5: Fetch and write data for "Report 8b = IEP Service Recs by Meal Status"
        results_byMealStatus = self.fetch_data_by_mealstatus(cursor)
        self.write_data_to_excel(ws, results_byMealStatus, start_row=53)

        # Step 6: Fetch and write data for "Report 8b = IEP Service Recs by Gender"
        results_byMealStatus = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_byMealStatus, start_row=60)

        # Step 7: Fetch and write data for "Report 8b = IEP Service Recs by ELL Status"
        results_byELLStatus = self.fetch_data_by_ellstatus(cursor)
        self.write_data_to_excel(ws, results_byELLStatus, start_row=68)
        
        # Step 8: Fetch and write data for "Report 8b = IEP Service Recs by Language"
        results_byLanguage = self.fetch_data_by_language(cursor)
        self.write_data_to_excel(ws, results_byLanguage, start_row=75)

        # Step 9: Fetch and write data for "Report 8b = IEP Service Recs by Grade Level"
        results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        self.write_data_to_excel(ws, results_byGradeLevel, start_row=84)

        # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Temporary Housing"
        results_byTempResFlag = self.fetch_data_by_tempResFlag(cursor)
        self.write_data_to_excel(ws, results_byTempResFlag, start_row=103)

        # Step 11: Fetch and write data for "Report 8b = IEP Service Recs by Foster Care Status"
        results_byFosterCareStatus = self.fetch_data_by_fosterCareStatus(cursor)
        self.write_data_to_excel(ws, results_byFosterCareStatus, start_row=111)
        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.main_report8b()                                                                  