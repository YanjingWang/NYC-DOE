import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import pyodbc


def set_fill_color(sheet, start_row, end_row, start_col, end_col, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.fill = fill


def set_border(sheet, range_list, border_style):
    for cell_range in range_list:
        for cell in sheet[cell_range]:
            cell.border = border_style


def set_columns_width(sheet, columns_width):
    for col, width in columns_width.items():
        sheet.column_dimensions[col].width = width


def connect_to_db(conn_str):
    return pyodbc.connect(conn_str)


def fetch_data_from_db(cursor, query):
    cursor.execute(query)
    return cursor.fetchall()


# Step 1: Create Excel Report Template
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Report 8b = IEP Service Recs"

# Setting white fill for entire sheet
set_fill_color(ws, 1, 120, 1, 26, "FFFFFF")

# Defining Border Styles
black_border_side = Side(style='thin', color='000000')
black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)

black_border_thickside = Side(style='thick', color='000000')
black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)

# Applying borders for report titles
set_border(ws, ['B1:L1', 'B3:N3'], black_border_thick)

# Setting Title and Subtitle
titles = {
    'B1': 'Report 8b IEP Service Recommendations Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; and Grade Level.',
    'B3': 'SY 2022-23 Students with IEP Recommended Services by District'
}
for cell, title in titles.items():
    ws[cell] = title
    ws[cell].font = Font(bold=True, size=12)
    ws[cell].alignment = Alignment(wrap_text=True)

# Adjusting column widths
column_widths = {
    'A': 5, 'B': 30, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 15,
    'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15, 'M': 15, 'N': 15
}
set_columns_width(ws, column_widths)

# Header setup
header_setup = [
    ('B4', 'District', 'B8CCE4', 'B4:B5'),
    ('C4', 'Related services only', 'E0F0F8', 'C4:D4'),
    ('E4', 'Special Education Teacher Support Services (SETSS)', 'E0F0F8', 'E4:F4'),
    ('G4', 'Integrated Co-Teaching Services', 'E0F0F8', 'G4:H4'),
    ('I4', 'Special Class in a Community School', 'E0F0F8', 'I4:J4'),
    ('K4', 'Special Class in a District 75 school', 'E0F0F8', 'K4:L4'),
    ('M4', 'Special Class in a Non-public School Placement', 'E0F0F8', 'M4:N4')
]
for cell, title, color, merge_range in header_setup:
    ws[cell] = title
    ws[cell].font = Font(bold=True, size=12)
    ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws.merge_cells(merge_range)

sub_headers = [
    ('C', '#', 'E0F0F8'),
    ('D', '%', 'E0F0F8'),
    ('E', '#', 'E0F0F8'),
    ('F', '%', 'E0F0F8'),
    ('G', '#', 'E0F0F8'),
    ('H', '%', 'E0F0F8'),
    ('I', '#', 'E0F0F8'),
    ('J', '%', 'E0F0F8'),
    ('K', '#', 'E0F0F8'),
    ('L', '%', 'E0F0F8'),
    ('M', '#', 'E0F0F8'),
    ('N', '%', 'E0F0F8')
]
for col, title, color in sub_headers:
    cell = f'{col}5'
    ws[cell] = title
    ws[cell].font = Font(bold=True, size=12)
    ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws.column_dimensions[col].width = 8

# Save workbook
wb.save("Report8b_byDistrict.xlsx")

# Step 2: Query Data from Database
# Connect to the Database
connection_string = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING'
connection = connect_to_db(connection_string)
cursor = connection.cursor()

# Execute query
data_query = '''
Select  ReportingDistrict,FORMAT(sum(RSOnly), '#,##0') as c1 ,
CONCAT(cast((sum(RSOnly)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c2 ,
FORMAT(sum(SETSS), '#,##0') as c3 ,
CONCAT(cast((sum(SETSS)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c4 ,
FORMAT(sum(ICT), '#,##0') as c5 ,
CONCAT(cast((sum(ICT)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c6 ,
FORMAT(sum(SpecialClass), '#,##0') as c7 ,
CONCAT(cast((sum(SpecialClass)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c8 ,
FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,
CONCAT(cast((sum(SpecialClassD75)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c10 ,
FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,
CONCAT(cast((sum(SpecialClassNPS)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c12  
FROM ##Report group by ReportingDistrict 
UNION ALL  
SELECT 'Total' as ReportingDistrict, FORMAT(SUM(RSOnly), '#,##0') as c1 , 
CONCAT(CAST((SUM(RSOnly)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c2 ,
FORMAT(SUM(SETSS), '#,##0') as c3 , 
CONCAT(CAST((SUM(SETSS)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c4 , 
FORMAT(SUM(ICT), '#,##0') as c5 , 
CONCAT(CAST((SUM(ICT)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c6 , 
FORMAT(SUM(SpecialClass), '#,##0') as c7 , 
CONCAT(CAST((SUM(SpecialClass)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c8 , 
FORMAT(SUM(SpecialClassD75), '#,##0') as c9 , 
CONCAT(CAST((SUM(SpecialClassD75)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c10 , 
FORMAT(SUM(SpecialClassNPS), '#,##0') as c11 , 
CONCAT(CAST((SUM(SpecialClassNPS)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c12 
FROM ##Report order by ReportingDistrict
'''
data = fetch_data_from_db(cursor, data_query)

# Step 3: Write data to Excel
# This is just an example. You'd have to integrate this step with your query data.
# rows = [
#     ("District 1", 100, 60, 80, 40, 70, 50, 60, 30, 50, 20, 40, 10),
#     ("District 2", 90, 55, 70, 35, 60, 45, 50, 25, 45, 15, 35, 5)
# ]
for index, row in enumerate(data, 6):  # Adjust this as per your needs
    for j, val in enumerate(row, 2):
        ws.cell(row=index, column=j).value = val
        ws.cell(row=index, column=j).border = black_border

# Save workbook
wb.save("Report8b_byDistrict_withData.xlsx")
