import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
import pyodbc

# Step 1: Create Excel Report Template

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Report 8b = IEP Service Recs"

# Set fill color for cells from A1 to Zn to white
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):  # max_row=1048576 is the last row in an Excel worksheet
    for cell in row:
        cell.fill = white_fill

# Create solid black border
black_border_side = Side(style='thin', color='000000')
black_border_thickside = Side(style='thick', color='000000')
black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
# Add report title and merge cells
ws['B1'] = 'Report 8b IEP Service Recommendations Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; and Grade Level.'
ws.merge_cells('B1:L1')  # Merge title cells
ws['B1'].border = black_border  # Applying the black border to the merged cells
ws['C1'].border = black_border  
ws['D1'].border = black_border  
ws['E1'].border = black_border  
ws['F1'].border = black_border  
ws['G1'].border = black_border  
ws['H1'].border = black_border  
ws['I1'].border = black_border  
ws['J1'].border = black_border  
ws['K1'].border = black_border  
ws['L1'].border = black_border  

ws['B3'] = 'SY 2022-23 Students with IEP Recommended Services by District'
ws.merge_cells('B3:N3')  # Merge subtitle cells
ws['B3'].border = black_border_thick  # Applying the black border to the merged cells
ws['C3'].border = black_border_thick  
ws['D3'].border = black_border_thick  
ws['E3'].border = black_border_thick  
ws['F3'].border = black_border_thick  
ws['G3'].border = black_border_thick  
ws['H3'].border = black_border_thick  
ws['I3'].border = black_border_thick  
ws['J3'].border = black_border_thick  
ws['K3'].border = black_border_thick  
ws['L3'].border = black_border_thick  
ws['M3'].border = black_border_thick  
ws['N3'].border = black_border_thick  


# Style and align the merged title and subtitle
for cell in ['B1', 'B3']:
    ws[cell].font = Font(bold=True, size=12)
    ws[cell].alignment = Alignment(wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15
ws.column_dimensions['L'].width = 15
ws.column_dimensions['M'].width = 15
ws.column_dimensions['N'].width = 15


# Header
header_font = Font(bold=True, size=12)
border_bottom = Border(bottom=Side(style='thin'))

ws['B4'] = 'District'
ws['B4'].font = header_font
ws['B4'].border = border_bottom
ws['B4'].alignment = Alignment(horizontal='center', vertical='center')  
ws['B4'].fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
ws.row_dimensions[4].height = 30
ws.merge_cells('B4:B5')

# More header formatting based on the attached image
columns = ['Related services only', 'Special Education Teacher Support Services (SETSS)',
           'Integrated Co-Teaching Services', 'Special Class in a Community School', 
           'Special Class in a District 75 school', 'Special Class in a Non-public School Placement']
column_letters = ['C', 'E', 'G', 'I', 'K', 'M']


for col, title in zip(column_letters, columns):
    ws[col + '4'] = title
    ws[col + '4'].font = header_font
    ws[col + '4'].border = border_bottom
    ws[col + '4'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws[col + '4'].fill = PatternFill(start_color="E0F0F8", end_color="E0F0F8", fill_type="solid")
    ws[col + '5'] = '#'
    ws[col + '5'].font = header_font
    ws[col + '5'].border = border_bottom
    ws[col + '5'].alignment = Alignment(horizontal='center', vertical='center') 
    ws[chr(ord(col) + 1) + '5'] = '%'
    ws[chr(ord(col) + 1) + '5'].font = header_font
    ws[chr(ord(col) + 1) + '5'].border = border_bottom
    ws[chr(ord(col) + 1) + '5'].alignment = Alignment(horizontal='center', vertical='center')
    ws[chr(ord(col) + 1) + '5'].fill = PatternFill(start_color="E0F0F8", end_color="E0F0F8", fill_type="solid")

# Merge header cells for '%' and '#' under each main column
for col in column_letters:
    ws.merge_cells(f'{col}4:{chr(ord(col) + 1)}4')



# # Apply borders to 'District', 'Related services only', 'Special Education Teacher Support Services (SETSS)','Integrated Co-Teaching Services', 'Special Class in a Community School', 'Special Class in a District 75 school', 'Special Class in a Non-public School Placement', # and % cells for consistency
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L','M','N']:
    ws[col + '4'].border = black_border
    ws[col + '5'].border = black_border


# Step 2: Connect to the database and fetch data

conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()
params = ('CC_StudentRegisterR814_0615', 'RPT_StudentRegister_0615')
cursor.execute("EXEC [dev].[USPCCAnnaulReport8b] @tableNameCCStudentRegisterR814=?, @tableNameRptStudentRegister0615=?", params)
query_byDistrict = """
Select  ReportingDistrict,FORMAT(sum(RSOnly), '#,##0') as c1 ,
CONCAT(cast((sum(RSOnly)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c2 ,
FORMAT(sum(SETSS), '#,##0') as c3 ,CONCAT(cast((sum(SETSS)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c4 ,
FORMAT(sum(ICT), '#,##0') as c5 ,CONCAT(cast((sum(ICT)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c6 ,
FORMAT(sum(SpecialClass), '#,##0') as c7 ,
CONCAT(cast((sum(SpecialClass)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c8 ,FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,
CONCAT(cast((sum(SpecialClassD75)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c10 ,FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,
CONCAT(cast((sum(SpecialClassNPS)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c12  FROM ##Report group by ReportingDistrict 
UNION ALL  SELECT 'Total' as ReportingDistrict, FORMAT(SUM(RSOnly), '#,##0') as c1 , 
CONCAT(CAST((SUM(RSOnly)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c2 , 
FORMAT(SUM(SETSS), '#,##0') as c3 , CONCAT(CAST((SUM(SETSS)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c4 , 
FORMAT(SUM(ICT), '#,##0') as c5 , CONCAT(CAST((SUM(ICT)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c6 , FORMAT(SUM(SpecialClass), '#,##0') as c7 , 
CONCAT(CAST((SUM(SpecialClass)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c8 , FORMAT(SUM(SpecialClassD75), '#,##0') as c9 , 
CONCAT(CAST((SUM(SpecialClassD75)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c10 , FORMAT(SUM(SpecialClassNPS), '#,##0') as c11 , 
CONCAT(CAST((SUM(SpecialClassNPS)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c12 FROM ##Report order by ReportingDistrict
"""
cursor.execute(query_byDistrict)
results = cursor.fetchall()
# cursor.execute("""
#     EXEC [dev].[USPCCAnnaulReport8b] @tableNameCCStudentRegisterR814 = 'CC_StudentRegisterR814_0615', 
#     @tableNameRptStudentRegister0615='RPT_StudentRegister_0615'
# """)

# Step 3: Write data to Excel

# for row_num, row_data in enumerate(results, start=6):
#     for col_num, value in enumerate(row_data, start=2):
#         ws.cell(row=row_num, column=col_num).value = value

for row_num, row_data in enumerate(results, start=6):
    ws.cell(row=row_num, column=2).value = row_data[0] # ReportingDistrict starts at column 'B'
    col_pointer = 3 # Start at column 'C'
    for i in range(1, len(row_data), 2):
        ws.cell(row=row_num, column=col_pointer).value = row_data[i] # Number data
        col_pointer += 1
        ws.cell(row=row_num, column=col_pointer).value = row_data[i+1] # Percentage data
        col_pointer += 1

# Apply black_border_no_bottom to all columns
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M','N']:
    ws[col + '6'].border = black_border_no_bottom
    ws[col + '7'].border = black_border_no_bottom
    ws[col + '8'].border = black_border_no_bottom
    ws[col + '9'].border = black_border_no_bottom
    ws[col + '10'].border = black_border_no_bottom
    ws[col + '11'].border = black_border_no_bottom
    ws[col + '12'].border = black_border_no_bottom
    ws[col + '13'].border = black_border_no_bottom
    ws[col + '14'].border = black_border_no_bottom
    ws[col + '15'].border = black_border_no_bottom
    ws[col + '16'].border = black_border_no_bottom
    ws[col + '17'].border = black_border_no_bottom
    ws[col + '18'].border = black_border_no_bottom
    ws[col + '19'].border = black_border_no_bottom
    ws[col + '20'].border = black_border_no_bottom
    ws[col + '21'].border = black_border_no_bottom
    ws[col + '22'].border = black_border_no_bottom
    ws[col + '23'].border = black_border_no_bottom
    ws[col + '24'].border = black_border_no_bottom
    ws[col + '25'].border = black_border_no_bottom
    ws[col + '26'].border = black_border_no_bottom
    ws[col + '27'].border = black_border_no_bottom
    ws[col + '28'].border = black_border_no_bottom
    ws[col + '29'].border = black_border_no_bottom
    ws[col + '30'].border = black_border_no_bottom
    ws[col + '31'].border = black_border_no_bottom
    ws[col + '32'].border = black_border_no_bottom
    ws[col + '33'].border = black_border_no_bottom
    ws[col + '34'].border = black_border_no_bottom
    ws[col + '35'].border = black_border_no_bottom
    ws[col + '36'].border = black_border_no_bottom
    ws[col + '37'].border = black_border_no_bottom
    ws[col + '38'].border = black_boarder_all


# Update font size for range B4:N38
for row in ws['B4':'N38']:
    for cell in row:
        cell.font = openpyxl.styles.Font(size=10)

# Update alignment for range C6:N38
for row in ws['C6':'N38']:
    for cell in row:
        if cell.value is not None:  # Ensure there is a value in the cell
            cell.value = str(cell.value) + ' '  # Prepend space to the value
        cell.alignment = openpyxl.styles.Alignment(horizontal='right')


# # Error ignoring style
# error_ignoring_style = openpyxl.styles.NamedStyle(name="ignore_error", number_format="@")
# error_ignoring_style.protection = openpyxl.styles.Protection(locked=True)
# error_ignoring_style.alignment = openpyxl.styles.Alignment(horizontal='general')
# for rule in error_ignoring_style.errorChecking:
#     error_ignoring_style.errorChecking[rule] = False

# # Apply style to all cells to ignore errors
# for row in ws.iter_rows():
#     for cell in row:
#         cell.style = error_ignoring_style


# Step 4: Save the report
# Save the report to the specified path
save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Report_8b_IEP_Service_Recs.xlsx'
wb.save(save_path)