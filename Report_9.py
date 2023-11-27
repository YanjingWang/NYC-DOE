import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Report 9 = Placement"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Report 9 = Placement")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Total Students with Initial IEP Meeting and Placement Notice',
                'Average # of School Days between IEP Meeting and Placement Notice']
        column_letters = ['C', 'D']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "B8CCE4"
        self.format_header(ws, 'B5', 'District', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B41', 'Race/Ethnicity', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B50', 'Meal Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B56', 'Gender', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B63', 'ELL Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B69', 'Language of Instruction', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B77', 'Grade Level', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B95', 'Temporary Housing Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B102', 'Foster Care Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_InitialReferralsR19_SY23_fix', 'INT_StudentDemographics_063023')
        cursor.execute("EXEC [dev].[USPCCAnnaulReport9] @tableNameCCInitialReferralsR19=?,  @tableNameINTStudentDemographics_0630=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_race(self,cursor):
        query_byRace = '''
        select EthnicityGroupCC, TOTAL_PWN,Average_Days   from (  select * from  (  Select   Ethnicity_sort as sort , EthnicityGroupCC, FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,
        FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report  group by EthnicityGroupCC, Ethnicity_sort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort 
        '''  # the byRace SQL query goes here
        cursor.execute(query_byRace)
        results_byRace = cursor.fetchall()
        return results_byRace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,cursor):
        query_byDistrict = '''
        select * from  (  Select    ReportingDistrict as sort , FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,
        FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report  group by ReportingDistrict  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byDistrict SQL query goes here
        cursor.execute(query_byDistrict)
        results_byDistrict = cursor.fetchall()
        return results_byDistrict

    def fetch_data_by_mealstatus(self,cursor):
        query_byMealStatus = '''
        select * from  (  Select    MealStatusGrouping as sort , FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,
        FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report  group by MealStatusGrouping  ) a  union all  select * from ##TotalRow  order by sort
        '''  # the byMealStatus SQL query goes here
        cursor.execute(query_byMealStatus)
        results_byMealStatus = cursor.fetchall()
        return results_byMealStatus
    
    def fetch_data_by_gender(self,cursor):
        query_byGender = '''
        select * from  (  Select    Gender as sort , FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,
        FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report  group by Gender  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byGender SQL query goes here
        cursor.execute(query_byGender)
        results_byGender = cursor.fetchall()
        return results_byGender
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byELLStatus = '''
        select * from  (  Select    ELLStatus as sort , FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,
        FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report  group by ELLStatus  ) a  union all  select * from ##TotalRow  order by sort  
        '''  # the byELLStatus SQL query goes here
        cursor.execute(query_byELLStatus)
        results_byELLStatus = cursor.fetchall()
        return results_byELLStatus
    
    def fetch_data_by_language(self,cursor):
        query_byLanguage = '''
        select OutcomeLanguageCC, TOTAL_PWN,Average_Days   from (  select * from  (  Select   Language_Sort as sort , OutcomeLanguageCC, 
        FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report  group by OutcomeLanguageCC, Language_Sort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort
        '''  # the byLanguage SQL query goes here
        cursor.execute(query_byLanguage)
        results_byLanguage = cursor.fetchall()
        return results_byLanguage
    
    def fetch_data_by_gradelevel(self,cursor):
        query_byGradeLevel = '''
        select GradeLevel, TOTAL_PWN,Average_Days   from (  select * from  (  Select   Grade_Sort as sort , GradeLevel, FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,
        FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report  group by GradeLevel, Grade_Sort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort  
        '''
        cursor.execute(query_byGradeLevel)
        results_byGradeLevel = cursor.fetchall()
        return results_byGradeLevel
    
    def fetch_data_by_tempResFlag(self,cursor):
        query_byTempResFlag = '''
        select TempResFlag, TOTAL_PWN,Average_Days   from (  select * from  (  Select   TempResFlagSort as sort ,  case when TempResFlag = 'Y' then 'Yes' else 'No'  end as TempResFlag, 
        FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report where TempResFlag in ('Y', 'N')  group by TempResFlag, TempResFlagSort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort 
        '''
        cursor.execute(query_byTempResFlag)
        results_byTempResFlag = cursor.fetchall()
        return results_byTempResFlag
    
    def fetch_data_by_fosterCareStatus(self,cursor):
        query_byFosterCareStatus = '''
        select FostercareFlag, TOTAL_PWN,Average_Days   from (  select * from  (  Select   FosterCareFlagSort as sort , FostercareFlag, FORMAT(count(StudentID) ,'#,##0') as TOTAL_PWN ,
        FORMAT(cast((sum(OutcomePlacementSchoolDays)*1.0)/count(StudentID)  as numeric(8,2)),'#,##0') as Average_Days  
        FROM ##Report  group by FostercareFlag, FosterCareFlagSort ) a  union all  select * from ##TotalRow_Sort  ) a order by sort 
        '''
        cursor.execute(query_byFosterCareStatus)
        results_byFosterCareStatus = cursor.fetchall()
        return results_byFosterCareStatus
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C5':'D38']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['C42':'D47']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['C51':'D53']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['C57':'D60']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['C64':'D66']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['C70':'D74']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['C78':'D91']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['C96':'D98']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['C103':'D105']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        # # Update alignment for range B6:B38, B43:B48, B53:B55, B60:B63, B68:B70, B75:B80, B86:B99, B105:B107, B113:B115
        # for row in ws['B6':'B38']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = ' ' + str(cell.value)  # Append space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='left')

        # for row in ws['B43':'B48']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = ' ' + str(cell.value)  # Append space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='left')
    def main_Report_9_Placement(self):
        title_cells = [
            {"cell": "B1", "value": "Report 9 Average Number of School Days from Initial IEP Meeting to Placement Notice Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; and Grade Level.", "merge_cells": "B1:D1"},
            

        ]

        subtitle_cells = [
            {"cell": "B4", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by District", "merge_cells": "B4:D4"},
            {"cell": "B40", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by  Race/Ethnicity", "merge_cells": "B40:D40"},
            {"cell": "B49", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by Meal Status ", "merge_cells": "B49:D49"},
            {"cell": "B55", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by Gender", "merge_cells": "B55:D55"},
            {"cell": "B62", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by English Language Learner (ELL) Status ", "merge_cells": "B62:D62"},
            {"cell": "B68", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by Language of Instruction", "merge_cells": "B68:D68"},
            {"cell": "B76", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by Grade Level", "merge_cells": "B76:D76"},
            {"cell": "B94", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by Temporary Housing Status", "merge_cells": "B94:D94"},
            {"cell": "B101", "value": "SY 2022-23 Average Number of School Days Between Initial IEP Meeting and Placement Notice by Foster Care Status", "merge_cells": "B101:D101"},
            

        ]

        column_widths = [5, 30, 70, 70]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_byRace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byRace, start_row=42)
        
        # Step 4: Fetch and write data for "Report 8b = IEP Service Recs by District"
        results_byDistrict = self.fetch_data_by_district(cursor)
        self.write_data_to_excel(ws, results_byDistrict, start_row=6)

        # Step 5: Fetch and write data for "Report 8b = IEP Service Recs by Meal Status"
        results_byMealStatus = self.fetch_data_by_mealstatus(cursor)
        self.write_data_to_excel(ws, results_byMealStatus, start_row=51)

        # Step 6: Fetch and write data for "Report 8b = IEP Service Recs by Gender"
        results_byMealStatus = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_byMealStatus, start_row=57)

        # Step 7: Fetch and write data for "Report 8b = IEP Service Recs by ELL Status"
        results_byELLStatus = self.fetch_data_by_ellstatus(cursor)
        self.write_data_to_excel(ws, results_byELLStatus, start_row=64)
        
        # Step 8: Fetch and write data for "Report 8b = IEP Service Recs by Language"
        results_byLanguage = self.fetch_data_by_language(cursor)
        self.write_data_to_excel(ws, results_byLanguage, start_row=70)

        # Step 9: Fetch and write data for "Report 8b = IEP Service Recs by Grade Level"
        results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        self.write_data_to_excel(ws, results_byGradeLevel, start_row=78)

        # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Temporary Housing"
        results_byTempResFlag = self.fetch_data_by_tempResFlag(cursor)
        self.write_data_to_excel(ws, results_byTempResFlag, start_row=96)

        # Step 11: Fetch and write data for "Report 8b = IEP Service Recs by Foster Care Status"
        results_byFosterCareStatus = self.fetch_data_by_fosterCareStatus(cursor)
        self.write_data_to_excel(ws, results_byFosterCareStatus, start_row=103)
        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.main_Report_9_Placement()                                                                  