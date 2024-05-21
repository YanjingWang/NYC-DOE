import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reports 1-4 = Initials"

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Total Students with Initial Referrals 7/1/2022 - 06/30/2023', 'Closed without IEP Meeting',
                'Student Determined Ineligible. IEP Meeting <= 60 Calendar Days from Date of Consent', 'Student Determined Ineligible. IEP Meeting > 60 Calendar Days from Date of Consent', 'Total Ineligible', 'Student Classified. IEP Meeting <= 60 Calendar Days from Date of Consent','Student Classified. IEP Meeting > 60 Calendar Days from Date of Consent','Total Classified','Total IEP Meetings Held (Ineligible + Classified)','Open and Awaiting Parental Consent as of 06/30/2023','Open and Parental Consent Received as of 06/30/2023']
        column_letters = ['C', 'D', 'E', 'F', 'G', 'H','I', 'J', 'K', 'L', 'M']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "E0F0F8"
        self.format_header(ws, 'B4', 'District', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B40', 'Race/Ethnicity', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B49', 'Meal Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B55', 'Gender', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B62', 'ELL Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B68', 'Language of Instruction', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B78', 'Grade Level', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B98', 'Temporary Housing Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B106', 'Foster Care Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_InitialReferralsR19_SY23', 'INT_StudentDemographics_063023')
        cursor.execute("EXEC [dbo].[USPCCAnnaulReport1to4] @tableNameCCInitialReferralsR19=?,  @tableNameINTStudentDemographics_0630=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_race(self,cursor,conn):
        # query_byRace = '''
        # select * from  (  Select   EthnicityGroupCC_sort as sort , EthnicityGroupCC	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        # FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        # FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        # FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as c8 	,
        # FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        # FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by EthnicityGroupCC, EthnicityGroupCC_sort ) a  union all  select * from ##TotalRow_Sort  order by sort 
        # '''  # the byRace SQL query goes here
        # cursor.execute(query_byRace)
        # results_byRace = cursor.fetchall()
        # print(results_byRace)
        # results_byRace = pd.DataFrame(results_byRace, columns=['EthnicityGroupCC', 'c1', 'c2', 'c3', 'c4', 'c5', 'c6', 'c7', 'c8','c9', 'c10'])
        query_byRace = '''
        select * from  (  Select   EthnicityGroupCC_sort as sort , EthnicityGroupCC	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by EthnicityGroupCC, EthnicityGroupCC_sort ) a  union all  select * from ##TotalRow_Sort  order by sort 
        '''
        df_byRace = pd.read_sql_query(query_byRace, conn)
        conn.close()
        results_byRace = df_byRace[['EthnicityGroupCC', 'c1', 'c2', 'c3', 'c4', 'c5', 'c6', 'c7', 'c8', 'c9', 'c10', 'c11']]
        print(results_byRace)
        return results_byRace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,cursor):
        query_byDistrict = '''
        select * from  (  Select    ReportingDistrict as sort 	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10	
        ,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by ReportingDistrict  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byDistrict SQL query goes here
        cursor.execute(query_byDistrict)
        results_byDistrict = cursor.fetchall()
        return results_byDistrict

    def fetch_data_by_mealstatus(self,cursor):
        query_byMealStatus = '''
        select * from  (  Select    MealStatusGrouping as sort 	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by MealStatusGrouping  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byMealStatus SQL query goes here
        cursor.execute(query_byMealStatus)
        results_byMealStatus = cursor.fetchall()
        return results_byMealStatus
    
    def fetch_data_by_gender(self,cursor):
        query_byGender = '''
        select * from  (  Select    GENDER as sort 	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by GENDER  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byGender SQL query goes here
        cursor.execute(query_byGender)
        results_byGender = cursor.fetchall()
        return results_byGender
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byELLStatus = '''
        select * from  (  Select    ELLStatus as sort 	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by ELLStatus  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byELLStatus SQL query goes here
        cursor.execute(query_byELLStatus)
        results_byELLStatus = cursor.fetchall()
        return results_byELLStatus
    
    def fetch_data_by_language(self,cursor):
        query_byLanguage = '''
        select * from  (  Select   Language_Sort as sort , LanguageOfInstructionCC2	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by LanguageOfInstructionCC2, Language_Sort ) a  union all  select * from ##TotalRow_Sort  order by sort 

        '''  # the byLanguage SQL query goes here
        cursor.execute(query_byLanguage)
        results_byLanguage = cursor.fetchall()
        return results_byLanguage
    
    def fetch_data_by_gradelevel(self,cursor):
        query_byGradeLevel = '''
        select * from  (  Select   GradeLevel_Sort as sort , GradeLevel	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by GradeLevel, GradeLevel_Sort ) a  union all  select * from ##TotalRow_Sort  order by sort
        '''
        cursor.execute(query_byGradeLevel)
        results_byGradeLevel = cursor.fetchall()
        return results_byGradeLevel
    
    def fetch_data_by_tempResFlag(self,cursor):
        query_byTempResFlag = '''
        select * from  (  Select   TempResFlagSort as sort ,  case when TempResFlag = 'Y' then 'Yes' else 'No'  end as TempResFlag	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final  where TempResFlag in ('Y', 'N')  group by TempResFlag, TempResFlagSort ) a  union all  select * from ##TotalRow_Sort  order by sort 
        '''
        cursor.execute(query_byTempResFlag)
        results_byTempResFlag = cursor.fetchall()
        return results_byTempResFlag
    
    def fetch_data_by_fosterCareStatus(self,cursor):
        query_byFosterCareStatus = '''
        select * from  (  Select   FosterCareFlagSort as sort , FosterCareFlag	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,
        FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,
        FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,
        FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,
        FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final   group by FosterCareFlag, FosterCareFlagSort ) a  union all  select * from ##TotalRow_Sort  order by sort
        '''
        cursor.execute(query_byFosterCareStatus)
        results_byFosterCareStatus = cursor.fetchall()
        return results_byFosterCareStatus
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row in enumerate(df.itertuples(index=False), start=start_row):
            for i, value in enumerate(row):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                cell = ws[col + str(row_num)]
                cell.value = value
                cell.border = black_border
                cell.alignment = Alignment(horizontal='right')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C5':'M37']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C41':'M46']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C50':'M52']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C56':'M59']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C63':'M65']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C69':'M74']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C79':'M92']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C99':'M101']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C107':'M109']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
        # # Update alignment for range B6:B38, B43:B48, B53:B55, B60:B63, B68:B70, B75:B80, B86:B99, B105:B107, B113:B115
        # for row in ws['B6':'B38']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = ' ' + str(cell.value)  # Append space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='left')

        # for row in ws['B43':'B48']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = ' ' + str(cell.value)  # Append space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='left')
    def main_Reports_1_4_Initials(self):
        title_cells = [
            {"cell": "B1", "value": "Report 8b IEP Service Recommendations Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; and Grade Level.", "merge_cells": "B1:M1"},
            

        ]

        subtitle_cells = [
            {"cell": "B3", "value": "SY 2022-23 Students with IEP Recommended Services by District", "merge_cells": "B3:M3"},
            {"cell": "B39", "value": "SY 2022-23 Students with IEP Recommended Services by Ethnicity", "merge_cells": "B39:M39"},
            {"cell": "B48", "value": "SY 2022-23 Students with IEP Recommended Services by Meal Status", "merge_cells": "B48:M48"},
            {"cell": "B54", "value": "SY 2022-23 Students with IEP Recommended Services by Gender", "merge_cells": "B54:M54"},
            {"cell": "B61", "value": "SY 2022-23 Students with IEP Recommended Services by ELL Status", "merge_cells": "B61:M61"},
            {"cell": "B67", "value": "SY 2022-23 Students with IEP Recommended Services  by Recommended Language of Instruction", "merge_cells": "B67:M67"},
            {"cell": "B77", "value": "SY 2022-23 Students with IEP Recommended Services  by Grade Level", "merge_cells": "B77:M77"},
            {"cell": "B97", "value": "SY 2022-23 Students with IEP Recommended Services  by Temporary Housing", "merge_cells": "B97:M97"},
            {"cell": "B105", "value": "SY 2022-23 Students with IEP Recommended Services  by Foster Care Status", "merge_cells": "B105:M105"},
            

        ]

        column_widths = [5, 30, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART' 
        results_byRace = self.fetch_data_by_race(cursor,conn=pyodbc.connect(conn_str))
        self.write_data_to_excel(ws, results_byRace, start_row=41)
        
        # # Step 4: Fetch and write data for "Report 8b = IEP Service Recs by District"
        # results_byDistrict = self.fetch_data_by_district(cursor)
        # self.write_data_to_excel(ws, results_byDistrict, start_row=5)

        # # Step 5: Fetch and write data for "Report 8b = IEP Service Recs by Meal Status"
        # results_byMealStatus = self.fetch_data_by_mealstatus(cursor)
        # self.write_data_to_excel(ws, results_byMealStatus, start_row=50)

        # # Step 6: Fetch and write data for "Report 8b = IEP Service Recs by Gender"
        # results_byMealStatus = self.fetch_data_by_gender(cursor)
        # self.write_data_to_excel(ws, results_byMealStatus, start_row=56)

        # # Step 7: Fetch and write data for "Report 8b = IEP Service Recs by ELL Status"
        # results_byELLStatus = self.fetch_data_by_ellstatus(cursor)
        # self.write_data_to_excel(ws, results_byELLStatus, start_row=63)
        
        # # Step 8: Fetch and write data for "Report 8b = IEP Service Recs by Language"
        # results_byLanguage = self.fetch_data_by_language(cursor)
        # self.write_data_to_excel(ws, results_byLanguage, start_row=75)

        # # Step 9: Fetch and write data for "Report 8b = IEP Service Recs by Grade Level"
        # results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        # self.write_data_to_excel(ws, results_byGradeLevel, start_row=86)

        # # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Temporary Housing"
        # results_byTempResFlag = self.fetch_data_by_tempResFlag(cursor)
        # self.write_data_to_excel(ws, results_byTempResFlag, start_row=105)

        # # Step 11: Fetch and write data for "Report 8b = IEP Service Recs by Foster Care Status"
        # results_byFosterCareStatus = self.fetch_data_by_fosterCareStatus(cursor)
        # self.write_data_to_excel(ws, results_byFosterCareStatus, start_row=113)
        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\CityCouncilAnnualReport.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.main_Reports_1_4_Initials()                                                                  