import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Report 12 = Program Services"
        ws = wb.create_sheet("Report 15a = Transportation")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=1615, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Curb to School',
                'Percent Curb to School',
                'Stop to School',
                'Percent Stop to School',
                'Unassigned',
                'Percent Unassigned']
        column_letters = ['C', 'D', 'E', 'F', 'G', 'H']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "B8CCE4"
        self.format_header(ws, 'B4', 'District', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B40', 'Race/Ethnicity ', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B49', 'Meal Status', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B55', 'Gender', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B62', 'ELL Status', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B68', 'Language of Instruction', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B76', 'Grade Level', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B93', 'Temporary Housing Status', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B99', 'Foster Care Status', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B105', 'School DBN', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_SpecialTransportation_061523','CC_StudentRegisterR814_061523')
        cursor.execute("EXEC [dbo].[USPCC_AnnaulReport13st] @tableNameCCSpecialTransportation =?,@tableNameStudentRegisterR814 =?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_district(self,cursor):
        query_bydistrict = '''
        select * from  ( Select  ReportingDistrict as sort ,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by ReportingDistrict  ) a  union all select 99 as  total, c1,c2,c3,c4,c5,c6 from ##totalRow13st order by sort 
        '''  # the bytab12 SQL query goes here
        cursor.execute(query_bydistrict)
        results_bydistrict = cursor.fetchall()
        return results_bydistrict
    
    def fetch_data_by_race(self,cursor):
        query_byrace = '''
        select EthnicityGroupCC, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select Ethnicity_sort as sort , EthnicityGroupCC,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by EthnicityGroupCC, Ethnicity_sort ) a  union all  select * from ##totalRow_Sort13st  ) a order by sort 
        '''
        cursor.execute(query_byrace)
        results_byrace = cursor.fetchall()
        return results_byrace
    
    def fetch_data_by_mealstatus(self,cursor):
        query_bymealstatus = '''
        select * from  ( Select  MealStatusGrouping as sort ,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by MealStatusGrouping  ) a  union all select * from ##totalRow13st  order by sort 
        '''
        cursor.execute(query_bymealstatus)
        results_bymealstatus = cursor.fetchall()
        return results_bymealstatus
    
    def fetch_data_by_gender(self,cursor):
        query_bygender = ''' 
        select * from  ( Select  Gender as sort ,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by Gender  ) a  union all select * from ##totalRow13st  order by sort 
        '''
        cursor.execute(query_bygender)
        results_bygender = cursor.fetchall()
        return results_bygender
        
    def fetch_data_by_gradelevel(self,cursor):
        query_bygradelevel = ''' 
        select GradeLevel, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select GradeSort as sort , GradeLevel,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by GradeLevel, GradeSort ) a  union all  select * from ##totalRow_Sort13st  ) a order by sort
        '''
        cursor.execute(query_bygradelevel)
        results_bygradelevel = cursor.fetchall()
        return results_bygradelevel
    
    def fetch_data_by_language(self,cursor):
        query_bylanguage = ''' 
        select OutcomeLanguageCC, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select OutcomeLanguageCCSort as sort , OutcomeLanguageCC,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by OutcomeLanguageCC, OutcomeLanguageCCSort ) a  union all  select * from ##totalRow_Sort13st  ) a order by sort 
        '''
        cursor.execute(query_bylanguage)
        results_bylanguage = cursor.fetchall()
        return results_bylanguage
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byellstatus = '''
        select * from  ( Select  ELLStatus as sort ,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by ELLStatus  ) a  union all select * from ##totalRow13st  order by sort 
        '''
        cursor.execute(query_byellstatus)
        results_byellstatus = cursor.fetchall()
        return results_byellstatus
    
    def fetch_data_by_tempstatus(self,cursor):
        query_bytempstatus = '''
        select TempResFlag, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select TempResFlagSort as sort , TempResFlag,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by TempResFlag, TempResFlagSort ) a  union all  select * from ##totalRow_Sort13st  ) a order by sort
        '''
        cursor.execute(query_bytempstatus)
        results_bytempstatus = cursor.fetchall()
        return results_bytempstatus
    
    def fetch_data_by_fostercarestaus(self,cursor):
        query_byfostercarestaus = '''
        select FostercareFlag, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select FosterCareFlagSort as sort , FostercareFlag,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by FostercareFlag, FosterCareFlagSort ) a  union all  select * from ##totalRow_Sort13st  ) a order by sort 
        '''
        cursor.execute(query_byfostercarestaus)
        results_byfostercarestaus = cursor.fetchall()
        return results_byfostercarestaus
    
    def fetch_data_by_school(self,cursor):
        query_byschool = '''
        select * from  ( Select  EnrolledDBN as sort ,FORMAT(Sum(CurbtoSchool) , '#,##0') as c1 ,concat(cast(Sum(CurbtoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c2 ,FORMAT(sum(StoptoSchool) , '#,##0') as c3 ,concat(cast(Sum(StoptoSchool)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c4 ,FORMAT(sum(Unassigned) , '#,##0') as c5 ,concat(cast(Sum(Unassigned)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%') as c6  FROM ##CCTotaltemp13st  group by EnrolledDBN  ) a  union all select * from ##totalRow13st  order by sort
        '''    
        cursor.execute(query_byschool)
        results_byschool = cursor.fetchall()
        return results_byschool

    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C5':'H37'] + ws['C41':'H46'] + ws['C50':'H52'] + ws['C56':'H59'] + ws['C63':'H65'] + ws['C69':'H73'] + ws['C77':'H90'] + ws['C94':'H96'] + ws['C100':'H102'] + ws['C106':'H1614']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                if isinstance(cell.value, str):
                    try:
                        cell.value = int(cell.value)
                    except ValueError:
                        # If the value cannot be converted to int, keep the original value
                        pass
        # Function to check if a string represents a valid number
        def is_number(s):
            try:
                float(s.replace(',', ''))  # Try converting after removing commas
                return True
            except ValueError:
                return False

        # Formatting specific cell ranges
        cell_ranges = ['C5:H37', 'C41:H46', 'C50:H52', 'C56:H59', 'C63:H65', 'C69:H73', 'C77:H90', 'C94:H96', 'C100:H102', 'C106:H1614']
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if isinstance(cell.value, str) and is_number(cell.value):
                        # Convert to float after removing commas
                        cell.value = float(cell.value.replace(',', ''))
                        # Apply number format with commas (optional)
                        cell.number_format = '#,##0'
        # for row in ws['C41':'H46']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C50':'H52']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C56':'H59']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C63':'H65']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C69':'H73']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C77':'H90']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C90':'H94']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C94':'H96']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C100':'H102']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')      

        # for row in ws['C106':'H1614']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')     

        for row in ws['B1': 'H1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B37':'H37'] + ws['B46':'H46'] + ws['B52':'H52'] + ws['B59':'H59'] + ws['B65':'H65'] + ws['B73':'H73'] + ws['B90':'H90'] + ws['B96':'H96'] + ws['B102':'H102'] + ws['B1614':'H1614']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['B3':'H3'] + ws['B39':'H39'] + ws['B48':'H48'] + ws['B54':'H54'] + ws['B61':'H61'] + ws['B67':'H67'] + ws['B75':'H75'] + ws['B92':'H92'] + ws['B98':'H98'] + ws['B104':'H104'] :
            for cell in row:
                cell.border = black_border_thick
                cell.font = Font(bold=True, size=12)

        # wrap text of B50 cell having 'Eligible for the Free/Reduced Price Lunch Program' to fit the cell
        ws['B50'].alignment = Alignment(wrap_text=True)
        # make cell B1 higher to fit the text
        ws.row_dimensions[1].height = 30           
    def Report_15_Transportation(self):
        title_cells = [
            {"cell": "B1", "value": "Report 15 Number & Percentage of Special Transportation Recommendations with Busing Assignment Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; Grade Level; Temp House Status; Foster Care Status and School.", "merge_cells": "B1:H1"},
            

        ]

        subtitle_cells = [
            {"cell": "B3", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by District", "merge_cells": "B3:H3"},
            {"cell": "B39", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by Race/Ethnicity", "merge_cells": "B39:H39"},    
            {"cell": "B48", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by Meal Status", "merge_cells": "B48:H48"}, 
            {"cell": "B54", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by Gender", "merge_cells": "B54:H54"},   
            {"cell": "B61", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by English Language Learner (ELL) Status", "merge_cells": "B61:H61"}, 
            {"cell": "B67", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by Recommended Language of Instruction", "merge_cells": "B67:H67"}, 
            {"cell": "B75", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by Grade Level", "merge_cells": "B75:H75"}, 
            {"cell": "B92", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by Temporary Housing Status", "merge_cells": "B92:H92"}, 
            {"cell": "B98", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by Foster Care Status", "merge_cells": "B98:H98"}, 
            {"cell": "B104", "value": "SY 2022-2023 Number & Percentage of Special Transportation Recommendations with Busing Assignment by School", "merge_cells": "B104:H104"}, 


        ]

        column_widths = [5, 70, 30, 30, 30, 30, 30, 30, 30]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_bydistrict= self.fetch_data_by_district(cursor)
        # replace 1 with 01, 2 with 02, 3 with 03, 4 with 04, 5 with 05, 6 with 06, 7 with 07, 8 with 08, 9 with 09 and 99 with Total
        results_bydistrict = [('0' + str(x[0]) if len(str(x[0])) == 1 else x[0], *x[1:]) for x in results_bydistrict]
        results_bydistrict = [('Total' if x[0] == 99 else x[0], *x[1:]) for x in results_bydistrict]
        self.write_data_to_excel(ws, results_bydistrict, start_row=5)

        # Step 6: fetch and write data by Race/Ethnicity
        results_byrace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byrace, start_row=41)

        results_bymealstatus = self.fetch_data_by_mealstatus(cursor)
        # replace Free or Reduced Price Meal to Eligible for the Free/Reduced Price Lunch Program
        results_bymealstatus = [('Eligible for the Free/Reduced Price Lunch Program' if x[0] == 'Free or Reduced Price Meal' else x[0], *x[1:]) for x in results_bymealstatus]
        self.write_data_to_excel(ws, results_bymealstatus, start_row=50)

        results_bygender = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_bygender, start_row=56)

        results_byellstatus = self.fetch_data_by_ellstatus(cursor)
        results_byellstatus = [('ELL' if x[0] == 'ELL' else ('NOT ELL' if x[0] == 'Non-Ell' else x[0]), *x[1:]) for x in results_byellstatus]
        self.write_data_to_excel(ws, results_byellstatus, start_row=63)

        results_bylanguage = self.fetch_data_by_language(cursor)
        results_bylanguage = [('English' if x[0] == 'ENGLISH' else ('Spanish' if x[0] == 'SPANISH' else ('Chinese' if x[0] == 'CHINESE' else ('Other' if x[0] == 'OTHER' else x[0]))), *x[1:]) for x in results_bylanguage]
        self.write_data_to_excel(ws, results_bylanguage, start_row=69)

        results_bygradelevel = self.fetch_data_by_gradelevel(cursor)
        results_bygradelevel = [(x[0].replace('0K', 'KG'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('01', '1'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('02', '2'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('03', '3'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('04', '4'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('05', '5'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('06', '6'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('07', '7'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('08', '8'), *x[1:]) for x in results_bygradelevel]
        results_bygradelevel = [(x[0].replace('09', '9'), *x[1:]) for x in results_bygradelevel]   
        self.write_data_to_excel(ws, results_bygradelevel, start_row=77)

        results_bytempstatus = self.fetch_data_by_tempstatus(cursor)
        results_bytempstatus = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_bytempstatus]   
        self.write_data_to_excel(ws, results_bytempstatus, start_row=94)

        results_byfostercarestaus = self.fetch_data_by_fostercarestaus(cursor)
        results_byfostercarestaus = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_byfostercarestaus]
        self.write_data_to_excel(ws, results_byfostercarestaus, start_row=100)

        results_byschool = self.fetch_data_by_school(cursor)
        self.write_data_to_excel(ws, results_byschool, start_row=106)

        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.Report_15_Transportation()                                                                  