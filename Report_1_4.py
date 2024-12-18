import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
import os
from copy import copy
class Solution:
    # Existing code...
    # Function to format headers
    def __init__(self):
        self.schoolyear = 'SY 2024-25'
        self.sqlsnapshottableschoolyear = '24'
        self.lastyear = '2023'
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Reports 1-4 = Initials"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Reports 1-4 = Initials")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = [f'Total Students with Initial Referrals 7/1/{self.lastyear} - 06/30/20{self.sqlsnapshottableschoolyear}', 'Closed without IEP Meeting',
                'Student Determined Ineligible. IEP Meeting <= 60 Calendar Days from Date of Consent', 'Student Determined Ineligible. IEP Meeting > 60 Calendar Days from Date of Consent', 'Total Ineligible', 'Student Classified. IEP Meeting <= 60 Calendar Days from Date of Consent','Student Classified. IEP Meeting > 60 Calendar Days from Date of Consent','Total Classified','Total IEP Meetings Held (Ineligible + Classified)',f'Open and Awaiting Parental Consent as of 06/30/20{self.sqlsnapshottableschoolyear}',f'Open and Parental Consent Received as of 06/30/20{self.sqlsnapshottableschoolyear}']
        column_letters = ['C', 'D', 'E', 'F', 'G', 'H','I', 'J', 'K', 'L', 'M']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "B8CCE4"
        self.format_header(ws, 'B4', 'District', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B40', 'Race/Ethnicity', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B49', 'Meal Status', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B55', 'Gender', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B62', 'ELL Status', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B68', 'Language of Instruction', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B78', 'Grade Level', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B98', 'Temporary Housing Status', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B106', 'Foster Care Status', columns, column_letters, 110, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_MART' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_InitialReferralsR19_SY'+self.sqlsnapshottableschoolyear)
        cursor.execute("EXEC [dbo].[USPCC_AnnaulReport1to4] @tableNameCCInitialReferralsR19=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_race(self,cursor):
        query_byRace = '''
        select EthnicityGroupCC, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11 from (  select * from  (  Select   EthnicityGroupCC_sort as sort , EthnicityGroupCC	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14   group by EthnicityGroupCC, EthnicityGroupCC_sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''  # the byRace SQL query goes here
        cursor.execute(query_byRace)
        results_byRace = cursor.fetchall()
        return results_byRace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,cursor):
        query_byDistrict = '''
        select * from  (  Select    ReportingDistrict as sort 	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14   group by ReportingDistrict  ) a  union all  select * from ##TotalRow  order by sort 

        '''  # the byDistrict SQL query goes here
        cursor.execute(query_byDistrict)
        results_byDistrict = cursor.fetchall()
        return results_byDistrict

    def fetch_data_by_mealstatus(self,cursor):
        query_byMealStatus = '''
        select * from  (  Select    MealStatusGrouping as sort 	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14   group by MealStatusGrouping  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byMealStatus SQL query goes here
        cursor.execute(query_byMealStatus)
        results_byMealStatus = cursor.fetchall()
        return results_byMealStatus
    
    def fetch_data_by_gender(self,cursor):
        query_byGender = '''
        select * from  (  Select    GENDER as sort 	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14   group by GENDER  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byGender SQL query goes here
        cursor.execute(query_byGender)
        results_byGender = cursor.fetchall()
        return results_byGender
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byELLStatus = '''
        select * from  (  Select    ELLStatus as sort 	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14   group by ELLStatus  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byELLStatus SQL query goes here
        cursor.execute(query_byELLStatus)
        results_byELLStatus = cursor.fetchall()
        return results_byELLStatus
    
    def fetch_data_by_language(self,cursor):
        query_byLanguage = '''
        select LanguageOfInstructionCC2, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11 from (  select * from  (  Select   Language_Sort as sort , LanguageOfInstructionCC2	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14   group by LanguageOfInstructionCC2, Language_Sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort  
        '''  # the byLanguage SQL query goes here
        cursor.execute(query_byLanguage)
        results_byLanguage = cursor.fetchall()
        return results_byLanguage
    
    def fetch_data_by_gradelevel(self,cursor):
        query_byGradeLevel = '''
        select GradeLevel, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11 from (  select * from  (  Select   GradeLevel_Sort as sort , GradeLevel	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14   group by GradeLevel, GradeLevel_Sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort  
        '''
        cursor.execute(query_byGradeLevel)
        results_byGradeLevel = cursor.fetchall()
        return results_byGradeLevel
    
    def fetch_data_by_tempResFlag(self,cursor):
        query_byTempResFlag = '''
        select STHFlag, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11 from (  select * from  (  Select   STHFlagSort as sort ,  case when STHFlag = 'Y' then 'Yes' else 'No'  end as STHFlag	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14  where STHFlag in ('Y', 'N')  group by STHFlag, STHFlagSort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''
        cursor.execute(query_byTempResFlag)
        results_byTempResFlag = cursor.fetchall()
        return results_byTempResFlag
    
    def fetch_data_by_fosterCareStatus(self,cursor):
        query_byFosterCareStatus = '''
        select FosterCareFlag, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11 from (  select * from  (  Select   FosterCareFlagSort as sort , FosterCareFlag	,FORMAT(sum(STUDENTS_WITH_REF), '#,##0') as c1     ,FORMAT(sum(CLOSED_WITHOUT_IEP), '#,##0') as c2 	,FORMAT(sum(INELIGIBLE_LESS_60), '#,##0') as c3     ,FORMAT(sum(INELIGIBLE_MORE_60), '#,##0') as c4 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0)), '#,##0') as c5     ,FORMAT(sum(CLASSIFIED_LESS_60), '#,##0') as c6     ,FORMAT(sum(CLASSIFIED_MORE_60), '#,##0') as c7     ,FORMAT(sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60), '#,##0') as C8 	,FORMAT(sum(COALESCE(INELIGIBLE_LESS_60,0) + COALESCE(INELIGIBLE_MORE_60,0) + COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0)), '#,##0') as c9 	,FORMAT(sum(TOTAL_AWAITING), '#,##0') as c10			,FORMAT(sum(TOTAL_OPEN), '#,##0') as c11  FROM ##Report_Final14   group by FosterCareFlag, FosterCareFlagSort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''
        cursor.execute(query_byFosterCareStatus)
        results_byFosterCareStatus = cursor.fetchall()
        return results_byFosterCareStatus
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38 and convert datatype string to number
        for row in ws['C5':'M37'] + ws['C41':'M46'] + ws['C50':'M52'] + ws['C56':'M59'] + ws['C63':'M65'] + ws['C69':'M74'] + ws['C79':'M92'] + ws['C99':'M101'] + ws['C107':'M109']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                if isinstance(cell.value, str):
                    try:
                        cell.value = int(cell.value)
                    except ValueError:
                        # If the value cannot be converted to int, keep the original value
                        pass
        # Function to check if a string represents a valid number
        def is_number(s):
            try:
                float(s.replace(',', ''))  # Try converting after removing commas
                return True
            except ValueError:
                return False

        # Formatting specific cell ranges
        cell_ranges = ['C5:M37', 'C41:M46', 'C50:M52', 'C56:M59', 'C63:M65', 'C69:M74', 'C79:M92', 'C99:M101', 'C107:M109']
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if isinstance(cell.value, str) and is_number(cell.value):
                        # Convert to float after removing commas
                        cell.value = float(cell.value.replace(',', ''))
                        # Apply number format with commas (optional)
                        cell.number_format = '#,##0'
        # for row in ws['C41':'M46']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C50':'M52']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C56':'M59']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C63':'M65']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C69':'M74']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C79':'M92']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C99':'M101']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C107':'M109']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['B1': 'M1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B37':'M37'] + ws['B46':'M46'] + ws['B52':'M52'] + ws['B59':'M59'] + ws['B65':'M65'] + ws['B74':'M74'] + ws['B92':'M92'] + ws['B101':'M101'] + ws['B109':'M109']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['B3':'M3'] + ws['B39':'M39'] + ws['B48':'M48'] + ws['B54':'M54'] + ws['B61':'M61'] + ws['B67':'M67'] + ws['B77':'M77'] + ws['B97':'M97'] + ws['B105':'M105']:
            for cell in row:
                cell.border = black_border_thick
                cell.font = Font(bold=True, size=12)

        for row in ws['G4':'G37'] + ws['G40':'G46'] + ws['G49':'G52'] + ws['G55':'G59'] + ws['G62':'G65'] + ws['G68':'G74'] + ws['G78':'G92'] + ws['G98':'G101'] + ws['G106':'G109']:
            for cell in row:
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        for row in ws['J4':'J37'] + ws['J40':'J46'] + ws['J49':'J52'] + ws['J55':'J59'] + ws['J62':'J65'] + ws['J68':'J74'] + ws['J78':'J92'] + ws['J98':'J101'] + ws['J106':'J109']:
            for cell in row:
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        for row in ws['E4':'F4'] + ws['H4':'I4'] + ws['E40':'F40'] + ws['H40':'I40'] + ws['E49':'F49'] + ws['H49':'I49'] + ws['E55':'F55'] + ws['H55':'I55'] + ws['E62':'F62'] + ws['H62':'I62'] + ws['E68':'F68'] + ws['H68':'I68'] + ws['E78':'F78'] + ws['H78':'I78'] + ws['E98':'F98'] + ws['H98':'I98'] + ws['E106':'F106'] + ws['H106':'I106']:
            for cell in row:
                # cell.fill = PatternFill(start_color='#DCE6F1', end_color='#DCE6F1', fill_type='solid')
                cell.fill = PatternFill(start_color='E0F0F8', end_color='E0F0F8', fill_type='solid')
                
        # make total column bold C, K, L ,M, D
        for row in ws['C5':'C37'] + ws['C41':'C46'] + ws['C50':'C52'] + ws['C56':'C59'] + ws['C63':'C65'] + ws['C69':'C74'] + ws['C79':'C92'] + ws['C99':'C101'] + ws['C107':'C109'] + ws['K5':'K37'] + ws['K41':'K46'] + ws['K50':'K52'] + ws['K56':'K59'] + ws['K63':'K65'] + ws['K69':'K74'] + ws['K79':'K92'] + ws['K99':'K101'] + ws['K107':'K109'] + ws['L5':'L37'] + ws['L41':'L46'] + ws['L50':'L52'] + ws['L56':'L59'] + ws['L63':'L65'] + ws['L69':'L74'] + ws['L79':'L92'] + ws['L99':'L101'] + ws['L107':'L109'] + ws['M5':'M37'] + ws['M41':'M46'] + ws['M50':'M52'] + ws['M56':'M59'] + ws['M63':'M65'] + ws['M69':'M74'] + ws['M79':'M92'] + ws['M99':'M101'] + ws['M107':'M109'] + ws['D5':'D37'] + ws['D41':'D46'] + ws['D50':'D52'] + ws['D56':'D59'] + ws['D63':'D65'] + ws['D69':'D74'] + ws['D79':'D92'] + ws['D99':'D101'] + ws['D107':'D109']:
            for cell in row:
                cell.font = Font(bold=True, size=12)
                
        # wrap text of B50 cell having 'Eligible for the Free/Reduced Price Lunch Program' to fit the cell
        ws['B50'].alignment = Alignment(wrap_text=True)
        
    def main_Reports_1_4_Initials(self):
        title_cells = [
            {"cell": "B1", "value": "Reports 1-4 Initial Referrals Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; Grade Level; Temp Housing Satus and Foster Care Status.", "merge_cells": "B1:M1"},
            

        ]

        subtitle_cells = [
            {"cell": "B3", "value": self.schoolyear + " Students with Initial Referrals by District", "merge_cells": "B3:M3"},
            {"cell": "B39", "value": self.schoolyear + " Students with Initial Referrals by Race/Ethnicity", "merge_cells": "B39:M39"},
            {"cell": "B48", "value": self.schoolyear + " Students with Initial Referrals by Meal Status", "merge_cells": "B48:M48"},
            {"cell": "B54", "value": self.schoolyear + " Students with Initial Referrals by Gender", "merge_cells": "B54:M54"},
            {"cell": "B61", "value": self.schoolyear + " Students with Initial Referrals by English Language Learner (ELL) Status", "merge_cells": "B61:M61"},
            {"cell": "B67", "value": self.schoolyear + " Students with Initial Referrals by Recommended Language of Instruction", "merge_cells": "B67:M67"},
            {"cell": "B77", "value": self.schoolyear + " Students with Initial Referrals by Grade Level", "merge_cells": "B77:M77"},
            {"cell": "B97", "value": self.schoolyear + " Students with Initial Referrals by Temporary Housing Status", "merge_cells": "B97:M97"},
            {"cell": "B105", "value": self.schoolyear + " Students with Initial Referrals by Foster Care Status", "merge_cells": "B105:M105"},
            

        ]

        column_widths = [5, 30, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_byRace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byRace, start_row=41)
        
        # Step 4: Fetch and write data for "Report 8b = IEP Service Recs by District"
        results_byDistrict = self.fetch_data_by_district(cursor)
        # # replace 01 as 1, 02 as 2, etc.
        # results_byDistrict = [(x[0].replace('01', '1'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('02', '2'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('03', '3'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('04', '4'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('05', '5'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('06', '6'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('07', '7'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('08', '8'), *x[1:]) for x in results_byDistrict]
        # results_byDistrict = [(x[0].replace('09', '9'), *x[1:]) for x in results_byDistrict]
        self.write_data_to_excel(ws, results_byDistrict, start_row=5)

        # Step 5: Fetch and write data for "Report 8b = IEP Service Recs by Meal Status"
        results_byMealStatus = self.fetch_data_by_mealstatus(cursor)
        # replace Free or Reduced Price Meal to Eligible for the Free/Reduced Price Lunch Program
        results_byMealStatus = [('Eligible for the Free/Reduced Price Lunch Program' if x[0] == 'Free or Reduced Price Meal' else x[0], *x[1:]) for x in results_byMealStatus]
        self.write_data_to_excel(ws, results_byMealStatus, start_row=50)

        # Step 6: Fetch and write data for "Report 8b = IEP Service Recs by Gender"
        results_byGender = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_byGender, start_row=56)

        # Step 7: Fetch and write data for "Report 8b = IEP Service Recs by ELL Status"
        results_byELLStatus = self.fetch_data_by_ellstatus(cursor)
        # replace 'ELL' with 'Ell' and 'NOT ELL' with 'Non-Ell' 
        results_byELLStatus = [('ELL' if x[0] == 'ELL' else ('NOT ELL' if x[0] == 'Non-Ell' else x[0]), *x[1:]) for x in results_byELLStatus]
        self.write_data_to_excel(ws, results_byELLStatus, start_row=63)
        
        # Step 8: Fetch and write data for "Report 8b = IEP Service Recs by Language"
        results_byLanguage = self.fetch_data_by_language(cursor)
        # replace 'ENGLISH' with 'English', 'SPANISH' with 'Spanish', 'CHINESE' with 'Chinese', 'OTHER' with 'Other', 'UNDETERMINED' with 'Undetermined*'
        results_byLanguage = [('English' if x[0] == 'ENGLISH' else ('Spanish' if x[0] == 'SPANISH' else ('Chinese' if x[0] == 'CHINESE' else ('Other' if x[0] == 'OTHER' else ('Undetermined*' if x[0] == 'UNDETERMINED' else x[0])))), *x[1:]) for x in results_byLanguage]
        self.write_data_to_excel(ws, results_byLanguage, start_row=69)

        # Step 9: Fetch and write data for "Report 8b = IEP Service Recs by Grade Level"
        results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        # replace 01 as 1, 02 as 2, etc.
        results_byGradeLevel = [(x[0].replace('0K', 'KG'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('01', '1'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('02', '2'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('03', '3'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('04', '4'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('05', '5'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('06', '6'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('07', '7'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('08', '8'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('09', '9'), *x[1:]) for x in results_byGradeLevel]        
        self.write_data_to_excel(ws, results_byGradeLevel, start_row=79)

        # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Temporary Housing"
        results_byTempResFlag = self.fetch_data_by_tempResFlag(cursor)
        self.write_data_to_excel(ws, results_byTempResFlag, start_row=99)

        # Step 11: Fetch and write data for "Report 8b = IEP Service Recs by Foster Care Status"
        results_byFosterCareStatus = self.fetch_data_by_fosterCareStatus(cursor)
        # replace 'YES' with 'Yes' and 'NO' with 'No'
        results_byFosterCareStatus = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_byFosterCareStatus]
        self.write_data_to_excel(ws, results_byFosterCareStatus, start_row=107)
        
        # insert text from cell B75 to O75
        ws['B75'] = f'''* The language of instruction recommended on the student's IEP is listed as "undetermined" if the student was determined to be ineligible for an IEP, the case was closed without an IEP meeting, or the case was open as of 6/30/20{self.sqlsnapshottableschoolyear}.'''
        # # Step 12: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

        # Step 13: Close the database connection
        cursor.close()

if __name__ == "__main__":
        Tab1_4 = Solution()
        Tab1_4.main_Reports_1_4_Initials()   
                                                    