import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
import os
class Solution:
    # Existing code...
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Report 14 = BIP School"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Report 14 = BIP District")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=6000, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Primary Program Type',
                'Students with Behavioral Intervention Plan',
                'Percent Students with Behavioral Intervention Plan',
                'Students without Behavioral Intervention Plan',
                'Percent Students without Behavioral Intervention Plan',]
        column_letters = ['C', 'D', 'E', 'F', 'G']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "B8CCE4"
        self.format_header(ws, 'B3', 'School District', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_StudentRegisterR814_061523')
        cursor.execute("[DEV].[USPCCAnnaulReport14c] @tableNameCCStudentRegisterR814=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_tab14(self,cursor):
        query_bytab14 = '''
        Select ReportingDistrict as 'School District' ,Case when PrimaryProgramType = 'No Active Program Services' then 'Related Services or Assistive Technology Only' else PrimaryProgramType end as PrimaryProgramType ,Case when BIP <> 0 and BIP <=5 then '<=5' when BIP >5 and NoBip <> 0 and NoBIP <= 5 then '>5' else BIP end as BIP ,PercentBIP ,
        Case when NoBIP <> 0 and NoBIP <=5 then '<=5' when NoBIP >5 and Bip <> 0 and BIP <= 5 then '>5' else NoBIP end as NoBIP ,PercentNoBIP from  (select distinct ReportingDistrict, PrimaryProgramType  ,cast(Sum(BIP) as varchar) as BIP , CONCAT(CAST(AVG(BIP * 100.00 ) as numeric(7,0)), '%') as 'PercentBIP'  ,cast(sum(NoBIP)as varchar) as NoBIP , CONCAT(CAST(AVG(NoBIP * 100.00 ) as numeric(7,0)), '%') as 'PercentNoBIP'  
        from(Select a.studentid  ,ReportingDistrict   ,PrimaryProgramType   ,case when BIPFlagCC = 'Y' then 1   else 0   end as 'BIP'   ,case when BIPFlagCC= 'N' then 1   else 0    end as 'NoBIP'   FROM SEO_MART.snap.CC_StudentRegisterR814_061523 as a) as a    
        group by  ReportingDistrict, PrimaryProgramType  ) as a union all  select * from (  select  'Total' as district, Null as  PrimaryProgramType  ,FORMAT(Sum(BIP) , '#,##0') as c1  , CONCAT(CAST(AVG(BIP * 100.00 ) as numeric(7,0)), '%')  as c2  ,FORMAT(sum(NoBIP), '#,##0') as c3  , CONCAT(CAST(AVG(NoBIP * 100.00 ) as numeric(7,0)), '%') as c4 
        from(Select a.studentid  ,PrimaryProgramType  ,case when BIPFlagCC = 'Y' then 1   else 0  end as 'BIP' ,case when BIPFlagCC= 'N' then 1   else 0   end as 'NoBIP' 
        FROM SEO_MART.snap.CC_StudentRegisterR814_061523 as a) as a   )  as total  order by 1,2 
        '''  # the bytab12 SQL query goes here
        cursor.execute(query_bytab14)
        results_bytab13 = cursor.fetchall()
        return results_bytab13

    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C4':'G6000']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['D4':'G132']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['B1': 'G1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B132':'G132'] :
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)
    def Report_14_BIP_District(self):
        title_cells = [
            {"cell": "B1", "value": "Report 14 Number and Percentage of Students with a Behavioral Intervention Plan by School", "merge_cells": "B1:G1"},
            

        ]

        subtitle_cells = [
            # {"cell": "B3", "value": "SY 2022-23 Inclusion of Students with IEPs", "merge_cells": "B3:D3"},           

        ]

        column_widths = [5, 30, 70, 30, 30, 30, 30]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by tab13"
        results_bytab14= self.fetch_data_by_tab14(cursor)
        self.write_data_to_excel(ws, results_bytab14, start_row=4)

        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

    def save_or_append_to_workbook(self, wb, save_path):
        # Check if the file already exists
        if os.path.exists(save_path):
            # Load the existing workbook
            book = openpyxl.load_workbook(save_path)
            # Get the active worksheet to copy its styles for the new worksheet
            active_sheet = book.active
            # Create a new worksheet by copying the active one
            new_sheet = book.copy_worksheet(active_sheet)
            # Set the title for the new worksheet
            new_sheet.title = wb.active.title
            # Now, copy the data from wb to new_sheet
            for row in wb.active.iter_rows():
                for cell in row:
                    new_sheet[cell.coordinate].value = cell.value
        else:
            # If the file does not exist, save the new workbook as it is
            book = wb
        # Save the workbook
        book.save(save_path)

if __name__ == "__main__":
        Tab14c = Solution()
        Tab14c.Report_14_BIP_District() 



                                                                        