import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    def __init__(self, datestamp='06/17/2024'):
        self.datestamp = datestamp
        self.lastrow = 20 #19
        self.ProcessedDate = '06-17-2024'
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Classification"
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\ELLCAP_Bilingual_Report_06-17-2024.xlsx')
        ws = wb.create_sheet("Classification")

        # # Set fill color for cells from A1 to Zn to white
        # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
        #     for cell in row:
        #         cell.fill = white_fill

        black_border,_,black_border_medium, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_medium

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['# of ELLs with Program Recommendations on IEPs', 
                   'Total',
                'ICT D1-32,79',
                'SC D1-32,79',
                # 'ICT D1-32,79',
                'SETSS D1-32,79',
                'Multiple Programs D1-32,79',
                'D75',
                'Total',
                'ICT D1-32,79',
                'SC D1-32,79',
                'SETSS D1-32,79',
                'Multiple Programs D1-32,79',
                'D75']
        column_letters = ['B','C', 'D', 'E', 'F', 'G', 'H','I', 'J', 'K', 'L', 'M','N']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "F2F2F2"
        column_fill_color = "F2F2F2"
        self.format_header(ws, 'A5', "Classification", columns, column_letters, 60, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_medium, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database and use self.ProcessedDate to filter the data
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_InitialReferralsR19_SY23')
        cursor.execute(
        '''     
        If Object_ID('tempdb..#BSEReg') is not null
        Drop Table #BSEReg 

        SELECT [StudentID]
            ,[OfficialClass]
            ,[EnrolledDBN]
            ,[AdminDistrict]
            ,[LEPFlag]
            ,[Classification]
            ,[LocationCategoryDescription]
            ,[GradeLevel]
            ,[ServiceLocation]
            ,[PSOutcomeLanguage]
            ,[PSOutcomeLanguageCC]
            ,[BilingualProgramRecommendation] AS [BilingualProgramRec]
            ,[IsBilingualSC] AS BilingualSC
            ,[IsBilingualICT] AS BilingualICT
            ,[SpecializedProgram]
            ,[ReceivingBilingualPrograms] AS ReceivingBilingual
            ,[ProcessedDate]
            ,[ProcessedDateTime]
            ,[SchoolYear]
            ,[IsBilingualSETSS] AS BilingualSETSS

            INTO #BSEReg
        From  [SEO_MART].[arch].[RPT_ELLCAPBilingualPS]
		Where ProcessedDate = '{self.ProcessedDate}



        ----- Create ELL CAP report 




        If Object_ID('tempdb..#Register') is not null
        Drop Table #Register 

        SELECT CAP.[StudentID]
            ,CAP.[OfficialClass]
            ,CAP.[EnrolledDBN]
            ,CAP.[AdminDistrict]
            ,CAP.[LEPFlag]
            ,CAP.[Classification]
            ,CAP.[LocationCategoryDescription]
            ,CAP.[GradeLevel]
            ,CAP.[ServiceLocation]
            ,CAP.[PSOutcomeLanguage]
            ,CAP.[PSOutcomeLanguageCC]
            ,case when CAP.GradeLevel = '0k' then 1
            when CAP.GradeLevel = '01' then 2
            when CAP.GradeLevel = '02' then 3
            when CAP.GradeLevel = '03' then 4
            when CAP.GradeLevel = '04' then 5
            when CAP.GradeLevel = '05' then 6
            when CAP.GradeLevel = '06' then 7
            when CAP.GradeLevel = '07' then 8
            when CAP.GradeLevel = '08' then 9
            when CAP.GradeLevel = '09' then 10
            when CAP.GradeLevel = '10' then 11
            when CAP.GradeLevel = '11' then 12
            when CAP.GradeLevel = '12' then 13
            end as GradeSort
            ,[BilingualProgramRec]
            ,[BilingualSC]
            ,[BilingualICT]
            ,[BilingualSETSS]
            ,CAP.[SpecializedProgram]
            ,ReceivingBilingual
            ,CAP.[ProcessedDate]
            ,CAP.[ProcessedDateTime]
            ,case when loc.boroughcode = 'O' then 'Q' else loc.boroughcode end as boroughcode
        into #Register
        FROM [SEO_MART].[arch].[RPT_PSProvisioningStudent] as CAP
        left join #BSEReg as NEWCAP on CAP.StudentID = NEWCAP.StudentID
        left join [SEO_MART].[dbo].[RPT_Locations] as loc on cap.enrolleddbn = loc.schooldbn 
        where CAP.ELLStatus = 'ELL' 
		and CAP.ProcessedDate = '{self.ProcessedDate}'
        '''
        )
        return cursor
    
    def fetch_data_by_Classification(self,cursor):
        query_byClassification = '''
        --Classification


        
        Select 
        a.[Classification]
        ,count(studentid) as 'ELLs'
        ,TotalBIL
        ,BilICT
        ,BilSC
        ,BilSETSS
        ,BilMulti
        ,D75
        ,TotalReceiving
        ,ReceivingBilICT
        ,ReceivingBilSC
        ,ReceivingBilSETSS
        ,ReceivingBilMulti
        ,ReceivingD75
        FROM #Register as a
        left join( select Classification
                    ,count(studentid) as 'TotalBIL'
                    FROM #Register
                    where [BilingualProgramRec] is not null 
                    group by Classification) as z on a.Classification = z.Classification
        left join( select Classification
                    ,count(studentid) as 'BilICT'
                    FROM #Register
                    where [BilingualProgramRec] = 'Integrated Co-Teaching Services'
                    and AdminDistrict <> 75
                    group by Classification) as b on a.Classification = b.Classification
        left join( select Classification
                    ,count(studentid) as 'BilSC'
                    FROM #Register
                    where [BilingualProgramRec] = 'Special Class'
                    and AdminDistrict <> 75
                    group by Classification) as c on a.Classification = c.Classification
        left join( select Classification
                    ,count(studentid) as 'BilSETSS'
                    FROM #Register
                    where [BilingualProgramRec] = 'SETSS'
                    and AdminDistrict <> 75
                    group by Classification) as J on a.Classification = J.Classification
        left join( select Classification
                    ,count(studentid) as 'BilMulti'
                    FROM #Register 
                    where [BilingualProgramRec] = 'Multiple Programs'
                    and AdminDistrict <> 75
                    group by Classification) as d on a.Classification = d.Classification
        left join( select Classification
                    ,count(studentid) as 'D75'
                    FROM #Register 
                    where [BilingualProgramRec] is not null
                    and AdminDistrict = 75
                    group by Classification) as e on a.Classification = e.Classification
        left join( select Classification
                    ,count(studentid) as 'TotalReceiving'
                    FROM #Register
                    where [BilingualProgramRec] is not null 
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by Classification) as Y on a.Classification = Y.Classification
        left join( select Classification
                    ,count(studentid) as 'ReceivingBilICT'
                    FROM #Register
                    where  [BilingualProgramRec] = 'Integrated Co-Teaching Services'
                    and AdminDistrict <> 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by Classification) as f on a.Classification = f.Classification
        left join( select Classification
                    ,count(studentid) as 'ReceivingBilSC'
                    FROM #Register
                    where [BilingualProgramRec] = 'Special Class'
                    and AdminDistrict <> 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by Classification) as g on a.Classification= g.Classification
        left join( select Classification
                    ,count(studentid) as 'ReceivingBilSETSS'
                    FROM #Register
                    where [BilingualProgramRec] = 'SETSS'
                    and AdminDistrict <> 75
                    and  ReceivingBilingual  = 'FULLY RECEIVING'
                    group by Classification) as K on a.Classification= K.Classification
        left join( select Classification
                    ,count(studentid) as 'ReceivingBilMulti'
                    FROM #Register 
                    where [BilingualProgramRec] = 'Multiple Programs'
                    and AdminDistrict <> 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by Classification) as h on a.Classification= h.Classification
        left join( select Classification
                    ,count(studentid) as 'ReceivingD75'
                    FROM #Register 
                    where [BilingualProgramRec] is not null
                    and AdminDistrict = 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by Classification) as i on a.Classification = i.Classification
        group by 
        a.Classification
        ,TotalBIL
        ,BilICT
        ,BilSC
        ,BilSETSS
        ,BilMulti
        ,D75
        ,TotalReceiving
        ,ReceivingBilICT
        ,ReceivingBilSC
        ,ReceivingBilSETSS
        ,ReceivingBilMulti
        ,ReceivingD75
        '''
        cursor.execute(query_byClassification)
        results_byDistrict= cursor.fetchall()
        return results_byDistrict  
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_border_right_side = Border(right=black_border_mediumside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_up_down = Border(top=black_border_thickside, bottom=black_border_thickside)
        black_border_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 1)  # +2 because data starts from column 'A'
                ws[col + str(row_num)].value = value
                # ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='right')  # Right align the data
        
        # Apply borders to all columns
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for col in ['B', 'H', 'N']:  # Columns B, H, N
                ws[col + str(row_num)].border = black_border_right_side   # Apply the right border

        # Update alignment for range C6:N38
        for row in ws['B6':'N18']:
            for cell in row:
                if cell.value is None:
                    cell.value = '-'  # Replace None with '-'
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        for row in ws['A1': 'N1']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['A5': 'N5']:
            for cell in row:
                if cell.value is not None:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for row in ws['A6': 'A'+str(self.lastrow)]:
            for cell in row:
                cell.font = Font(bold=True, size=12)

        for row in ws['A'+str(self.lastrow):'N'+str(self.lastrow)]:
            # make font bold
            for cell in row:
                cell.font = Font(bold=True, size=12)

        fill_color = "F2F2F2"  # Color for the Total columns and row
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        ws.auto_filter.ref = "A5:N5"
        
        # Fill 'Total' columns with the specified color
        for col in ['C', 'I']:  # Columns F and L are the 'Total' columns
            for row_num in range(start_row, ws.max_row + 1):
                ws[col + str(row_num)].fill = fill
        
        # Fill 'Total' row with the specified color
        for row in ws.iter_rows(min_col=1, max_col=ws.max_column, min_row=self.lastrow, max_row=self.lastrow):
            for cell in row:
                cell.fill = fill
        # add border to the last row
        for cell in ws['A'+str(self.lastrow):'B'+str(self.lastrow)][0]:  # Only one row, so it's safe to use [0]
            cell.border = black_border_up_down
        for cell in ws['C'+str(self.lastrow):'N'+str(self.lastrow)][0]:  # Only one row, so it's safe to use [0]
            cell.border = black_border_all_medium  

        cell_ranges = ['B6:N'+str(self.lastrow) ]
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, int or float):
                        try:
                            cell.number_format = '#,##0'  # Apply comma format
                        except ValueError:
                            # If the value cannot be converted to int, keep the original value
                            print("Int converting Error")
                            pass
        for row in ws['C4':'H4'] + ws['I4':'N4']:
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',wrap_text=True)

        # change the row height
        ws.row_dimensions[4].height = 40 
        # fill A6 as NULL
        ws['A6'] = 'NULL'
        # insert datestamp to the second row then merge cell A2:Q2 then bold the font
        ws['A2'] = 'As of ' + self.datestamp 
        ws.merge_cells('A2:N2')
        ws['A2'].font = Font(bold=True, size=12)         
                              
    def Report_Grade(self):
        title_cells = [
            {"cell": "A1", "value": "ELLs with IEPs and Bilingual ICT or SC IEP Program Recommendations by District for SY 23-24", "merge_cells": "A1:N1"},
        ]

        subtitle_cells = [
            {"cell": "C4", "value": "# of ELLs with IEPs with Bilingual Program Recommendations", "merge_cells": "C4:H4"},
            {"cell": "I4", "value": "# of ELLs with IEPs with BSE Recommendation Served in a Bilingual Class with a Bilingual Teacher", "merge_cells": "I4:N4"},            

        ]

        column_widths = [40, 20, 10, 15, 15, 15, 15, 15, 10, 15, 15, 15, 15, 15]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        # fill subtitle_cells with color D0CECE
        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # # Step 3: Fetch and write data for "Report DBN"
        # self.fetch_data_by_BESReg(conn)
        # self.fetch_data_by_Register(conn)
        results_bytab1 = self.fetch_data_by_Classification(cursor)
        # add Total row to calculate the total number of ELLs at the end of the list
        total = ['Total']
        for i in range(1, 14):
            # if data = Null, set it to 0; if data is string, set it to NULL
            if results_bytab1 and all(isinstance(row[i], int) for row in results_bytab1):
                total.append(sum([row[i] for row in results_bytab1 if row[i] is not None]))
            elif results_bytab1 and all(isinstance(row[i], str) for row in results_bytab1):
                total.append('')
            else:
                # if the data is mixed with data and nonetype,calculate the sum of the data and set the nonetype to 0
                total.append(sum([row[i] if row[i] is not None else 0 for row in results_bytab1]))
        results_bytab1.append(total)

        self.write_data_to_excel(ws, results_bytab1, start_row=6)
        

        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\ELLCAP_Bilingual_Report_06-17-2024.xlsx'
        wb.save(save_path)
        cursor.close()

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.Report_Grade()                                                                  