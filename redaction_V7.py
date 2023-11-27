import openpyxl

def mask_value_initial(val):
    if isinstance(val, (int, float)) and 0 < val <= 5:
        return '<=5'
    return val

def mask_value_secondary(val):
    if isinstance(val, (int, float)) and 0 < val <= 5:
        return '<=5'
    elif isinstance(val, (int, float)) and val > 5:
        return '>5'
    return val

def mask_range(ws, start_row, start_col, end_row, end_col):
    # Step 1: Initial Masking
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = mask_value_initial(cell.value)

    # Step 2: Secondary Masking
    # pairs = [(5, 6), (8, 9), (7, 10), (4, 11, 12, 13)]
    pairs = [(5, 6), (8, 9), (7, 10)]
    for pair in pairs:
        for row_num in range(start_row, end_row + 1):
            masked_cols = [col for col in pair if ws.cell(row=row_num, column=col).value == '<=5']
            if len(masked_cols) == 1:
                other_cols = [col for col in pair if col not in masked_cols]
                for col in other_cols:
                    ws.cell(row=row_num, column=col).value = mask_value_secondary(ws.cell(row=row_num, column=col).value)

    for col in ws.iter_cols(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        valid_vals = [cell.value for cell in col if isinstance(cell.value, (int, float)) and cell.value != '<=5']
        if sum([cell.value == '<=5' for cell in col]) == 1 and valid_vals:
            min_val = min(valid_vals)
            for cell in col:
                if cell.value == min_val:
                    cell.value = mask_value_secondary(cell.value)

    # Step 3: Masking based on Subtotals
    mask_dependencies = [
        (5, 6, 7),
        (8, 9, 10),
        (7, 10, 11),
        (11, 12, 13, 4, 3)
    ]

    # for deps in mask_dependencies:
    #     result_col = deps[-1]
    #     for row_num in range(start_row, end_row + 1):
    #         if any([ws.cell(row=row_num, column=col).value == '<=5' for col in deps[:-1]]) and ws.cell(row=row_num, column=result_col).value != '<=5':
    #             ws.cell(row=row_num, column=result_col).value = '>5'

def mask_excel_file(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb['Reports 1-4 = Initials']

    # Mask data for the specific ranges
    ranges = [
        (5, 3, 37, 13), (41, 3, 46, 13), (50, 3, 52, 13), (56, 3, 58, 13),
        (62, 3, 75, 13), (79, 3, 92, 13)
    ]

    for r in ranges:
        mask_range(ws, *r)

    # Save the modified workbook
    wb.save(filename)

# Call the function with your filename
mask_excel_file('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Annual Special Education Data Report Unredacted SY21.xlsx')
