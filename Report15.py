import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Report 12 = Program Services"
        ws = wb.create_sheet("Report 15 = Related Services")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Full Encounter',
                'Percent Full Encounter',
                'Partial Encounter',
                'Percent Partial Encounter',
                'No Encounter',
                'Percent No Encounter']
        column_letters = ['C', 'D', 'E', 'F', 'G', 'H']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "B8CCE4"
        self.format_header(ws, 'B4', 'Related Services Recommendation Type', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B16', 'Race/Ethnicity ', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B25', 'Meal Status', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B31', 'Gender', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B38', 'ELL Status', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B44', 'Language of Instruction', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B52', 'Grade Level', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B69', ' Temporary Housing Status', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B75', 'Foster Care Status', columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_RSMandateR13_061523')
        cursor.execute("EXEC [dev].[USPCCAnnaulReport13] @tableNameCCRSMandateR13=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_program(self,cursor):
        query_byProgram = '''
        select * from  (  select distinct  MandatesBilingual  ,  FORMAT(Sum(FullEncounter) , '#,##0') as c1  ,concat(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2  ,FORMAT(sum(PartialEncounter) , '#,##0') as c3  ,concat(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4  ,FORMAT(sum(NoEncounter) , '#,##0') as c5  ,concat(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  from ##CCTotaltemp13 a  group by MandatesBilingual    ) cityide  union all  select * from (  select distinct 'Total' MandatesBilingual  ,  FORMAT(Sum(FullEncounter) , '#,##0') as c1  ,concat(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2  ,FORMAT(sum(PartialEncounter) , '#,##0') as c3  ,concat(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4  ,FORMAT(sum(NoEncounter) , '#,##0') as c5  ,concat(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  from ##CCTotaltemp13 a)  as total 
        '''  # the bytab12 SQL query goes here
        cursor.execute(query_byProgram)
        results_byprogram = cursor.fetchall()
        return results_byprogram
    
    def fetch_data_by_race(self,cursor):
        query_byrace = '''
        select EthnicityGroupCC, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select Ethnicity_sort as sort , EthnicityGroupCC,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by EthnicityGroupCC, Ethnicity_sort ) a  union all  select * from ##TotalRow_Sort13  ) a order by sort 
        '''
        cursor.execute(query_byrace)
        results_byrace = cursor.fetchall()
        return results_byrace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,cursor):
        query_bydistrict = '''
        select * from  ( Select  ReportingDistrict as sort ,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by ReportingDistrict  ) a  union all  select * from ##TotalRow13  order by sort  
        '''  # the byDistrict SQL query goes here
        cursor.execute(query_bydistrict)
        results_bydistrict = cursor.fetchall()
        return results_bydistrict
    
    def fetch_data_by_mealstatus(self,cursor):
        query_bymealstatus = '''
        select * from  ( Select  MealStatusGrouping as sort ,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by MealStatusGrouping  ) a  union all  select * from ##TotalRow13  order by sort  
        '''
        cursor.execute(query_bymealstatus)
        results_bymealstatus = cursor.fetchall()
        return results_bymealstatus
    
    def fetch_data_by_gender(self,cursor):
        query_bygender = ''' 
        select * from  ( Select  Gender as sort ,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by Gender  ) a  union all  select * from ##TotalRow13  order by sort  
        '''
        cursor.execute(query_bygender)
        results_bygender = cursor.fetchall()
        return results_bygender
        
    def fetch_data_by_gradelevel(self,cursor):
        query_bygradelevel = ''' 
        select GradeLevel, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select GradeSort as sort , GradeLevel,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*100/nullif(Count(studentid),0) as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*100/nullif(Count(studentid),0) as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*100/nullif(Count(studentid),0) as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by GradeLevel, GradeSort ) a  union all  select * from ##TotalRow_Sort13  ) a order by sort 
        '''
        cursor.execute(query_bygradelevel)
        results_bygradelevel = cursor.fetchall()
        return results_bygradelevel
    
    def fetch_data_by_language(self,cursor):
        query_bylanguage = ''' 
        select RecommendedLanguage, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select RecommendedLanguageSort as sort , RecommendedLanguage,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by RecommendedLanguage, RecommendedLanguageSort ) a  union all  select * from ##TotalRow_Sort13  ) a order by sort 
        '''
        cursor.execute(query_bylanguage)
        results_bylanguage = cursor.fetchall()
        return results_bylanguage
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byellstatus = '''
        select * from  ( Select  ELLStatus as sort ,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by ELLStatus  ) a  union all  select * from ##TotalRow13  order by sort 
        '''
        cursor.execute(query_byellstatus)
        results_byellstatus = cursor.fetchall()
        return results_byellstatus
    
    def fetch_data_by_tempstatus(self,cursor):
        query_bytempstatus = '''
        select TempResFlag, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select TempResFlagSort as sort , TempResFlag,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by TempResFlag, TempResFlagSort ) a  union all  select * from ##TotalRow_Sort13  ) a order by sort 
        '''
        cursor.execute(query_bytempstatus)
        results_bytempstatus = cursor.fetchall()
        return results_bytempstatus
    
    def fetch_data_by_fostercarestaus(self,cursor):
        query_byfostercarestaus = '''
        select FostercareFlag, c1,c2,c3,c4,c5,c6 from (  select * from  ( Select FosterCareFlagSort as sort , FostercareFlag,FORMAT(Sum(FullEncounter) , '#,##0') as c1 ,CONCAT(cast(Sum(FullEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c2 ,FORMAT(sum(PartialEncounter) , '#,##0') as c3 ,CONCAT(cast(Sum(PartialEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c4 ,FORMAT(sum(NoEncounter) , '#,##0') as c5 ,CONCAT(cast(Sum(NoEncounter)*1.0/nullif(Count(studentid),0)*100 as numeric(7)), '%')  as c6  FROM ##CCTotaltemp13  group by FostercareFlag, FosterCareFlagSort ) a  union all  select * from ##TotalRow_Sort13  ) a order by sort 
        '''
        cursor.execute(query_byfostercarestaus)
        results_byfostercarestaus = cursor.fetchall()
        return results_byfostercarestaus
    

    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C5':'H13'] + ws['C17':'H22'] + ws['C26':'H28'] + ws['C32':'H35'] + ws['C39':'H41'] + ws['C45':'H49'] + ws['C53':'H66'] + ws['C70':'H72'] + ws['C76':'H78']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                if isinstance(cell.value, str):
                    try:
                        cell.value = int(cell.value)
                    except ValueError:
                        # If the value cannot be converted to int, keep the original value
                        pass
        # Function to check if a string represents a valid number
        def is_number(s):
            try:
                float(s.replace(',', ''))  # Try converting after removing commas
                return True
            except ValueError:
                return False

        # Formatting specific cell ranges
        cell_ranges = ['C5:H13', 'C17:H22', 'C26:H28', 'C32:H35', 'C39:H41', 'C45:H49', 'C53:H66', 'C70:H72', 'C76:H78']
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if isinstance(cell.value, str) and is_number(cell.value):
                        # Convert to float after removing commas
                        cell.value = float(cell.value.replace(',', ''))
                        # Apply number format with commas (optional)
                        cell.number_format = '#,##0'
        # for row in ws['C17':'H22']:
        #     for cell in row:
        #         if cell.value is not None:  # Ensure there is a value in the cell
        #             cell.value = str(cell.value) + ''  # Prepend space to the value
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C26':'H28']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C32':'H35']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C39':'H41']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C45':'H49']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C53':'H66']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C70':'H72']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # for row in ws['C76':'H78']:
        #     for cell in row:
        #         if cell.value is not None:
        #             cell.value = str(cell.value) + ''
        #         cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['B1': 'H1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B13':'H13'] + ws['B22':'H22'] + ws['B28':'H28'] + ws['B35':'H35'] + ws['B41':'H41'] + ws['B49':'H49'] + ws['B66':'H66'] + ws['B72':'H72'] + ws['B78':'H78']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['B3':'H3'] + ws['B15':'H15'] + ws['B24':'H24'] + ws['B30':'H30'] + ws['B37':'H37'] + ws['B43':'H43'] + ws['B51':'H51'] + ws['B68':'H68'] + ws['B74':'H74']:
            for cell in row:
                cell.border = black_border_thick
                cell.font = Font(bold=True, size=12)

        # wrap text of B26 cell having 'Eligible for the Free/Reduced Price Lunch Program' to fit the cell
        ws['B26'].alignment = Alignment(wrap_text=True)
        # make cell B1 higher to fit the text
        ws.row_dimensions[1].height = 30   
    def Report_Report_15_Related_Services(self):
        title_cells = [
            {"cell": "B1", "value": "Report 15 Number & Percentage of Related Service Recommendations with Encounter Recorded Disaggregated by: Service Recommendation; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of Instruction; Grade Level; Temp House Status; Foster Care Status.", "merge_cells": "B1:H1"},
            

        ]

        subtitle_cells = [
            {"cell": "B3", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded", "merge_cells": "B3:H3"},
            {"cell": "B15", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded by Race/Ethnicity", "merge_cells": "B15:H15"},    
            {"cell": "B24", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded by Meal Status", "merge_cells": "B24:H24"}, 
            {"cell": "B30", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded by Gender", "merge_cells": "B30:H30"},   
            {"cell": "B37", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded by English Language Learner (ELL) Status", "merge_cells": "B37:H37"}, 
            {"cell": "B43", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded by Recommended Language of Instruction", "merge_cells": "B43:H43"}, 
            {"cell": "B51", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded by Grade Level", "merge_cells": "B51:H51"}, 
            {"cell": "B68", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded by Temporary Housing Status", "merge_cells": "B68:H68"}, 
            {"cell": "B74", "value": "SY 2022-23 Number & Percentage of Related Service Recommendations with Encounter Recorded by Foster Care Status", "merge_cells": "B74:H74"}, 


        ]

        column_widths = [5, 40, 30, 30, 30, 30, 30, 30, 30]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_byprogram= self.fetch_data_by_program(cursor)
        sort_order= {'Counseling Services': 1, 'Counseling Services Bilingual': 2, 'Hearing Education Services': 3, 'Occupational Therapy': 4, 'Physical Therapy': 5, 'Speech-Language Therapy': 6, 'Speech-Language Therapy Bilingual': 7, 'Vision Education Services': 8, 'Total': 9}
        sort_results_byprogram = sorted(results_byprogram, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, sort_results_byprogram, start_row=5)

        # Step 6: fetch and write data by Race/Ethnicity
        results_byrace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byrace, start_row=17)

        results_bymealstatus = self.fetch_data_by_mealstatus(cursor)
        # replace Free or Reduced Price Meal to Eligible for the Free/Reduced Price Lunch Program
        results_bymealstatus = [('Eligible for the Free/Reduced Price Lunch Program' if x[0] == 'Free or Reduced Price Meal' else x[0], *x[1:]) for x in results_bymealstatus]
        self.write_data_to_excel(ws, results_bymealstatus, start_row=26)

        results_bygender = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_bygender, start_row=32)

        results_byellstatus = self.fetch_data_by_ellstatus(cursor)
        results_byellstatus = [('ELL' if x[0] == 'ELL' else ('NOT ELL' if x[0] == 'Non-Ell' else x[0]), *x[1:]) for x in results_byellstatus]
        self.write_data_to_excel(ws, results_byellstatus, start_row=39)

        results_bylanguage = self.fetch_data_by_language(cursor)
        results_bylanguage = [('English' if x[0] == 'ENGLISH' else ('Spanish' if x[0] == 'SPANISH' else ('Chinese' if x[0] == 'CHINESE' else ('Other' if x[0] == 'OTHER' else x[0]))), *x[1:]) for x in results_bylanguage]
        self.write_data_to_excel(ws, results_bylanguage, start_row=45)

        results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        results_byGradeLevel = [(x[0].replace('0K', 'KG'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('01', '1'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('02', '2'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('03', '3'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('04', '4'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('05', '5'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('06', '6'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('07', '7'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('08', '8'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('09', '9'), *x[1:]) for x in results_byGradeLevel]   
        # # Define a dictionary for sorting order
        # sort_order= {'KG': 1, '1': 2, '2': 3, '3': 4, '4': 5, '5': 6, '6': 7, '7': 8, '8': 9, '9': 10, '10': 11, '11': 12, '12': 13, 'Totals': 14}
        # # Sort the list using a lambda function that references the sort_order dictionary KG,01,02...12
        # sort_results_byGradeLevel = sorted(results_byGradeLevel, key=lambda x: sort_order[x[0]])
        self.write_data_to_excel(ws, results_byGradeLevel, start_row=53)

        results_bytempstatus = self.fetch_data_by_tempstatus(cursor)
        results_bytempstatus = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_bytempstatus]        
        self.write_data_to_excel(ws, results_bytempstatus, start_row=70)

        results_byfostercarestaus = self.fetch_data_by_fostercarestaus(cursor)
        results_byfostercarestaus = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_byfostercarestaus]
        self.write_data_to_excel(ws, results_byfostercarestaus, start_row=76)

        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.Report_Report_15_Related_Services()                                                                  