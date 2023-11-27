import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
import pyodbc

# Step 1: Create Excel Report Template

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Report 8b = IEP Service Recs"

# Set fill color for cells from A1 to Zn to white
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):  # max_row=1048576 is the last row in an Excel worksheet
    for cell in row:
        cell.fill = white_fill

# Create solid black border
black_border_side = Side(style='thin', color='000000')
black_border_thickside = Side(style='thick', color='000000')
black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
# Add report title and merge cells

ws['B40'] = 'SY 2022-23 Students with IEP Recommended Services by Ethnicity'
ws.merge_cells('B40:N40')  # Merge subtitle cells
ws['B40'].border = black_border_thick  # Applying the black border to the merged cells
ws['C40'].border = black_border_thick  
ws['D40'].border = black_border_thick  
ws['E40'].border = black_border_thick  
ws['F40'].border = black_border_thick  
ws['G40'].border = black_border_thick  
ws['H40'].border = black_border_thick  
ws['I40'].border = black_border_thick  
ws['J40'].border = black_border_thick  
ws['K40'].border = black_border_thick  
ws['L40'].border = black_border_thick  
ws['M40'].border = black_border_thick  
ws['N40'].border = black_border_thick  


# Style and align the merged title and subtitle
for cell in ['B1', 'B3']:
    ws[cell].font = Font(bold=True, size=12)
    ws[cell].alignment = Alignment(wrap_text=True)

# Adjust column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15
ws.column_dimensions['L'].width = 15
ws.column_dimensions['M'].width = 15
ws.column_dimensions['N'].width = 15


# Header
header_font = Font(bold=True, size=12)
border_bottom = Border(bottom=Side(style='thin'))

ws['B41'] = 'Ethnicity'
ws['B41'].font = header_font
ws['B41'].border = border_bottom
ws['B41'].alignment = Alignment(horizontal='center', vertical='center')  
ws['B41'].fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
ws.row_dimensions[41].height = 30
ws.merge_cells('B41:B42')

# More header formatting based on the attached image
columns = ['Related services only', 'Special Education Teacher Support Services (SETSS)',
           'Integrated Co-Teaching Services', 'Special Class in a Community School', 
           'Special Class in a District 75 school', 'Special Class in a Non-public School Placement']
column_letters = ['C', 'E', 'G', 'I', 'K', 'M']


for col, title in zip(column_letters, columns):
    ws[col + '41'] = title
    ws[col + '41'].font = header_font
    ws[col + '41'].border = border_bottom
    ws[col + '41'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws[col + '41'].fill = PatternFill(start_color="E0F0F8", end_color="E0F0F8", fill_type="solid")
    ws[col + '42'] = '#'
    ws[col + '42'].font = header_font
    ws[col + '42'].border = border_bottom
    ws[col + '42'].alignment = Alignment(horizontal='center', vertical='center') 
    ws[chr(ord(col) + 1) + '42'] = '%'
    ws[chr(ord(col) + 1) + '42'].font = header_font
    ws[chr(ord(col) + 1) + '42'].border = border_bottom
    ws[chr(ord(col) + 1) + '42'].alignment = Alignment(horizontal='center', vertical='center')
    ws[chr(ord(col) + 1) + '42'].fill = PatternFill(start_color="E0F0F8", end_color="E0F0F8", fill_type="solid")

# Merge header cells for '%' and '#' under each main column
for col in column_letters:
    ws.merge_cells(f'{col}41:{chr(ord(col) + 1)}41')



# # Apply borders to 'District', 'Related services only', 'Special Education Teacher Support Services (SETSS)','Integrated Co-Teaching Services', 'Special Class in a Community School', 'Special Class in a District 75 school', 'Special Class in a Non-public School Placement', # and % cells for consistency
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L','M','N']:
    ws[col + '41'].border = black_border
    ws[col + '42'].border = black_border


# Step 2: Connect to the database and fetch data

conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()
params = ('CC_StudentRegisterR814_0615', 'RPT_StudentRegister_0615')
cursor.execute("EXEC [dev].[USPCCAnnaulReport8b] @tableNameCCStudentRegisterR814=?, @tableNameRptStudentRegister0615=?", params)
query_byRace = '''
WITH CTE_byRace AS (
SELECT EthnicityGroupCC, 
       FORMAT(sum(RSOnly), '#,##0') as c1,
	   CONCAT(cast((sum(RSOnly)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c2 ,
	   FORMAT(sum(SETSS), '#,##0') as c3 ,
	   CONCAT(cast((sum(SETSS)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c4 ,
	   FORMAT(sum(ICT), '#,##0') as c5 ,
	   CONCAT(cast((sum(ICT)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c6 ,
	   FORMAT(sum(SpecialClass), '#,##0') as c7 ,
	   CONCAT(cast((sum(SpecialClass)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c8 ,
	   FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,
	   CONCAT(cast((sum(SpecialClassD75)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c10 ,
	   FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,
       CONCAT(CAST((SUM(SpecialClassNPS)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c12,
       EthnicityGroupCC_sort -- Add this column here
FROM ##Report 
GROUP BY EthnicityGroupCC, EthnicityGroupCC_sort -- Group by this column as well

UNION ALL

SELECT 'Total' as EthnicityGroupCC,
       FORMAT(SUM(RSOnly), '#,##0') as c1,
	CONCAT(cast((sum(RSOnly)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c2 ,
	FORMAT(sum(SETSS), '#,##0') as c3 ,
	CONCAT(cast((sum(SETSS)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c4 ,
	FORMAT(sum(ICT), '#,##0') as c5 ,
	CONCAT(cast((sum(ICT)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c6 ,
	FORMAT(sum(SpecialClass), '#,##0') as c7 ,
	CONCAT(cast((sum(SpecialClass)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c8 ,
	FORMAT(sum(SpecialClassD75), '#,##0') as c9 ,
	CONCAT(cast((sum(SpecialClassD75)*100.0)/(select count(*) from ##Report) as numeric(7,1)), '%') as c10 ,
	FORMAT(sum(SpecialClassNPS), '#,##0') as c11 ,
       CONCAT(CAST((SUM(SpecialClassNPS)*100.0)/(SELECT COUNT(*) FROM ##Report) AS numeric(7,1)), '%') as c12,
       6 as EthnicityGroupCC_sort -- This is for the 'Total' row, you can use NULL or any suitable placeholder
FROM ##Report 

--ORDER BY EthnicityGroupCC_sort -- Now, you can order by this column
)
SELECT EthnicityGroupCC,c1,c2,c3,c4,c5,c6,c7,c8,c9,c10,c11,c12 FROM CTE_byRace
ORDER BY EthnicityGroupCC_sort
'''
cursor.execute(query_byRace)
results = cursor.fetchall()

# Step 3: Write data to Excel


for row_num, row_data in enumerate(results, start=43):
    ws.cell(row=row_num, column=2).value = row_data[0] # ReportingDistrict starts at column 'B'
    col_pointer = 3 # Start at column 'C'
    for i in range(1, len(row_data), 2):
        ws.cell(row=row_num, column=col_pointer).value = row_data[i] # Number data
        col_pointer += 1
        ws.cell(row=row_num, column=col_pointer).value = row_data[i+1] # Percentage data
        col_pointer += 1

# Apply black_border_no_bottom to all columns
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M','N']:
    ws[col + '43'].border = black_border_no_bottom
    ws[col + '44'].border = black_border_no_bottom
    ws[col + '45'].border = black_border_no_bottom
    ws[col + '46'].border = black_border_no_bottom
    ws[col + '47'].border = black_border_no_bottom
    ws[col + '48'].border = black_boarder_all


# Update font size for range B4:N38
for row in ws['B41':'N48']:
    for cell in row:
        cell.font = openpyxl.styles.Font(size=10)

# Update alignment for range C6:N38
for row in ws['C43':'N48']:
    for cell in row:
        if cell.value is not None:  # Ensure there is a value in the cell
            cell.value = str(cell.value) + ' '  # Prepend space to the value
        cell.alignment = openpyxl.styles.Alignment(horizontal='right')


# Step 4: Save the report
# Save the report to the specified path
save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Report_8b_IEP_Service_Recs.xlsx'
wb.save(save_path)