"""
Reports 1-4 = Initials subtotal calculation:
E + F = G, H + I = J, G + J = K, D + K + L + M = C
Reports 5-7 = Reevaluations subtotal calculation:
E + F = G, H + I = J, G + J = K, D + K + L  = C
Report 8 = Registers subtotal calculation:
C + D + E + F = G, H + I + J + K = L, G + L = M
SWDs by School : C4+ ...+ C1602 = C1603
Report 8a = Disability class subtotal calculation:
C + D +E + F + G + H + I + J + K + L + M + N + O = P 
Report 9 = Placement: C6+...+C37 = C38, D6+...+D37 = D38, C42+...C46 = C47, D42+...D46 = D47, C51+C52 = C53, D51+D52 = D53, C57+C58 = C59, D57+D58 = D59, C63+C64 = C65, 
D63+D64 = D65, C69+C72 = C73, D69+D72 = D73, C77+C89=D90, D77+D89=D90
Report 10 = LRE-MRE: C6+...C37=C38,D6+...D37=D38,E6+...E37=E38,F6+...F37=F38, C42+...C46=C47,D42+...D46=D47,E42+...E46=E47,F42+...F46=F47,C51+C52=C53,D51+D52=D53,E51+E52=E53,F51+F52=F53,E57+E58=E59,F57+F58=F59,C63+C64=C65,D63+D64=D65,E63+E64=E65,F63+F64=F65,C69+C72=C73,D69+D72=D73,E69+E72=E73,F69+F72=F73,C77+...+C89=C90,D77+...+D89=D90,E77+...+E89=E90,F77+...+F89=F90

Redaction rules: 
Column based masking:
Initial mask: find all data 0<data<=5 and make them as <=5
Then since there are Total calculations for every column, iterate each column if there is only one <=5 in this column to mask the smallest data of rest unmasked data for this column as >5. if min_val == 0 then skip it and mask the next smallest value

For percentage redaction, if the number is masked as `<=5` or `>5` then the adjcent percentage should be masked as `*`.

100% percentage sum redaction, if the sum of the percentage is 100% then mask the smallest numeric and percentage within the same row. Since
Tab `Report 10 = IEP Service Recs`:  D+F+H+J+L+N = 100%
Tab `Report 14 = Programs`: D+F+H = 100%
Tab `Report 14a = Bilingual Programs`: D+F+H = 100%
Tab `Report 15 = Related Services`: D+F+H = 100%
Tab `Report 15a = Transportation`: D+F+H = 100%
Tab `Report 16 = BIP`: D+F = 100% I want to change percentage redaction rule: if there is one percentage cell is redacted as `*` and its adjacent numeric cell is masked as `<=5` or `>5` , then we redact the smallest numeric cell for the rest of the same row of that `*` redacted cell belongs to based on its value, if its value >5 then mask it as `>5` if its value <=5 then mask it as `<=5` and then mask its adjacent percentage cell. For example, check attached picture, for Tab `Report 10 = IEP Service Recs` since   D21+F21+H21+J21+L21+N21 = 100% and M21 is redacted as `<=5` and N21 is redacted as `*`, we need to mask K21 as `<=5` and its adjacent percentage cell * to avoid back track. 

For overall column in every range check, if there is only one masked cell (`<=5` or `>5`) in that column then it's underredaction, we need to mask the smallest value of the rest cell in that column.For example, in tab `Report 9 = Disability class`, N102 is the only `>5 ` of its column within its range it belongs to, in order to avoid the back track we need to redact the smallest value of this column which is N100 (0), N102's range is (100, 3, 102, 16) according to redaction_config_SY24.py

For related services only Bilingual recommendations can be partially encountered. Partially encountered in that case means it was a bilingual recommendation that was provided in English. For the non-bilingual recommendations there is no such thing as "partial encounter" which is why we say "N/A" 

There is different redaction logic for bilingual and non-bilingual related services recommendations 

for the bilingual recommendations the logic is the same as the program redaction logic, for the non-bilingual recommendations (also called "monolingual recommendations") the masking value is redacted, but the percentages can remain since there are only two possible values  
"""
import openpyxl
import os, shutil
from redaction_config_triennial_SY24 import TRIENNIAL_REPORTS_CONFIG_SY24
from redaction_config_triennial_SY240402 import TRIENNIAL_REPORTS_CONFIG_SY240402
class Solution:
    def copyonefile(src,dst):
        shutil.copy(src,dst)
        print('copying one file from {0} to {1} is compelte'.format(src,dst)) 
    mylocalCCfolder = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\CCUnredacted'   
    # copyonefile(r'R:\SEO Analytics\Reporting\City Council\City Council SY24\Triennial Reports\Non-Redacted City Council Triennial Report SY24.xlsx', mylocalCCfolder)
    copyonefile(r'R:\SEO Analytics\Reporting\City Council\City Council SY24\04.02.24 Triannual Report\Non-Redacted City Council Triennial Report_04022024.xlsx', mylocalCCfolder)
    # copyonefile(r'R:\SEO Analytics\Reporting\City Council\City Council SY24\Triennial Reports\Non-Redacted City Council Triennial Report SY24.xlsx',"C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop")
    copyonefile(r'R:\SEO Analytics\Reporting\City Council\City Council SY24\04.02.24 Triannual Report\Non-Redacted City Council Triennial Report_04022024.xlsx',"C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop")
    # percentage redaction including 0% to 100% redaction
    def is_percentage(self, cell):
        # Check if cell's format is for percentage
        if '%' in cell.number_format:
            return True
        # Specifically check for 100% values, which might be stored as 1.0
        if cell.value == 1.0 and '%' in cell.number_format:
            return True
        return False



    
    def mask_value_initial(self,val):
        if isinstance(val, (int, float)) and 0 < val <= 5:
            return '<=5'
        return val

    def mask_value_secondary(self,val):
        if isinstance(val, (int, float)) and 0 < val <= 5:
            return '<=5'
        elif isinstance(val, (int, float)) and val > 5:
            return '>5'
        return val
    
    def mask_value_zero(self,val):
        if isinstance(val, (int, float)) and val == 0:
            return '<=5'
        return val

    def initial_mask(self,ws, start_row, start_col, end_row, end_col):
        # Step 1: Initial Masking
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                # do not mask percentage and 100%
                if not self.is_percentage(cell):
                    cell.value = self.mask_value_initial(cell.value)
        
        # Mask smallest if only one '<=5' in column
        for col in ws.iter_cols(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            valid_vals = [cell.value for cell in col if isinstance(cell.value, (int, float)) and cell.value != '<=5']
            if sum([cell.value == '<=5' for cell in col]) == 1 and valid_vals:
                min_val = min(valid_vals)
                # if min_val == 0 then mask it as <=5 if not then mask it as >5
                if min_val == 0:
                    for cell in col:
                        if cell.value == min_val:
                            cell.value = '<=5'
                            break
                for cell in col:
                    if cell.value == min_val:
                        cell.value = '>5'
                        break

    def PS_column_masking(self,ws, start_row, start_col, end_row, end_col, numeric_percentage_pairs):
        """Full receiving: 1.case when fullyreceiving <> 0 and FullyReceiving <= 5 then '<=5';2.when (notreceiving <=5 and notreceiving <> 0 and FullyReceiving = 0 and PartiallyReceiving <> 0 and PartiallyReceiving >5) then '<=5';3.when (PartiallyReceiving <=5 and PartiallyReceiving <> 0 and FullyReceiving = 0 and NotReceiving <> 0 and NotReceiving >5) then '<=5';else fullyreceiving 
        #  Partial receiving: 1.case when Partiallyreceiving <> 0 and partiallyreceiving <= 5 then '<=5'; 2.when (Partiallyreceiving >5 and notreceiving <> 0 and notreceiving <=5 and fullyreceiving >5) then '>5'; 3.when (notreceiving <=5 and notreceiving <> 0 and PartiallyReceiving = 0 and FullyReceiving <> 0 and FullyReceiving >5) then '<=5'; 4.when (FullyReceiving <=5 and fullyreceiving <> 0 and PartiallyReceiving = 0 and NotReceiving <> 0 and NotReceiving >5) then '<=5' else partiallyreceiving 
        #  No receiving: 1.case when NotReceiving <> 0 and notreceiving <= 5 then '<=5'; 2.when (Partiallyreceiving >5 and notreceiving >5 and fullyreceiving <> 0 and FullyReceiving <=5) then '>5'; 3. when (Partiallyreceiving <=5 and partiallyreceiving <> 0 and notreceiving >5 and fullyreceiving <> 0 and FullyReceiving >5) then '>5'; 4. when (Partiallyreceiving <=5 and partiallyreceiving <> 0 and notreceiving = 0 and fullyreceiving <> 0 and FullyReceiving >5) then '<=5'; 5. when (FullyReceiving <=5 and FullyReceiving <> 0 and notreceiving = 0 and PartiallyReceiving <> 0 and PartiallyReceiving >5) then '<=5' else notreceiving """
        full_receiving_col = numeric_percentage_pairs[0][0]
        partial_receiving_col = numeric_percentage_pairs[1][0]
        not_receiving_col = numeric_percentage_pairs[2][0]
        for row in range(start_row, end_row + 1):
            fully_receiving_cell = ws.cell(row=row, column=full_receiving_col)
            partial_receiving_cell = ws.cell(row=row, column=partial_receiving_col)
            not_receiving_cell = ws.cell(row=row, column=not_receiving_col)
            # fully receiving
            if fully_receiving_cell.value != 0 and fully_receiving_cell.value <= 5:
                fully_receiving_cell.value = '<=5'
            elif not_receiving_cell.value <= 5 and not_receiving_cell.value != 0 and fully_receiving_cell.value == 0 and partial_receiving_cell.value != 0 and partial_receiving_cell.value > 5:
                fully_receiving_cell.value = '<=5'
            elif partial_receiving_cell.value <= 5 and partial_receiving_cell.value != 0 and fully_receiving_cell.value == 0 and not_receiving_cell.value != 0 and not_receiving_cell.value > 5:
                fully_receiving_cell.value = '<=5'
            # partial receiving
            elif partial_receiving_cell.value != 0 and partial_receiving_cell.value <= 5:
                partial_receiving_cell.value = '<=5'
            elif partial_receiving_cell.value > 5 and not_receiving_cell.value != 0 and not_receiving_cell.value <= 5 and fully_receiving_cell.value > 5:
                partial_receiving_cell.value = '>5'
            elif not_receiving_cell.value <= 5 and not_receiving_cell.value != 0 and partial_receiving_cell.value == 0 and fully_receiving_cell.value != 0 and fully_receiving_cell.value > 5:
                partial_receiving_cell.value = '<=5'
            elif fully_receiving_cell.value <= 5 and fully_receiving_cell.value != 0 and partial_receiving_cell.value == 0 and not_receiving_cell.value != 0 and not_receiving_cell.value > 5:
                partial_receiving_cell.value = '<=5'
            # not receiving
            elif not_receiving_cell.value != 0 and not_receiving_cell.value <= 5:
                not_receiving_cell.value = '<=5'
            elif partial_receiving_cell.value > 5 and not_receiving_cell.value > 5 and fully_receiving_cell.value !=0 and fully_receiving_cell.value <= 5:
                not_receiving_cell.value = '>5'
            elif partial_receiving_cell.value <= 5 and partial_receiving_cell.value != 0 and not_receiving_cell.value > 5 and fully_receiving_cell.value != 0 and fully_receiving_cell.value > 5:
                not_receiving_cell.value = '>5'
            elif partial_receiving_cell.value <= 5 and partial_receiving_cell.value != 0 and not_receiving_cell.value == 0 and fully_receiving_cell.value != 0 and fully_receiving_cell.value > 5:
                not_receiving_cell.value = '<=5'
            elif fully_receiving_cell.value <= 5 and fully_receiving_cell.value != 0 and not_receiving_cell.value == 0 and partial_receiving_cell.value != 0 and partial_receiving_cell.value > 5:
                not_receiving_cell.value = '<=5'

    def RS_column_masking(self,ws, start_row, start_col, end_row, end_col, numeric_percentage_pairs):
        """ Full encounter: 
             1.case when (FullEncounter <> 0 and FullEncounter <= 5) then '<=5'; 
             2. when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (NoEncounter <=5 and NoEncounter <> 0 and FullEncounter = 0 and PartialEncounter <> 0 and PartialEncounter >5)) then '<=5'; 
             3.when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (PartialEncounter <=5 and PartialEncounter <> 0 and FullEncounter = 0 and NoEncounter <> 0 and NoEncounter >5)) then '<=5'; 
             4. when (recommendation_type_cell.value not in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and NoEncounter <=5 and NoEncounter <> 0 and FullEncounter <> 0) then '>5' else FullEncounter 
             Partial encounter:
            1.case when PartialEncounter <> 0 and PartialEncounter <= 5 then '<=5'; 
            2.when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (PartialEncounter >5 and NoEncounter <> 0 and NoEncounter <=5 and FullEncounter >5)) then '>5'; 
            3.when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (NoEncounter <=5 and NoEncounter <> 0 and PartialEncounter = 0 and FullEncounter <> 0 and FullEncounter >5)) then '<=5'; 
            4.when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (FullEncounter <=5 and FullEncounter <> 0 and PartialEncounter = 0 and NoEncounter <> 0 and NoEncounter >5)) then '<=5'; 
            5.when (recommendation_type_cell.value not in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual')) then 'N/A' else PartialEncounter 
            No encounter:
            1.case when NoEncounter <> 0 and NoEncounter <= 5 then '<=5'; 
            2.when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (PartialEncounter >5 and NoEncounter >5 and FullEncounter <> 0 and FullEncounter <=5)) then '>5'; 
            3.when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (PartialEncounter <=5 and PartialEncounter <> 0 and NoEncounter >5 and FullEncounter <> 0 and FullEncounter >5)) then '>5'; 
            4.when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (PartialEncounter <=5 and PartialEncounter <> 0 and NoEncounter = 0 and FullEncounter <> 0 and FullEncounter >5)) then '<=5'; 
            5.when (recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and (FullEncounter <=5 and FullEncounter <> 0 and NoEncounter = 0 and PartialEncounter <> 0 and PartialEncounter >5)) then '<=5'; 
            6.when (recommendation_type_cell.value not in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and FullEncounter <=5 and FullEncounter <> 0 and NoEncounter <> 0) then '>5' else NoEncounter """
        full_encounter_col = numeric_percentage_pairs[0][0]
        partial_encounter_col = numeric_percentage_pairs[1][0]
        no_encounter_col = numeric_percentage_pairs[2][0]
        for row_num in range(start_row, end_row + 1):
            if report == 'RS Delivery by Supt':
                recommendation_type_cell = ws.cell(row=row_num, column=3) # Assuming column C contains the recommendation type
            elif report == 'RS Delivery by District' or report == 'RS Delivery by School':
                recommendation_type_cell = ws.cell(row=row_num, column=2)  # Assuming column B contains the recommendation type
            full_encounter_cell = ws.cell(row=row_num, column=full_encounter_col)
            partial_encounter_cell = ws.cell(row=row_num, column=partial_encounter_col)
            no_encounter_cell = ws.cell(row=row_num, column=no_encounter_col)
            # Full encounter
            if full_encounter_cell.value != 0 and full_encounter_cell.value <= 5:
                full_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and no_encounter_cell.value <=5 and no_encounter_cell.value != 0 and full_encounter_cell.value == 0 and partial_encounter_cell.value != 0 and partial_encounter_cell.value > 5:
                full_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and partial_encounter_cell.value <=5 and partial_encounter_cell.value != 0 and full_encounter_cell.value == 0 and no_encounter_cell.value != 0 and no_encounter_cell.value > 5:
                full_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value not in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and no_encounter_cell.value <=5 and no_encounter_cell.value != 0 and full_encounter_cell.value != 0:
                full_encounter_cell.value = '>5'
            # Partial encounter
            elif partial_encounter_cell.value != 0 and partial_encounter_cell.value <= 5:
                partial_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and partial_encounter_cell.value >5 and no_encounter_cell.value != 0 and no_encounter_cell.value <=5 and full_encounter_cell.value >5:
                partial_encounter_cell.value = '>5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and no_encounter_cell.value <=5 and no_encounter_cell.value != 0 and partial_encounter_cell.value == 0 and full_encounter_cell.value != 0 and full_encounter_cell.value >5:
                partial_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and full_encounter_cell.value <=5 and full_encounter_cell.value != 0 and partial_encounter_cell.value == 0 and no_encounter_cell.value != 0 and no_encounter_cell.value >5:
                partial_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value not in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual'):
                partial_encounter_cell.value = 'N/A'
            # No encounter
            elif no_encounter_cell.value != 0 and no_encounter_cell.value <= 5:
                no_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and partial_encounter_cell.value >5 and no_encounter_cell.value >5 and full_encounter_cell.value != 0 and full_encounter_cell.value <=5:
                no_encounter_cell.value = '>5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and partial_encounter_cell.value <=5 and partial_encounter_cell.value != 0 and no_encounter_cell.value >5 and full_encounter_cell.value != 0 and full_encounter_cell.value >5:
                no_encounter_cell.value = '>5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and partial_encounter_cell.value <=5 and partial_encounter_cell.value != 0 and no_encounter_cell.value == 0 and full_encounter_cell.value != 0 and full_encounter_cell.value >5:
                no_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and full_encounter_cell.value <=5 and full_encounter_cell.value != 0 and no_encounter_cell.value == 0 and partial_encounter_cell.value != 0 and partial_encounter_cell.value >5:
                no_encounter_cell.value = '<=5'
            elif recommendation_type_cell.value not in ('Counseling Services Bilingual','Speech-Language Therapy Bilingual') and full_encounter_cell.value <=5 and full_encounter_cell.value != 0 and no_encounter_cell.value != 0:
                no_encounter_cell.value = '>5'







    def redact_percentage_based_on_number_byPS(self, ws, numeric_col, perc_col, start_row, end_row):
        for row in range(start_row, end_row + 1):
            number_cell = ws.cell(row=row, column=numeric_col)
            percentage_cell = ws.cell(row=row, column=perc_col)
            # Check if the numeric cell is masked and the percentage cell is not '100%'
            if number_cell.value in ['<=5', '>5']:
                # If the percentage is not '100%', redact it as '*'
                if not (percentage_cell.value == '100%' or percentage_cell.value == 1.0):
                    percentage_cell.value = '*'
             

    def apply_percentage_redaction_byPS(self,ws, config, start_row, end_row):
        for numeric_col, perc_col in config['numeric_percentage_pairs']:
            self.redact_percentage_based_on_number_byPS(ws, numeric_col, perc_col, start_row, end_row)

            
    def mask_smallest_numeric_and_percentage_byPS(self, ws, start_row, end_row, numeric_percentage_pairs):
        for row_num in range(start_row, end_row + 1):
            # Initialize lists to hold numeric and percentage cells
            numeric_cells = []
            percentage_cells = []
            full_receiving_col = numeric_percentage_pairs[0][0]  # Assuming the first column in the pair is the full receiving column

            # Gather numeric and percentage cells for the current row based on the provided pairs
            for numeric_col, perc_col in numeric_percentage_pairs:
                numeric_cell = ws.cell(row=row_num, column=numeric_col)
                percentage_cell = ws.cell(row=row_num, column=perc_col)
                if isinstance(numeric_cell.value, (int, float)):
                    numeric_cells.append(numeric_cell)
                if percentage_cell.value == '*':
                    percentage_cells.append(percentage_cell)

            # Check if there's at least one percentage cell marked with '*' and at least two numeric cells
            if len(numeric_cells) >= 2 and any(percentage_cell.value == '*' for percentage_cell in percentage_cells):
                # Filter out numeric cells that are already masked and the full receiving column cell
                unmasked_numeric_cells = [cell for cell in numeric_cells if cell.value not in ['<=5', '>5'] and cell.column != full_receiving_col]

                # If there are any unmasked numeric cells that are not in the full receiving column
                if len(unmasked_numeric_cells) == 1:
                    # Mask those cells based on their values
                    for cell in unmasked_numeric_cells:
                        if cell.value <= 5:
                            cell.value = '<=5'
                        else:
                            cell.value = '>5'
                        # Mask the adjacent percentage cell
                        adjacent_percentage_cell = ws.cell(row=row_num, column=cell.column + 1)
                        if adjacent_percentage_cell.value != '100%' and adjacent_percentage_cell.value != 1.0:
                            adjacent_percentage_cell.value = '*'
                        # print(ws.title, f"Masking non-full column cell {cell.coordinate} as {'<=5' if cell.value == '<=5' else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")
                elif len(unmasked_numeric_cells) >= 2:
                    # mask the smallest value of unmasked numeric cells
                    min_val = min([cell.value for cell in unmasked_numeric_cells])
                    max_val = max([cell.value for cell in unmasked_numeric_cells])
                    if min_val != max_val:
                        for cell in unmasked_numeric_cells:
                            if cell.value == min_val:
                                cell.value = '<=5' if min_val <= 5 else '>5'
                                # Mask the adjacent percentage cell
                                adjacent_percentage_cell = ws.cell(row=row_num, column=cell.column + 1)
                                if adjacent_percentage_cell.value != '100%' and adjacent_percentage_cell.value != 1.0:
                                    adjacent_percentage_cell.value = '*'
                                # print(ws.title, f"Masking non-full column cell {cell.coordinate} as {'<=5' if cell.value == '<=5' else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")
                    else:
                        # mask the first cell of unmasked numeric cells
                        cell = unmasked_numeric_cells[0]
                        cell.value = '<=5' if min_val <= 5 else '>5'
                        # Mask the adjacent percentage cell
                        adjacent_percentage_cell = ws.cell(row=row_num, column=cell.column + 1)
                        if adjacent_percentage_cell.value != '100%' and adjacent_percentage_cell.value != 1.0:
                            adjacent_percentage_cell.value = '*'
                        # print(ws.title, f"Masking non-full column cell {cell.coordinate} as {'<=5' if cell.value == '<=5' else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")


    def mask_smallest_numeric_and_percentage_byRS(self, ws, start_row, end_row,numeric_percentage_pairs):
        for row_num in range(start_row, end_row + 1):
            if report == 'RS Delivery by Supt':
                recommendation_type_cell = ws.cell(row=row_num, column=3) # Assuming column C contains the recommendation type
            elif report == 'RS Delivery by District' or report == 'RS Delivery by School':
                recommendation_type_cell = ws.cell(row=row_num, column=2)  # Assuming column B contains the recommendation type
            numeric_cells = [ws.cell(row=row_num, column=col) for col, _ in numeric_percentage_pairs]
            percentage_cells = [ws.cell(row=row_num, column=col) for _, col in numeric_percentage_pairs]
            

            if recommendation_type_cell.value:
                if "Bilingual" in recommendation_type_cell.value:
                    # For bilingual, mask the smallest unmasked numeric cell and its adjacent percentage cell marked with '*'
                    self.mask_smallest_numeric_and_percentage_byPS(ws, start_row, end_row, numeric_percentage_pairs)
                    # self.mask_cells_bilingual(ws, start_row, end_row, numeric_percentage_pairs)
                else:
                    # For non-bilingual, mask the smallest unmasked numeric cell
                    self.mask_cells_non_bilingual(ws, numeric_cells,numeric_percentage_pairs)
                  
            # Additional check for conditions where we need to unmask the <=5 value
            # We need to find the cell with the 100% value
            hundred_percent_cells = [cell for cell in percentage_cells if cell.value == '100%' or cell.value == 1.0]
            for hundred_percent_cell in hundred_percent_cells:
                # Check for 'N/A' and '0%' in the same row
                if 'N/A' in [cell.value for cell in numeric_cells] and any(val in ('0%', 0.0) for val in [cell.value for cell in percentage_cells]):
                    # Get the corresponding numeric cell
                    hundred_percent_numeric_cell = ws.cell(row=row_num, column=hundred_percent_cell.column - 1)
                    if hundred_percent_numeric_cell.value in ['<=5', '>5']:
                        # Restore the original value for the numeric cell adjacent to the '0%' cell
                        zero_percent_cell = ws.cell(row=row_num, column=hundred_percent_cell.column + 4)
                        if zero_percent_cell.value == '0%' or zero_percent_cell.value == 0.0:
                            masked_numeric_cell = ws.cell(row=row_num, column=zero_percent_cell.column - 1)
                            original_value = 0
                            if isinstance(original_value, (int, float)) and original_value <= 5:
                                masked_numeric_cell.value = original_value
                                # print(ws.title, f"Restoring cell {masked_numeric_cell.coordinate} to original value {original_value} due to impossible backtrack scenario.")                    

    def mask_cells_non_bilingual(self, ws, numeric_cells,numeric_percentage_pairs):
        # Identify the full receiving column from the numeric_percentage_pairs
        full_receiving_col = numeric_percentage_pairs[0][0]
        # Filter out cells that are not integers or floats
        masked_numeric_cells = [cell for cell in numeric_cells if cell.value in ['<=5', '>5']]
        unmasked_numeric_cells = [cell for cell in numeric_cells if isinstance(cell.value, (int, float)) and cell.value not in ['<=5', '>5']]
        # Only proceed if there are two or more numeric cells to compare
        if len(masked_numeric_cells) == 1 and unmasked_numeric_cells:
            # Find the smallest numeric cell by value
            smallest_numeric_cell = min(unmasked_numeric_cells, key=lambda cell: cell.value)
            # Check if the smallest unmasked numeric cell is in the full receiving column
            if smallest_numeric_cell.column == full_receiving_col:
                # Find the smallest numeric cell that is not in the full receiving column
                non_full_receiving_numeric_cells = [cell for cell in unmasked_numeric_cells if cell.column != full_receiving_col]
                if len(non_full_receiving_numeric_cells) >=2:
                    smallest_non_full_receiving_numeric_cell = min(non_full_receiving_numeric_cells, key=lambda cell: float(cell.value))
                    # Mask it based on its value
                    smallest_non_full_receiving_numeric_cell.value = '<=5' if float(smallest_non_full_receiving_numeric_cell.value) <= 5 else '>5'
                else:
                    # mask the smallest value of the rest cell in full receiving column
                    smallest_numeric_cell = min(unmasked_numeric_cells, key=lambda cell: cell.value)
                    smallest_numeric_cell.value = '<=5' if smallest_numeric_cell.value <= 5 else '>5'
            else:
                # Find the smallest numeric cell by value
                smallest_numeric_cell = min(unmasked_numeric_cells, key=lambda cell: cell.value)
                # Mask it based on its value
                smallest_numeric_cell.value = '<=5' if smallest_numeric_cell.value <= 5 else '>5'
                # print(ws.title, f"Masking cell {smallest_numeric_cell.coordinate} as {'<=5' if smallest_numeric_cell.value == '<=5' else '>5'}")
        else:
            # print("No unmasked numeric cells found to mask.")
            pass
                                
    # def mask_cells_non_bilingual(self, ws, numeric_cells,numeric_percentage_pairs):
    #     # Identify the full receiving column from the numeric_percentage_pairs
    #     full_receiving_col = numeric_percentage_pairs[0][0]
    #     # Filter out cells that are not integers or floats
    #     masked_numeric_cells = [cell for cell in numeric_cells if cell.value in ['<=5', '>5']]
    #     unmasked_numeric_cells = [cell for cell in numeric_cells if isinstance(cell.value, (int, float)) and cell.value not in ['<=5', '>5'] and cell.column != full_receiving_col]
    #     # Only proceed if there are two or more numeric cells to compare
    #     if len(masked_numeric_cells) == 1 and unmasked_numeric_cells:
    #         if len(unmasked_numeric_cells) == 1:
    #             for cell in unmasked_numeric_cells:
    #                 if cell.value <= 5:
    #                     cell.value = '<=5'
    #                 else:
    #                     cell.value = '>5'
    #         elif len(unmasked_numeric_cells) >= 2:
    #             # mask the smallest value of the rest cell in that column
    #             min_val = min([cell.value for cell in unmasked_numeric_cells])
    #             max_val = max([cell.value for cell in unmasked_numeric_cells])
    #             if min_val != max_val:
    #                 for cell in unmasked_numeric_cells:
    #                     if cell.value == min_val:
    #                         cell.value = '<=5' if min_val <= 5 else '>5'
    #                         print(ws.title, f"Masking cell {cell.coordinate} as {'<=5' if cell.value == '<=5' else '>5'}")
    #             else:
    #                 # mask the first cell of unmasked numeric cells
    #                 cell = unmasked_numeric_cells[0]
    #                 cell.value = '<=5' if min_val <= 5 else '>5'
    #                 print(ws.title, f"Masking cell {cell.coordinate} as {'<=5' if cell.value == '<=5' else '>5'}")
    #     else:
    #         # print("No unmasked numeric cells found to mask.")
    #         pass
       
    def apply_na_redaction(self, ws, na_redaction_config, bilingual_percent_config):
        # Get the numeric and percentage column pairs from the configuration
        numeric_percentage_pairs = bilingual_percent_config['mask_RS_bilingual_percent']

        # Iterate through each row for redaction based on the configuration
        for start_row, recommendation_type_col, partial_encounter_col, percent_partial_col in na_redaction_config['NA_Partcial_Encounter_Redaction']:
            for row in range(start_row, ws.max_row + 1):
                recommendation_type_cell = ws.cell(row=row, column=recommendation_type_col)
                
                # Apply redaction rules for bilingual recommendations
                if recommendation_type_cell.value and "Bilingual" in recommendation_type_cell.value:
                    # Iterate through each numeric and percentage column pair
                    for numeric_col, perc_col in numeric_percentage_pairs:
                        numeric_cell = ws.cell(row=row, column=numeric_col)
                        percentage_cell = ws.cell(row=row, column=perc_col)
                        # print(ws.title, f"Row {row}, Numeric cell: {numeric_cell.coordinate}, Percentage cell: {percentage_cell.coordinate}")
                        # Mask the percentage cell if the numeric cell is masked
                        if numeric_cell.value in ['<=5', '>5']:
                            # If the percentage is not '100%', redact it as '*'
                            if not (percentage_cell.value == '100%' or percentage_cell.value == 1.0):
                                percentage_cell.value = '*'

                # Handle non-bilingual recommendations
                else:
                    partial_encounter_cell = ws.cell(row=row, column=partial_encounter_col)
                    percent_partial_cell = ws.cell(row=row, column=percent_partial_col)
                    # Redact as 'N/A' if applicable
                    if partial_encounter_cell.value == 0:
                        partial_encounter_cell.value = 'N/A'
                    if percent_partial_cell.value in ['0%', 0]:
                        percent_partial_cell.value = 'N/A'


    def mask_by_samecategory_byPS(self, ws, cross_tab_config, unredacted_ws):
        for config in cross_tab_config:
            primary_type_col, full_receiving_col, percent_full_receiving_col = config

        category_dict = {}

        for row in range(3, ws.max_row + 1):  # Assuming row 1 is the header and row 2 is the category
            category = ws.cell(row=row, column=primary_type_col).value
            full_receiving_value = ws.cell(row=row, column=full_receiving_col).value

            if category is None:
                continue

            if category not in category_dict:
                category_dict[category] = {'rows': [], 'values': []}

            category_dict[category]['rows'].append(row)
            category_dict[category]['values'].append(full_receiving_value)

        for category, info in category_dict.items():
            masked_cells = [value for value in info['values'] if value in ['<=5', '>5']]
            if len(masked_cells) >= 2:
                continue
            elif masked_cells.count('<=5') == 1:
                unmasked_values_indices = [index for index, value in enumerate(info['values']) if value not in ['<=5', '>5', None] and value != 0]
                smallest_unmasked_value = None
                smallest_unmasked_index = None

                for index in unmasked_values_indices:
                    row = info['rows'][index]
                    # Check for other masked cells in the same row
                    if any(ws.cell(row=row, column=col).value in ['<=5', '>5'] for col in range(1, ws.max_column + 1)):
                        value = info['values'][index]
                        if smallest_unmasked_value is None or value < smallest_unmasked_value:
                            smallest_unmasked_value = value
                            smallest_unmasked_index = index

                if smallest_unmasked_value is not None:
                    smallest_row = info['rows'][smallest_unmasked_index]
                    ws.cell(row=smallest_row, column=full_receiving_col).value = '<=5' if smallest_unmasked_value <= 5 else '>5'
                    adjacent_percentage_cell = ws.cell(row=smallest_row, column=percent_full_receiving_col)
                    adjacent_percentage_cell.value = '*'
                    # print(ws.title, f"Masking same category cell having other masked cell in the same row {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if smallest_unmasked_value <= 5 else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")
                else:
                    # mask the smallest value of the rest cell in that column
                    unmasked_values = [value for value in info['values'] if value not in ['<=5', '>5', None]]
                    if unmasked_values:
                        min_val = min(unmasked_values)
                        for index, value in enumerate(info['values']):
                            if value == min_val:
                                ws.cell(row=info['rows'][index], column=full_receiving_col).value = '<=5' if min_val <= 5 else '>5'
                                adjacent_percentage_cell = ws.cell(row=info['rows'][index], column=percent_full_receiving_col)
                                adjacent_percentage_cell.value = '*'
                                # print(ws.title, f"Masking same category cell {ws.cell(row=info['rows'][index], column=full_receiving_col).coordinate} as {'<=5' if min_val <= 5 else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")

            elif masked_cells.count('>5') == 1:
                gt5_index = info['values'].index('>5')
                gt5_row = info['rows'][gt5_index]
                numeric_original_value = unredacted_ws.cell(row=gt5_row, column=full_receiving_col).value
                ws.cell(row=gt5_row, column=full_receiving_col).value = numeric_original_value
                # Unmask the adjacent percentage cell
                adjacent_percentage_cell = ws.cell(row=gt5_row, column=percent_full_receiving_col)
                percent_original_value = unredacted_ws.cell(row=gt5_row, column=percent_full_receiving_col).value
                ws.cell(row=gt5_row, column=percent_full_receiving_col).value = percent_original_value
                # print(ws.title, f"Unmasking cell {ws.cell(row=gt5_row, column=full_receiving_col).coordinate} with original value {numeric_original_value} and {adjacent_percentage_cell.coordinate} with original value {percent_original_value}")


    def mask_by_samecategoryanddistrict_byPS(self, ws, cross_tab_config, unredacted_ws):
        for config in cross_tab_config:
            district_col, primary_type_col, full_receiving_col, percent_full_receiving_col = config

        # Create a dictionary to hold categories, districts, and their respective rows and values
        category_district_dict = {}

        # Iterate through the worksheet and populate the dictionary
        for row in range(3, ws.max_row + 1):
            school_dbn = ws.cell(row=row, column=district_col).value
            category = ws.cell(row=row, column=primary_type_col).value
            full_receiving_value = ws.cell(row=row, column=full_receiving_col).value

            if school_dbn is None or category is None:
                continue

            # Extract the district number from the school DBN
            district = school_dbn[:2]

            key = (district, category)
            if key not in category_district_dict:
                category_district_dict[key] = {'rows': [], 'values': []}

            category_district_dict[key]['rows'].append(row)
            category_district_dict[key]['values'].append(full_receiving_value)

        # Iterate over the categories and districts and apply the masking rule
        for key, info in category_district_dict.items():
            district, category = key
            masked_cells = [value for value in info['values'] if value in ['<=5', '>5']]
            if len(masked_cells) >= 2:
                continue
            elif masked_cells.count('<=5') == 1:
                smallest_unmasked_value = None
                smallest_unmasked_index = None
                for index, value in enumerate(info['values']):
                    # skip 0 values
                    if value not in ['<=5', '>5', None, 0]:
                        row = info['rows'][index]
                        # Check for other masked cells in the same row
                        if any(ws.cell(row=row, column=col).value in ['<=5', '>5'] for col in range(1, ws.max_column + 1)):
                            if smallest_unmasked_value is None or value < smallest_unmasked_value:
                                smallest_unmasked_value = value
                                smallest_unmasked_index = index

                if smallest_unmasked_value is not None:
                    smallest_row = info['rows'][smallest_unmasked_index]
                    ws.cell(row=smallest_row, column=full_receiving_col).value = '<=5' if smallest_unmasked_value <= 5 else '>5'
                    adjacent_percentage_cell = ws.cell(row=smallest_row, column=percent_full_receiving_col)
                    adjacent_percentage_cell.value = '*'
                    # print(f"Masked {category} in district {district} on row {smallest_row} with value {'<=5' if smallest_unmasked_value <= 5 else '>5'}")
                    print(ws.title, f"Masking same category and district non-zero cell having other masked cell in the same row {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if ws.cell(row=smallest_row, column=full_receiving_col).value == '<=5' else '>5'}")
                else:
                    # mask the smallest non-zero value of the rest cell in that column
                    unmasked_values = [value for value in info['values'] if value not in ['<=5', '>5', None] and value != 0]
                    if unmasked_values:
                        min_val = min(unmasked_values)
                        for index, value in enumerate(info['values']):
                            if value == min_val:
                                ws.cell(row=info['rows'][index], column=full_receiving_col).value = '<=5' if min_val <= 5 else '>5'
                                adjacent_percentage_cell = ws.cell(row=info['rows'][index], column=percent_full_receiving_col)
                                adjacent_percentage_cell.value = '*'
                                # print(f"Masked {category} in district {district} on row {info['rows'][index]} with value {'<=5' if min_val <= 5 else '>5'}")
                                print(ws.title, f"Masking same category and district non-zero cell {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if ws.cell(row=smallest_row, column=full_receiving_col).value == '<=5' else '>5'}")

            elif masked_cells.count('>5') == 1:
                gt5_index = info['values'].index('>5')
                gt5_row = info['rows'][gt5_index]
                original_value = unredacted_ws.cell(row=gt5_row, column=full_receiving_col).value
                ws.cell(row=gt5_row, column=full_receiving_col).value = original_value
                print(f"Unmasked {category} in district {district} on row {gt5_row} to original value {original_value}")
        





    def mask_by_samecategory_byRS(self, ws, cross_tab_config, unredacted_ws):
        for config in cross_tab_config:
            primary_type_col, full_receiving_col, percent_full_receiving_col = config

        category_dict = {}

        for row in range(3, ws.max_row + 1):  # Assuming row 1 is the header and row 2 is the category
            category = ws.cell(row=row, column=primary_type_col).value
            full_receiving_value = ws.cell(row=row, column=full_receiving_col).value

            if category is None:
                continue

            if category not in category_dict:
                category_dict[category] = {'rows': [], 'values': []}

            category_dict[category]['rows'].append(row)
            category_dict[category]['values'].append(full_receiving_value)

        for category, info in category_dict.items():
            masked_cells = [value for value in info['values'] if value in ['<=5', '>5']]
            if len(masked_cells) >= 2:
                continue
            elif masked_cells.count('<=5') == 1:
                unmasked_values_indices = [index for index, value in enumerate(info['values']) if value not in ['<=5', '>5', None] and value != 0]
                smallest_unmasked_value = None
                smallest_unmasked_index = None

                for index in unmasked_values_indices:
                    row = info['rows'][index]
                    # Check for other masked cells in the same row
                    if any(ws.cell(row=row, column=col).value in ['<=5', '>5'] for col in range(1, ws.max_column + 1)):
                        value = info['values'][index]
                        if smallest_unmasked_value is None or value < smallest_unmasked_value:
                            smallest_unmasked_value = value
                            smallest_unmasked_index = index

                if smallest_unmasked_value is not None:
                    smallest_row = info['rows'][smallest_unmasked_index]
                    ws.cell(row=smallest_row, column=full_receiving_col).value = '<=5' if smallest_unmasked_value <= 5 else '>5'
                    # Mask the adjacent percentage cell if RS Program Type is "Bilingual"
                    if ws.cell(row=smallest_row, column=primary_type_col).value == "Bilingual":
                        adjacent_percentage_cell = ws.cell(row=smallest_row, column=percent_full_receiving_col)
                        adjacent_percentage_cell.value = '*'
                        # print(ws.title, f"Masking same category cell having other masked cell in the same row {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if smallest_unmasked_value <= 5 else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")
                    else:
                        print(ws.title, f"Masking same category cell having other masked cell in the same row {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if smallest_unmasked_value <= 5 else '>5'}")
                else:
                    # mask the smallest value of the rest cell in that column
                    unmasked_values = [value for value in info['values'] if value not in ['<=5', '>5', None]]
                    if unmasked_values:
                        min_val = min(unmasked_values)
                        for index, value in enumerate(info['values']):
                            if value == min_val:
                                ws.cell(row=info['rows'][index], column=full_receiving_col).value = '<=5' if min_val <= 5 else '>5'
                                # Mask the adjacent percentage cell if RS Program Type is "Bilingual"
                                if ws.cell(row=info['rows'][index], column=primary_type_col).value == "Bilingual":
                                    adjacent_percentage_cell = ws.cell(row=info['rows'][index], column=percent_full_receiving_col)
                                    adjacent_percentage_cell.value = '*'
                                    # print(ws.title, f"Masking same category cell {ws.cell(row=info['rows'][index], column=full_receiving_col).coordinate} as {'<=5' if min_val <= 5 else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")
                                else:
                                    print(ws.title, f"Masking same category cell {ws.cell(row=info['rows'][index], column=full_receiving_col).coordinate} as {'<=5' if min_val <= 5 else '>5'}")

            elif masked_cells.count('>5') == 1:
                gt5_index = info['values'].index('>5')
                gt5_row = info['rows'][gt5_index]
                numeric_original_value = unredacted_ws.cell(row=gt5_row, column=full_receiving_col).value
                ws.cell(row=gt5_row, column=full_receiving_col).value = numeric_original_value
                # Unmask the adjacent percentage cell
                adjacent_percentage_cell = ws.cell(row=gt5_row, column=percent_full_receiving_col)
                percent_original_value = unredacted_ws.cell(row=gt5_row, column=percent_full_receiving_col).value
                ws.cell(row=gt5_row, column=percent_full_receiving_col).value = percent_original_value
                # print(ws.title, f"Unmasking cell {ws.cell(row=gt5_row, column=full_receiving_col).coordinate} with original value {numeric_original_value} and {adjacent_percentage_cell.coordinate} with original value {percent_original_value}")

    def mask_by_samecategoryanddistrict_byRS(self, ws, cross_tab_config, unredacted_ws):
        for config in cross_tab_config:
            district_col, primary_type_col, full_receiving_col, percent_full_receiving_col = config

        # Create a dictionary to hold categories, districts, and their respective rows and values
        category_district_dict = {}

        # Iterate through the worksheet and populate the dictionary
        for row in range(3, ws.max_row + 1):
            school_dbn = ws.cell(row=row, column=district_col).value
            category = ws.cell(row=row, column=primary_type_col).value
            full_receiving_value = ws.cell(row=row, column=full_receiving_col).value

            if school_dbn is None or category is None:
                continue

            # Extract the district number from the school DBN
            district = school_dbn[:2]

            key = (district, category)
            if key not in category_district_dict:
                category_district_dict[key] = {'rows': [], 'values': []}

            category_district_dict[key]['rows'].append(row)
            category_district_dict[key]['values'].append(full_receiving_value)

        # Iterate over the categories and districts and apply the masking rule
        for key, info in category_district_dict.items():
            district, category = key
            masked_cells = [value for value in info['values'] if value in ['<=5', '>5']]
            if len(masked_cells) >= 2:
                continue
            elif masked_cells.count('<=5') == 1:
                smallest_unmasked_value = None
                smallest_unmasked_index = None
                for index, value in enumerate(info['values']):
                    if value not in ['<=5', '>5', None]:
                        row = info['rows'][index]
                        # Check for other masked cells in the same row
                        if any(ws.cell(row=row, column=col).value in ['<=5', '>5'] for col in range(1, ws.max_column + 1)):
                            if smallest_unmasked_value is None or value < smallest_unmasked_value:
                                smallest_unmasked_value = value
                                smallest_unmasked_index = index

                if smallest_unmasked_value is not None:
                    smallest_row = info['rows'][smallest_unmasked_index]
                    ws.cell(row=smallest_row, column=full_receiving_col).value = '<=5' if smallest_unmasked_value <= 5 else '>5'
                    # Mask the adjacent percentage cell if RS Program Type is "Bilingual"
                    if ws.cell(row=smallest_row, column=primary_type_col).value == "Bilingual":
                        adjacent_percentage_cell = ws.cell(row=smallest_row, column=percent_full_receiving_col)
                        adjacent_percentage_cell.value = '*'
                        # print(f"Masked {category} in district {district} on row {smallest_row} with value {'<=5' if smallest_unmasked_value <= 5 else '>5'}")
                        # print(ws.title, f"Masking same category and district cell having other masked cell in the same row {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if ws.cell(row=smallest_row, column=full_receiving_col).value == '<=5' else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")
                    else:
                        print(ws.title, f"Masking same category and district cell having other masked cell in the same row {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if smallest_unmasked_value <= 5 else '>5'}")
                else:
                    # mask the smallest value of the rest cell in that column
                    unmasked_values = [value for value in info['values'] if value not in ['<=5', '>5', None] and value != 0]
                    if unmasked_values:
                        min_val = min(unmasked_values)
                        for index, value in enumerate(info['values']):
                            if value == min_val:
                                ws.cell(row=info['rows'][index], column=full_receiving_col).value = '<=5' if min_val <= 5 else '>5'
                                # Mask the adjacent percentage cell if RS Program Type is "Bilingual"
                                if ws.cell(row=info['rows'][index], column=primary_type_col).value == "Bilingual":
                                    adjacent_percentage_cell = ws.cell(row=info['rows'][index], column=percent_full_receiving_col)
                                    adjacent_percentage_cell.value = '*'
                                    # print(f"Masked {category} in district {district} on row {info['rows'][index]} with value {'<=5' if min_val <= 5 else '>5'}")
                                    # print(ws.title, f"Masking same category and district cell {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if ws.cell(row=smallest_row, column=full_receiving_col).value == '<=5' else '>5'} and {adjacent_percentage_cell.coordinate} as '*'")
                                else:
                                    print(ws.title, f"Masking same category and district cell {ws.cell(row=smallest_row, column=full_receiving_col).coordinate} as {'<=5' if min_val <= 5 else '>5'}")

            elif masked_cells.count('>5') == 1:
                gt5_index = info['values'].index('>5')
                gt5_row = info['rows'][gt5_index]
                original_value = unredacted_ws.cell(row=gt5_row, column=full_receiving_col).value
                ws.cell(row=gt5_row, column=full_receiving_col).value = original_value
                print(f"Unmasked {category} in district {district} on row {gt5_row} to original value {original_value}")


    def highlight_overredaction(self, ws, groups, ranges, unredacted_ws):
        for group in groups:
            for (start_row, start_col, end_row, end_col) in ranges:  # Extracting the range from the tuple
                for row_num in range(start_row, end_row + 1):
                    gt5_cells = []
                    lte5_exists = False
                    for col_index in group:
                        cell = ws.cell(row=row_num, column=col_index)
                        if cell.value == '>5':
                            gt5_cells.append(cell)
                        elif cell.value == '<=5':
                            lte5_exists = True

                    # Check if the first '>5' is not the only one in its column within the range
                    if gt5_cells and (len(gt5_cells) > 2 or (len(gt5_cells) == 2 and lte5_exists)):
                        first_gt5_cell = gt5_cells[0]
                        col_values = [ws.cell(row=r, column=first_gt5_cell.column).value for r in range(start_row, end_row + 1)]
                        if col_values.count('>5') > 1:
                            # self.green_cell(first_gt5_cell) # Highligh the overredaction cell in green
                            # Get the original value from the unredacted worksheet
                            original_value = unredacted_ws.cell(row=first_gt5_cell.row, column=first_gt5_cell.column).value
                            # Unmask the cell by setting its value to the original value
                            first_gt5_cell.value = original_value
                            print(ws.title, f"Unmasking overredacted cell {first_gt5_cell.coordinate} with original value")

    def restore_percentage_cells_if_excessive_redaction(self, ws, numeric_percentage_pairs, unredacted_ws):
        for row_num in range(1, ws.max_row + 1):
            # Initialize a counter for masked numeric cells in the row
            masked_numeric_count = 0
            masked_numeric_cells = []
            full_col = numeric_percentage_pairs[0][0]
            partial_column = numeric_percentage_pairs[1][0]
            no_column = numeric_percentage_pairs[2][0]
            # Iterate through each numeric and percentage column pair
            for numeric_col, perc_col in numeric_percentage_pairs:
                numeric_cell = ws.cell(row=row_num, column=numeric_col)
                percentage_cell = ws.cell(row=row_num, column=perc_col)

                # Check if the numeric cell is masked as <=5 or >5 and add it to the list
                if numeric_cell.value in ['<=5', '>5']:
                    masked_numeric_count += 1
                    masked_numeric_cells.append(numeric_cell)

            # If more than three numeric cells are masked in the row
            if masked_numeric_count >= 3:
                # if there are more than two `<=5` in masked numric cells, then restore the percentage cells
                if len([cell for cell in masked_numeric_cells if cell.value == '<=5']) >= 2:
                    # Iterate again to restore the original percentage values where applicable
                    for numeric_col, perc_col in numeric_percentage_pairs:
                        percentage_cell = ws.cell(row=row_num, column=perc_col)
                        if percentage_cell.value == '*':
                            # Retrieve the original value from the unredacted worksheet
                            original_percentage_value = unredacted_ws.cell(row=row_num, column=perc_col).value
                            # Restore the original percentage value
                            percentage_cell.value = original_percentage_value
                            # print(f"Restored original value for cell {percentage_cell.coordinate} in row {row_num}")

                # if there are more than two `>5` in masked numric cells, then restore the first `>5` cell not in the full column to its original value and its adjacent percentage cell. For example, (`<=5`, `*`, `>5`, `*` ,`>5`, `*`) or (`>5`, `*`, `>5`, `*` ,`<=5`, `*`) or (`>5`, `*`, `>5`, `*` ,`>5`, `*`) we restored the `>5` cell in partial column and its adjacent percentage cell; (`>5`, `*`, `<=5`, `*` ,`>5`, `*`) we restored the `>5` cell in no column and its adjacent percentage cell
                elif len([cell for cell in masked_numeric_cells if cell.value == '>5']) >= 2:
                    # Find the first `>5` cell not in the full column
                    first_gt5_cell = next((cell for cell in masked_numeric_cells if cell.value == '>5' and cell.column != full_col), None)
                    if first_gt5_cell:
                        # Retrieve the original value from the unredacted worksheet
                        original_numeric_value = unredacted_ws.cell(row=row_num, column=first_gt5_cell.column).value
                        # Restore the original numeric value
                        first_gt5_cell.value = original_numeric_value
                        # Unmask the adjacent percentage cell
                        adjacent_percentage_cell = ws.cell(row=row_num, column=numeric_percentage_pairs[masked_numeric_cells.index(first_gt5_cell)][1])
                        percent_original_value = unredacted_ws.cell(row=row_num, column=adjacent_percentage_cell.column).value
                        adjacent_percentage_cell.value = percent_original_value
                        # print(f"Unmasked cell {first_gt5_cell.coordinate} with original value {original_numeric_value} and {adjacent_percentage_cell.coordinate} with original value {percent_original_value}")

    def mask_percent_having0_and_100_percent(self, ws, numeric_percentage_pairs, unredacted_ws):
        for row_num in range(1, ws.max_row + 1):
            # Initialize a counter for masked numeric cells in the row
            masked_numeric_count = 0
            # Create a list to store the numeric cells and their corresponding percentage cells
            numeric_and_percentage_cells = []
            
            # Iterate through each numeric and percentage column pair
            for numeric_col, perc_col in numeric_percentage_pairs:
                numeric_cell = ws.cell(row=row_num, column=numeric_col)
                percentage_cell = ws.cell(row=row_num, column=perc_col)
                
                # Add the cells to the list as a tuple
                numeric_and_percentage_cells.append((numeric_cell, percentage_cell))
                
                # Check if the numeric cell is masked as <=5 or >5 and increment the count
                if numeric_cell.value in ['<=5', '>5']:
                    masked_numeric_count += 1

            # If there are exactly two masked numeric cells
            if masked_numeric_count == 2:
                # Check if one percentage is 0% and another is 100%
                zero_percent_cells = [perc_cell for num_cell, perc_cell in numeric_and_percentage_cells if perc_cell.value in ('0%', 0.0)]
                hundred_percent_cells = [perc_cell for num_cell, perc_cell in numeric_and_percentage_cells if perc_cell.value in ('100%', 1.0)]

                if zero_percent_cells and hundred_percent_cells:
                    # Mask the 0% percentage cell with the masked value of its adjacent numeric cell
                    for zero_percent_cell in zero_percent_cells:
                        adjacent_numeric_cell = ws.cell(row=row_num, column=zero_percent_cell.column - 1)
                        if adjacent_numeric_cell.value in ['<=5', '>5']:
                            # Mask the 0% percentage cell with '*'
                            zero_percent_cell.value = '*'
                            # print(ws.title, f"Masking 0% cell {zero_percent_cell.coordinate} as '*' due to adjacent numeric cell {adjacent_numeric_cell.coordinate} being masked.")

                    for hundred_percent_cell in hundred_percent_cells:
                        adjacent_numeric_cell = ws.cell(row=row_num, column=hundred_percent_cell.column - 1)
                        if adjacent_numeric_cell.value in ['<=5', '>5']:
                            # Mask the 100% percentage cell with '*'
                            hundred_percent_cell.value = '*'
                            # print(ws.title, f"Masking 100% cell {hundred_percent_cell.coordinate} as '*' due to adjacent numeric cell {adjacent_numeric_cell.coordinate} being masked.")
                
                # Check if there are two `<=5` and one `N/A` in the same row,if so, unmask the `<=5`'s percentage cell
                if len([cell for cell in numeric_and_percentage_cells if cell[0].value == '<=5']) == 2 and len([cell for cell in numeric_and_percentage_cells if cell[0].value == 'N/A']) == 1:
                    # Retrieve the original value of the `<=5`'s percentage cell from the unredacted worksheet
                    original_value = unredacted_ws.cell(row=row_num, column=numeric_and_percentage_cells[0][1].column).value
                    numeric_and_percentage_cells[0][1].value = original_value
                    original_value = unredacted_ws.cell(row=row_num, column=numeric_and_percentage_cells[2][1].column).value
                    numeric_and_percentage_cells[2][1].value = original_value
                    # print(ws.title, f"Unmasking percent cell {numeric_and_percentage_cells[0][1].coordinate} with original value {original_value}")

            # If there are two 0% percentage cells and one masked numeric cell, unmask masked numeric cell's percentage cell
            if masked_numeric_count == 1:
                zero_percent_cells = [perc_cell for num_cell, perc_cell in numeric_and_percentage_cells if perc_cell.value in ('0%', 0.0)]
                if len(zero_percent_cells) == 2:
                    # unmask the masked numeric cell's percentage cell
                    for numeric_cell, percentage_cell in numeric_and_percentage_cells:
                        if numeric_cell.value in ['<=5', '>5']:
                            # Retrieve the original value of its adjacent percent cell from the unredacted worksheet
                            original_value = unredacted_ws.cell(row=row_num, column=percentage_cell.column).value
                            # Restore the original value
                            percentage_cell.value = original_value

    def unmask_numeric_having0_and_100_percent_non_bilingual(self, ws, numeric_percentage_pairs, unredacted_ws): 
        for row_num in range(1, ws.max_row + 1):
            # Initialize a counter for masked numeric cells in the row
            masked_numeric_count = 0
            # Create a list to store the numeric cells and their corresponding percentage cells
            numeric_and_percentage_cells = []
            
            # Iterate through each numeric and percentage column pair
            for numeric_col, perc_col in numeric_percentage_pairs:
                numeric_cell = ws.cell(row=row_num, column=numeric_col)
                percentage_cell = ws.cell(row=row_num, column=perc_col)
                
                # Add the cells to the list as a tuple
                numeric_and_percentage_cells.append((numeric_cell, percentage_cell))
                
                # Check if the numeric cell is masked as <=5 or >5 and increment the count
                if numeric_cell.value in ['<=5', '>5']:
                    masked_numeric_count += 1

            # If there are exactly two masked numeric cells and one `N/A` in the same row
            if masked_numeric_count == 2 and len([cell for cell in numeric_and_percentage_cells if cell[0].value == 'N/A']) == 1:
                # Check if one percentage is 0% and another is 100%
                zero_percent_cells = [perc_cell for num_cell, perc_cell in numeric_and_percentage_cells if perc_cell.value in ('0%', 0.0)]
                hundred_percent_cells = [perc_cell for num_cell, perc_cell in numeric_and_percentage_cells if perc_cell.value in ('100%', 1.0)]

                if zero_percent_cells and hundred_percent_cells:
                    # restore the original value of the `<=5`
                    for numeric_cell, percentage_cell in numeric_and_percentage_cells:
                        if numeric_cell.value in ['<=5'] and percentage_cell.value in ['0%', 0.0]:
                            # Retrieve the original value of its adjacent percent cell from the unredacted worksheet
                            original_value = unredacted_ws.cell(row=row_num, column=numeric_cell.column).value
                            # Restore the original value
                            numeric_cell.value = original_value
                            # print(ws.title, f"Unmasking percent cell {percentage_cell.coordinate} with original value {original_value}")



            

                


    def mask_0_percent_in_full(self, ws, numeric_percentage_pairs):
        for row_num in range(1, ws.max_row + 1):
            # Initialize a counter for masked numeric cells in the row
            masked_numeric_count = 0
            full_receiving_col_index = numeric_percentage_pairs[0][0]
            percent_full_receiving_col_index = numeric_percentage_pairs[0][1]
            # Gather cell objects for the full receiving column and its corresponding percentage
            full_receiving_cell = ws.cell(row=row_num, column=full_receiving_col_index)
            percent_full_receiving_cell = ws.cell(row=row_num, column=percent_full_receiving_col_index)

            # Iterate through each numeric and percentage column pair
            for numeric_col, _ in numeric_percentage_pairs:
                numeric_cell = ws.cell(row=row_num, column=numeric_col)
                # Check if the numeric cell is masked as <=5 or >5
                if numeric_cell.value in ['<=5', '>5']:
                    masked_numeric_count += 1

            # Check conditions and mask as necessary
            if (masked_numeric_count >= 3 and 
                full_receiving_cell.value == '<=5' and 
                percent_full_receiving_cell.value in ('0%', 0.0)):
                # Mask percent_full_receiving_cell as `*`
                percent_full_receiving_cell.value = '*'
                print(ws.title, f"Masking Percent cell {percent_full_receiving_cell.coordinate} as '*'")

    def overredaction_0(self, ws, numeric_percentage_pairs, unredacted_ws):
        for row_num in range(1, ws.max_row + 1):
            # Initialize a counter for masked numeric cells in the row
            masked_numeric_count = 0
            masked_numeric_cells = []
            full_col = numeric_percentage_pairs[0][0]
            partial_column = numeric_percentage_pairs[1][0]
            no_column = numeric_percentage_pairs[2][0]
            # Iterate through each numeric and percentage column pair
            for numeric_col, perc_col in numeric_percentage_pairs:
                numeric_cell = ws.cell(row=row_num, column=numeric_col)
                percentage_cell = ws.cell(row=row_num, column=perc_col)

                # Check if the numeric cell is masked as <=5 or >5 and add it to the list
                if numeric_cell.value in ['<=5', '>5']:
                    masked_numeric_count += 1
                    masked_numeric_cells.append(numeric_cell)

            # If more than three numeric cells are masked in the row
            if masked_numeric_count >= 3:
                if len([cell for cell in masked_numeric_cells if cell.value == '<=5']) >= 2:
                    # if there is a `<=5` and its percentage cell is 0%, then restore the masked `<=5` cell to its original value 
                    if (ws.cell(row=row_num, column=full_col).value == '<=5' and ws.cell(row=row_num, column=numeric_percentage_pairs[0][1]).value in ('0%', 0.0)):
                        # Retrieve the original value from the unredacted worksheet
                        original_value = unredacted_ws.cell(row=row_num, column=full_col).value
                        # Restore the original value
                        ws.cell(row=row_num, column=full_col).value = original_value
                        print(ws.title, f"Unmasking full cell {ws.cell(row=row_num, column=full_col).coordinate} with original value {original_value}")
                    elif (ws.cell(row=row_num, column=partial_column).value == '<=5' and ws.cell(row=row_num, column=numeric_percentage_pairs[1][1]).value in ('0%', 0.0)):
                        # Retrieve the original value from the unredacted worksheet
                        original_value = unredacted_ws.cell(row=row_num, column=partial_column).value
                        # Restore the original value
                        ws.cell(row=row_num, column=partial_column).value = original_value
                        print(ws.title, f"Unmasking partial cell {ws.cell(row=row_num, column=partial_column).coordinate} with original value {original_value}")
                    elif (ws.cell(row=row_num, column=no_column).value == '<=5' and ws.cell(row=row_num, column=numeric_percentage_pairs[2][1]).value in ('0%', 0.0)):
                        # Retrieve the original value from the unredacted worksheet
                        original_value = unredacted_ws.cell(row=row_num, column=no_column).value
                        # Restore the original value
                        ws.cell(row=row_num, column=no_column).value = original_value
                        print(ws.title, f"Unmasking no cell {ws.cell(row=row_num, column=no_column).coordinate} with original value {original_value}")


    def mask_excel_file(self,filename,tab_name,configurations,unredacted_filename):
        # Load the workbooks
        wb = openpyxl.load_workbook(filename)
        unredacted_wb = openpyxl.load_workbook(unredacted_filename, data_only=True)
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        try:
            ws = wb[tab_name]
            unredacted_ws = unredacted_wb[tab_name]  # Define unredacted_ws here
        except KeyError:
            print(f"Warning: Worksheet {tab_name} does not exist in the file {filename}. Skipping...") #SWDs by School is not in SY23
            return

        # # Convert string to int where possible
        # for r in configurations['ranges']:
        #     for row in ws.iter_rows(min_row=r[0], max_row=r[2], min_col=r[1], max_col=r[3]):
        #         for cell in row:
        #             if isinstance(cell.value, str):
        #                 try:
        #                     cell.value = int(cell.value)
        #                 except ValueError:
        #                     # If the value cannot be converted to int, keep the original value
        #                     pass
                        
        # 1. Mask data for the specific ranges
        for r in configurations['ranges']:
            if 'numeric_percentage_pairs' in configurations and 'PS_flag' in configurations and configurations['PS_flag'] == True:
                numeric_percentage_pairs = configurations['numeric_percentage_pairs']
                self.PS_column_masking(ws, *r, numeric_percentage_pairs)
            if 'numeric_percentage_pairs' in configurations and 'RS_flag' in configurations and configurations['RS_flag'] == True:
                numeric_percentage_pairs = configurations['numeric_percentage_pairs']
                self.RS_column_masking(ws, *r, numeric_percentage_pairs)

        for r in configurations['ranges']:
            self.initial_mask(ws, *r)

        # 2. Apply N/A redaction based on new configuration if the key exists
        if 'NA_Partcial_Encounter_Redaction' in configurations:
            self.apply_na_redaction(ws, configurations, configurations)


                                   
        # 3. Apply the percentage redaction based on configuration if the key exists for PS reports
        for r in configurations['ranges']:
            if 'numeric_percentage_pairs' in configurations and 'PS_flag' in configurations and configurations['PS_flag'] == True:
                self.apply_percentage_redaction_byPS(ws, configurations, r[0], r[2])

        for r in configurations['ranges']:
            if 'numeric_percentage_pairs' in configurations and 'RS_flag' in configurations and configurations['RS_flag'] == True:
                for row_num in range(r[0], r[2] + 1):
                    if report == 'RS Delivery by Supt':
                        recommendation_type_cell = ws.cell(row=row_num, column=3)
                        # print(f'RS Delivery by Supt Row {row_num} - Recommendation Type: {recommendation_type_cell.value}')
                    elif report == 'RS Delivery by District' or report == 'RS Delivery by School':
                        recommendation_type_cell = ws.cell(row=row_num, column=2)
                    
                    # Now we check the value of recommendation_type_cell
                    if recommendation_type_cell.value:
                        if "Bilingual" in recommendation_type_cell.value:
                            # self.apply_percentage_redaction_byPS(ws, configurations, r[0], r[2])
                            # print(f"Applying {ws.title} for row {row_num} percentage redaction")
                            pass


        # 4. Apply the 100% percentage sum redaction based on configuration if the key exists
        for r in configurations['ranges']:
            if '100_percentage_sum' in configurations and 'PS_flag' in configurations and configurations['PS_flag'] == True:
                # Retrieve the correct numeric_percentage_pairs for the current tab
                numeric_percentage_pairs = configurations['numeric_percentage_pairs']
                # Call the function with the correct pairs
                self.mask_smallest_numeric_and_percentage_byPS(ws, r[0], r[2], numeric_percentage_pairs)
            if '100_percentage_sum' in configurations and 'RS_flag' in configurations and configurations['RS_flag'] == True:
                # Retrieve the correct numeric_percentage_pairs for the current tab
                numeric_percentage_pairs = configurations['numeric_percentage_pairs']
                # Call the function with the correct pairs
                self.mask_smallest_numeric_and_percentage_byRS(ws, r[0], r[2], numeric_percentage_pairs)  
    
        # 5. Apply mask by category redaction based on new configuration if the key exists
        if 'mask_by_category' in configurations and 'mask_by_district' not in configurations and 'PS_flag' in configurations and configurations['PS_flag'] == True:
            self.mask_by_samecategory_byPS(ws, configurations['mask_by_category'], unredacted_ws)

        # 6. Apply mask by district redaction based on new configuration if the key exists
        if 'mask_by_district' in configurations and 'mask_by_category' in configurations and 'PS_flag' in configurations and configurations['PS_flag'] == True:
            self.mask_by_samecategoryanddistrict_byPS(ws, configurations['mask_by_district'], unredacted_ws)

        # Apply mask by category redaction based on new configuration if the key exists
        if 'mask_by_category' in configurations and 'mask_by_district' not in configurations and 'RS_flag' in configurations and configurations['RS_flag'] == True:
            self.mask_by_samecategory_byRS(ws, configurations['mask_by_category'], unredacted_ws)

        # Apply mask by category and district redaction based on new configuration if the key exists
        if 'mask_by_category' in configurations and 'mask_by_district' in configurations and 'RS_flag' in configurations and configurations['RS_flag'] == True:
            self.mask_by_samecategoryanddistrict_byRS(ws, configurations['mask_by_district'], unredacted_ws)

        # 7. unmask the adjacent percentage cells
        for r in configurations['ranges']:
            self.restore_percentage_cells_if_excessive_redaction(ws, configurations['numeric_percentage_pairs'], unredacted_ws) 
            self.mask_0_percent_in_full(ws, configurations['numeric_percentage_pairs'])
            self.mask_percent_having0_and_100_percent(ws, configurations['numeric_percentage_pairs'], unredacted_ws)
            self.overredaction_0(ws, configurations['numeric_percentage_pairs'], unredacted_ws)
        for r in configurations['ranges']:
            if 'numeric_percentage_pairs' in configurations and 'RS_flag' in configurations and configurations['RS_flag'] == True:
                self.unmask_numeric_having0_and_100_percent_non_bilingual(ws, configurations['numeric_percentage_pairs'], unredacted_ws)
        # # Mask underredacted columns
        # for r in configurations['ranges']:
        #     self.check_and_mask_underredacted_columns(ws, r[0], r[2], r[1], r[3])
        # Save the modified workbook
        wb.save(filename) # Adjust the range if necessary
        wb.close()
    def unmask_green_cells(self, redacted_filename, unredacted_filename, tab_name):
        # Load both workbooks
        redacted_wb = openpyxl.load_workbook(redacted_filename, data_only=True)
        unredacted_wb = openpyxl.load_workbook(unredacted_filename, data_only=True)

        # Access the specific tab in both workbooks
        redacted_ws = redacted_wb[tab_name]
        unredacted_ws = unredacted_wb[tab_name]
        print(f"Unmasking green cells in {tab_name}...")
        # Iterate through the worksheet and find green cells
        for row in redacted_ws.iter_rows():
            for cell in row:
                if cell.fill.start_color.index == '00ff15':  # Check if the cell is highlighted in green
                    # Get the corresponding cell from the unredacted worksheet
                    original_cell = unredacted_ws[cell.coordinate]
                    print(f"Found green cell {cell.coordinate} with value {cell.value}")
                    # Replace the value in the redacted worksheet with the original value
                    cell.value = original_cell.value
                    print(f"Unmasking cell {cell.coordinate} with value {cell.value}")
                    # Remove the green fill
                    cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                else:
                    print("No green cells found in this worksheet")
                    continue

        # Save the modified redacted workbook
        redacted_wb.save(redacted_filename)

    # Helper function to print cell values, types, and formats
    def print_cell_formats(file_name, sheet_name, cell_range):
        wb = openpyxl.load_workbook(file_name)
        ws = wb[sheet_name]
        for row in ws[cell_range]:
            for cell in row:
                print(f"Cell {cell.coordinate} - Value: {cell.value} - Type: {type(cell.value)} - Format: {cell.number_format}")


    def check_and_mask_underredacted_columns(self, ws, start_row, end_row, start_col, end_col):
        # Iterate over columns within the specified range
        for col_index in range(start_col, end_col + 1):
            column_cells = [ws.cell(row=row_index, column=col_index) for row_index in range(start_row, end_row + 1)]
            masked_cells = [cell for cell in column_cells if cell.value in ['<=5', '>5']]
            unmasked_cells = [cell for cell in column_cells if cell not in masked_cells and isinstance(cell.value, (int, float))]

            # If there is only one masked cell in the column
            if len(masked_cells) == 1 and unmasked_cells:
                # Find the smallest unmasked value
                smallest_unmasked_cell = min(unmasked_cells, key=lambda cell: cell.value)
                # Mask it based on its value
                if 0 <= smallest_unmasked_cell.value <= 5:
                    smallest_unmasked_cell.value = '<=5'
                elif smallest_unmasked_cell.value > 5:
                    smallest_unmasked_cell.value = '>5'
                # Optionally highlight the cell
                # self.yellow_cell(smallest_unmasked_cell)
                print(ws.title,f"Underredaction: Masking cell {smallest_unmasked_cell.coordinate} with {'<=5' if smallest_unmasked_cell.value == '<=5' else '>5'}")

    def check_and_mask_underredacted_rows(self, ws, start_row, end_row, start_col, end_col):
        # Iterate over rows within the specified range
        for row_index in range(start_row, end_row + 1):
            row_cells = [ws.cell(row=row_index, column=col_index) for col_index in range(start_col, end_col + 1)]
            masked_cells = [cell for cell in row_cells if cell.value in ['<=5', '>5']]
            unmasked_cells = [cell for cell in row_cells if cell not in masked_cells and isinstance(cell.value, (int, float))]

            # If there is only one masked cell in the row
            if len(masked_cells) == 1 and unmasked_cells:
                # Find the smallest unmasked value
                smallest_unmasked_cell = min(unmasked_cells, key=lambda cell: cell.value)
                # Mask it based on its value
                if 0 <= smallest_unmasked_cell.value <= 5:
                    smallest_unmasked_cell.value = '<=5'
                elif smallest_unmasked_cell.value > 5:
                    smallest_unmasked_cell.value = '>5'
                # Optionally highlight the cell
                # self.yellow_cell(smallest_unmasked_cell)
                print(ws.title,f"Underredaction: Masking cell {smallest_unmasked_cell.coordinate} with {'<=5' if smallest_unmasked_cell.value == '<=5' else '>5'}")




# Call the function with your filename
if __name__ == "__main__":
    processor = Solution()
    # #SY24
    # filename_SY24 = 'C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted City Council Triennial Report SY24.xlsx'
    # unredacted_filename_SY24 = 'C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\CityCouncil\\CCUnredacted\\Non-Redacted City Council Triennial Report SY24.xlsx'
    # for report, config in TRIENNIAL_REPORTS_CONFIG_SY24.items():
    #     processor.mask_excel_file(filename_SY24, report, config, unredacted_filename_SY24)
    # redacted_filenames_SY24 = [ 'C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted City Council Triennial Report SY24.xlsx']
    # unredacted_filenames_SY24 = ['C:\\Users\\Ywang36\OneDrive - NYCDOE\\Desktop\\CityCouncil\\CCUnredacted\\Non-Redacted City Council Triennial Report SY24.xlsx']
    # for redacted_file, unredacted_file in zip(redacted_filenames_SY24, unredacted_filenames_SY24):
    #     redacted_wb = openpyxl.load_workbook(redacted_file, data_only=True)
    #     unredacted_wb = openpyxl.load_workbook(unredacted_file, data_only=True)

    #     for report, config in TRIENNIAL_REPORTS_CONFIG_SY24.items():
    #         if 'groups' in config and 'ranges' in config:  # Ensure both 'groups' and 'ranges' keys exist
    #             ws = redacted_wb[report]
    #             unredacted_ws = unredacted_wb[report]
    #             processor.highlight_overredaction(ws, config['groups'], config['ranges'], unredacted_ws)

    #     # Save the redacted workbook after unmasking green cells
    #     redacted_wb.save(redacted_file)
    #     redacted_wb.close()

    #SY24 04022024
    filename_SY24_04022024 = 'C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted City Council Triennial Report_04022024.xlsx'
    unredacted_filename_SY24_04022024 = 'C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\CityCouncil\\CCUnredacted\\Non-Redacted City Council Triennial Report_04022024.xlsx'
    for report, config in TRIENNIAL_REPORTS_CONFIG_SY240402.items(): 
        processor.mask_excel_file(filename_SY24_04022024, report, config, unredacted_filename_SY24_04022024)
    redacted_filenames_SY24_04022024 = ['C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Non-Redacted City Council Triennial Report_04022024.xlsx']
    unredacted_filenames_SY24_04022024 = ['C:\\Users\\Ywang36\OneDrive - NYCDOE\\Desktop\\CityCouncil\\CCUnredacted\\Non-Redacted City Council Triennial Report_04022024.xlsx']
    for redacted_file, unredacted_file in zip(redacted_filenames_SY24_04022024, unredacted_filenames_SY24_04022024):
        redacted_wb = openpyxl.load_workbook(redacted_file, data_only=True)
        unredacted_wb = openpyxl.load_workbook(unredacted_file, data_only=True)

        for report, config in TRIENNIAL_REPORTS_CONFIG_SY240402.items():
            if 'groups' in config and 'ranges' in config:  # Ensure both 'groups' and 'ranges' keys exist
                ws = redacted_wb[report]
                unredacted_ws = unredacted_wb[report]
                processor.highlight_overredaction(ws, config['groups'], config['ranges'], unredacted_ws)

        # Save the redacted workbook after unmasking green cells
        redacted_wb.save(redacted_file)
        redacted_wb.close()