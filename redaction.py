import openpyxl

def mask_range(ws, start_row, start_col, end_row, end_col):
    # Step 1: Mask data <= 5
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value <= 5:
                cell.value = '<=5'

    # Step 2: If only one '<=5' in a row, mask smallest value in the row
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        star_count = sum([1 for cell in row if cell.value == '<=5'])
        if star_count == 1:
            min_val = min([cell.value for cell in row if isinstance(cell.value, (int, float))])
            for cell in row:
                if cell.value == min_val:
                    cell.value = '<=5'
                    break

    # Step 3: If only one '<=5' in a column, mask smallest value in the column
    for col in ws.iter_cols(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        col_values = [cell.value for cell in col]
        star_count = col_values.count('<=5')
        if star_count == 1:
            min_val = min([cell.value for cell in col if isinstance(cell.value, (int, float))])
            for cell in col:
                if cell.value == min_val:
                    cell.value = '<=5'
                    break
def mask_special_cells(ws):
    # Adjust E85 if F85 is masked
    if isinstance(ws['F85'].value, (int, float)) and ws['F85'].value <= 5:
        ws['E85'].value = '>5'

def mask_column_totals(ws, start_col, end_col):
    for col in ws.iter_cols(min_col=start_col, max_col=end_col):
        masked_cells = [cell for cell in col if cell.value == '<=5']
        if len(masked_cells) == 1:
            # Mask the next largest value in the column
            numeric_cells = [cell for cell in col if isinstance(cell.value, (int, float))]
            numeric_cells.sort(key=lambda x: x.value)
            for cell in numeric_cells:
                if cell.value != masked_cells[0].value:
                    cell.value = '>5'
                    break

def mask_excel_file(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    
    # Find the sheet named 'Report 8 = Registers'
    # ws = wb['Report 8 = Registers']
    ws = wb['Reports 1-4 = Initials']
    
    # Mask data for the specific ranges
    # ranges = [(5, 3, 37, 13), (41, 3, 46, 13), (50, 3, 52, 13), (56, 3, 58, 13), (62, 3, 75, 13), (79, 3, 92, 13)]
    ranges = [(5, 3, 37, 13), (41, 3, 46, 13), (50, 3, 52, 13), (56, 3, 58, 13),
                        (62, 3, 75, 13), (79, 3, 92, 13)]
    for r in ranges:
        mask_range(ws, *r)
    
    # Save the modified workbook
    wb.save(filename)

# Call the function with your filename
mask_excel_file('C:\\Users\\Ywang36\\OneDrive - NYCDOE\\Desktop\\Annual Special Education Data Report Unredacted SY21.xlsx')
