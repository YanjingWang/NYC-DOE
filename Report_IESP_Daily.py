import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc, time
import os
from copy import copy
import re
from openpyxl.utils.exceptions import IllegalCharacterError
import win32com.client as win32

class Solution:
    # Existing code...
    # Function to format headers
    def __init__(self):
        # format MM.DD.YYYY
        self.date = time.strftime("%m.%d.%Y")
        # self.date = '10.24.2024'
        self.lastrow = 62045 #5789 #22768
        self.filepath = rf'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\IESPDaily\IESP PNI Mandates with PA and EA_{self.date}.xlsx'

    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        wb = openpyxl.load_workbook(rf'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\IESPDaily\IESP PNI Mandates with PA and EA_{self.date}.xlsx')
        # open existing sheet 'TimelyPNI Students w IESP or SP'
        ws = wb['TimelyPNI Students w IESP or SP']       
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def apply_border_to_cell(self, cell):
        # Define a thin black border
        thin_black_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        cell.border = thin_black_border

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_Reporting' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        cursor.execute("EXEC[RAJI].[USPDIM_PNI]")
        time.sleep(5)
        cursor.execute("EXEC[RAJI].[USPDIM_IEPNPS]")    
        time.sleep(60)    
        cursor.execute("EXEC[RAJI].[USPDIM_StudentRegisterNPS]")
        time.sleep(30)
        cursor.execute("EXEC[RAJI].[USPDIM_PAAgenciesNStaff]") 
        time.sleep(10)
        cursor.execute("EXEC[RAJI].[USPDIM_PAMandatesNPS]")
        time.sleep(60)
        cursor.execute("EXEC[RAJI].[USPDIM_PAAssignmentsNPS]")
        time.sleep(30)
        cursor.execute("EXEC[RAJI].[USPDIM_EAServiceRecordsNPS]")
        time.sleep(60*2)
        cursor.execute("EXEC[RAJI].[USPDIM_NPSProvisioning]")
        time.sleep(30)
        return cursor
    
    # Function to execute non-query SQL commands
    def execute_non_query(self, cursor, sql):
        try:
            cursor.execute(sql)
            cursor.commit()
        except Exception as e:
            print(f"Error executing SQL: {e}")
    
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def report_1(self,cursor):

        # Non-query part of the SQL
        non_query_part1 = '''
        drop table if exists #StudentRegister_IESP;
        select * into #StudentRegister_IESP from SEO_REPORTING.Raji.DIM_StudentRegisterNPS 
        where ETLCurrentRow = 'Y' and isnull(PNIStatus, '') <> '' 
        and isnull(LatestRecommendPlacementDesc, '') not in (
                'Board of Cooperative Educational Services',
                'NYS Supported Non Public School – 4201 - Day',
                'NYS Supported Non Public School – 4201 - Residential',
                'NYSED-Approved Non Public School - Day',
                'NYSED-Approved Non Public School - Placed by ACS',
                'NYSED-Approved Non Public School - Residential'
        );
        create index idx_StudentRegister on #StudentRegister_IESP(StudentID);
        '''
        # Execute the non-query part
        self.execute_non_query(cursor, non_query_part1)

        non_query_part2 = '''
        drop table if exists #PNI
        select pni.StudentID, pni.DateNotIntentRcvd, pni.PNISchoolDBN, pni.PNIArticSchoolDBN,
            l.LocationName as PNISchoolName, l.PrincipalEmail as PNISchoolPrincipalEmail, l.PrincipalPhoneNumber as PNISchoolPrincipalPhone,
            al.LocationName as PNIArticSchoolName, al.PrincipalEmail as PNIArticSchoolPrincipalEmail, al.PrincipalPhoneNumber as PNIArticSchoolPrincipalPhone
        into #PNI
        from
        (
        select StudentID, DateNotIntentRcvd, PNISchoolDBN, PNIArticSchoolDBN,
            row_number() over (partition by StudentID order by PNIDocumentIDT desc) as rnk 
        from SEO_REPORTING.Raji.DIM_PNI 
        where ETLCurrentRow = 'Y'
        ) pni  
            left join SEO_MART.dbo.RPT_Locations l 
                on pni.PNISchoolDBN = l.SchoolDBN 
            left join SEO_MART.dbo.RPT_Locations al
                on pni.PNIArticSchoolDBN = al.SchoolDBN
        where pni.rnk = 1
        '''
        self.execute_non_query(cursor, non_query_part2)

        non_query_part3 = '''
        drop table if exists #IEPDocs
        select * into #IEPDocs
        from SEO_REPORTING.Raji.DIM_IEPNPS 
        where ETLCurrentRow = 'Y' and OutcomeTypeDesc in ('IEP', 'IESP', 'CSP', 'SP')
        and StudentID in (select StudentID from #StudentRegister_IESP)
        create index idx_StudentID on #IEPDocs(StudentID)
        '''
        self.execute_non_query(cursor, non_query_part3)

        non_query_part4 = '''
        drop table if exists #Report1 
        select  
        r.PNIStatus as StudentCategory
        , case when r.DateNotIntentRcvd is null or r.DateNotIntentRcvd = '01-01-1900' then '' else convert(varchar(10), r.DateNotIntentRcvd, 121) end as DateNotIntentRcvd
        , r.StudentID
        , r.FirstName
        , r.LastName
        , r.LatestClassification as Classification
        , case when r.LatestOutcomeDocumentIDT = iep.OutcomeDocumentIDT 
            or r.LatestOutcomeTypeDesc in ('Discharge', 'Declass', 'Ineligible', 'CaseClose')
            then 'Y' else 'N' end as LatestOutcome
        , r.LatestOutcomeTypeDesc LatestOutcomeTypeDesc
        , r.LatestOutcomeDate as 'LatestIEPDate'
        , iep.OutcomeDate as 'IEPDate'
        ,iep.OutcomeEffectiveDate
        , iep.OutcomeTypeDesc
        , iep.OutcomeDocumentIDT
        , iep.RecommendPlacementDesc
        , iep.ProcessStageDesc
        ,case when r.LatestOutcomeDocumentIDT = iep.OutcomeDocumentIDT and prov.RecommendedStartDate is not null 
            and cast(getdate() as date) between prov.RecommendedStartDate and isnull(prov.RecommendedEndDate, '12-31-9999') 
            then 'Y' else 'N' end as 'ActiveMandate'
        , prov.RecommendedMandateType
        , prov.RecommendedGroupType
        , prov.RecommendedLocation
        , prov.RecommendedLanguage
        , prov.RecommendedFrequency
        , prov.RecommendedDuration
        , prov.RecommendedStartDate
        , prov.RecommendedEndDate
        , isnull(PAMandateStatus, '') as PAMandateStatus
        , case when PAFirstAttendDate is null or PAFirstAttendDate = '01-01-1900' then '' else convert(varchar(10), PAFirstAttendDate, 121) end as PAFirstAttendDate
        , case when PAEarliestAssignmentDate is null or PAEarliestAssignmentDate = '01-01-1900' then '' else convert(varchar(10), PAEarliestAssignmentDate, 121) end as PAEarliestAssignmentDate
        , isnull(PACurrentAssignmentStatus, '') as PACurrentAssignmentStatus
        , isnull(PAFirstAttendDelayReason, '') as PAFirstAttendDelayReason
        , isnull(PAAgencyName, '') as PAAgencyName
        , isnull(PAProviderName, '') as PAProviderName
        , isnull(EAFullEncoutnersCount, 0) as 'EAFullEncounters'
        , isnull(EAPartialEncountersCount, 0) as 'EAPartialEncounters'
        , isnull(EAStudentAbsenceCount, 0) as 'StudentAbsences'
        , case when EAEarliestFullServiceDate is null or EAEarliestFullServiceDate = '01-01-1900' then '' else convert(varchar(10), EAEarliestFullServiceDate, 121) end as EAEarliestFullServiceDate
        , case when EAEarliestPartialServiceDate is null or EAEarliestPartialServiceDate = '01-01-1900' then '' else convert(varchar(10), EAEarliestPartialServiceDate, 121) end as EAEarliestPartialServiceDate
        , case when EALatestFullServiceDate is null or EALatestFullServiceDate = '01-01-1900' then '' else convert(varchar(10), EALatestFullServiceDate, 121) end as EALatestFullServiceDate
        , case when EALatestPartialServiceDate is null or EALatestPartialServiceDate = '01-01-1900' then '' else convert(varchar(10), EALatestPartialServiceDate, 121) end as EALatestPartialServiceDate
        , r.ProfileIDT
        , r.GradeLevel
        , r.BirthDate
        , isnull(r.HousingType, '') as STH
        , r.EnrolledDBN
        , l.LocationName
        , isnull(l.PrincipalEmail, '') as PrincipalEmail
        , isnull(l.PrincipalPhoneNumber, '') as PrincipalPhoneNumber
        , isnull(l.Address1 + ' ', '') + isnull(l.Address2, '') as EnrolledDBNAddress
        , isnull(l.City, '') as EnrolledDBNCity
        , isnull(l.ZipCode, '') as EnrolledDBNZipcode
        , isnull(r.AdminDistrict, '') as AdminDistrict
        , isnull(l.CSE, '') as CSE
        , isnull(r.HomeDistrict, '') as HomeDistrict
        , isnull(pni.PNISchoolDBN, '') as PNISchoolDBN
        , isnull(pni.PNISchoolName, '') as PNISchoolName
        , isnull(pni.PNISchoolPrincipalEmail, '') as PNISchoolPrincipalEmail
        , isnull(pni.PNISchoolPrincipalPhone, '') as PNISchoolPrincipalPhone
        , isnull(pni.PNIArticSchoolDBN, '') as PNIArticSchoolDBN
        , isnull(pni.PNIArticSchoolName, '') as PNIArticSchoolName
        , isnull(pni.PNIArticSchoolPrincipalEmail, '') as PNIArticSchoolPrincipalEmail
        , isnull(pni.PNIArticSchoolPrincipalPhone, '') as PNIArticSchoolPrincipalPhone
        into #Report1
        from #StudentRegister_IESP r 
            left join SEO_MART.dbo.RPT_Locations l 
                on r.EnrolledDBN = l.SchoolDBN
            left join #PNI pni 
                on r.StudentID = pni.StudentID
                and pni.DateNotIntentRcvd = r.DateNotIntentRcvd
            left join #IEPDocs iep 
                on r.StudentID = iep.StudentID 
            left join SEO_REPORTING.Raji.DIM_NPSProvisioning prov 
                on prov.StudentID = r.StudentID 
                and prov.OutcomeDocumentIDT = iep.OutcomeDocumentIDT
                and prov.ETLCurrentRow = 'Y'
        where isnull(r.PNIStatus, '') in ('TimelyPNI', 'KIP', 'TimelyConsent-Initials', 'TimelyPNI-Initials'
        , 'TimelyPNI-CaseCloseB4PNI', 'TimelyPNI-CaseCloseAfterPNI', 'TimelyPNI-NoIEP', 'KIP-CaseCloseB4PNI', 
        'KIP-NoIEP', 'CannotDetermine-NoIEP', 'CannotDetermine-CaseCloseB4PNI', 'CannotDetermine-CaseCloseAfterPNI'
        )
        and r.LatestOutcomeTypeDesc in ('IESP', 'SP', 'Discharge', 'Ineligible', 'CaseClose', 'Declass')
        and 
        (
        prov.RecommendedMandateType is null OR
        prov.RecommendedMandateType in (
        'Audiology Services',
        --Behavior management/support plan - 
        'Counseling Services',
        'Counseling by a Social Worker / Psychologist',
        --'Extra time to complete assignments - ',
        'Hearing Education Services',
        --'Highlighted work - ',
        --'Instructional materials in alternative formats  - ',
        'Interpreting Services - Oral Transliterator',
        'Interpreting Services - Sign Language Interpreter',
        'Medical Services',
        'Occupational Therapy',
        'Orientation and Mobility Services',
        'Other',
        'Paraprofessional - Behavior Support',
        'Paraprofessional - Health',
        'Paraprofessional - Orientation and Mobility',
        'Paraprofessional - Physical Assistance',
        'Paraprofessional - Special Transportation',
        'Paraprofessional - Toileting',
        'Parent Counseling and Training',
        'Physical Therapy',
        'Psychological Services',
        'School Health Services',
        'School Nurse Services',
        'School Social Work',
        'SETSS',
        --'Special seating arrangements - ',
        'Speech-Language Therapy',
        --'Study guide outlines of key concepts - ',
        'Vision Education Services'
        )
        )
        order by 1
        '''
        self.execute_non_query(cursor, non_query_part4)

        # Query part of the SQL
        query_by_report_1 = '''
        select * from #Report1
        '''  # the byRace SQL query goes here
        cursor.execute(query_by_report_1)
        results_by_report1 = cursor.fetchall()
        return results_by_report1

    def report_2(self,cursor):
        non_query_part1 = '''
    --Report 2: KIP or TimelyPNI kids with IEP or CSP in NPS
        drop table if exists #Report2 
        select  
        r.PNIStatus as StudentCategory
        , case when r.DateNotIntentRcvd is null or r.DateNotIntentRcvd = '01-01-1900' then '' else convert(varchar(10), r.DateNotIntentRcvd, 121) end as DateNotIntentRcvd
        , r.StudentID
        , r.FirstName
        , r.LastName
        , r.LatestClassification as Classification
        , case when r.LatestOutcomeDocumentIDT = iep.OutcomeDocumentIDT 
            or r.LatestOutcomeTypeDesc in ('Discharge', 'Declass', 'Ineligible', 'CaseClose')
            then 'Y' else 'N' end as LatestOutcome
        , r.LatestOutcomeTypeDesc LatestOutcomeTypeDesc
        , r.LatestOutcomeDate as 'LatestIEPDate'
        , iep.OutcomeDate as 'IEPDate'
        ,iep.OutcomeEffectiveDate
        , iep.OutcomeTypeDesc
        , iep.OutcomeDocumentIDT
        , iep.RecommendPlacementDesc
        , iep.ProcessStageDesc
        ,case when r.LatestOutcomeDocumentIDT = iep.OutcomeDocumentIDT and prov.RecommendedStartDate is not null 
            and cast(getdate() as date) between prov.RecommendedStartDate and isnull(prov.RecommendedEndDate, '12-31-9999') 
            then 'Y' else 'N' end as 'ActiveMandate'
        , prov.RecommendedMandateType
        , prov.RecommendedGroupType
        , prov.RecommendedLocation
        , prov.RecommendedLanguage
        , prov.RecommendedFrequency
        , prov.RecommendedDuration
        , prov.RecommendedStartDate
        , prov.RecommendedEndDate
        , isnull(PAMandateStatus, '') as PAMandateStatus
        , case when PAFirstAttendDate is null or PAFirstAttendDate = '01-01-1900' then '' else convert(varchar(10), PAFirstAttendDate, 121) end as PAFirstAttendDate
        , case when PAEarliestAssignmentDate is null or PAEarliestAssignmentDate = '01-01-1900' then '' else convert(varchar(10), PAEarliestAssignmentDate, 121) end as PAEarliestAssignmentDate
        , isnull(PACurrentAssignmentStatus, '') as PACurrentAssignmentStatus
        , isnull(PAFirstAttendDelayReason, '') as PAFirstAttendDelayReason
        , isnull(PAAgencyName, '') as PAAgencyName
        , isnull(PAProviderName, '') as PAProviderName
        , isnull(EAFullEncoutnersCount, 0) as 'EAFullEncounters'
        , isnull(EAPartialEncountersCount, 0) as 'EAPartialEncounters'
        , isnull(EAStudentAbsenceCount, 0) as 'StudentAbsences'
        , case when EAEarliestFullServiceDate is null or EAEarliestFullServiceDate = '01-01-1900' then '' else convert(varchar(10), EAEarliestFullServiceDate, 121) end as EAEarliestFullServiceDate
        , case when EAEarliestPartialServiceDate is null or EAEarliestPartialServiceDate = '01-01-1900' then '' else convert(varchar(10), EAEarliestPartialServiceDate, 121) end as EAEarliestPartialServiceDate
        , case when EALatestFullServiceDate is null or EALatestFullServiceDate = '01-01-1900' then '' else convert(varchar(10), EALatestFullServiceDate, 121) end as EALatestFullServiceDate
        , case when EALatestPartialServiceDate is null or EALatestPartialServiceDate = '01-01-1900' then '' else convert(varchar(10), EALatestPartialServiceDate, 121) end as EALatestPartialServiceDate
        , r.ProfileIDT
        , r.GradeLevel
        , r.BirthDate
        , isnull(r.HousingType, '') as STH
        , r.EnrolledDBN
        , l.LocationName
        , isnull(l.PrincipalEmail, '') as PrincipalEmail
        , isnull(l.PrincipalPhoneNumber, '') as PrincipalPhoneNumber
        , isnull(l.Address1 + ' ', '') + isnull(l.Address2, '') as EnrolledDBNAddress
        , isnull(l.City, '') as EnrolledDBNCity
        , isnull(l.ZipCode, '') as EnrolledDBNZipcode
        , isnull(r.AdminDistrict, '') as AdminDistrict
        , isnull(l.CSE, '') as CSE
        , isnull(r.HomeDistrict, '') as HomeDistrict
        , isnull(pni.PNISchoolDBN, '') as PNISchoolDBN
        , isnull(pni.PNISchoolName, '') as PNISchoolName
        , isnull(pni.PNISchoolPrincipalEmail, '') as PNISchoolPrincipalEmail
        , isnull(pni.PNISchoolPrincipalPhone, '') as PNISchoolPrincipalPhone
        , isnull(pni.PNIArticSchoolDBN, '') as PNIArticSchoolDBN
        , isnull(pni.PNIArticSchoolName, '') as PNIArticSchoolName
        , isnull(pni.PNIArticSchoolPrincipalEmail, '') as PNIArticSchoolPrincipalEmail
        , isnull(pni.PNIArticSchoolPrincipalPhone, '') as PNIArticSchoolPrincipalPhone
        into #Report2
        from #StudentRegister_IESP r 
            left join SEO_MART.dbo.RPT_Locations l 
                on r.EnrolledDBN = l.SchoolDBN
            left join #PNI pni 
                on r.StudentID = pni.StudentID
                and pni.DateNotIntentRcvd = r.DateNotIntentRcvd
            left join #IEPDocs iep 
                on r.StudentID = iep.StudentID 
            left join SEO_REPORTING.Raji.DIM_NPSProvisioning prov 
                on prov.StudentID = r.StudentID 
                and prov.OutcomeDocumentIDT = iep.OutcomeDocumentIDT
                and prov.ETLCurrentRow = 'Y'
        where isnull(r.PNIStatus, '') in ('TimelyPNI', 'KIP', 'TimelyConsent-Initials', 'TimelyPNI-Initials'
        , 'TimelyPNI-CaseCloseB4PNI', 'TimelyPNI-CaseCloseAfterPNI', 'TimelyPNI-NoIEP', 'KIP-CaseCloseB4PNI', 
        'KIP-NoIEP', 'CannotDetermine-NoIEP', 'CannotDetermine-CaseCloseB4PNI', 'CannotDetermine-CaseCloseAfterPNI'
        )
        and r.LatestOutcomeTypeDesc in ('IEP', 'CSP', 'Discharge', 'Ineligible', 'CaseClose', 'Declass')
        and 
        (
        prov.RecommendedMandateType is null OR
        prov.RecommendedMandateType in (
        'Audiology Services',
        --Behavior management/support plan - 
        'Counseling Services',
        'Counseling by a Social Worker / Psychologist',
        --'Extra time to complete assignments - ',
        'Hearing Education Services',
        --'Highlighted work - ',
        --'Instructional materials in alternative formats  - ',
        'Interpreting Services - Oral Transliterator',
        'Interpreting Services - Sign Language Interpreter',
        'Medical Services',
        'Occupational Therapy',
        'Orientation and Mobility Services',
        'Other',
        'Paraprofessional - Behavior Support',
        'Paraprofessional - Health',
        'Paraprofessional - Orientation and Mobility',
        'Paraprofessional - Physical Assistance',
        'Paraprofessional - Special Transportation',
        'Paraprofessional - Toileting',
        'Parent Counseling and Training',
        'Physical Therapy',
        'Psychological Services',
        'School Health Services',
        'School Nurse Services',
        'School Social Work',
        'SETSS',
        --'Special seating arrangements - ',
        'Speech-Language Therapy',
        --'Study guide outlines of key concepts - ',
        'Vision Education Services'
        )
        )
        order by 1
        '''
        self.execute_non_query(cursor, non_query_part1)
        
        query_by_report_2 = '''
        select * from #Report2         
        '''
        cursor.execute(query_by_report_2)
        results_by_report2 = cursor.fetchall()
        return results_by_report2
    
    def report_3(self,cursor):
        non_query_part1 = '''

        --BEGIN: Report 3 --------

        drop table if exists #ParentsLanguage

        create table #ParentsLanguage(StudentID int, Parent2Language INT, Parent1SpokenLanguage INT, Parent2SpokenLanguage INT)
        '''
        self.execute_non_query(cursor, non_query_part1)

        non_query_part2 = '''
        insert into #ParentsLanguage
        select '0' as StudentID, '0' as Parent2Language, '0' Parent1SpokenLanguage, '0' as Parent2SpokenLanguage 
        '''
        self.execute_non_query(cursor, non_query_part2)

        non_query_part3 = '''
            drop table if exists #contactInfo
            select 
                S.StudentID,
                rtrim(ltrim(isnull(S.FirstName, ''))) + ' ' + rtrim(ltrim(isnull(S.LastName, ''))) as StudentName
                ,case when S.RedAlert = 0 then 'N' else 'Y' end as 'RedAlertStudent'
                ,case when S.CarterStudent = 0 then 'N' else 'Y' end as 'CarterStudent'
                --,Borough
                --,DistrictHome
                , rtrim(ltrim(isnull(S.Address, ''))) + ' ' + rtrim(ltrim(isnull(S.ApartmentNumber, ''))) + ' ' + rtrim(ltrim(isnull(S.StudentCity, ''))) + ' ' + rtrim(ltrim(isnull(S.ParentState, ''))) + ' ' + rtrim(ltrim(isnull(S.ZipCode, ''))) as StudentAddress

                , rtrim(ltrim(isnull(S.ParentFirstName, ''))) + ' ' + rtrim(ltrim(isnull(S.ParentLastName, ''))) as PrimaryContactName
                , rtrim(ltrim(isnull(S.Parent1Street, ''))) + ' ' + rtrim(ltrim(isnull(S.Parent1ApartmentNum, ''))) + ' ' + rtrim(ltrim(isnull(S.ParentCity1, ''))) + ' ' + rtrim(ltrim(isnull(S.ParentState, ''))) + ' ' + rtrim(ltrim(isnull(S.Parent1Zip, ''))) 
                    as PrimaryContactAddress
                , rtrim(ltrim(isnull(S.Parent1Email, ''))) as PrimaryContactEmail
                , 'Home:' + case when isnull(S.Parent1phone1, '') <> '' then S.Parent1phone1 else 'NA' end + ';' + 
                'Cell:' + case when isnull(S.Parent1Phone3, '') <> '' then S.Parent1Phone3 else 'NA' end + ';' + 
                'Work:' + case when isnull(S.Parent1Phone2, '') <> '' then S.Parent1Phone2 else 'NA' end 
                as PrimaryContactPhone
                , 'Written:' + case when isnull(S.Parent1LanguageDesc, '') <> '' then rtrim(ltrim(isnull(S.Parent1LanguageDesc, ''))) else 'NA' end + '; '  +
                    'Spoken:' + case when isnull(priSp.[Description], '') <> '' then rtrim(ltrim(isnull(priSp.[Description], ''))) else 'NA' end 	
                    as PrimaryContactLanguage  

                , rtrim(ltrim(isnull(S.ParentFirstName2, ''))) + ' ' + rtrim(ltrim(isnull(S.ParentLastName2, ''))) as SecondaryContactName
                , rtrim(ltrim(isnull(S.Parent2Street, ''))) + ' ' + rtrim(ltrim(isnull(S.Parent2ApartmentNum, ''))) + ' ' + rtrim(ltrim(isnull(S.ParentCity2, ''))) + ' ' + rtrim(ltrim(isnull(ST.Code, ''))) + ' ' + rtrim(ltrim(isnull(S.Parent2Zip, ''))) 
                    as SecondaryContactAddress
                , rtrim(ltrim(isnull(S.Parent2Email, ''))) as SecondaryContactEmail
                , 'Home:' + case when isnull(S.Parent2Phone1, '') <> '' then S.Parent2Phone1 else 'NA' end + '; ' + 
                'Cell:' + case when isnull(S.Parent2Phone3, '') <> '' then S.Parent2Phone3 else 'NA' end + '; ' + 
                'Work:' + case when isnull(S.Parent2Phone2, '') <> '' then S.Parent2Phone2 else 'NA' end 
                as SecondaryContactPhone
                , 'Written:' + case when isnull(secWr.Description, '') <> '' then rtrim(ltrim(isnull(secWr.Description, ''))) else 'NA' end + '; ' +
                    'Spoken:' + case when isnull(secSp.[Description], '') <> '' then rtrim(ltrim(isnull(secSp.[Description], ''))) else 'NA' end 	
                    as SecondaryContactLanguage  
            into #contactInfo
            --select Parent2State
            from SEO_MART.dbo.DWH_SESISStudents S
                left join (select IDT, Code from SEO_MART.dbo.lk_USStateCodes) ST 
                    on S.Parent2State = ST.IDT
                left join #ParentsLanguage pl on S.StudentID = pl.StudentID 
                left join SEO_MART.dbo.lk_LanguageCodes secWr with(nolock)
                    on pl.Parent2Language = secWr.IDT
                    and secWr.SourceName = 'SESIS' 
                left join SEO_MART.dbo.lk_LanguageCodes priSp with(nolock)
                    on pl.Parent1SpokenLanguage = priSp.IDT
                    and priSp.SourceName = 'SESIS' 
                left join SEO_MART.dbo.lk_LanguageCodes secSp with(nolock)
                    on pl.Parent2SpokenLanguage = secSp.IDT
                    and secSp.SourceName = 'SESIS' 
            where S.StudentID in (select StudentID from #StudentRegister_IESP
                where isnull(PNIStatus, '') in ('KIP', 'TimelyPNI', 'TimelyPNI-Initials', 'TimelyConsent-Initials')
            )
        '''
        self.execute_non_query(cursor, non_query_part3)

        query_by_report_3 = '''
            select * from #contactInfo  
        '''
        cursor.execute(query_by_report_3)
        results_by_report3 = cursor.fetchall()
        return results_by_report3
    
    def report_4(self,cursor):
        non_query_part1 = '''
        --Report 3: UnTimelyPNI kids with IEP, IESP, CSP or SP in NPS
            drop table if exists #Report3 
            select  
            r.PNIStatus as StudentCategory
            , case when r.DateNotIntentRcvd is null or r.DateNotIntentRcvd = '01-01-1900' then '' else convert(varchar(10), r.DateNotIntentRcvd, 121) end as DateNotIntentRcvd
            , r.StudentID
            , r.FirstName
            , r.LastName
            , r.LatestClassification as Classification
            , case when r.LatestOutcomeDocumentIDT = iep.OutcomeDocumentIDT 
                or r.LatestOutcomeTypeDesc in ('Discharge', 'Declass', 'Ineligible', 'CaseClose')
                then 'Y' else 'N' end as LatestOutcome
            , r.LatestOutcomeTypeDesc LatestOutcomeTypeDesc
            , r.LatestOutcomeDate as 'LatestIEPDate'
            , iep.OutcomeDate as 'IEPDate'
            ,iep.OutcomeEffectiveDate
            , iep.OutcomeTypeDesc
            , iep.OutcomeDocumentIDT
            , iep.RecommendPlacementDesc
            , iep.ProcessStageDesc
            ,case when r.LatestOutcomeDocumentIDT = iep.OutcomeDocumentIDT and prov.RecommendedStartDate is not null 
                and cast(getdate() as date) between prov.RecommendedStartDate and isnull(prov.RecommendedEndDate, '12-31-9999') 
                then 'Y' else 'N' end as 'ActiveMandate'
            , prov.RecommendedMandateType
            , prov.RecommendedGroupType
            , prov.RecommendedLocation
            , prov.RecommendedLanguage
            , prov.RecommendedFrequency
            , prov.RecommendedDuration
            , prov.RecommendedStartDate
            , prov.RecommendedEndDate
            , isnull(PAMandateStatus, '') as PAMandateStatus
            , case when PAFirstAttendDate is null or PAFirstAttendDate = '01-01-1900' then '' else convert(varchar(10), PAFirstAttendDate, 121) end as PAFirstAttendDate
            , case when PAEarliestAssignmentDate is null or PAEarliestAssignmentDate = '01-01-1900' then '' else convert(varchar(10), PAEarliestAssignmentDate, 121) end as PAEarliestAssignmentDate
            , isnull(PACurrentAssignmentStatus, '') as PACurrentAssignmentStatus
            , isnull(PAFirstAttendDelayReason, '') as PAFirstAttendDelayReason
            , isnull(PAAgencyName, '') as PAAgencyName
            , isnull(PAProviderName, '') as PAProviderName
            , isnull(EAFullEncoutnersCount, 0) as 'EAFullEncounters'
            , isnull(EAPartialEncountersCount, 0) as 'EAPartialEncounters'
            , isnull(EAStudentAbsenceCount, 0) as 'StudentAbsences'
            , case when EAEarliestFullServiceDate is null or EAEarliestFullServiceDate = '01-01-1900' then '' else convert(varchar(10), EAEarliestFullServiceDate, 121) end as EAEarliestFullServiceDate
            , case when EAEarliestPartialServiceDate is null or EAEarliestPartialServiceDate = '01-01-1900' then '' else convert(varchar(10), EAEarliestPartialServiceDate, 121) end as EAEarliestPartialServiceDate
            , case when EALatestFullServiceDate is null or EALatestFullServiceDate = '01-01-1900' then '' else convert(varchar(10), EALatestFullServiceDate, 121) end as EALatestFullServiceDate
            , case when EALatestPartialServiceDate is null or EALatestPartialServiceDate = '01-01-1900' then '' else convert(varchar(10), EALatestPartialServiceDate, 121) end as EALatestPartialServiceDate
            , r.ProfileIDT
            , r.GradeLevel
            , r.BirthDate
            , isnull(r.HousingType, '') as STH
            , r.EnrolledDBN
            , l.LocationName
            , isnull(l.PrincipalEmail, '') as PrincipalEmail
            , isnull(l.PrincipalPhoneNumber, '') as PrincipalPhoneNumber
            , isnull(l.Address1 + ' ', '') + isnull(l.Address2, '') as EnrolledDBNAddress
            , isnull(l.City, '') as EnrolledDBNCity
            , isnull(l.ZipCode, '') as EnrolledDBNZipcode
            , isnull(r.AdminDistrict, '') as AdminDistrict
            , isnull(l.CSE, '') as CSE
            , isnull(r.HomeDistrict, '') as HomeDistrict
            , isnull(pni.PNISchoolDBN, '') as PNISchoolDBN
            , isnull(pni.PNISchoolName, '') as PNISchoolName
            , isnull(pni.PNISchoolPrincipalEmail, '') as PNISchoolPrincipalEmail
            , isnull(pni.PNISchoolPrincipalPhone, '') as PNISchoolPrincipalPhone
            , isnull(pni.PNIArticSchoolDBN, '') as PNIArticSchoolDBN
            , isnull(pni.PNIArticSchoolName, '') as PNIArticSchoolName
            , isnull(pni.PNIArticSchoolPrincipalEmail, '') as PNIArticSchoolPrincipalEmail
            , isnull(pni.PNIArticSchoolPrincipalPhone, '') as PNIArticSchoolPrincipalPhone
            into #Report3
            --select count(*)
            from #StudentRegister_IESP r 
                left join SEO_MART.dbo.RPT_Locations l 
                    on r.EnrolledDBN = l.SchoolDBN
                left join #PNI pni 
                    on r.StudentID = pni.StudentID
                    and pni.DateNotIntentRcvd = r.DateNotIntentRcvd
                left join #IEPDocs iep 
                    on r.StudentID = iep.StudentID 
                left join SEO_REPORTING.Raji.DIM_NPSProvisioning prov 
                    on prov.StudentID = r.StudentID 
                    and prov.OutcomeDocumentIDT = iep.OutcomeDocumentIDT
                    and prov.ETLCurrentRow = 'Y'
            where isnull(r.PNIStatus, '') = 'UnTimelyPNI'
            --and r.LatestOutcomeTypeDesc in ('IEP', 'CSP', 'Discharge', 'Ineligible', 'CaseClose', 'Declass')
            and 
            (
            prov.RecommendedMandateType is null OR
            prov.RecommendedMandateType in (
            'Audiology Services',
            --Behavior management/support plan - 
            'Counseling Services',
            'Counseling by a Social Worker / Psychologist',
            --'Extra time to complete assignments - ',
            'Hearing Education Services',
            --'Highlighted work - ',
            --'Instructional materials in alternative formats  - ',
            'Interpreting Services - Oral Transliterator',
            'Interpreting Services - Sign Language Interpreter',
            'Medical Services',
            'Occupational Therapy',
            'Orientation and Mobility Services',
            'Other',
            'Paraprofessional - Behavior Support',
            'Paraprofessional - Health',
            'Paraprofessional - Orientation and Mobility',
            'Paraprofessional - Physical Assistance',
            'Paraprofessional - Special Transportation',
            'Paraprofessional - Toileting',
            'Parent Counseling and Training',
            'Physical Therapy',
            'Psychological Services',
            'School Health Services',
            'School Nurse Services',
            'School Social Work',
            'SETSS',
            --'Special seating arrangements - ',
            'Speech-Language Therapy',
            --'Study guide outlines of key concepts - ',
            'Vision Education Services'
            )
            )
            order by 1
        '''
        self.execute_non_query(cursor, non_query_part1)
        
        query_by_report_4 = '''
            select * From #Report3
        '''
        cursor.execute(query_by_report_4)
        results_by_report4 = cursor.fetchall()
        return results_by_report4
    
    # Step 3: Write data to Excel for "Report 1", "Report 2", "Report 3", and "Report 4"
    def sanitize_value(self,value):
        # Remove any illegal characters
        if isinstance(value, str):
            # Use regex to remove illegal characters that Excel does not accept
            return re.sub(r"[\x00-\x1F]+", "", value)
        return value

    def write_data_to_excel(self, ws, data, start_cell):
        for r_idx, row in enumerate(data, start=start_cell[1]):
            for c_idx, value in enumerate(row, start=start_cell[0]):
                try:
                    # Sanitize the value before writing it to the Excel cell
                    ws.cell(row=r_idx, column=c_idx).value = self.sanitize_value(value)
                    self.apply_border_to_cell(ws.cell(row=r_idx, column=c_idx)) # Apply the black border to the cell
                except IllegalCharacterError:
                    # Handle the case where there's an illegal character
                    print(f"Illegal character found in row {r_idx}, column {c_idx}. Skipping the value.")
 
    def main_IESPDaily(self):

        cursor = self.connect_to_database()

        # Load the Excel workbook
        wb = openpyxl.load_workbook(rf'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\IESPDaily\IESP PNI Mandates with PA and EA_{self.date}.xlsx')

        # Fetch data for reports
        results_by_report1 = self.report_1(cursor)
        results_by_report2 = self.report_2(cursor)
        results_by_report3 = self.report_3(cursor)
        results_by_report4 = self.report_4(cursor)

        # Write data to specific sheets and ranges
        self.write_data_to_excel(wb['TimelyPNI Students w IESP or SP'], results_by_report1, (1, 2))  # A2:BH62046
        print('Report 1 done')
        self.write_data_to_excel(wb['TimelyPNI Students w IEP or CSP'], results_by_report2, (1, 2))  # A2:AZ5789
        print('Report 2 done')
        self.write_data_to_excel(wb['ContactInfo'], results_by_report3, (1, 2))  # A2:O66769
        print('Report 3 done')
        self.write_data_to_excel(wb['UntimelyPNI'], results_by_report4, (1, 2))  # A2:BH6953
        print('Report 4 done')

        # # refresh all pivot tables
        # # Start an instance of Excel
        # excel_app = win32.Dispatch('Excel.Application')

        # # Optional: Make Excel visible if you want to see what's happening
        # # excel_app.Visible = True

        # # Open the workbook
        # workbook = excel_app.Workbooks.Open(self.filepath)

        # # Refresh all pivot tables in all sheets
        # for sheet in workbook.Sheets:
        #     for pivot_table in sheet.PivotTables():
        #         pivot_table.PivotCache().Refresh()

        # # Save the workbook after refreshing
        # workbook.Save()

        # # Close the workbook and quit Excel
        # workbook.Close(SaveChanges=True)
        # excel_app.Quit()

        # print("Pivot tables refreshed successfully!")

        # Save the Excel file
        save_path = rf'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\IESPDaily\IESP PNI Mandates with PA and EA_{self.date}.xlsx'
        wb.save(save_path)

        # Close the database connection
        cursor.close()

    def refresh_pivot_table(self):
        # Start an instance of Excel
        excel_app = win32.Dispatch('Excel.Application')

        # Optional: Make Excel visible if you want to see what's happening
        excel_app.Visible = True

        # Open the workbook
        workbook = excel_app.Workbooks.Open(self.filepath)

        # Refresh all pivot tables in all sheets
        for sheet in workbook.Sheets:
            for pivot_table in sheet.PivotTables():
                pivot_table.PivotCache().Refresh()

        # Save the workbook after refreshing
        workbook.Save()

        # Close the workbook and quit Excel
        workbook.Close(SaveChanges=True)
        excel_app.Quit()

        print("Pivot tables refreshed successfully!")

if __name__ == "__main__":
        Tab2_5 = Solution()
        Tab2_5.main_IESPDaily()   
        Tab2_5.refresh_pivot_table()
                                                    