import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
class Solution:
    # Existing code...
    # Function to format headers
    def __init__(self, datestamp="04022024",date="April 2, 2024"):
        self.datestamp = datestamp
        self.date = date
        self.lastrow = 35
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print(col + cell_number)

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print(col + cell_number)



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted City Council Triennial Report_CW.xlsx')
        ws = wb.create_sheet("Transportation by District")

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=self.lastrow, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['School District',
                'Curb to School',
                'Percent Curb to School',
                'Stop to School',
                'Percent Stop to School',
                'Unassigned',
                'Percent Unassigned']
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "D9E1F2"
        column_fill_color = "D9E1F2"
        self.format_header(ws, 'A2', "School District", columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        # self.format_header(ws, 'A10', "Primary Program Type", columns, column_letters, 30, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        # params = ('CC_PSStudentR12_061523')
        cursor.execute("EXEC [Mike].[USPCCTriannualReportSTDistrictLevel]")
        return cursor

    def fetch_data_by_tab9(self,cursor):
        query_bytab9 = '''
        Select * 
        from ##STDistrict
        Order by 
        case when [School District]='Total' then
        'zzzzz'
        End
        ,[School District]
        '''  # the byRace SQL query goes here
        cursor.execute(query_bytab9)
        results_bytab9 = cursor.fetchall()
        return results_bytab9
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 1)  # +2 because data starts from column 'A'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range 
        for row in ws['A3':'A'+str(self.lastrow)]:
            for cell in row:
                if cell.value is not None:
                    cell.value = str(cell.value) + ''
                    cell.font = Font(bold=True, size=12)
                cell.alignment = openpyxl.styles.Alignment(horizontal='left')
        for row in ws['B3':'G'+str(self.lastrow)]:  
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        for row in ws['B1': 'G1']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['A'+str(self.lastrow):'G'+str(self.lastrow)]: 
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['A1': 'G1']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)
                
        cell_ranges = ['B3:B'+str(self.lastrow), 'D3:D'+str(self.lastrow), 'F3:F'+str(self.lastrow) ]
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        try:
                            cell.value = int(cell.value)
                            cell.number_format = '#,##0'  # Apply comma format
                            print("Int converting")
                        except ValueError:
                            # If the value cannot be converted to int, keep the original value
                            print("Int converting Error")
                            pass
        cell_ranges = ['C3:C'+str(self.lastrow), 'E3:E'+str(self.lastrow),'G3:G'+str(self.lastrow)]
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = '0%'  # Apply percentage format
                            print("Float converting")
                        except ValueError:
                            # If the value cannot be converted to int, keep the original value
                            print("Float converting Error")
                            pass
    def Report_Transportation_by_District(self):
        title_cells = [
            {"cell": "A1", "value": self.date+ " Number & Percentage of Special Transportation Recommendations with Busing Assignment", "merge_cells": "A1:G1"},
            

        ]

        subtitle_cells = [
            # {"cell": "B12", "value": "October 31, 2022 Number & Percentage of Students Receiving Bilingual Recommended Special Education Programs ", "merge_cells": "A9:G9"},            

        ]

        column_widths = [40, 20, 20, 20, 20, 20, 20]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_bytab9= self.fetch_data_by_tab9(cursor)
        self.write_data_to_excel(ws, results_bytab9, start_row=3)

        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted City Council Triennial Report_CW.xlsx'
        # # save path ended with self.datestamp 04022024
        # save_path = save_path[:-5] + self.datestamp + ".xlsx"
        wb.save(save_path)

if __name__ == "__main__":
        Tab9 = Solution()
        Tab9.Report_Transportation_by_District()                                                                  