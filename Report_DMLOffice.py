import pandas as pd
from sqlalchemy import create_engine, text
import urllib
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
class DMLOfficeRport1:
    def __init__(self, lastrow=186933, sqlschoolstartdate='2023-09-07', sqlschoolenddate='2024-07-01', schoolyear='SY 2023-2024',datestamp='6.30.24',sqlimplementationdate='06-30-2024',sqlprocesseddate='2024-06-30'):
        self.lastrow = lastrow
        self.sqlschoolstartdate = sqlschoolstartdate
        self.schoolyear = schoolyear
        self.datestamp = datestamp
        self.sqlimplementationdate = sqlimplementationdate
        self.sqlprocesseddate = sqlprocesseddate
        self.sqlschoolenddate = sqlschoolenddate
    def create_excel_report(self):
        # Database connection details
        server = 'ES00VPADOSQL180,51433'
        database = 'SEO_MART'
        username = 'your_username'
        password = 'your_password'

        # Connection string
        conn_str = (
            'DRIVER=SQL Server;'
            'SERVER=' + server + ';'
            'DATABASE=' + database + ';'
            # 'UID=' + username + ';'
            # 'PWD=' + password
        )
        params = urllib.parse.quote_plus(conn_str)

        # Create engine
        engine = create_engine(f'mssql+pyodbc:///?odbc_connect={params}')

        # SQL query
        query= """
            SELECT 
                E.[Student ID],
                E.[First Name],
                E.[Last Name],
                E.[Grade Level],
                E.[IEP Flag (ATS)],
                CASE 
                    WHEN I.Classification IS NULL AND I.OutcomeDocumentType = 'CSP' THEN 'CSP Awaiting'
                    ELSE I.Classification
                END AS Classification,
                I.OutcomeDocumentType,
                I.InactiveDate
            FROM [SEO_REPORTING].[dbo].[FinalELLDataetSY24] E
            LEFT JOIN [SEO_MART].[snap].[RPT_SESISActiveIEP_063024] I
                ON I.StudentID = E.[Student ID]
                --AND I.ProjectedIEPImplementationDate = '2024-06-30'
                AND (I.InactiveDate IS NULL OR ('"""+self.sqlschoolstartdate+"""' <= I.InactiveDate AND I.InactiveDate <= '"""+self.sqlschoolenddate+"""'))
        """

        # Establish a connection
        with engine.connect() as connection:
            # Load data into pandas DataFrame
            df = pd.read_sql(text(query), connection)

        # Convert the InactiveDate column to the desired format
        df['InactiveDate'] = pd.to_datetime(df['InactiveDate']).dt.strftime('%d/%m/%Y')

        # Specify the path for the output file
        output_file = self.schoolyear+'_Final_ELL_Dataset_SEO.xlsx'

        # Save the DataFrame to an Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='SEO Data as of '+self.datestamp)

        # Load the workbook and select the active sheet
        workbook = load_workbook(output_file)
        sheet = workbook['SEO Data as of '+self.datestamp]

        # Define the border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply the border to the specified range
        for row in sheet.iter_rows(min_row=2, max_row=self.lastrow, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border

        # Apply the border between E and F column medium border
        for row in range(1, self.lastrow+1):
            sheet['E' + str(row)].border = Border(right=Side(style='medium'))

        # Define the font style
        font_style = Font(name='Times New Roman', size=10)

        # Apply the font and border to the specified range
        for row in sheet.iter_rows(min_row=1, max_row=self.lastrow, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
                cell.font = font_style
                
        # Define the fill style for duplicates
        pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

        # Apply the border to the specified range and fill duplicates with pink color
        student_ids = {}
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.border = thin_border
                if cell.value in student_ids:
                    student_ids[cell.value].append(cell)
                else:
                    student_ids[cell.value] = [cell]

        # Highlight duplicates in pink
        for cells in student_ids.values():
            if len(cells) > 1:
                for cell in cells:
                    cell.fill = pink_fill

        # Set the width of columns A to H as 10, 20, 20, 10, 20, 20, 40, 20
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 40
        sheet.column_dimensions['H'].width = 20
        
        # Set A1 to H1 to bold and center and add filter to the first row
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            sheet[col + '1'].font = sheet[col + '1'].font.copy(bold=True)
            sheet[col + '1'].alignment = sheet[col + '1'].alignment.copy(horizontal='center')


        # Set A1 to Alastrow to center, D1 to Dlastrow to center
        for col in ['A', 'D', 'E']:
            for row in range(1, self.lastrow+1):
                sheet[col + str(row)].alignment = sheet[col + str(row)].alignment.copy(horizontal='center')

        # Set column 'GradeLevel' as if KG, 01, 02, 03, 04, 05, 06, 07, 08, 09, 10, 11, 12 as 0K, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12
        for row in range(2, self.lastrow+1):
            if sheet['D' + str(row)].value == 'KG':
                sheet['D' + str(row)].value = '0K'
            elif sheet['D' + str(row)].value == '01':
                sheet['D' + str(row)].value = '1'
            elif sheet['D' + str(row)].value == '02':
                sheet['D' + str(row)].value = '2'
            elif sheet['D' + str(row)].value == '03':
                sheet['D' + str(row)].value = '3'
            elif sheet['D' + str(row)].value == '04':
                sheet['D' + str(row)].value = '4'
            elif sheet['D' + str(row)].value == '05':
                sheet['D' + str(row)].value = '5'
            elif sheet['D' + str(row)].value == '06':
                sheet['D' + str(row)].value = '6'
            elif sheet['D' + str(row)].value == '07':
                sheet['D' + str(row)].value = '7'
            elif sheet['D' + str(row)].value == '08':
                sheet['D' + str(row)].value = '8'
            elif sheet['D' + str(row)].value == '09':
                sheet['D' + str(row)].value = '9'
            elif sheet['D' + str(row)].value == '10':
                sheet['D' + str(row)].value = '10'
            elif sheet['D' + str(row)].value == '11':
                sheet['D' + str(row)].value = '11'
            elif sheet['D' + str(row)].value == '12':
                sheet['D' + str(row)].value = '12'

        # Set column 'IEP Flag (ATS)' as if N as '-'
        for row in range(2, self.lastrow+1):
            if sheet['E' + str(row)].value == 'N':
                sheet['E' + str(row)].value = '-'

        # Set column 'Classification' as if blank cell as NULL
        for row in range(2, self.lastrow+1):
            if sheet['F' + str(row)].value == '' or sheet['F' + str(row)].value is None:
                sheet['F' + str(row)].value = 'NULL'

        # Set column 'OutcomeDocumentType' as if blank cell as NULL
        for row in range(2, self.lastrow+1):
            if sheet['G' + str(row)].value == '' or sheet['G' + str(row)].value is None:
                sheet['G' + str(row)].value = 'NULL'

        # Set column 'InactiveDate' as if blank cell as NULL
        for row in range(2, self.lastrow+1):
            if sheet['H' + str(row)].value == '' or sheet['H' + str(row)].value is None:
                sheet['H' + str(row)].value = 'NULL'

        # Set column 'InactiveDate' as if containes 2023 as 2023, 2024 as 2024, else as NULL
        for row in range(2, self.lastrow + 1):
            cell = sheet['H' + str(row)]
            if cell.value is None or cell.value == '':
                cell.value = 'NULL'
            # elif '2023' in cell.value:
            #     cell.value = '2023'
            # elif '2024' in cell.value:
            #     cell.value = '2024'
            # else:
            #     cell.value = 'NULL'

        # Add the filter for column A to H
        sheet.auto_filter.ref = 'A1:H'+str(self.lastrow + 1)

        # Save the changes to the workbook
        workbook.save(output_file)

        print(f"Excel report with borders has been created: {output_file}")

if __name__ == '__main__':
    report = DMLOfficeRport1()
    report.create_excel_report()
