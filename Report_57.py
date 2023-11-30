import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc
import os
class Solution:
    # Existing code...
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center')  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style



    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx')
        ws = wb.create_sheet("Reports 5-7 = Reevaluations")
        # wb = openpyxl.Workbook()
        # ws = wb.active
        # ws.title = "Reports 5-7 = Reevaluations"

        # Set fill color for cells from A1 to Zn to white
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
            for cell in row:
                cell.fill = white_fill

        black_border, black_border_thick, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_thick

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['Total Students with Reevaluation Referrals 7/1/2022 - 06/30/2023', 'Closed without IEP Meeting',
                'Student Declassified. IEP Meeting  <= 60 Calendar Days from Date of Referral', 'Student Declassified. IEP Meeting > 60 Calendar Days from Date of Referral', 'Total Declassified', 'Student Classified. IEP Meeting <= 60 Calendar Days from Date of Referral','Student Classified. IEP Meeting > 60 Calendar Days from Date of Referral','Total Eligible','Total IEP Meetings Held (Declassified + Eligible)','Open and Awaiting Parental Consent as of 06/30/2023','Total Open as of 06/30/2023']
        column_letters = ['C', 'D', 'E', 'F', 'G', 'H','I', 'J', 'K', 'L']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "B8CCE4"
        column_fill_color = "B8CCE4"
        self.format_header(ws, 'B4', 'District', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B40', 'Race/Ethnicity', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B49', 'Meal Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B55', 'Gender', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B62', 'ELL Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B68', 'Language of Instruction', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B76', 'Grade Level', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B94', 'Temporary Housing Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)
        self.format_header(ws, 'B101', 'Foster Care Status', columns, column_letters, 80, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_ReevalReferralsR510_SY23_Fix', 'INT_StudentDemographics_063023')
        cursor.execute("EXEC [dev].[USPCCAnnaulReport5to7] @tableNameCCReevalReferralsR510=?,  @tableNameINTStudentDemographics_0630=?", params)
        return cursor
    # Fetch data for "Report 8b = IEP Service Recs by Race"
    def fetch_data_by_race(self,cursor):
        query_byRace = '''
        select EthnicityGroupCC, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 from (  select * from  ( Select   EthnicityGroupCC_sort as sort , EthnicityGroupCC ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,
        sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,
        sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,
        sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final   group by EthnicityGroupCC, EthnicityGroupCC_sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''  # the byRace SQL query goes here
        cursor.execute(query_byRace)
        results_byRace = cursor.fetchall()
        return results_byRace

    # Fetch data for "Report 8b = IEP Service Recs by District"
    def fetch_data_by_district(self,cursor):
        query_byDistrict = '''
        select * from  ( Select    ReportingDistrict as sort  ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,
        sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,
        sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final   group by ReportingDistrict  ) a  union all  select * from ##TotalRow  order by sort  
        '''  # the byDistrict SQL query goes here
        cursor.execute(query_byDistrict)
        results_byDistrict = cursor.fetchall()
        return results_byDistrict

    def fetch_data_by_mealstatus(self,cursor):
        query_byMealStatus = '''
        select * from  ( Select    MealStatusGrouping as sort  ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,
        sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,
        sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,
        sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final   group by MealStatusGrouping  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byMealStatus SQL query goes here
        cursor.execute(query_byMealStatus)
        results_byMealStatus = cursor.fetchall()
        return results_byMealStatus
    
    def fetch_data_by_gender(self,cursor):
        query_byGender = '''
        select * from  ( Select    GENDER as sort  ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,
        sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,
        sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,
        sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final   group by GENDER  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byGender SQL query goes here
        cursor.execute(query_byGender)
        results_byGender = cursor.fetchall()
        return results_byGender
    
    def fetch_data_by_ellstatus(self,cursor):
        query_byELLStatus = '''
        select * from  ( Select    ELLStatus as sort  ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,
        sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,
        sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,
        sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final   group by ELLStatus  ) a  union all  select * from ##TotalRow  order by sort 
        '''  # the byELLStatus SQL query goes here
        cursor.execute(query_byELLStatus)
        results_byELLStatus = cursor.fetchall()
        return results_byELLStatus
    
    def fetch_data_by_language(self,cursor):
        query_byLanguage = '''
        select OutcomeLanguageCC, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 from (  select * from  ( Select   Language_Sort as sort , OutcomeLanguageCC ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,
        sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,
        sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,
        sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final   group by OutcomeLanguageCC, Language_Sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''  # the byLanguage SQL query goes here
        cursor.execute(query_byLanguage)
        results_byLanguage = cursor.fetchall()
        return results_byLanguage
    
    def fetch_data_by_gradelevel(self,cursor):
        query_byGradeLevel = '''
        select GradeLevel, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 from (  select * from  ( Select   GradeLevel_Sort as sort , GradeLevel ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,
        sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,
        sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,
        sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final   group by GradeLevel, GradeLevel_Sort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''
        cursor.execute(query_byGradeLevel)
        results_byGradeLevel = cursor.fetchall()
        return results_byGradeLevel
    
    def fetch_data_by_tempResFlag(self,cursor):
        query_byTempResFlag = '''
        select TempResFlag, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 from (  select * from  ( Select   TempResFlagSort as sort ,  case when TempResFlag = 'Y' then 'Yes' else 'No'  end as TempResFlag ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,
        sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,
        sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,
        sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final  where TempResFlag in ('Y', 'N')  group by TempResFlag, TempResFlagSort ) a  union all  select * from ##TotalRow_Sort ) a order by sort
        '''
        cursor.execute(query_byTempResFlag)
        results_byTempResFlag = cursor.fetchall()
        return results_byTempResFlag
    
    def fetch_data_by_fosterCareStatus(self,cursor):
        query_byFosterCareStatus = '''
        select FosterCareFlag, c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 from (  select * from  ( Select   FosterCareFlagSort as sort , FosterCareFlag ,sum(STUDENTS_WITH_REF) as c1  ,sum(CLOSED_WITHOUT_IEP) as c2   ,
        sum(Declass_LESS_60) as c3   ,sum(Declass_MORE_60) as c4  ,sum(COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c5  ,sum(CLASSIFIED_LESS_60) as c6  ,
        sum(CLASSIFIED_MORE_60) as c7  ,sum(CLASSIFIED_LESS_60 + CLASSIFIED_MORE_60) as  c8  ,
        sum(COALESCE(CLASSIFIED_LESS_60,0) + COALESCE(CLASSIFIED_MORE_60,0) + COALESCE(Declass_LESS_60,0) + COALESCE(Declass_MORE_60,0)) as c9   ,sum(TOTAL_OPEN) as c10  
        FROM ##Report_Final   group by FosterCareFlag, FosterCareFlagSort ) a  union all  select * from ##TotalRow_Sort ) a order by sort 
        '''
        cursor.execute(query_byFosterCareStatus)
        results_byFosterCareStatus = cursor.fetchall()
        return results_byFosterCareStatus
    
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_boarder_medium = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_boarder_medium, left=black_boarder_medium, right=black_boarder_medium, bottom=black_boarder_medium)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 2)  # +2 because data starts from column 'B'
                ws[col + str(row_num)].value = value
                ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply borders to all columns
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_no_bottom

        # Update alignment for range C6:N38
        for row in ws['C5':'L37']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C41':'L46']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C50':'L52']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C56':'L59']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C63':'L65']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C69':'L73']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C77':'L90']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C95':'L97']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        for row in ws['C102':'L104']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
        for row in ws['B1': 'L1']:
            for cell in row:
                cell.border = black_border
                cell.font = Font(bold=True, size=12)

        for row in ws['B37': 'L37'] + ws['B46': 'L46'] + ws['B52': 'L52'] + ws['B59': 'L59'] + ws['B65': 'L65'] + ws['B73': 'L73'] + ws['B90': 'L90'] + ws['B97': 'L97'] + ws['B104': 'L104']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['B3':'L3'] + ws['B39':'L39'] + ws['B48':'L48'] + ws['B54':'L54'] + ws['B61':'L61'] + ws['B67':'L67'] + ws['B75':'L75'] + ws['B93':'L93'] + ws['B100':'L100']:
            for cell in row:
                cell.border = black_border_thick
                cell.font = Font(bold=True, size=12)

        for row in ws['G4':'G37'] + ws['G40':'G46'] + ws['G49':'G52'] + ws['G55':'G59'] + ws['G62':'G65'] + ws['G68':'G74'] + ws['G76':'G90'] + ws['G94':'G97'] + ws['G101':'G104']:
            for cell in row:
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        for row in ws['J4':'J37'] + ws['J40':'J46'] + ws['J49':'J52'] + ws['J55':'J59'] + ws['J62':'J65'] + ws['J68':'J74'] + ws['J76':'J90'] + ws['J94':'J97'] + ws['J101':'J104']:
            for cell in row:
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        for row in ws['E4':'F4'] + ws['H4':'I4'] + ws['E40':'F40'] + ws['H40':'I40'] + ws['E49':'F49'] + ws['H49':'I49'] + ws['E55':'F55'] + ws['H55':'I55'] + ws['E62':'F62'] + ws['H62':'I62'] + ws['E68':'F68'] + ws['H68':'I68'] + ws['E76':'F76'] + ws['H76':'I76'] + ws['E94':'F94'] + ws['H94':'I94'] + ws['E101':'F101'] + ws['H101':'I101']:
            for cell in row:
                # cell.fill = PatternFill(start_color='#DCE6F1', end_color='#DCE6F1', fill_type='solid')
                cell.fill = PatternFill(start_color='E0F0F8', end_color='E0F0F8', fill_type='solid')
    def main_Reports_5_7_Reevaluations(self):
        title_cells = [
            {"cell": "B1", "value": "Reports 5-7 Reevaluation Referrals Disaggregated by: District; Race/Ethnicity; Meal Status; Gender; ELL Status; Recommended Language of  nstruction; and Grade Level.", "merge_cells": "B1:L1"},
            

        ]

        subtitle_cells = [
            {"cell": "B3", "value": "SY 2022-23 Students with IEP Recommended Services by District", "merge_cells": "B3:L3"},
            {"cell": "B39", "value": "SY 2022-23 Students with IEP Recommended Services by Ethnicity", "merge_cells": "B39:L39"},
            {"cell": "B48", "value": "SY 2022-23 Students with IEP Recommended Services by Meal Status", "merge_cells": "B48:L48"},
            {"cell": "B54", "value": "SY 2022-23 Students with IEP Recommended Services by Gender", "merge_cells": "B54:L54"},
            {"cell": "B61", "value": "SY 2022-23 Students with IEP Recommended Services by ELL Status", "merge_cells": "B61:L61"},
            {"cell": "B67", "value": "SY 2022-23 Students with IEP Recommended Services  by Recommended Language of Instruction", "merge_cells": "B67:L67"},
            {"cell": "B75", "value": "SY 2022-23 Students with IEP Recommended Services  by Grade Level", "merge_cells": "B75:L75"},
            {"cell": "B93", "value": "SY 2022-23 Students with IEP Recommended Services  by Temporary Housing", "merge_cells": "B93:L93"},
            {"cell": "B100", "value": "SY 2022-23 Students with IEP Recommended Services  by Foster Care Status", "merge_cells": "B100:L100"},
            

        ]

        column_widths = [5, 30, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # Step 3: Fetch and write data for "Report 8b = IEP Service Recs by Race"
        results_byRace = self.fetch_data_by_race(cursor)
        self.write_data_to_excel(ws, results_byRace, start_row=41)
        
        # Step 4: Fetch and write data for "Report 8b = IEP Service Recs by District"
        results_byDistrict = self.fetch_data_by_district(cursor)
        # replace 01 as 1, 02 as 2, etc.
        results_byDistrict = [(x[0].replace('01', '1'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('02', '2'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('03', '3'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('04', '4'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('05', '5'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('06', '6'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('07', '7'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('08', '8'), *x[1:]) for x in results_byDistrict]
        results_byDistrict = [(x[0].replace('09', '9'), *x[1:]) for x in results_byDistrict]
        self.write_data_to_excel(ws, results_byDistrict, start_row=5)

        # Step 5: Fetch and write data for "Report 8b = IEP Service Recs by Meal Status"
        results_byMealStatus = self.fetch_data_by_mealstatus(cursor)
        self.write_data_to_excel(ws, results_byMealStatus, start_row=50)

        # Step 6: Fetch and write data for "Report 8b = IEP Service Recs by Gender"
        results_byMealStatus = self.fetch_data_by_gender(cursor)
        self.write_data_to_excel(ws, results_byMealStatus, start_row=56)

        # Step 7: Fetch and write data for "Report 8b = IEP Service Recs by ELL Status"
        results_byELLStatus = self.fetch_data_by_ellstatus(cursor)
        # replace 'ELL' with 'Ell' and 'NOT ELL' with 'Non-Ell'
        results_byELLStatus = [('Ell' if x[0] == 'ELL' else ('Non-ELL' if x[0] == 'NOT ELL' else x[0]), *x[1:]) for x in results_byELLStatus]
        self.write_data_to_excel(ws, results_byELLStatus, start_row=63)
        
        # Step 8: Fetch and write data for "Report 8b = IEP Service Recs by Language"
        results_byLanguage = self.fetch_data_by_language(cursor)
        # replace 'English' with 'English', 'Spanish' with 'Spanish', 'Chinese' with 'Chinese', 'Other' with 'Other'
        results_byLanguage = [('English' if x[0] == 'ENGLISH' else ('Spanish' if x[0] == 'SPANISH' else ('Chinese' if x[0] == 'CHINESE' else ('Other' if x[0] == 'OTHER' else x[0]))), *x[1:]) for x in results_byLanguage]
        self.write_data_to_excel(ws, results_byLanguage, start_row=69)

        # Step 9: Fetch and write data for "Report 8b = IEP Service Recs by Grade Level"
        results_byGradeLevel = self.fetch_data_by_gradelevel(cursor)
        # replace 01 as 1, 02 as 2, etc.
        results_byGradeLevel = [(x[0].replace('0K', 'KG'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('01', '1'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('02', '2'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('03', '3'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('04', '4'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('05', '5'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('06', '6'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('07', '7'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('08', '8'), *x[1:]) for x in results_byGradeLevel]
        results_byGradeLevel = [(x[0].replace('09', '9'), *x[1:]) for x in results_byGradeLevel]    
        self.write_data_to_excel(ws, results_byGradeLevel, start_row=77)

        # Step 10: Fetch and write data for "Report 8b = IEP Service Recs by Temporary Housing"
        results_byTempResFlag = self.fetch_data_by_tempResFlag(cursor)
        self.write_data_to_excel(ws, results_byTempResFlag, start_row=95)

        # Step 11: Fetch and write data for "Report 8b = IEP Service Recs by Foster Care Status"
        results_byFosterCareStatus = self.fetch_data_by_fosterCareStatus(cursor)
        # replace 'YES' with 'Yes' and 'NO' with 'No'
        results_byFosterCareStatus = [('Yes' if x[0] == 'Y' else ('No' if x[0] == 'N' else x[0]), *x[1:]) for x in results_byFosterCareStatus]
        self.write_data_to_excel(ws, results_byFosterCareStatus, start_row=102)
        # Step 9: Save the combined report
        save_path = r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted Annual Special Education Data Report.xlsx'
        wb.save(save_path)

    def save_or_append_to_workbook(self, wb, save_path):
        # Check if the file already exists
        if os.path.exists(save_path):
            # Load the existing workbook
            book = openpyxl.load_workbook(save_path)
            # Get the active worksheet to copy its styles for the new worksheet
            active_sheet = book.active
            # Create a new worksheet by copying the active one
            new_sheet = book.copy_worksheet(active_sheet)
            # Set the title for the new worksheet
            new_sheet.title = wb.active.title
            # Now, copy the data from wb to new_sheet
            for row in wb.active.iter_rows():
                for cell in row:
                    new_sheet[cell.coordinate].value = cell.value
        else:
            # If the file does not exist, save the new workbook as it is
            book = wb
        # Save the workbook
        book.save(save_path)

if __name__ == "__main__":
        Tab5_7 = Solution()
        Tab5_7.main_Reports_5_7_Reevaluations() 








