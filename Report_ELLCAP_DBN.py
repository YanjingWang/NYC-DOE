import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, colors
from openpyxl.utils import get_column_letter
import pyodbc, urllib
from sqlalchemy import create_engine, text
from datetime import datetime
class Solution:
    def __init__(self):
        # Database connection details
        server = 'ES00VPADOSQL180,51433'
        database = 'SEO_MART'
        username = 'your_username'
        password = 'your_password'

        # Connection string
        conn_str = (
            'DRIVER=SQL Server;'
            'SERVER=' + server + ';'
            'DATABASE=' + database + ';'
            # 'UID=' + username + ';'
            # 'PWD=' + password
        )
        params = urllib.parse.quote_plus(conn_str)

        # Create engine
        self.engine = create_engine(f'mssql+pyodbc:///?odbc_connect={params}')
        self.datestamp = datetime.strptime(self.get_ProcessedDate(), '%m-%d-%Y').strftime('%m/%d/%Y') # '10/28/2024'
        self.lastrow = 1551 # 1543 #1541
        self.ProcessedDate = self.get_ProcessedDate() #'10-28-2024'
        self.schoolyear = self.get_schoolyear() #'SY 24-25'
    # Function to format headers
    def get_column_index_from_string(self, column_letter):
        return openpyxl.utils.column_index_from_string(column_letter)
    def format_header(self,ws, header_start_cell, header_title, columns, column_letters, row_height, header_fill_color, column_fill_color, border_style, font_style):
        # Set title, font, border, alignment, fill, row dimensions, and merge cells for the main header
        ws[header_start_cell] = header_title
        ws[header_start_cell].font = font_style
        ws[header_start_cell].border = border_style
        ws[header_start_cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  
        ws[header_start_cell].fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        ws.row_dimensions[int(header_start_cell[1:])].height = row_height
        
        # Apply formatting to the sub headers
        for col, title in zip(column_letters, columns):
            cell_number = str(int(header_start_cell[1:])) #don't +3 here
            ws[col + cell_number] = title
            ws[col + cell_number].font = font_style
            ws[col + cell_number].border = border_style
            ws[col + cell_number].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws[col + cell_number].fill = PatternFill(start_color=column_fill_color, end_color=column_fill_color, fill_type="solid")
            print('Subheader'+ f'{col}{cell_number}')

        # Apply borders to all the cells in the header
        for col in [header_start_cell[0]] + column_letters + [chr(ord(c)) for c in column_letters]:
            cell_number = str(int(header_start_cell[1:]))
            ws[col + cell_number].border = border_style
            # ws[col + str(int(cell_number)-1)].border = border_style
            print('Header'+ f'{col}{cell_number}')

    # get self.ProcessedDate from SQL Query
    def get_ProcessedDate(self):
        query = '''
        SELECT TOP 1 ProcessedDate FROM [SEO_MART].[arch].[RPT_PSProvisioningStudent] WHERE ProcessedDate = (SELECT MAX (ProcessedDate) FROM [SEO_MART].[arch].[RPT_PSProvisioningStudent])
        '''
        with self.engine.connect() as connection:
            result = connection.execute(text(query)).fetchone()._mapping
            # Store the formatted date in a new variable
            processed_date = datetime.strptime(result['ProcessedDate'], '%Y-%m-%d').strftime('%m-%d-%Y')
            print(f"Processed Date: {processed_date}")
            return processed_date
        
    def get_schoolyear(self):
        query ='''
        SELECT TOP 1 SchoolYear FROM [SEO_MART].[arch].[RPT_PSProvisioningStudent] WHERE ProcessedDate = (SELECT MAX (ProcessedDate) FROM [SEO_MART].[arch].[RPT_PSProvisioningStudent])
        '''
        with self.engine.connect() as connection:
            result = connection.execute(text(query)).fetchone()._mapping
            print(f"School Year: {result['SchoolYear']}")
            # Format the school year from '2024-2025' to 'SY 24-25'
            formatted_schoolyear = 'SY ' + result['SchoolYear'][2:4] + '-' + result['SchoolYear'][7:9]
            return formatted_schoolyear

    # Create Excel Report Template
    def create_excel_report_template(self, title_cells, subtitle_cells, column_widths):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DBN"
        # wb = openpyxl.load_workbook(r'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\Non-Redacted City Council Triennial Report_CW.xlsx')
        # ws = wb.create_sheet("Program Delivery")

        # # Set fill color for cells from A1 to Zn to white
        # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # for row in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=26):
        #     for cell in row:
        #         cell.fill = white_fill

        black_border, _,black_border_medium, _, _ = self.create_border_styles()

        # Add report title and merge cells
        for cell_info in title_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border

        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.value = cell_info["value"]
            ws.merge_cells(cell_info["merge_cells"])  # Merge cells outside the loop
            cell.border = black_border_medium #F4:G4 and L4:Q4

        # Style and align the merged title and subtitle
        for cell_info in title_cells + subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(wrap_text=True)

        # Adjust column widths
        for col, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = width

        # Call the header formatting function for each header section
        columns = ['School District', 
                   'School Name',
                'School Level', 
                '# of ELLs with Program Recommendations on IEPs',
                'Total',
                'ICT D1-32,79',
                'SC D1-32,79',
                'SETSS D1-32,79',
                'Multiple Programs D1-32,79',
                'D75',
                'Total',
                'ICT D1-32,79',
                'SC D1-32,79',
                'SETSS D1-32,79',
                'Multiple Programs D1-32,79',
                'D75']
        column_letters = ['B','C', 'D', 'E', 'F', 'G', 'H','I', 'J', 'K', 'L', 'M','N','O','P','Q']
        # You need to pass the correct parameters to the format_header function
        # For example, for the 'District' header starting at row 4
        # ... You would repeat the above line for each section (Ethnicity, Meal Status, Gender) with the appropriate start_row
        # Define the styles outside of the function calls to avoid recreation every time
        header_font = Font(bold=True, size=12)
        border_bottom_thin = Border(bottom=Side(style='thin'))
        black_border_thinside = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_thinside, left=black_border_thinside, right=black_border_thinside, bottom=black_border_thinside)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        black_border_no_bottom = Border(left=black_border_mediumside, right=black_border_mediumside)
        black_boarder_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        header_fill_color = "F2F2F2"
        column_fill_color = "F2F2F2"
        self.format_header(ws, 'A5', "DBN", columns, column_letters, 60, header_fill_color, column_fill_color,  black_boarder_all_medium, header_font)

        
        # Deleting the default created sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        return wb, ws



    def create_border_styles(self):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')

        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)

        return black_border, black_border_thick, black_border_medium, black_border_no_bottom, black_boarder_all

    # Step 2: Connect to the database
    def connect_to_database(self):
        conn_str = 'DRIVER=SQL SERVER;SERVER=ES00VPADOSQL180,51433;DATABASE=SEO_REPORTING' #;UID=your_username;PWD=your_password
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        params = ('CC_InitialReferralsR19_SY23')
        cursor.execute(
        f'''     
        If Object_ID('tempdb..#BSEReg') is not null
        Drop Table #BSEReg 

        SELECT [StudentID]
            ,[OfficialClass]
            ,[EnrolledDBN]
            ,[AdminDistrict]
            ,[LEPFlag]
            ,[Classification]
            ,[LocationCategoryDescription]
            ,[GradeLevel]
            ,[ServiceLocation]
            ,[PSOutcomeLanguage]
            ,[PSOutcomeLanguageCC]
            ,[BilingualProgramRecommendation] AS [BilingualProgramRec]
            ,[IsBilingualSC] AS BilingualSC
            ,[IsBilingualICT] AS BilingualICT
            ,[SpecializedProgram]
            ,[ReceivingBilingualPrograms] AS ReceivingBilingual
            ,[ProcessedDate]
            ,[ProcessedDateTime]
            ,[SchoolYear]
            ,[IsBilingualSETSS] AS BilingualSETSS

            INTO #BSEReg
        From  [SEO_MART].[arch].[RPT_ELLCAPBilingualPS]
		Where ProcessedDate = '{self.ProcessedDate}'



        ----- Create ELL CAP report 




        If Object_ID('tempdb..#Register') is not null
        Drop Table #Register 

        SELECT CAP.[StudentID]
            ,CAP.[OfficialClass]
            ,CAP.[EnrolledDBN]
            ,CAP.[AdminDistrict]
            ,CAP.[LEPFlag]
            ,CAP.[Classification]
            ,CAP.[LocationCategoryDescription]
            ,CAP.[GradeLevel]
            ,CAP.[ServiceLocation]
            ,CAP.[PSOutcomeLanguage]
            ,CAP.[PSOutcomeLanguageCC]
            ,case when CAP.GradeLevel = '0k' then 1
            when CAP.GradeLevel = '01' then 2
            when CAP.GradeLevel = '02' then 3
            when CAP.GradeLevel = '03' then 4
            when CAP.GradeLevel = '04' then 5
            when CAP.GradeLevel = '05' then 6
            when CAP.GradeLevel = '06' then 7
            when CAP.GradeLevel = '07' then 8
            when CAP.GradeLevel = '08' then 9
            when CAP.GradeLevel = '09' then 10
            when CAP.GradeLevel = '10' then 11
            when CAP.GradeLevel = '11' then 12
            when CAP.GradeLevel = '12' then 13
            end as GradeSort
            ,[BilingualProgramRec]
            ,[BilingualSC]
            ,[BilingualICT]
            ,[BilingualSETSS]
            ,CAP.[SpecializedProgram]
            ,ReceivingBilingual
            ,CAP.[ProcessedDate]
            ,CAP.[ProcessedDateTime]
            ,case when loc.boroughcode = 'O' then 'Q' else loc.boroughcode end as boroughcode
        into #Register
        FROM [SEO_MART].[arch].[RPT_PSProvisioningStudent] as CAP
        left join #BSEReg as NEWCAP on CAP.StudentID = NEWCAP.StudentID
        left join [SEO_MART].[dbo].[RPT_Locations] as loc on cap.enrolleddbn = loc.schooldbn 
        where CAP.ELLStatus = 'ELL' 
		and CAP.ProcessedDate = '{self.ProcessedDate}'
        '''
        )
        return cursor
    
    def fetch_data_by_SchoolDBN(self,cursor):
        query_bySchoolDBN = '''
        ---SchoolDBN

        Select a.enrolledDbn
        ,[AdminDistrict]
        ,[ServiceLocation]
        ,LocationCategoryDescription as [School Level]
        ,count(studentid) as 'ELLs'
        ,TotalBIL
        ,BilICT
        ,BilSC
        ,BilSETSS
        ,BilMulti
        ,D75
        ,TotalReceiving
        ,ReceivingBilICT
        ,ReceivingBilSC
        ,ReceivingBilSETSS
        ,ReceivingBilMulti
        ,ReceivingD75
        FROM #Register as a
        left join( select enrolleddbn
                    ,count(studentid) as 'TotalBIL'
                    FROM #Register
                    where [BilingualProgramRec] is not null 
                    group by EnrolledDBN) as z on a.EnrolledDBN = z.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'BilICT'
                    FROM #Register
                    where [BilingualProgramRec] = 'Integrated Co-Teaching Services'
                    and AdminDistrict <> 75
                    group by EnrolledDBN) as b on a.EnrolledDBN = b.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'BilSC'
                    FROM #Register
                    where [BilingualProgramRec] = 'Special Class'
                    and AdminDistrict <> 75
                    group by EnrolledDBN) as c on a.EnrolledDBN = c.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'BilMulti'
                    FROM #Register 
                    where [BilingualProgramRec]= 'Multiple Programs'
                    and AdminDistrict <> 75
                    group by EnrolledDBN) as d on a.EnrolledDBN = d.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'D75'
                    FROM #Register 
                    where [BilingualProgramRec] is not null
                    and AdminDistrict = 75
                    group by EnrolledDBN) as e on a.EnrolledDBN = e.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'TotalReceiving'
                    FROM #Register
                    where [BilingualProgramRec] is not null 
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by EnrolledDBN) as Y on a.EnrolledDBN = Y.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'ReceivingBilICT'
                    FROM #Register
                    where [BilingualProgramRec] = 'Integrated Co-Teaching Services'
                    and AdminDistrict <> 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by EnrolledDBN) as f on a.EnrolledDBN = f.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'ReceivingBilSC'
                    FROM #Register
                    where [BilingualProgramRec] = 'Special Class'
                    and AdminDistrict <> 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by EnrolledDBN) as g on a.EnrolledDBN = g.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'ReceivingBilMulti'
                    FROM #Register 
                    where [BilingualProgramRec] = 'Multiple Programs'
                    and AdminDistrict <> 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by EnrolledDBN) as h on a.EnrolledDBN = h.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'ReceivingD75'
                    FROM #Register 
                    where [BilingualProgramRec] is not null
                    and AdminDistrict = 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by EnrolledDBN) as i on a.EnrolledDBN = i.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'BilSETSS'
                    FROM #Register
                    where [BilingualProgramRec] = 'SETSS'
                    and AdminDistrict <> 75
                    group by EnrolledDBN) as j on a.EnrolledDBN = j.EnrolledDBN
        left join( select enrolleddbn
                    ,count(studentid) as 'ReceivingBilSETSS'
                    FROM #Register
                    where [BilingualProgramRec] = 'SETSS'
                    and AdminDistrict <> 75
                    and ReceivingBilingual = 'FULLY RECEIVING'
                    group by EnrolledDBN) as k on a.EnrolledDBN = k.EnrolledDBN
        group by a.enrolledDbn
        ,[AdminDistrict]
        ,[ServiceLocation]
        ,LocationCategoryDescription
        ,TotalBIL
        ,BilICT
        ,BilSC
        ,BilSETSS
        ,BilMulti
        ,D75
        ,TotalReceiving
        ,ReceivingBilICT
        ,ReceivingBilSC
        ,ReceivingBilSETSS
        ,ReceivingBilMulti
        ,ReceivingD75
        '''
        cursor.execute(query_bySchoolDBN)
        results_bySchoolDBN= cursor.fetchall()
        return results_bySchoolDBN  
    # Step 3: Write data to Excel for "Report 8b = IEP Service Recs by Race"
    def write_data_to_excel(self, ws, data, start_row):
        black_border_side = Side(style='thin', color='000000')
        black_border_thickside = Side(style='thick', color='000000')
        black_border_mediumside = Side(style='medium', color='000000')
        black_border = Border(top=black_border_side, left=black_border_side, right=black_border_side, bottom=black_border_side)
        black_border_thick = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_boarder_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        black_border_no_bottom = Border(left=black_border_thickside, right=black_border_thickside)
        black_border_right_side = Border(right=black_border_mediumside)
        black_boarder_all = Border(top=black_border_thickside, left=black_border_thickside, right=black_border_thickside, bottom=black_border_thickside)
        black_border_all_medium = Border(top=black_border_mediumside, left=black_border_mediumside, right=black_border_mediumside, bottom=black_border_mediumside)
        black_border_up_down = Border(top=black_border_mediumside, bottom=black_border_mediumside)
        # Write data to Excel starting from row B5
        for row_num, row_data in enumerate(data, start=start_row):  # Adjusted start_row here
            for i, value in enumerate(row_data):
                col = get_column_letter(i + 1)  # +2 because data starts from column 'A'
                ws[col + str(row_num)].value = value
                # ws[col + str(row_num)].border = black_border
                ws[col + str(row_num)].alignment = Alignment(horizontal='left')  # Right align the data
        
        # Apply black_border_right_side borders to E, K, Q columns
        for col in ['E', 'K', 'Q']:
            for row_num in range(start_row, start_row + len(data)):
                ws[col + str(row_num)].border = black_border_right_side


        # Update alignment for range C6:N38
        for row in ws['A5':'Q5']:
            for cell in row:
                if cell.value is not None:  # Ensure there is a value in the cell
                    cell.value = str(cell.value) + ''  # Prepend space to the value
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                # add filter to the header from A5 to Q5
                ws.auto_filter.ref = "A5:Q5"

        for row in ws['F6':'Q'+str(self.lastrow)]:
            for cell in row:
                if cell.value is not None: 
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        for row in ws['B6':'B'+str(self.lastrow)]:
            for cell in row:
                if cell.value is not None:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        for row in ws['F6':'Q'+str(self.lastrow-1)]:
            for cell in row:
                if cell.value is None:
                    cell.value = '-'  # Replace None with '-'
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        for row in ws['C6':'D'+str(self.lastrow)]:
            for cell in row:
                if cell.value is not None:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
        for row in ws['E6':'E'+str(self.lastrow)]:
            for cell in row:
                if cell.value is not None:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        for row in ws['A6':'A'+str(self.lastrow)]:
            for cell in row:
                cell.font = Font(bold=True, size=12)
        for row in ws['A1': 'Q1']:
            for cell in row:
                cell.border = black_boarder_all
                cell.font = Font(bold=True, size=12)

        for row in ws['A'+ str(self.lastrow):'Q'+ str(self.lastrow)]:
            # make font bold
            for cell in row:
                cell.font = Font(bold=True, size=12)

        for row in ws['F4':'K4'] + ws['L4':'Q4']:
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',wrap_text=True)

        fill_color = "F2F2F2"  # Color for the Total columns and row
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Fill 'Total' columns with the specified color
        for col in ['F', 'L']:  # Columns F and L are the 'Total' columns
            for row_num in range(start_row, ws.max_row + 1):
                ws[col + str(row_num)].fill = fill
        
        # Fill 'Total' row with the specified color
        for row in ws.iter_rows(min_col=1, max_col=ws.max_column, min_row=self.lastrow, max_row=self.lastrow):
            for cell in row:
                cell.fill = fill
        # add border to the last row
        for cell in ws['A'+str(self.lastrow):'E'+str(self.lastrow)][0]:  # Only one row, so it's safe to use [0]
            cell.border = black_border_up_down
        for cell in ws['F'+str(self.lastrow):'Q'+str(self.lastrow)][0]:  # Only one row, so it's safe to use [0]
            cell.border = black_border_all_medium
                
        cell_ranges = ['A6:Q'+str(self.lastrow)]
        for cell_range in cell_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        try:
                            cell.number_format = '#,##0'  # Apply comma format
                        except ValueError:
                            # If the value cannot be converted to int, keep the original value
                            print("Int converting Error")
                            pass

        # change the row height
        ws.row_dimensions[4].height = 40  
        # insert datestamp to the second row then merge cell A2:Q2 then bold the font
        ws['A2'] = 'As of ' + self.datestamp 
        ws.merge_cells('A2:Q2') 
        ws['A2'].font = Font(bold=True, size=12) 
                              
    def Report_DBN(self):
        title_cells = [
            {"cell": "A1", "value": "ELLs with IEPs and Bilingual ICT or SC IEP Program Recommendations by DBN for "+self.schoolyear, "merge_cells": "A1:Q1"},
        ]

        subtitle_cells = [
            {"cell": "F4", "value": "# of ELLs with IEPs with Bilingual Program Recommendations", "merge_cells": "F4:K4"},
            {"cell": "L4", "value": "# of ELLs with IEPs with BSE Recommendation Served in a Bilingual Class with a Bilingual Teacher", "merge_cells": "L4:Q4"},            

        ]

        column_widths = [10, 15, 65, 30, 30, 10, 15, 15, 15, 15, 10, 10, 15, 15, 15, 20, 10]
        # Step 1: Create Excel Report Template
        wb, ws = self.create_excel_report_template(title_cells, subtitle_cells, column_widths)
        # fill subtitle_cells with color D0CECE
        for cell_info in subtitle_cells:
            cell = ws[cell_info["cell"]]
            cell.fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
        
        # Step 2: Connect to the database
        cursor = self.connect_to_database()
        
        # # Step 3: Fetch and write data for "Report DBN"
        results_bytab1 = self.fetch_data_by_SchoolDBN(cursor)
        # add Total row to calculate the total number of ELLs at the end of the list
        total = ['Total']
        for i in range(1, 17):
            # if data = Null, set it to 0; if data is string, set it to NULL
            if results_bytab1 and all(isinstance(row[i], int) for row in results_bytab1):
                total.append(sum([row[i] for row in results_bytab1 if row[i] is not None]))
            elif results_bytab1 and all(isinstance(row[i], str) for row in results_bytab1):
                total.append('')
            else:
                # if the data is mixed with data and nonetype,calculate the sum of the data and set the nonetype to 0
                total.append(sum([row[i] if row[i] is not None else 0 for row in results_bytab1]))
        results_bytab1.append(total)

        self.write_data_to_excel(ws, results_bytab1, start_row=6)
        

        # Step 9: Save the combined report
        save_path = rf'C:\Users\Ywang36\OneDrive - NYCDOE\Desktop\CityCouncil\ELLCAP_Bilingual_Report_{self.ProcessedDate}.xlsx'
        wb.save(save_path)
        cursor.close()

if __name__ == "__main__":
        Tab1 = Solution()
        Tab1.Report_DBN()                                                                  